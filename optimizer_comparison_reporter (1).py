# -*- coding: utf-8 -*-
"""
PFAZ 2: Optimizer Comparison Reporter
======================================

Kapsamlı optimizer karşılaştırma ve analiz sistemi

Features:
- Adam, SGD, RMSprop, AdamW comparison
- Learning rate effect analysis
- Batch size impact study
- Convergence speed analysis
- Excel comprehensive reports

Author: Nuclear Physics AI Training Pipeline
Version: 1.0.0
Date: 2025-10-15
"""

import numpy as np
import pandas as pd
from pathlib import Path
import json
import logging
from typing import Dict, List, Tuple, Optional
from datetime import datetime
from collections import defaultdict
import warnings
warnings.filterwarnings('ignore')

try:
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import seaborn as sns
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False
    logging.warning("Matplotlib/Seaborn not available")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logging.warning("openpyxl not available")

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


# ============================================================================
# OPTIMIZER ANALYZER
# ============================================================================

class OptimizerAnalyzer:
    """Analyze optimizer performance from training results"""
    
    def __init__(self):
        self.optimizer_stats = defaultdict(lambda: {
            'count': 0,
            'r2_scores': [],
            'rmse_scores': [],
            'mae_scores': [],
            'training_times': [],
            'convergence_epochs': []
        })
        
        self.learning_rate_stats = defaultdict(lambda: {
            'count': 0,
            'r2_scores': [],
            'rmse_scores': []
        })
        
        self.batch_size_stats = defaultdict(lambda: {
            'count': 0,
            'r2_scores': [],
            'rmse_scores': []
        })
    
    def extract_optimizer(self, config: Dict) -> str:
        """Extract optimizer name from config"""
        return config.get('optimizer', 'unknown')
    
    def extract_learning_rate(self, config: Dict) -> float:
        """Extract learning rate from config"""
        return config.get('learning_rate', 0.0)
    
    def extract_batch_size(self, config: Dict) -> int:
        """Extract batch size from config"""
        return config.get('batch_size', 0)
    
    def add_result(self, config: Dict, metrics: Dict):
        """Add training result for analysis"""
        
        optimizer = self.extract_optimizer(config)
        learning_rate = self.extract_learning_rate(config)
        batch_size = self.extract_batch_size(config)
        
        # Extract metrics
        val_metrics = metrics.get('val', {})
        r2 = val_metrics.get('r2', val_metrics.get('r2_avg', 0))
        rmse = val_metrics.get('rmse', val_metrics.get('rmse_avg', 0))
        mae = val_metrics.get('mae', val_metrics.get('mae_avg', 0))
        training_time = metrics.get('training_time', 0)
        
        # Convergence epochs
        convergence_epochs = len(metrics.get('history', {}).get('val_loss', []))
        if convergence_epochs == 0:
            convergence_epochs = metrics.get('epochs_trained', 0)
        
        # Update optimizer stats
        self.optimizer_stats[optimizer]['count'] += 1
        self.optimizer_stats[optimizer]['r2_scores'].append(r2)
        self.optimizer_stats[optimizer]['rmse_scores'].append(rmse)
        self.optimizer_stats[optimizer]['mae_scores'].append(mae)
        self.optimizer_stats[optimizer]['training_times'].append(training_time)
        self.optimizer_stats[optimizer]['convergence_epochs'].append(convergence_epochs)
        
        # Update learning rate stats
        lr_key = f"{learning_rate:.6f}"
        self.learning_rate_stats[lr_key]['count'] += 1
        self.learning_rate_stats[lr_key]['r2_scores'].append(r2)
        self.learning_rate_stats[lr_key]['rmse_scores'].append(rmse)
        
        # Update batch size stats
        bs_key = str(batch_size)
        self.batch_size_stats[bs_key]['count'] += 1
        self.batch_size_stats[bs_key]['r2_scores'].append(r2)
        self.batch_size_stats[bs_key]['rmse_scores'].append(rmse)
    
    def get_optimizer_summary(self) -> pd.DataFrame:
        """Get summary statistics per optimizer"""
        
        data = []
        
        for optimizer, stats in self.optimizer_stats.items():
            if stats['count'] == 0:
                continue
            
            row = {
                'Optimizer': optimizer,
                'Count': stats['count'],
                'Avg_R2': np.mean(stats['r2_scores']),
                'Std_R2': np.std(stats['r2_scores']),
                'Best_R2': np.max(stats['r2_scores']),
                'Worst_R2': np.min(stats['r2_scores']),
                'Avg_RMSE': np.mean(stats['rmse_scores']),
                'Std_RMSE': np.std(stats['rmse_scores']),
                'Avg_MAE': np.mean(stats['mae_scores']),
                'Avg_Training_Time': np.mean(stats['training_times']),
                'Avg_Convergence_Epochs': np.mean(stats['convergence_epochs'])
            }
            
            data.append(row)
        
        df = pd.DataFrame(data)
        
        if not df.empty:
            df = df.sort_values('Avg_R2', ascending=False)
        
        return df
    
    def get_learning_rate_summary(self) -> pd.DataFrame:
        """Get summary statistics per learning rate"""
        
        data = []
        
        for lr_key, stats in self.learning_rate_stats.items():
            if stats['count'] == 0:
                continue
            
            row = {
                'Learning_Rate': float(lr_key),
                'Count': stats['count'],
                'Avg_R2': np.mean(stats['r2_scores']),
                'Std_R2': np.std(stats['r2_scores']),
                'Best_R2': np.max(stats['r2_scores']),
                'Avg_RMSE': np.mean(stats['rmse_scores']),
                'Std_RMSE': np.std(stats['rmse_scores'])
            }
            
            data.append(row)
        
        df = pd.DataFrame(data)
        
        if not df.empty:
            df = df.sort_values('Learning_Rate')
        
        return df
    
    def get_batch_size_summary(self) -> pd.DataFrame:
        """Get summary statistics per batch size"""
        
        data = []
        
        for bs_key, stats in self.batch_size_stats.items():
            if stats['count'] == 0:
                continue
            
            try:
                batch_size = int(bs_key)
            except:
                continue
            
            row = {
                'Batch_Size': batch_size,
                'Count': stats['count'],
                'Avg_R2': np.mean(stats['r2_scores']),
                'Std_R2': np.std(stats['r2_scores']),
                'Best_R2': np.max(stats['r2_scores']),
                'Avg_RMSE': np.mean(stats['rmse_scores']),
                'Std_RMSE': np.std(stats['rmse_scores'])
            }
            
            data.append(row)
        
        df = pd.DataFrame(data)
        
        if not df.empty:
            df = df.sort_values('Batch_Size')
        
        return df


# ============================================================================
# OPTIMIZER COMPARISON REPORTER
# ============================================================================

class OptimizerComparisonReporter:
    """
    Main optimizer comparison and reporting system
    
    Features:
    - Load training results
    - Analyze optimizer performance
    - Compare learning rates and batch sizes
    - Generate comprehensive Excel reports
    - Create visualization charts
    """
    
    def __init__(self, output_dir: str = 'optimizer_comparison'):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        self.analyzer = OptimizerAnalyzer()
        self.results = []
        
        logger.info("=" * 80)
        logger.info("OPTIMIZER COMPARISON REPORTER INITIALIZED")
        logger.info("=" * 80)
        logger.info(f"Output directory: {self.output_dir}")
        logger.info("=" * 80)
    
    def load_training_results(self, 
                             trained_models_dir: Path,
                             configs_file: Path) -> List[Dict]:
        """
        Load training results
        
        Args:
            trained_models_dir: Directory with trained models
            configs_file: Path to training_configs_50.json
        
        Returns:
            List of training results
        """
        logger.info("Loading training results...")
        
        # Load configs
        with open(configs_file, 'r') as f:
            configs = json.load(f)
        
        # Create config lookup
        config_lookup = {c['id']: c for c in configs}
        
        # Find all metrics files
        metrics_files = list(trained_models_dir.glob('**/metrics_*.json'))
        
        logger.info(f"Found {len(metrics_files)} training result files")
        
        results = []
        
        for metrics_file in metrics_files:
            try:
                with open(metrics_file, 'r') as f:
                    metrics = json.load(f)
                
                # Extract identifiers
                config_id = metrics_file.stem.replace('metrics_', '')
                
                if config_id not in config_lookup:
                    logger.warning(f"Config not found: {config_id}")
                    continue
                
                config = config_lookup[config_id]
                
                result = {
                    'config_id': config_id,
                    'config': config,
                    'metrics': metrics,
                    'file_path': metrics_file
                }
                
                results.append(result)
                
                # Add to analyzer
                self.analyzer.add_result(config, metrics)
            
            except Exception as e:
                logger.error(f"Failed to load {metrics_file}: {e}")
        
        logger.info(f"Successfully loaded {len(results)} training results")
        
        self.results = results
        return results
    
    def generate_comparison_analysis(self) -> Dict[str, pd.DataFrame]:
        """Generate all comparison analyses"""
        
        logger.info("\n" + "=" * 80)
        logger.info("GENERATING COMPARISON ANALYSES")
        logger.info("=" * 80)
        
        analyses = {}
        
        # Optimizer comparison
        logger.info("Analyzing optimizer performance...")
        analyses['optimizer_summary'] = self.analyzer.get_optimizer_summary()
        
        # Learning rate analysis
        logger.info("Analyzing learning rate effects...")
        analyses['learning_rate_summary'] = self.analyzer.get_learning_rate_summary()
        
        # Batch size analysis
        logger.info("Analyzing batch size effects...")
        analyses['batch_size_summary'] = self.analyzer.get_batch_size_summary()
        
        # Best configurations
        logger.info("Identifying best configurations...")
        analyses['best_configs'] = self._get_best_configs()
        
        logger.info("✅ Analysis completed\n")
        
        return analyses
    
    def _get_best_configs(self, top_n: int = 10) -> pd.DataFrame:
        """Get top N best configurations"""
        
        data = []
        
        for result in self.results:
            config = result['config']
            metrics = result['metrics']
            
            val_metrics = metrics.get('val', {})
            r2 = val_metrics.get('r2', val_metrics.get('r2_avg', 0))
            rmse = val_metrics.get('rmse', val_metrics.get('rmse_avg', 0))
            mae = val_metrics.get('mae', val_metrics.get('mae_avg', 0))
            
            row = {
                'Config_ID': config['id'],
                'Optimizer': config.get('optimizer', 'unknown'),
                'Learning_Rate': config.get('learning_rate', 0),
                'Batch_Size': config.get('batch_size', 0),
                'Architecture': str(config.get('architecture', [])),
                'R2': r2,
                'RMSE': rmse,
                'MAE': mae,
                'Training_Time': metrics.get('training_time', 0)
            }
            
            data.append(row)
        
        df = pd.DataFrame(data)
        
        if not df.empty:
            df = df.sort_values('R2', ascending=False).head(top_n)
        
        return df
    
    def generate_charts(self):
        """Generate visualization charts"""
        
        if not MATPLOTLIB_AVAILABLE:
            logger.warning("Matplotlib not available - skipping charts")
            return
        
        logger.info("Generating comparison charts...")
        
        charts_dir = self.output_dir / 'charts'
        charts_dir.mkdir(exist_ok=True)
        
        # Chart 1: Optimizer comparison
        self._create_optimizer_comparison_chart(charts_dir)
        
        # Chart 2: Learning rate effect
        self._create_learning_rate_chart(charts_dir)
        
        # Chart 3: Batch size effect
        self._create_batch_size_chart(charts_dir)
        
        logger.info(f"✅ Charts saved to {charts_dir}")
    
    def _create_optimizer_comparison_chart(self, output_dir: Path):
        """Create optimizer comparison bar chart"""
        
        summary = self.analyzer.get_optimizer_summary()
        
        if summary.empty:
            return
        
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 6))
        
        # R2 comparison
        ax1.bar(summary['Optimizer'], summary['Avg_R2'], 
                yerr=summary['Std_R2'], capsize=5, alpha=0.7, color='skyblue', edgecolor='black')
        ax1.set_xlabel('Optimizer', fontsize=12, fontweight='bold')
        ax1.set_ylabel('Average R² Score', fontsize=12, fontweight='bold')
        ax1.set_title('Optimizer Performance Comparison (R²)', fontsize=14, fontweight='bold')
        ax1.grid(axis='y', alpha=0.3)
        ax1.tick_params(axis='x', rotation=45)
        
        # RMSE comparison
        ax2.bar(summary['Optimizer'], summary['Avg_RMSE'],
                yerr=summary['Std_RMSE'], capsize=5, alpha=0.7, color='lightcoral', edgecolor='black')
        ax2.set_xlabel('Optimizer', fontsize=12, fontweight='bold')
        ax2.set_ylabel('Average RMSE', fontsize=12, fontweight='bold')
        ax2.set_title('Optimizer Performance Comparison (RMSE)', fontsize=14, fontweight='bold')
        ax2.grid(axis='y', alpha=0.3)
        ax2.tick_params(axis='x', rotation=45)
        
        plt.tight_layout()
        plt.savefig(output_dir / 'optimizer_comparison.png', dpi=150, bbox_inches='tight')
        plt.close()
    
    def _create_learning_rate_chart(self, output_dir: Path):
        """Create learning rate effect chart"""
        
        summary = self.analyzer.get_learning_rate_summary()
        
        if summary.empty:
            return
        
        fig, ax = plt.subplots(figsize=(10, 6))
        
        ax.errorbar(summary['Learning_Rate'], summary['Avg_R2'],
                    yerr=summary['Std_R2'], fmt='o-', capsize=5, 
                    markersize=8, linewidth=2, alpha=0.7)
        
        ax.set_xlabel('Learning Rate', fontsize=12, fontweight='bold')
        ax.set_ylabel('Average R² Score', fontsize=12, fontweight='bold')
        ax.set_title('Learning Rate Effect on Performance', fontsize=14, fontweight='bold')
        ax.set_xscale('log')
        ax.grid(True, alpha=0.3)
        
        plt.tight_layout()
        plt.savefig(output_dir / 'learning_rate_effect.png', dpi=150, bbox_inches='tight')
        plt.close()
    
    def _create_batch_size_chart(self, output_dir: Path):
        """Create batch size effect chart"""
        
        summary = self.analyzer.get_batch_size_summary()
        
        if summary.empty:
            return
        
        fig, ax = plt.subplots(figsize=(10, 6))
        
        ax.errorbar(summary['Batch_Size'], summary['Avg_R2'],
                    yerr=summary['Std_R2'], fmt='s-', capsize=5,
                    markersize=8, linewidth=2, alpha=0.7, color='green')
        
        ax.set_xlabel('Batch Size', fontsize=12, fontweight='bold')
        ax.set_ylabel('Average R² Score', fontsize=12, fontweight='bold')
        ax.set_title('Batch Size Effect on Performance', fontsize=14, fontweight='bold')
        ax.grid(True, alpha=0.3)
        
        plt.tight_layout()
        plt.savefig(output_dir / 'batch_size_effect.png', dpi=150, bbox_inches='tight')
        plt.close()
    
    def generate_excel_report(self, filename: str = 'Optimizer_Comparison.xlsx'):
        """Generate comprehensive Excel report"""
        
        excel_path = self.output_dir / filename
        
        logger.info(f"Generating Excel report: {excel_path}")
        
        # Generate analyses
        analyses = self.generate_comparison_analysis()
        
        if OPENPYXL_AVAILABLE:
            self._generate_formatted_excel(excel_path, analyses)
        else:
            self._generate_simple_excel(excel_path, analyses)
        
        logger.info(f"✅ Excel report saved: {excel_path}")
    
    def _generate_simple_excel(self, excel_path: Path, analyses: Dict):
        """Generate simple Excel without formatting"""
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            for sheet_name, df in analyses.items():
                if not df.empty:
                    df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
    
    def _generate_formatted_excel(self, excel_path: Path, analyses: Dict):
        """Generate formatted Excel with styling"""
        
        wb = Workbook()
        wb.remove(wb.active)
        
        # Define styles
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        # Sheet 1: Optimizer Summary
        if 'optimizer_summary' in analyses and not analyses['optimizer_summary'].empty:
            ws = wb.create_sheet('Optimizer_Summary')
            df = analyses['optimizer_summary']
            
            # Write headers
            headers = df.columns.tolist()
            ws.append(headers)
            
            # Style headers
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Write data
            for _, row in df.iterrows():
                ws.append(row.tolist())
            
            # Auto-adjust columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Sheet 2: Learning Rate Analysis
        if 'learning_rate_summary' in analyses and not analyses['learning_rate_summary'].empty:
            ws = wb.create_sheet('Learning_Rate_Analysis')
            df = analyses['learning_rate_summary']
            
            headers = df.columns.tolist()
            ws.append(headers)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            for _, row in df.iterrows():
                ws.append(row.tolist())
            
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 20)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Sheet 3: Batch Size Analysis
        if 'batch_size_summary' in analyses and not analyses['batch_size_summary'].empty:
            ws = wb.create_sheet('Batch_Size_Analysis')
            df = analyses['batch_size_summary']
            
            headers = df.columns.tolist()
            ws.append(headers)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            for _, row in df.iterrows():
                ws.append(row.tolist())
            
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 20)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Sheet 4: Best Configurations
        if 'best_configs' in analyses and not analyses['best_configs'].empty:
            ws = wb.create_sheet('Best_Configurations')
            df = analyses['best_configs']
            
            headers = df.columns.tolist()
            ws.append(headers)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Highlight top 3
            highlight_fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
            
            for idx, row in df.iterrows():
                row_data = row.tolist()
                ws.append(row_data)
                
                # Highlight top 3 rows
                if idx < 3:
                    for cell in ws[ws.max_row]:
                        cell.fill = highlight_fill
            
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(excel_path)


# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    """Main execution for testing"""
    
    print("\n" + "=" * 80)
    print("PFAZ 2: OPTIMIZER COMPARISON REPORTER - TEST")
    print("=" * 80)
    
    # Initialize reporter
    reporter = OptimizerComparisonReporter(output_dir='test_optimizer_comparison')
    
    # Create mock data for testing
    print("\nCreating mock training results...")
    
    test_models_dir = Path('test_trained_models')
    test_configs_file = Path('test_configs.json')
    
    # Create mock configs
    mock_configs = [
        {
            'id': 'TRAIN_001',
            'optimizer': 'adam',
            'learning_rate': 0.001,
            'batch_size': 32,
            'architecture': [256, 128, 64]
        },
        {
            'id': 'TRAIN_002',
            'optimizer': 'sgd',
            'learning_rate': 0.01,
            'batch_size': 64,
            'architecture': [256, 128, 64]
        },
        {
            'id': 'TRAIN_003',
            'optimizer': 'adam',
            'learning_rate': 0.0001,
            'batch_size': 16,
            'architecture': [512, 256, 128]
        }
    ]
    
    with open(test_configs_file, 'w') as f:
        json.dump(mock_configs, f)
    
    # Create mock metrics
    for config in mock_configs:
        metrics_dir = test_models_dir / 'MM_75nuclei' / 'DNN' / config['id']
        metrics_dir.mkdir(parents=True, exist_ok=True)
        
        mock_metrics = {
            'val': {
                'r2': np.random.uniform(0.7, 0.95),
                'rmse': np.random.uniform(0.05, 0.15),
                'mae': np.random.uniform(0.03, 0.10)
            },
            'training_time': np.random.uniform(100, 500),
            'epochs_trained': np.random.randint(50, 100)
        }
        
        with open(metrics_dir / f"metrics_{config['id']}.json", 'w') as f:
            json.dump(mock_metrics, f)
    
    # Load results
    print("\nLoading training results...")
    reporter.load_training_results(test_models_dir, test_configs_file)
    
    # Generate analyses
    print("\nGenerating comparison analyses...")
    analyses = reporter.generate_comparison_analysis()
    
    # Print summaries
    for name, df in analyses.items():
        print(f"\n{name.upper()}:")
        print(df.to_string(index=False))
    
    # Generate charts
    print("\nGenerating charts...")
    reporter.generate_charts()
    
    # Generate Excel report
    print("\nGenerating Excel report...")
    reporter.generate_excel_report()
    
    print("\n✅ OPTIMIZER COMPARISON TEST COMPLETED!")
    print(f"Output directory: {reporter.output_dir}")


if __name__ == "__main__":
    main()
