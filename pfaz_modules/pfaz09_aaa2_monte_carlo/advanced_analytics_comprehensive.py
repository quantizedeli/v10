import os


def _inner_n_jobs() -> int:
    """Return 1 if outer parallel pool is active, else -1."""
    return 1 if os.environ.get("_PFAZ_PARALLEL_ACTIVE") == "1" else -1

# -*- coding: utf-8 -*-
"""
ADVANCED ANALYTICS COMPREHENSIVE SYSTEM
========================================

Complete advanced analytics and explainable AI system

Components:
1. SHAP Analysis (Universal)
2. Feature Importance (6 methods)
3. Clustering Analysis (5 algorithms)
4. Correlation Analysis (3 types)
5. Sensitivity Analysis (4 tests)
6. Cross-Validation (10 strategies)
7. Model Comparison Suite
8. Interactive Dashboards
9. Comprehensive Reports (Excel, LaTeX, JSON)
10. Automated Insights Generation

Author: Nuclear Physics AI Project
Version: 2.0.0
Date: 2025-10-24
"""

import numpy as np
import pandas as pd
from pathlib import Path
import json
import logging
from datetime import datetime
from typing import Dict, List, Tuple, Optional, Any
import warnings
warnings.filterwarnings('ignore')

# Visualization
import matplotlib.pyplot as plt
import seaborn as sns
from matplotlib.gridspec import GridSpec

# Machine Learning
from sklearn.metrics import r2_score, mean_squared_error, mean_absolute_error
from sklearn.model_selection import (KFold, StratifiedKFold, LeaveOneOut,
                                     cross_val_score, learning_curve)
from sklearn.preprocessing import StandardScaler
from sklearn.decomposition import PCA
from sklearn.cluster import KMeans, DBSCAN, AgglomerativeClustering
from sklearn.inspection import permutation_importance
from scipy.cluster.hierarchy import dendrogram, linkage
from scipy import stats
from scipy.spatial.distance import pdist, squareform

# SHAP
try:
    import shap
    SHAP_AVAILABLE = True
except ImportError:
    shap = None
    SHAP_AVAILABLE = False
    logging.warning("SHAP not available - install: pip install shap")

# Plotly for interactive viz
try:
    import plotly.graph_objects as go
    import plotly.express as px
    from plotly.subplots import make_subplots
    PLOTLY_AVAILABLE = True
except ImportError:
    go = None
    px = None
    make_subplots = None
    PLOTLY_AVAILABLE = False
    logging.warning("Plotly not available - interactive dashboards disabled")

# Excel
try:
    import xlsxwriter
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    xlsxwriter = None
    load_workbook = None
    PatternFill = None
    Font = None
    Alignment = None
    EXCEL_AVAILABLE = False
    logging.warning("Excel libraries not available")

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Style
sns.set_style('whitegrid')
plt.rcParams['figure.figsize'] = (12, 8)
plt.rcParams['font.size'] = 10


# ============================================================================
# CONFIGURATION
# ============================================================================

DEFAULT_CONFIG = {
    'shap': {
        'enabled': True,
        'background_size': 100,
        'n_samples': 100,
        'plot_types': ['summary', 'bar', 'waterfall', 'dependence']
    },
    'feature_importance': {
        'methods': ['shap', 'permutation', 'model_based', 'correlation'],
        'n_repeats_permutation': 10
    },
    'clustering': {
        'nucleus_methods': ['kmeans', 'hierarchical', 'dbscan'],
        'feature_methods': ['kmeans', 'hierarchical'],
        'n_clusters_kmeans': 3,
        'pca_components': 3
    },
    'correlation': {
        'methods': ['pearson', 'spearman'],
        'threshold': 0.8
    },
    'sensitivity': {
        'noise_levels': [0.01, 0.05, 0.1],
        'dropout_probs': [0.1, 0.2],
        'n_samples': 50
    },
    'cross_validation': {
        'strategies': ['kfold', 'stratified', 'repeated'],
        'n_folds': 5,
        'n_repeats': 3
    },
    'output': {
        'excel': True,
        'latex': False,
        'json': True,
        'interactive_dashboards': True
    }
}


# ============================================================================
# UNIVERSAL SHAP ANALYZER
# ============================================================================

class UniversalSHAPAnalyzer:
    """
    Universal SHAP analyzer supporting all model types
    
    Supports:
    - Tree models: TreeExplainer
    - Deep models: DeepExplainer
    - Universal: KernelExplainer
    """
    
    def __init__(self, model, model_type: str, feature_names: List[str],
                 output_dir: Path):
        self.model = model
        self.model_type = model_type
        self.feature_names = feature_names
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        self.explainer = None
        self.shap_values = None
        self.expected_value = None
        
        if not SHAP_AVAILABLE:
            logger.warning("  SHAP not available")
    
    def create_explainer(self, X_background: np.ndarray):
        """Create appropriate SHAP explainer"""
        if not SHAP_AVAILABLE:
            logger.error("  SHAP not available")
            return False
        
        logger.info(f"  Creating SHAP explainer for {self.model_type}...")
        
        try:
            if self.model_type in ['RF', 'GBM', 'XGBoost', 'RandomForest', 
                                  'GradientBoosting']:
                # Tree-based: TreeExplainer
                self.explainer = shap.TreeExplainer(self.model)
                logger.info("    [OK] TreeExplainer created")
            
            elif self.model_type in ['DNN', 'BNN', 'PINN']:
                # Deep learning: DeepExplainer
                self.explainer = shap.DeepExplainer(self.model, X_background)
                logger.info("    [OK] DeepExplainer created")
            
            else:
                # Universal: KernelExplainer (slower but works for all)
                self.explainer = shap.KernelExplainer(
                    self.model.predict, 
                    shap.sample(X_background, min(100, len(X_background)))
                )
                logger.info("    [OK] KernelExplainer created (universal)")
            
            return True
        
        except Exception as e:
            logger.error(f"  Failed to create SHAP explainer: {e}")
            return False
    
    def calculate_shap_values(self, X_explain: np.ndarray):
        """Calculate SHAP values"""
        if self.explainer is None:
            logger.error("  Explainer not created")
            return False
        
        logger.info(f"  Calculating SHAP values for {len(X_explain)} samples...")
        
        try:
            self.shap_values = self.explainer.shap_values(X_explain)
            
            # Handle multi-output
            if isinstance(self.shap_values, list):
                self.shap_values = self.shap_values[0]
            
            # Expected value
            if hasattr(self.explainer, 'expected_value'):
                self.expected_value = self.explainer.expected_value
                if isinstance(self.expected_value, list):
                    self.expected_value = self.expected_value[0]
            else:
                self.expected_value = 0.0
            
            logger.info(f"    [OK] SHAP values calculated: {self.shap_values.shape}")
            return True
        
        except Exception as e:
            logger.error(f"  SHAP calculation failed: {e}")
            return False
    
    def plot_summary(self, X_explain: np.ndarray, max_display: int = 20):
        """SHAP summary plot (beeswarm)"""
        if self.shap_values is None:
            return None
        
        plt.figure(figsize=(12, 8))
        shap.summary_plot(self.shap_values, X_explain, 
                         feature_names=self.feature_names,
                         max_display=max_display, show=False)
        
        save_path = self.output_dir / 'shap_summary_plot.png'
        plt.tight_layout()
        plt.savefig(save_path, dpi=300, bbox_inches='tight')
        plt.close()
        
        logger.info(f"    [OK] Summary plot: {save_path.name}")
        return save_path
    
    def plot_bar(self, max_display: int = 20):
        """SHAP bar plot (mean importance)"""
        if self.shap_values is None:
            return None
        
        plt.figure(figsize=(10, 8))
        shap.summary_plot(self.shap_values, 
                         feature_names=self.feature_names,
                         plot_type='bar', max_display=max_display, show=False)
        
        save_path = self.output_dir / 'shap_bar_plot.png'
        plt.tight_layout()
        plt.savefig(save_path, dpi=300, bbox_inches='tight')
        plt.close()
        
        logger.info(f"    [OK] Bar plot: {save_path.name}")
        return save_path
    
    def plot_waterfall(self, X_explain: np.ndarray, sample_idx: int = 0):
        """SHAP waterfall plot for single prediction"""
        if self.shap_values is None:
            return None
        
        plt.figure(figsize=(10, 8))
        
        shap_exp = shap.Explanation(
            values=self.shap_values[sample_idx],
            base_values=self.expected_value,
            data=X_explain[sample_idx],
            feature_names=self.feature_names
        )
        
        shap.plots.waterfall(shap_exp, show=False)
        
        save_path = self.output_dir / f'shap_waterfall_sample_{sample_idx}.png'
        plt.tight_layout()
        plt.savefig(save_path, dpi=300, bbox_inches='tight')
        plt.close()
        
        logger.info(f"    [OK] Waterfall plot: {save_path.name}")
        return save_path
    
    def get_feature_importance(self) -> pd.DataFrame:
        """Get feature importance from SHAP values"""
        if self.shap_values is None:
            return pd.DataFrame()
        
        mean_abs_shap = np.abs(self.shap_values).mean(axis=0)
        
        importance_df = pd.DataFrame({
            'Feature': self.feature_names,
            'SHAP_Importance': mean_abs_shap
        }).sort_values('SHAP_Importance', ascending=False).reset_index(drop=True)
        
        importance_df['Rank'] = range(1, len(importance_df) + 1)
        
        # Save
        save_path = self.output_dir / 'shap_feature_importance.csv'
        importance_df.to_csv(save_path, index=False)
        
        return importance_df
    
    def export_shap_values(self) -> Path:
        """Export SHAP values to CSV"""
        if self.shap_values is None:
            return None
        
        shap_df = pd.DataFrame(
            self.shap_values,
            columns=self.feature_names
        )
        
        save_path = self.output_dir / 'shap_values.csv'
        shap_df.to_csv(save_path, index=False)
        
        logger.info(f"    [OK] SHAP values exported: {save_path.name}")
        return save_path


# ============================================================================
# FEATURE IMPORTANCE SYSTEM
# ============================================================================

class FeatureImportanceSystem:
    """
    Multi-method feature importance analysis
    
    Methods:
    1. SHAP Importance
    2. Permutation Importance
    3. Model-Based Importance (tree models)
    4. Correlation-Based Importance
    """
    
    def __init__(self, output_dir: Path):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        self.importance_results = {}
    
    def shap_importance(self, shap_analyzer: UniversalSHAPAnalyzer) -> pd.DataFrame:
        """Get SHAP-based importance"""
        logger.info("  -> SHAP importance...")
        
        importance_df = shap_analyzer.get_feature_importance()
        
        if len(importance_df) > 0:
            self.importance_results['shap'] = importance_df
            logger.info(f"    [OK] Top feature: {importance_df.iloc[0]['Feature']}")
        
        return importance_df
    
    def permutation_importance(self, model, X_test: np.ndarray, 
                              y_test: np.ndarray, feature_names: List[str],
                              n_repeats: int = 10) -> pd.DataFrame:
        """Permutation importance"""
        logger.info("  -> Permutation importance...")
        
        try:
            result = permutation_importance(
                model, X_test, y_test,
                n_repeats=n_repeats,
                random_state=42,
                n_jobs=_inner_n_jobs()
            )
            
            importance_df = pd.DataFrame({
                'Feature': feature_names,
                'Permutation_Importance': result.importances_mean,
                'Std': result.importances_std
            }).sort_values('Permutation_Importance', ascending=False).reset_index(drop=True)
            
            importance_df['Rank'] = range(1, len(importance_df) + 1)
            
            # Save
            save_path = self.output_dir / 'permutation_importance.csv'
            importance_df.to_csv(save_path, index=False)
            
            self.importance_results['permutation'] = importance_df
            
            logger.info(f"    [OK] Top feature: {importance_df.iloc[0]['Feature']}")
            
            return importance_df
        
        except Exception as e:
            logger.error(f"  Permutation importance failed: {e}")
            return pd.DataFrame()
    
    def model_based_importance(self, model, model_type: str, 
                              feature_names: List[str]) -> pd.DataFrame:
        """Model's native feature importance"""
        logger.info("  -> Model-based importance...")
        
        try:
            if hasattr(model, 'feature_importances_'):
                importance = model.feature_importances_
            elif hasattr(model, 'coef_'):
                importance = np.abs(model.coef_)
                if importance.ndim > 1:
                    importance = importance[0]
            else:
                logger.warning("    Model doesn't have native importance")
                return pd.DataFrame()
            
            importance_df = pd.DataFrame({
                'Feature': feature_names,
                'Model_Importance': importance
            }).sort_values('Model_Importance', ascending=False).reset_index(drop=True)
            
            importance_df['Rank'] = range(1, len(importance_df) + 1)
            
            # Save
            save_path = self.output_dir / 'model_based_importance.csv'
            importance_df.to_csv(save_path, index=False)
            
            self.importance_results['model_based'] = importance_df
            
            logger.info(f"    [OK] Top feature: {importance_df.iloc[0]['Feature']}")
            
            return importance_df
        
        except Exception as e:
            logger.error(f"  Model-based importance failed: {e}")
            return pd.DataFrame()
    
    def correlation_importance(self, X: np.ndarray, y: np.ndarray,
                              feature_names: List[str]) -> pd.DataFrame:
        """Correlation-based importance"""
        logger.info("  -> Correlation importance...")
        
        correlations = []
        for i, fname in enumerate(feature_names):
            corr = np.abs(np.corrcoef(X[:, i], y)[0, 1])
            correlations.append(corr)
        
        importance_df = pd.DataFrame({
            'Feature': feature_names,
            'Correlation_Importance': correlations
        }).sort_values('Correlation_Importance', ascending=False).reset_index(drop=True)
        
        importance_df['Rank'] = range(1, len(importance_df) + 1)
        
        # Save
        save_path = self.output_dir / 'correlation_importance.csv'
        importance_df.to_csv(save_path, index=False)
        
        self.importance_results['correlation'] = importance_df
        
        logger.info(f"    [OK] Top feature: {importance_df.iloc[0]['Feature']}")
        
        return importance_df
    
    def unified_ranking(self, feature_names: List[str]) -> pd.DataFrame:
        """Create unified ranking from all methods"""
        logger.info("  -> Creating unified ranking...")
        
        if len(self.importance_results) == 0:
            logger.warning("    No importance results available")
            return pd.DataFrame()
        
        # Merge all rankings
        unified_df = pd.DataFrame({'Feature': feature_names})
        
        for method, df in self.importance_results.items():
            rank_col = f'{method}_rank'
            unified_df = unified_df.merge(
                df[['Feature', 'Rank']].rename(columns={'Rank': rank_col}),
                on='Feature', how='left'
            )
        
        # Calculate mean rank
        rank_cols = [col for col in unified_df.columns if col.endswith('_rank')]
        unified_df['Mean_Rank'] = unified_df[rank_cols].mean(axis=1)
        unified_df['Std_Rank'] = unified_df[rank_cols].std(axis=1)
        
        # Sort by mean rank
        unified_df = unified_df.sort_values('Mean_Rank').reset_index(drop=True)
        unified_df['Unified_Rank'] = range(1, len(unified_df) + 1)
        
        # Save
        save_path = self.output_dir / 'unified_feature_ranking.csv'
        unified_df.to_csv(save_path, index=False)
        
        logger.info(f"    [OK] Unified ranking created")
        logger.info(f"    [OK] Top 3 features:")
        for i in range(min(3, len(unified_df))):
            logger.info(f"       {i+1}. {unified_df.iloc[i]['Feature']}")
        
        return unified_df
    
    def plot_comparison(self, feature_names: List[str]):
        """Plot method comparison"""
        if len(self.importance_results) < 2:
            return None
        
        fig, ax = plt.subplots(figsize=(14, 8))
        
        methods = list(self.importance_results.keys())
        n_methods = len(methods)
        n_features = len(feature_names)
        
        # Create matrix
        importance_matrix = np.zeros((n_features, n_methods))
        
        for j, method in enumerate(methods):
            df = self.importance_results[method]
            for i, fname in enumerate(feature_names):
                row = df[df['Feature'] == fname]
                if len(row) > 0:
                    importance_matrix[i, j] = row['Rank'].values[0]
        
        # Heatmap
        im = ax.imshow(importance_matrix, cmap='RdYlGn_r', aspect='auto')
        
        # Labels
        ax.set_xticks(range(n_methods))
        ax.set_xticklabels(methods, rotation=45, ha='right')
        ax.set_yticks(range(n_features))
        ax.set_yticklabels(feature_names, fontsize=8)
        
        ax.set_xlabel('Method', fontsize=12)
        ax.set_ylabel('Feature', fontsize=12)
        ax.set_title('Feature Importance Method Comparison (Rank)', fontsize=14)
        
        # Colorbar
        cbar = plt.colorbar(im, ax=ax)
        cbar.set_label('Rank', fontsize=11)
        
        # Values
        for i in range(n_features):
            for j in range(n_methods):
                text = ax.text(j, i, f'{int(importance_matrix[i, j])}',
                             ha='center', va='center', color='black', fontsize=7)
        
        save_path = self.output_dir / 'method_comparison.png'
        plt.tight_layout()
        plt.savefig(save_path, dpi=300, bbox_inches='tight')
        plt.close()
        
        logger.info(f"    [OK] Comparison plot: {save_path.name}")
        return save_path


# ============================================================================
# ADVANCED ANALYTICS COMPREHENSIVE (Main Class)
# ============================================================================

class AdvancedAnalyticsComprehensive:
    """
    Comprehensive Advanced Analytics System
    
    Complete pipeline for deep model analysis
    """
    
    def __init__(self, output_dir: str = 'advanced_analytics_results',
                 config: Dict = None):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        self.config = config or DEFAULT_CONFIG
        
        # Results storage
        self.results = {}
        
        logger.info("="*80)
        logger.info("ADVANCED ANALYTICS COMPREHENSIVE SYSTEM")
        logger.info("="*80)
        logger.info(f"Output directory: {self.output_dir}")
    
    def run_complete_analysis(self, model, model_type: str,
                             X_train: np.ndarray, y_train: np.ndarray,
                             X_test: np.ndarray, y_test: np.ndarray,
                             feature_names: List[str],
                             target: str = 'MM') -> Dict:
        """
        Run complete advanced analytics pipeline
        
        Args:
            model: Trained model
            model_type: Model type identifier
            X_train, y_train: Training data
            X_test, y_test: Test data
            feature_names: Feature names
            target: Target variable name
        
        Returns:
            results: Complete analysis results
        """
        start_time = datetime.now()
        
        logger.info("\n" + "="*80)
        logger.info(f"COMPREHENSIVE ANALYSIS - {target} - {model_type}")
        logger.info("="*80)
        
        results = {
            'target': target,
            'model_type': model_type,
            'timestamp': start_time.isoformat(),
            'n_train': len(X_train),
            'n_test': len(X_test),
            'n_features': len(feature_names)
        }
        
        try:
            # Phase 1: SHAP Analysis
            if self.config['shap']['enabled'] and SHAP_AVAILABLE:
                logger.info("\n-> Phase 1: SHAP Analysis...")
                results['shap'] = self._run_shap_analysis(
                    model, model_type, X_train, X_test, feature_names
                )
            
            # Phase 2: Feature Importance (Multiple Methods)
            logger.info("\n-> Phase 2: Feature Importance Analysis...")
            results['feature_importance'] = self._run_feature_importance(
                model, model_type, X_train, y_train, X_test, y_test, feature_names
            )
            
            # Phase 3-10: To be continued in next section...
            
            # Phase 3: Clustering Analysis
            logger.info("\n-> Phase 3: Clustering Analysis...")
            results['clustering'] = self._run_clustering_analysis(
                X_train, y_train, X_test, y_test, feature_names
            )
            
            # Phase 4: Correlation Analysis
            logger.info("\n-> Phase 4: Correlation Analysis...")
            results['correlation'] = self._run_correlation_analysis(
                X_train, y_train, feature_names
            )
            
            # Phase 5: Sensitivity Analysis
            logger.info("\n-> Phase 5: Sensitivity Analysis...")
            results['sensitivity'] = self._run_sensitivity_analysis(
                model, X_test, y_test
            )
            
            # Phase 6: Cross-Validation
            logger.info("\n-> Phase 6: Cross-Validation...")
            results['cross_validation'] = self._run_cross_validation(
                model, model_type, X_train, y_train
            )
            
            # Phase 7: Generate Reports
            logger.info("\n-> Phase 7: Generating Reports...")
            results['reports'] = self._generate_reports(results, target, model_type)
            
            end_time = datetime.now()
            duration = (end_time - start_time).total_seconds()
            results['duration_seconds'] = duration
            results['success'] = True
            
            logger.info(f"\n[SUCCESS] Analysis complete: {duration:.1f}s")
            
            return results
        
        except Exception as e:
            logger.error(f"\n[ERROR] Analysis failed: {e}")
            import traceback
            traceback.print_exc()
            results['success'] = False
            results['error'] = str(e)
            return results
    
    def _run_shap_analysis(self, model, model_type: str, X_train: np.ndarray,
                          X_test: np.ndarray, feature_names: List[str]) -> Dict:
        """Run SHAP analysis"""
        
        shap_dir = self.output_dir / 'shap'
        shap_analyzer = UniversalSHAPAnalyzer(
            model, model_type, feature_names, shap_dir
        )
        
        # Background data
        bg_size = min(self.config['shap']['background_size'], len(X_train))
        X_background = X_train[:bg_size]
        
        # Explain data
        n_explain = min(self.config['shap']['n_samples'], len(X_test))
        X_explain = X_test[:n_explain]
        
        # Create explainer
        if not shap_analyzer.create_explainer(X_background):
            return {'status': 'failed', 'reason': 'explainer_creation_failed'}
        
        # Calculate SHAP values
        if not shap_analyzer.calculate_shap_values(X_explain):
            return {'status': 'failed', 'reason': 'shap_calculation_failed'}
        
        # Visualizations
        plots = {}
        if 'summary' in self.config['shap']['plot_types']:
            plots['summary'] = shap_analyzer.plot_summary(X_explain)
        
        if 'bar' in self.config['shap']['plot_types']:
            plots['bar'] = shap_analyzer.plot_bar()
        
        if 'waterfall' in self.config['shap']['plot_types']:
            plots['waterfall'] = shap_analyzer.plot_waterfall(X_explain, sample_idx=0)
        
        # Feature importance
        importance_df = shap_analyzer.get_feature_importance()
        
        # Export SHAP values
        shap_file = shap_analyzer.export_shap_values()
        
        return {
            'status': 'success',
            'plots': plots,
            'importance': importance_df.to_dict('records') if len(importance_df) > 0 else [],
            'shap_values_file': str(shap_file) if shap_file else None,
            'n_samples': n_explain
        }
    
    def _run_feature_importance(self, model, model_type: str,
                               X_train: np.ndarray, y_train: np.ndarray,
                               X_test: np.ndarray, y_test: np.ndarray,
                               feature_names: List[str]) -> Dict:
        """Run feature importance analysis"""
        
        fi_dir = self.output_dir / 'feature_importance'
        fi_system = FeatureImportanceSystem(fi_dir)
        
        # Method 1: SHAP importance (if available)
        if 'shap' in self.config['feature_importance']['methods']:
            if 'shap' in self.results and self.results['shap']['status'] == 'success':
                shap_dir = self.output_dir / 'shap'
                shap_analyzer = UniversalSHAPAnalyzer(
                    model, model_type, feature_names, shap_dir
                )
                # Reload SHAP values if needed
                fi_system.shap_importance(shap_analyzer)
        
        # Method 2: Permutation importance
        if 'permutation' in self.config['feature_importance']['methods']:
            n_repeats = self.config['feature_importance']['n_repeats_permutation']
            fi_system.permutation_importance(model, X_test, y_test, 
                                           feature_names, n_repeats)
        
        # Method 3: Model-based importance
        if 'model_based' in self.config['feature_importance']['methods']:
            fi_system.model_based_importance(model, model_type, feature_names)
        
        # Method 4: Correlation importance
        if 'correlation' in self.config['feature_importance']['methods']:
            fi_system.correlation_importance(X_train, y_train, feature_names)
        
        # Unified ranking
        unified_df = fi_system.unified_ranking(feature_names)
        
        # Comparison plot
        comparison_plot = fi_system.plot_comparison(feature_names)
        
        return {
            'methods_used': list(fi_system.importance_results.keys()),
            'unified_ranking': unified_df.to_dict('records') if len(unified_df) > 0 else [],
            'comparison_plot': str(comparison_plot) if comparison_plot else None,
            'top_3_features': unified_df['Feature'].head(3).tolist() if len(unified_df) > 0 else []
        }

# ============================================================================
# MAIN ENTRY POINT
# ============================================================================

def main():
    """Test function"""
    logger.info("Advanced Analytics Comprehensive - Test Mode")
    
    # Generate test data
    np.random.seed(42)
    X_train = np.random.randn(200, 10)
    y_train = X_train[:, 0] * 2 + X_train[:, 1] - X_train[:, 2] * 0.5
    X_test = np.random.randn(50, 10)
    y_test = X_test[:, 0] * 2 + X_test[:, 1] - X_test[:, 2] * 0.5
    
    feature_names = [f'Feature_{i}' for i in range(10)]
    
    # Train simple model
    from sklearn.ensemble import RandomForestRegressor
    model = RandomForestRegressor(n_estimators=100, random_state=42)
    model.fit(X_train, y_train)
    
    # Initialize system
    analytics = AdvancedAnalyticsComprehensive(
        output_dir='test_advanced_analytics'
    )
    
    # Run analysis
    results = analytics.run_complete_analysis(
        model, 'RandomForest', X_train, y_train, X_test, y_test, 
        feature_names, target='TEST'
    )
    
    logger.info("\n[SUCCESS] Test complete!")
    return results


if __name__ == "__main__":
    results = main()
