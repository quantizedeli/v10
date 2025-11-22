# -*- coding: utf-8 -*-
"""
PFAZ 3: ANFIS All Nuclei Predictor
===================================

Predict on ALL nuclei datasets using trained ANFIS models

Features:
- Load ALL nuclei datasets (MM, QM, MM_QM, Beta_2)
- Load all trained ANFIS models
- Multi-model predictions
- Delta calculations (Pred - Exp)
- Best ANFIS model identification
- Multi-sheet Excel generation (4 sheets)
- Summary statistics

Author: Nuclear Physics AI Training Pipeline
Version: 1.0.0
Date: 2025-10-15
"""

import numpy as np
import pandas as pd
from pathlib import Path
import json
import logging
from typing import Dict, List, Optional
import warnings
warnings.filterwarnings('ignore')

try:
    import joblib
    JOBLIB_AVAILABLE = True
except ImportError:
    JOBLIB_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


# ============================================================================
# ANFIS MODEL LOADER
# ============================================================================

class ANFISModelLoader:
    """Load trained ANFIS models"""
    
    def __init__(self, models_dir: Path):
        self.models_dir = Path(models_dir)
        self.loaded_models = {}
        
    def load_all_models(self) -> Dict:
        """Load all trained ANFIS models"""
        
        model_files = list(self.models_dir.glob('**/model_*.pkl'))
        
        logger.info(f"Found {len(model_files)} ANFIS model files")
        
        for model_file in model_files:
            try:
                # Extract identifiers from path
                config_id = model_file.stem.replace('model_', '')
                dataset_name = model_file.parent.parent.name
                
                model_id = f"{dataset_name}_{config_id}"
                
                # Load model
                model = joblib.load(model_file)
                
                self.loaded_models[model_id] = {
                    'model': model,
                    'config_id': config_id,
                    'dataset_name': dataset_name,
                    'model_path': model_file
                }
                
            except Exception as e:
                logger.warning(f"Failed to load {model_file}: {e}")
        
        logger.info(f"Successfully loaded {len(self.loaded_models)} ANFIS models")
        
        return self.loaded_models


# ============================================================================
# ANFIS ALL NUCLEI PREDICTOR
# ============================================================================

class ANFISAllNucleiPredictor:
    """
    ANFIS All Nuclei Predictor
    
    Features:
    - Load ALL nuclei datasets
    - Predict with all ANFIS models
    - Generate comprehensive Excel
    - Calculate deltas and identify best models
    """
    
    def __init__(self,
                 datasets_dir: str,
                 models_dir: str,
                 output_dir: str = 'anfis_all_nuclei_predictions'):
        
        self.datasets_dir = Path(datasets_dir)
        self.models_dir = Path(models_dir)
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        self.model_loader = ANFISModelLoader(self.models_dir)
        self.predictions = {}
        
        logger.info("=" * 80)
        logger.info("ANFIS ALL NUCLEI PREDICTOR INITIALIZED")
        logger.info("=" * 80)
        logger.info(f"Datasets dir: {self.datasets_dir}")
        logger.info(f"Models dir: {self.models_dir}")
        logger.info(f"Output dir: {self.output_dir}")
        logger.info("=" * 80)
    
    def load_all_datasets(self) -> Dict[str, pd.DataFrame]:
        """Load ALL nuclei datasets"""
        
        datasets = {}
        
        # Dataset names
        dataset_names = [
            'MM_ALL_267nuclei',
            'QM_ALL_219nuclei',
            'MM_QM_ALL_219nuclei',
            'Beta_2_ALL_267nuclei'
        ]
        
        for name in dataset_names:
            dataset_file = None
            
            # Try different extensions
            for ext in ['.csv', '.xlsx']:
                potential_file = self.datasets_dir / f"{name}{ext}"
                if potential_file.exists():
                    dataset_file = potential_file
                    break
            
            if dataset_file is None:
                logger.warning(f"Dataset not found: {name}")
                continue
            
            # Load dataset
            if dataset_file.suffix == '.csv':
                df = pd.read_csv(dataset_file)
            else:
                df = pd.read_excel(dataset_file)
            
            datasets[name] = df
            logger.info(f"Loaded {name}: {len(df)} nuclei")
        
        logger.info(f"Total datasets loaded: {len(datasets)}")
        
        return datasets
    
    def predict_all_nuclei(self):
        """Predict on all nuclei with all models"""
        
        logger.info("\n" + "=" * 80)
        logger.info("PREDICTING ON ALL NUCLEI DATASETS")
        logger.info("=" * 80)
        
        # Load models
        models = self.model_loader.load_all_models()
        
        if not models:
            logger.error("No ANFIS models found!")
            return
        
        # Load datasets
        datasets = self.load_all_datasets()
        
        if not datasets:
            logger.error("No datasets found!")
            return
        
        # Predict on each dataset
        for dataset_name, df in datasets.items():
            logger.info(f"\nProcessing {dataset_name}...")
            
            # Identify features and target
            target_col = self._identify_target(dataset_name)
            
            if target_col not in df.columns:
                logger.warning(f"Target {target_col} not found in {dataset_name}")
                continue
            
            # Identify feature columns
            feature_cols = [col for col in df.columns 
                          if col not in ['NUCLEUS', 'A', 'Z', 'N', target_col]]
            
            X = df[feature_cols].values
            y_exp = df[target_col].values
            
            # Create result dataframe
            result_df = df[['NUCLEUS', 'A', 'Z', 'N']].copy() if 'NUCLEUS' in df.columns else df[['A', 'Z', 'N']].copy()
            result_df[f'Experimental_{target_col}'] = y_exp
            
            # Predict with each model
            model_predictions = {}
            
            for model_id, model_info in models.items():
                try:
                    model = model_info['model']
                    
                    # Predict
                    y_pred = model.predict(X)
                    
                    # Calculate delta
                    delta = y_pred - y_exp
                    abs_error = np.abs(delta)
                    
                    # Store predictions
                    result_df[f'{model_id}_Pred'] = y_pred
                    result_df[f'{model_id}_Delta'] = delta
                    result_df[f'{model_id}_AbsError'] = abs_error
                    
                    model_predictions[model_id] = {
                        'predictions': y_pred,
                        'delta': delta,
                        'abs_error': abs_error,
                        'mae': np.mean(abs_error),
                        'rmse': np.sqrt(np.mean(delta**2)),
                        'r2': 1 - np.sum(delta**2) / np.sum((y_exp - np.mean(y_exp))**2)
                    }
                
                except Exception as e:
                    logger.warning(f"Prediction failed for {model_id}: {e}")
            
            # Identify best model per nucleus
            if model_predictions:
                best_models = []
                best_preds = []
                best_deltas = []
                
                for i in range(len(result_df)):
                    min_error = float('inf')
                    best_model = None
                    best_pred = None
                    best_delta = None
                    
                    for model_id, pred_info in model_predictions.items():
                        error = pred_info['abs_error'][i]
                        if error < min_error:
                            min_error = error
                            best_model = model_id
                            best_pred = pred_info['predictions'][i]
                            best_delta = pred_info['delta'][i]
                    
                    best_models.append(best_model)
                    best_preds.append(best_pred)
                    best_deltas.append(best_delta)
                
                result_df['Best_ANFIS_Model'] = best_models
                result_df[f'Best_Pred_{target_col}'] = best_preds
                result_df[f'Best_Delta_{target_col}'] = best_deltas
            
            # Store predictions
            self.predictions[dataset_name] = {
                'dataframe': result_df,
                'target': target_col,
                'model_stats': model_predictions
            }
            
            logger.info(f"[SUCCESS] {dataset_name}: {len(models)} models, {len(df)} nuclei")
        
        logger.info("\n" + "=" * 80)
        logger.info(f"PREDICTIONS COMPLETED FOR {len(self.predictions)} DATASETS")
        logger.info("=" * 80 + "\n")
    
    def _identify_target(self, dataset_name: str) -> str:
        """Identify target column from dataset name"""
        if 'MM_QM' in dataset_name:
            return 'Q'
        elif 'MM' in dataset_name:
            return 'MM'
        elif 'QM' in dataset_name:
            return 'Q'
        elif 'Beta_2' in dataset_name:
            return 'Beta_2'
        else:
            return 'MM'  # Default
    
    def generate_excel_report(self, filename: str = 'ANFIS_All_Nuclei_Predictions.xlsx'):
        """Generate comprehensive Excel report"""
        
        if not self.predictions:
            logger.warning("No predictions to export")
            return
        
        excel_path = self.output_dir / filename
        
        logger.info(f"Generating Excel report: {excel_path}")
        
        if OPENPYXL_AVAILABLE:
            self._generate_formatted_excel(excel_path)
        else:
            self._generate_simple_excel(excel_path)
        
        logger.info(f"[SUCCESS] Excel report saved: {excel_path}")
    
    def _generate_simple_excel(self, excel_path: Path):
        """Generate simple Excel without formatting"""
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            for dataset_name, pred_data in self.predictions.items():
                df = pred_data['dataframe']
                sheet_name = dataset_name[:31]  # Excel limit
                df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    def _generate_formatted_excel(self, excel_path: Path):
        """Generate formatted Excel with styling"""
        
        wb = Workbook()
        wb.remove(wb.active)
        
        # Styles
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        best_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        
        for dataset_name, pred_data in self.predictions.items():
            df = pred_data['dataframe']
            sheet_name = dataset_name[:31]
            
            ws = wb.create_sheet(sheet_name)
            
            # Write headers
            headers = df.columns.tolist()
            ws.append(headers)
            
            # Style headers
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Write data
            for _, row in df.iterrows():
                ws.append(row.tolist())
            
            # Highlight best predictions
            if 'Best_ANFIS_Model' in df.columns:
                best_col_idx = df.columns.get_loc('Best_ANFIS_Model') + 1
                for row_idx in range(2, len(df) + 2):
                    ws.cell(row_idx, best_col_idx).fill = best_fill
            
            # Auto-adjust columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(excel_path)
    
    def save_csv_predictions(self):
        """Save predictions as CSV files"""
        
        csv_dir = self.output_dir / 'csv'
        csv_dir.mkdir(exist_ok=True)
        
        for dataset_name, pred_data in self.predictions.items():
            df = pred_data['dataframe']
            csv_file = csv_dir / f"{dataset_name}_predictions.csv"
            df.to_csv(csv_file, index=False)
            logger.info(f"CSV saved: {csv_file}")
    
    def generate_summary_statistics(self) -> pd.DataFrame:
        """Generate summary statistics"""
        
        summary_data = []
        
        for dataset_name, pred_data in self.predictions.items():
            model_stats = pred_data['model_stats']
            
            for model_id, stats in model_stats.items():
                summary_data.append({
                    'Dataset': dataset_name,
                    'ANFIS_Model': model_id,
                    'MAE': stats['mae'],
                    'RMSE': stats['rmse'],
                    'R2': stats['r2']
                })
        
        summary_df = pd.DataFrame(summary_data)
        
        # Save summary
        summary_file = self.output_dir / 'anfis_prediction_summary.csv'
        summary_df.to_csv(summary_file, index=False)
        
        logger.info(f"Summary statistics saved: {summary_file}")
        
        return summary_df


# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    """Main execution for testing"""
    
    print("\n" + "=" * 80)
    print("PFAZ 3: ANFIS ALL NUCLEI PREDICTOR - TEST")
    print("=" * 80)
    
    predictor = ANFISAllNucleiPredictor(
        datasets_dir='generated_datasets',
        models_dir='test_trained_anfis',
        output_dir='test_anfis_predictions'
    )
    
    # Predict
    print("\nPredicting on all nuclei...")
    predictor.predict_all_nuclei()
    
    # Generate Excel
    print("\nGenerating Excel report...")
    predictor.generate_excel_report()
    
    # Save CSVs
    print("\nSaving CSV files...")
    predictor.save_csv_predictions()
    
    # Summary statistics
    print("\nGenerating summary statistics...")
    summary = predictor.generate_summary_statistics()
    print(summary.head())
    
    print("\n[SUCCESS] TEST COMPLETED!")


if __name__ == "__main__":
    main()
