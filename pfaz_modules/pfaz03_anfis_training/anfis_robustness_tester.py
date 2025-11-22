# -*- coding: utf-8 -*-
"""
PFAZ 3: ANFIS Robustness Tester
================================

ANFIS-specific robustness testing with iterative training

Features:
- Iterative ANFIS training with outlier removal
- Robustness score calculation
- Performance tracking across iterations
- Excel comprehensive reports
- Outlier analysis

Author: Nuclear Physics AI Training Pipeline
Version: 1.0.0
Date: 2025-10-15
"""

import numpy as np
import pandas as pd
from pathlib import Path
import json
import logging
from typing import Dict, List, Tuple, Optional
from dataclasses import dataclass
import warnings
warnings.filterwarnings('ignore')

try:
    import joblib
    JOBLIB_AVAILABLE = True
except ImportError:
    JOBLIB_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


# ============================================================================
# DATA STRUCTURES
# ============================================================================

@dataclass
class RobustnessIteration:
    """Single robustness iteration result"""
    iteration: int
    n_samples: int
    n_outliers_removed: int
    outlier_indices: List[int]
    train_r2: float
    train_rmse: float
    train_mae: float
    val_r2: float
    val_rmse: float
    val_mae: float
    test_r2: float
    test_rmse: float
    test_mae: float
    robustness_score: float


# ============================================================================
# SIMPLE ANFIS TRAINER (for robustness testing)
# ============================================================================

class SimpleANFISTrainer:
    """Simplified ANFIS for robustness testing"""
    
    def __init__(self, n_mfs: int = 3):
        self.n_mfs = n_mfs
        self.centers = None
        self.sigmas = None
        self.consequent_params = None
        
    def _initialize_fis(self, n_features: int):
        """Initialize fuzzy inference system"""
        self.centers = np.random.randn(self.n_mfs, n_features)
        self.sigmas = np.ones((self.n_mfs, n_features)) * 0.5
        self.consequent_params = np.random.randn(self.n_mfs, n_features + 1) * 0.1
        
    def _membership_function(self, x: np.ndarray, center: np.ndarray, sigma: np.ndarray) -> float:
        """Gaussian membership function"""
        return np.exp(-0.5 * np.sum(((x - center) / sigma) ** 2))
    
    def _forward_pass(self, X: np.ndarray) -> np.ndarray:
        """Forward pass"""
        n_samples = X.shape[0]
        predictions = np.zeros(n_samples)
        
        for i in range(n_samples):
            x = X[i]
            mu = np.array([
                self._membership_function(x, self.centers[j], self.sigmas[j])
                for j in range(self.n_mfs)
            ])
            mu_sum = np.sum(mu) + 1e-10
            w = mu / mu_sum
            x_aug = np.append(x, 1.0)
            y = np.sum([w[j] * np.dot(self.consequent_params[j], x_aug) 
                       for j in range(self.n_mfs)])
            predictions[i] = y
        
        return predictions
    
    def train(self, X_train: np.ndarray, y_train: np.ndarray,
             max_epochs: int = 50, learning_rate: float = 0.01) -> Dict:
        """Train ANFIS"""
        n_features = X_train.shape[1]
        self._initialize_fis(n_features)
        
        for epoch in range(max_epochs):
            y_pred = self._forward_pass(X_train)
            train_loss = np.mean((y_train - y_pred) ** 2)
            
            # Simple gradient update
            for j in range(self.n_mfs):
                grad = -2 * np.mean((y_train - y_pred).reshape(-1, 1) * 
                                   np.column_stack([X_train, np.ones(len(X_train))]), axis=0)
                self.consequent_params[j] -= learning_rate * grad
        
        # Final metrics
        y_pred_final = self._forward_pass(X_train)
        
        from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
        
        return {
            'train_mae': float(mean_absolute_error(y_train, y_pred_final)),
            'train_rmse': float(np.sqrt(mean_squared_error(y_train, y_pred_final))),
            'train_r2': float(r2_score(y_train, y_pred_final))
        }
    
    def predict(self, X: np.ndarray) -> np.ndarray:
        """Make predictions"""
        return self._forward_pass(X)


# ============================================================================
# ANFIS ROBUSTNESS TESTER
# ============================================================================

class ANFISRobustnessTester:
    """
    ANFIS Robustness Tester
    
    Features:
    - Iterative training with outlier removal
    - Robustness score calculation
    - Performance tracking
    - Excel reports
    """
    
    def __init__(self,
                 output_dir: str = 'anfis_robustness_analysis',
                 max_iterations: int = 5,
                 outlier_threshold: float = 3.0):
        """
        Initialize ANFIS Robustness Tester
        
        Args:
            output_dir: Output directory
            max_iterations: Maximum robustness iterations
            outlier_threshold: Standard deviations for outlier detection
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        self.max_iterations = max_iterations
        self.outlier_threshold = outlier_threshold
        
        self.robustness_results = {}
        
        logger.info("=" * 80)
        logger.info("ANFIS ROBUSTNESS TESTER INITIALIZED")
        logger.info("=" * 80)
        logger.info(f"Output directory: {self.output_dir}")
        logger.info(f"Max iterations: {self.max_iterations}")
        logger.info(f"Outlier threshold: {self.outlier_threshold} std")
        logger.info("=" * 80)
    
    def test_robustness(self,
                       X_train: np.ndarray,
                       y_train: np.ndarray,
                       X_val: np.ndarray,
                       y_val: np.ndarray,
                       X_test: np.ndarray,
                       y_test: np.ndarray,
                       config_name: str = 'ANFIS',
                       n_mfs: int = 3) -> List[RobustnessIteration]:
        """
        Test ANFIS robustness with iterative outlier removal
        
        Returns:
            List of RobustnessIteration results
        """
        
        logger.info(f"\n" + "=" * 80)
        logger.info(f"ANFIS ROBUSTNESS TEST: {config_name}")
        logger.info("=" * 80)
        
        iterations = []
        
        # Create copies for iterative removal
        X_train_iter = X_train.copy()
        y_train_iter = y_train.copy()
        all_indices = np.arange(len(y_train))
        remaining_indices = all_indices.copy()
        
        for iteration in range(self.max_iterations):
            logger.info(f"\nIteration {iteration + 1}/{self.max_iterations}")
            logger.info(f"Training samples: {len(X_train_iter)}")
            
            # Train ANFIS
            trainer = SimpleANFISTrainer(n_mfs=n_mfs)
            train_metrics = trainer.train(X_train_iter, y_train_iter, max_epochs=50)
            
            # Predictions
            y_train_pred = trainer.predict(X_train_iter)
            y_val_pred = trainer.predict(X_val)
            y_test_pred = trainer.predict(X_test)
            
            # Calculate errors
            train_errors = y_train_iter - y_train_pred
            
            # Detect outliers
            error_std = np.std(train_errors)
            error_mean = np.mean(train_errors)
            
            outlier_mask = np.abs(train_errors - error_mean) > (self.outlier_threshold * error_std)
            outlier_indices_in_current = np.where(outlier_mask)[0]
            
            # Map to original indices
            outlier_indices_original = remaining_indices[outlier_indices_in_current].tolist()
            
            # Calculate all metrics
            from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
            
            val_mae = float(mean_absolute_error(y_val, y_val_pred))
            val_rmse = float(np.sqrt(mean_squared_error(y_val, y_val_pred)))
            val_r2 = float(r2_score(y_val, y_val_pred))
            
            test_mae = float(mean_absolute_error(y_test, y_test_pred))
            test_rmse = float(np.sqrt(mean_squared_error(y_test, y_test_pred)))
            test_r2 = float(r2_score(y_test, y_test_pred))
            
            # Robustness score (higher is better)
            # Based on: test_r2, consistency (val vs test), and outlier ratio
            consistency_score = 1.0 - abs(val_r2 - test_r2)
            outlier_ratio = len(outlier_indices_in_current) / len(X_train_iter)
            robustness_score = (test_r2 * 0.6) + (consistency_score * 0.3) + ((1 - outlier_ratio) * 0.1)
            
            # Store iteration result
            iter_result = RobustnessIteration(
                iteration=iteration + 1,
                n_samples=len(X_train_iter),
                n_outliers_removed=len(outlier_indices_in_current),
                outlier_indices=outlier_indices_original,
                train_r2=train_metrics['train_r2'],
                train_rmse=train_metrics['train_rmse'],
                train_mae=train_metrics['train_mae'],
                val_r2=val_r2,
                val_rmse=val_rmse,
                val_mae=val_mae,
                test_r2=test_r2,
                test_rmse=test_rmse,
                test_mae=test_mae,
                robustness_score=robustness_score
            )
            
            iterations.append(iter_result)
            
            logger.info(f"Train R²: {train_metrics['train_r2']:.4f}")
            logger.info(f"Val R²:   {val_r2:.4f}")
            logger.info(f"Test R²:  {test_r2:.4f}")
            logger.info(f"Outliers detected: {len(outlier_indices_in_current)}")
            logger.info(f"Robustness score: {robustness_score:.4f}")
            
            # Remove outliers for next iteration
            if len(outlier_indices_in_current) == 0 or iteration == self.max_iterations - 1:
                logger.info(f"\nStopping: No more outliers or max iterations reached")
                break
            
            # Keep only non-outliers
            keep_mask = ~outlier_mask
            X_train_iter = X_train_iter[keep_mask]
            y_train_iter = y_train_iter[keep_mask]
            remaining_indices = remaining_indices[keep_mask]
        
        logger.info("\n" + "=" * 80)
        logger.info(f"ROBUSTNESS TEST COMPLETED: {len(iterations)} iterations")
        logger.info("=" * 80 + "\n")
        
        # Store results
        self.robustness_results[config_name] = iterations
        
        return iterations
    
    def get_best_iteration(self, config_name: str) -> RobustnessIteration:
        """Get best iteration based on robustness score"""
        
        if config_name not in self.robustness_results:
            raise ValueError(f"No results for {config_name}")
        
        iterations = self.robustness_results[config_name]
        best = max(iterations, key=lambda x: x.robustness_score)
        
        return best
    
    def generate_excel_report(self, filename: str = 'ANFIS_Robustness_Report.xlsx'):
        """Generate comprehensive Excel report"""
        
        if not self.robustness_results:
            logger.warning("No robustness results to export")
            return
        
        excel_path = self.output_dir / filename
        
        logger.info(f"Generating Excel report: {excel_path}")
        
        if OPENPYXL_AVAILABLE:
            self._generate_formatted_excel(excel_path)
        else:
            self._generate_simple_excel(excel_path)
        
        logger.info(f"[SUCCESS] Excel report saved: {excel_path}")
    
    def _generate_simple_excel(self, excel_path: Path):
        """Generate simple Excel"""
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            for config_name, iterations in self.robustness_results.items():
                data = []
                
                for iter_result in iterations:
                    data.append({
                        'Iteration': iter_result.iteration,
                        'N_Samples': iter_result.n_samples,
                        'N_Outliers_Removed': iter_result.n_outliers_removed,
                        'Train_R2': iter_result.train_r2,
                        'Train_RMSE': iter_result.train_rmse,
                        'Val_R2': iter_result.val_r2,
                        'Val_RMSE': iter_result.val_rmse,
                        'Test_R2': iter_result.test_r2,
                        'Test_RMSE': iter_result.test_rmse,
                        'Robustness_Score': iter_result.robustness_score
                    })
                
                df = pd.DataFrame(data)
                sheet_name = config_name[:31]
                df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    def _generate_formatted_excel(self, excel_path: Path):
        """Generate formatted Excel with styling"""
        
        wb = Workbook()
        wb.remove(wb.active)
        
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        best_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        
        for config_name, iterations in self.robustness_results.items():
            sheet_name = config_name[:31]
            ws = wb.create_sheet(sheet_name)
            
            # Headers
            headers = [
                'Iteration', 'N_Samples', 'N_Outliers_Removed',
                'Train_R2', 'Train_RMSE', 'Val_R2', 'Val_RMSE',
                'Test_R2', 'Test_RMSE', 'Robustness_Score'
            ]
            ws.append(headers)
            
            # Style headers
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Add data
            best_idx = -1
            best_score = -1
            
            for idx, iter_result in enumerate(iterations):
                row = [
                    iter_result.iteration,
                    iter_result.n_samples,
                    iter_result.n_outliers_removed,
                    round(iter_result.train_r2, 4),
                    round(iter_result.train_rmse, 4),
                    round(iter_result.val_r2, 4),
                    round(iter_result.val_rmse, 4),
                    round(iter_result.test_r2, 4),
                    round(iter_result.test_rmse, 4),
                    round(iter_result.robustness_score, 4)
                ]
                ws.append(row)
                
                if iter_result.robustness_score > best_score:
                    best_score = iter_result.robustness_score
                    best_idx = idx + 2  # +2 for header and 1-indexing
            
            # Highlight best iteration
            if best_idx > 0:
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(best_idx, col_idx).fill = best_fill
            
            # Auto-adjust columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 20)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(excel_path)
    
    def save_summary_report(self):
        """Save summary report as JSON"""
        
        report_file = self.output_dir / 'anfis_robustness_summary.json'
        
        summary = {
            'max_iterations': self.max_iterations,
            'outlier_threshold': self.outlier_threshold,
            'configs_tested': len(self.robustness_results),
            'results': {}
        }
        
        for config_name, iterations in self.robustness_results.items():
            best = self.get_best_iteration(config_name)
            
            summary['results'][config_name] = {
                'total_iterations': len(iterations),
                'best_iteration': best.iteration,
                'best_robustness_score': best.robustness_score,
                'best_test_r2': best.test_r2,
                'initial_samples': iterations[0].n_samples,
                'final_samples': best.n_samples,
                'total_outliers_removed': sum(iter.n_outliers_removed for iter in iterations)
            }
        
        with open(report_file, 'w') as f:
            json.dump(summary, f, indent=2)
        
        logger.info(f"Summary report saved: {report_file}")


# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    """Test execution"""
    
    print("\n" + "=" * 80)
    print("PFAZ 3: ANFIS ROBUSTNESS TESTER - TEST")
    print("=" * 80)
    
    # Create test data
    np.random.seed(42)
    n_samples = 200
    n_features = 4
    
    X = np.random.randn(n_samples, n_features)
    y = X[:, 0] * 2 + X[:, 1] * 1.5 - X[:, 2] + np.random.randn(n_samples) * 0.5
    
    # Add some outliers
    outlier_indices = np.random.choice(n_samples, size=10, replace=False)
    y[outlier_indices] += np.random.randn(10) * 5
    
    # Split data
    n_train = int(0.7 * n_samples)
    n_val = int(0.15 * n_samples)
    
    X_train, y_train = X[:n_train], y[:n_train]
    X_val, y_val = X[n_train:n_train+n_val], y[n_train:n_train+n_val]
    X_test, y_test = X[n_train+n_val:], y[n_train+n_val:]
    
    # Initialize tester
    tester = ANFISRobustnessTester(
        output_dir='test_anfis_robustness',
        max_iterations=5,
        outlier_threshold=3.0
    )
    
    # Test robustness
    print("\nTesting robustness...")
    iterations = tester.test_robustness(
        X_train, y_train,
        X_val, y_val,
        X_test, y_test,
        config_name='CFG001_Grid_2MF',
        n_mfs=2
    )
    
    # Get best iteration
    best = tester.get_best_iteration('CFG001_Grid_2MF')
    print(f"\nBest iteration: {best.iteration}")
    print(f"Robustness score: {best.robustness_score:.4f}")
    print(f"Test R²: {best.test_r2:.4f}")
    
    # Generate reports
    print("\nGenerating Excel report...")
    tester.generate_excel_report()
    
    print("\nSaving summary...")
    tester.save_summary_report()
    
    print("\n[SUCCESS] TEST COMPLETED!")


if __name__ == "__main__":
    main()
