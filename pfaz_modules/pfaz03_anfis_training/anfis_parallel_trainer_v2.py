"""
PFAZ 3: ANFIS Parallel Trainer V2
==================================

Adaptive Neuro-Fuzzy Inference System (ANFIS) parallel trainer
Integrated with the nuclear physics AI pipeline

Features:
- Parallel ANFIS training across multiple datasets
- Simplified Python-based ANFIS (no MATLAB required)
- Multiple configuration support
- Progress tracking
- Comprehensive logging

Author: Nuclear Physics AI Training Pipeline
Version: 2.0.0
Date: 2025-11-22
"""

import numpy as np
import pandas as pd
from pathlib import Path
import json
import logging
import time
from datetime import datetime
from typing import Dict, List, Tuple, Optional, Any
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from itertools import product as iproduct
import warnings
warnings.filterwarnings('ignore')
warnings.filterwarnings('ignore', category=UserWarning, module='sklearn')
warnings.filterwarnings('ignore', message='.*sklearn.utils.parallel.*')

from sklearn.metrics import r2_score, mean_squared_error, mean_absolute_error
from sklearn.preprocessing import StandardScaler

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Minimum val_R2 to save a model .pkl (Poor: <0.5, Failed: <0 — not saved)
R2_MIN_SAVE_THRESHOLD = 0.5


# ============================================================================
# DATA STRUCTURES
# ============================================================================

@dataclass
class ANFISTrainingJob:
    """Single ANFIS training job specification"""
    job_id: str
    config: Dict
    dataset_path: Path
    dataset_name: str
    output_dir: Path
    training_mode: str = 'standard'  # 'pilot', 'advanced', 'advanced_reduced'

@dataclass
class ANFISTrainingResult:
    """ANFIS training result data structure"""
    job_id: str
    config_id: str
    dataset_name: str
    success: bool
    metrics: Optional[Dict] = None
    model_path: Optional[Path] = None
    training_time: Optional[float] = None
    error_message: Optional[str] = None
    training_mode: str = 'standard'


# ============================================================================
# SIMPLE PYTHON ANFIS IMPLEMENTATION
# ============================================================================

def _adaptive_n_mfs(n_inputs: int, n_train: int, requested_mfs: int) -> int:
    """Choose safe n_mfs so that n_rules stays feasible relative to training size.
    Rule-of-thumb: n_rules < n_train / 3  (need at least 3 samples per rule)
    """
    for mfs in range(requested_mfs, 1, -1):
        n_rules = mfs ** n_inputs
        if n_rules < max(4, n_train / 3):
            return mfs
    return 2  # minimum


class TakagiSugenoANFIS:
    """
    First-order Takagi-Sugeno ANFIS (Python implementation, no MATLAB needed).

    Architecture:
    - Grid partition: evenly-spaced MF centers per input
    - MF types: gaussian, bell, triangle, trapezoid
    - SubClust approximation: KMeans-based rule center init
    - Type-1 consequents (linear per rule): p_r * [1, x1..xn]
    - Hybrid learning:
        (a) LSE (ridge) for consequent params -- closed form
        (b) L-BFGS-B for premise params (optional, enabled when n_samples large enough)
    - Internal StandardScaler normalization
    - Early stopping on validation RMSE
    """

    def __init__(self, n_inputs: int, n_mfs: int = 3,
                 mf_type: str = 'gaussian',
                 method: str = 'grid',
                 max_iter: int = 300,
                 patience: int = 30,
                 alpha: float = 1e-2,
                 use_gradient: bool = True,
                 gpu_enabled: bool = False):
        self.n_inputs = n_inputs
        self.n_mfs = n_mfs
        self.mf_type = mf_type.lower()
        self.method = method.lower()
        self.max_iter = max_iter
        self.patience = patience
        self.alpha = alpha          # L2 regularization for LSE
        self.use_gradient = use_gradient
        self.n_rules = n_mfs ** n_inputs if method == 'grid' else n_mfs
        self.scaler = StandardScaler()
        # Parameters (set during fit)
        self.centers = None   # (n_inputs, n_mfs) for grid; (n_rules, n_inputs) for subclust
        self.spreads = None   # same shape
        self.consequent = None  # (n_rules, n_inputs + 1)
        self._is_fitted = False
        # GPU (PyTorch LBFGS premise optimization)
        self.gpu_enabled = gpu_enabled
        self._torch_device = None  # belirlenir fit() icinde

    # ------------------------------------------------------------------ MFs --
    def _gaussian(self, x, c, s):
        s = np.abs(s) + 1e-6
        return np.exp(-0.5 * ((x - c) / s) ** 2)

    def _bell(self, x, c, s):
        s = np.abs(s) + 1e-6
        return 1.0 / (1.0 + np.abs((x - c) / s) ** 4)

    def _triangle(self, x, c, s):
        s = np.abs(s) + 1e-6
        return np.maximum(0.0, 1.0 - np.abs(x - c) / s)

    def _trapezoid(self, x, c, s):
        s = np.abs(s) + 1e-6
        half = s * 0.5
        left  = (x - (c - s)) / (half + 1e-6)
        right = ((c + s) - x) / (half + 1e-6)
        return np.maximum(0.0, np.minimum(1.0, np.minimum(left, right)))

    def _eval_mf(self, x, c, s):
        if self.mf_type == 'gaussian':
            return self._gaussian(x, c, s)
        elif self.mf_type == 'bell':
            return self._bell(x, c, s)
        elif self.mf_type in ('triangle', 'trimf'):
            return self._triangle(x, c, s)
        else:  # trapezoid / trapmf
            return self._trapezoid(x, c, s)

    # -------------------------------------------------------- Initialization --
    def _init_grid(self, X_norm):
        n_inputs = X_norm.shape[1]
        centers = np.zeros((n_inputs, self.n_mfs))
        spreads = np.zeros((n_inputs, self.n_mfs))
        for i in range(n_inputs):
            lo, hi = X_norm[:, i].min(), X_norm[:, i].max()
            if abs(hi - lo) < 1e-8:
                lo, hi = lo - 1.0, hi + 1.0
            centers[i] = np.linspace(lo, hi, self.n_mfs)
            spreads[i] = np.full(self.n_mfs, (hi - lo) / (2 * self.n_mfs + 1e-6))
        self.centers = centers  # (n_inputs, n_mfs)
        self.spreads = spreads

    def _init_subclust(self, X_norm):
        from sklearn.cluster import KMeans
        k = min(self.n_rules, max(2, len(X_norm) // 4))
        self.n_rules = k
        km = KMeans(n_clusters=k, random_state=42, n_init=5)
        km.fit(X_norm)
        self.centers = km.cluster_centers_   # (k, n_inputs)
        # Spread = mean half-distance to nearest centre
        dists = np.sqrt(((X_norm[:, None, :] - self.centers[None, :, :]) ** 2).sum(axis=2))
        nearest = dists.min(axis=1)
        spread_val = max(np.mean(nearest) * 0.5, 0.1)
        self.spreads = np.full((k, X_norm.shape[1]), spread_val)

    # ------------------------------------------------ Firing strengths --
    def _firing_strengths(self, X_norm):
        n = X_norm.shape[0]
        if self.method == 'subclust':
            fs = np.ones((n, self.n_rules))
            for r in range(self.n_rules):
                for i in range(self.n_inputs):
                    fs[:, r] *= self._eval_mf(X_norm[:, i], self.centers[r, i], self.spreads[r, i])
        else:  # grid
            rule_combos = list(iproduct(range(self.n_mfs), repeat=self.n_inputs))
            fs = np.ones((n, len(rule_combos)))
            for r, combo in enumerate(rule_combos):
                for i, j in enumerate(combo):
                    fs[:, r] *= self._eval_mf(X_norm[:, i], self.centers[i, j], self.spreads[i, j])
        total = fs.sum(axis=1, keepdims=True)
        fn = fs / np.where(total < 1e-10, 1.0, total)
        return fs, fn

    # ------------------------------------------------ LSE for consequents --
    def _lse_consequents(self, X_norm, y, fn):
        n = X_norm.shape[0]
        n_rules = fn.shape[1]
        X_aug = np.column_stack([np.ones(n), X_norm])  # (n, n_inputs+1)
        # A: (n, n_rules * (n_inputs+1))
        A = np.zeros((n, n_rules * (self.n_inputs + 1)))
        for r in range(n_rules):
            w = fn[:, r:r + 1]
            A[:, r * (self.n_inputs + 1):(r + 1) * (self.n_inputs + 1)] = w * X_aug
        # Ridge regression: (A^T A + alpha*I) theta = A^T y
        ATA = A.T @ A + self.alpha * np.eye(A.shape[1])
        ATy = A.T @ y.ravel()
        try:
            theta = np.linalg.solve(ATA, ATy)
        except np.linalg.LinAlgError:
            theta, _, _, _ = np.linalg.lstsq(A, y.ravel(), rcond=None)
        return theta.reshape(n_rules, self.n_inputs + 1)

    def _infer(self, X_norm, fn):
        n = X_norm.shape[0]
        X_aug = np.column_stack([np.ones(n), X_norm])
        rule_out = X_aug @ self.consequent.T   # (n, n_rules)
        return (fn * rule_out).sum(axis=1)

    # ------------------------------------------------ Gradient optimization --
    def _pack(self):
        if self.method == 'subclust':
            return np.concatenate([self.centers.ravel(), self.spreads.ravel()])
        return np.concatenate([self.centers.ravel(), self.spreads.ravel()])

    def _unpack(self, v, n_inputs):
        sz = self.n_rules * n_inputs if self.method == 'subclust' else n_inputs * self.n_mfs
        c_flat = v[:sz].reshape(self.centers.shape)
        s_flat = np.abs(v[sz:]).reshape(self.spreads.shape) + 1e-4
        return c_flat, s_flat

    def _objective(self, v, X_norm, y):
        c, s = self._unpack(v, X_norm.shape[1])
        saved_c, saved_s = self.centers, self.spreads
        self.centers, self.spreads = c, s
        _, fn = self._firing_strengths(X_norm)
        self.consequent = self._lse_consequents(X_norm, y, fn)
        y_pred = self._infer(X_norm, fn)
        self.centers, self.spreads = saved_c, saved_s
        return float(np.mean((y.ravel() - y_pred) ** 2))

    # ----------------------------------------------------------- fit / predict --
    def fit(self, X_train, y_train, X_val=None, y_val=None):
        y_train = y_train.ravel()
        if y_val is not None:
            y_val = y_val.ravel()

        # Normalize
        X_norm = self.scaler.fit_transform(X_train)
        X_val_norm = self.scaler.transform(X_val) if X_val is not None else None

        # Initialize MF parameters
        if self.method == 'subclust':
            self._init_subclust(X_norm)
        else:
            self._init_grid(X_norm)

        # Initial LSE pass
        _, fn = self._firing_strengths(X_norm)
        self.consequent = self._lse_consequents(X_norm, y_train, fn)

        # Gradient refinement: GPU (PyTorch LBFGS) veya CPU (scipy L-BFGS-B)
        if self.use_gradient and len(X_train) >= 10:
            _used_gpu = False
            if self.gpu_enabled:
                _used_gpu = self._gradient_torch(X_norm, y_train, X_val_norm, y_val)
            if not _used_gpu:
                self._gradient_scipy(X_norm, y_train, X_val_norm, y_val)

        self._is_fitted = True
        return self

    # ------------------------------------------------ Gradient helpers --

    def _gradient_scipy(self, X_norm, y_train, X_val_norm, y_val):
        """Scipy L-BFGS-B ile one parametresi optimizasyonu (CPU)."""
        try:
            from scipy.optimize import minimize
            best = {'val': float('inf'), 'v': self._pack(), 'c': self.consequent.copy()}
            _patience_count = [0]

            def callback(v):
                if X_val_norm is None:
                    return
                c, s = self._unpack(v, X_norm.shape[1])
                sc, ss = self.centers, self.spreads
                self.centers, self.spreads = c, s
                _, fn_v = self._firing_strengths(X_val_norm)
                self.consequent = self._lse_consequents(X_norm, y_train,
                                                        self._firing_strengths(X_norm)[1])
                y_pred_v = self._infer(X_val_norm, fn_v)
                val_loss = float(np.mean((y_val - y_pred_v) ** 2))
                self.centers, self.spreads = sc, ss
                if val_loss < best['val']:
                    best.update({'val': val_loss, 'v': v.copy(), 'c': self.consequent.copy()})
                    _patience_count[0] = 0
                else:
                    _patience_count[0] += 1

            minimize(self._objective, self._pack(), args=(X_norm, y_train),
                     method='L-BFGS-B', callback=callback,
                     options={'maxiter': self.max_iter, 'ftol': 1e-10, 'gtol': 1e-8})

            self.centers, self.spreads = self._unpack(best['v'], X_norm.shape[1])
            _, fn_best = self._firing_strengths(X_norm)
            self.consequent = self._lse_consequents(X_norm, y_train, fn_best)
        except Exception as e:
            logger.debug(f"Scipy L-BFGS-B hatasi: {e}")
            _, fn = self._firing_strengths(X_norm)
            self.consequent = self._lse_consequents(X_norm, y_train, fn)

    def _gradient_torch(self, X_norm, y_train, X_val_norm, y_val) -> bool:
        """
        PyTorch LBFGS ile one parametresi optimizasyonu (GPU/CPU).
        Basarisizsa False dondurur; ana kod scipy fallback'e gecer.
        """
        try:
            import torch
            dev = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
            if dev.type == 'cpu' and not self.gpu_enabled:
                return False  # GPU isteniyordu, bulunamadi -> scipy'a bırak

            dtype = torch.float32
            Xt = torch.tensor(X_norm, dtype=dtype, device=dev)
            yt = torch.tensor(y_train.ravel(), dtype=dtype, device=dev)

            # Baslangic parametrelerini tensore cevir (ogrenilecek)
            c0 = torch.tensor(self.centers.copy(), dtype=dtype, device=dev)
            s0 = torch.tensor(self.spreads.copy(), dtype=dtype, device=dev)
            centers_p = torch.nn.Parameter(c0)
            spreads_p = torch.nn.Parameter(s0)

            def _torch_forward(c, s, X_t):
                """PyTorch tensorleriyle tam ANFIS forward pass."""
                n = X_t.shape[0]
                eps = torch.tensor(1e-6, dtype=dtype, device=dev)
                if self.method == 'subclust':
                    # s shape: (n_rules, n_inputs)
                    # c shape: (n_rules, n_inputs)
                    diff = X_t.unsqueeze(1) - c.unsqueeze(0)  # (n, n_rules, n_inputs)
                    if self.mf_type == 'gaussian':
                        mf = torch.exp(-0.5 * (diff / (s.abs().unsqueeze(0) + eps)) ** 2)
                    else:
                        mf = torch.exp(-0.5 * (diff / (s.abs().unsqueeze(0) + eps)) ** 2)
                    fs = mf.prod(dim=2)  # (n, n_rules)
                else:
                    # Grid: c shape (n_inputs, n_mfs), s shape (n_inputs, n_mfs)
                    from itertools import product as iproduct2
                    combos = list(iproduct2(range(self.n_mfs), repeat=self.n_inputs))
                    fs_list = []
                    for combo in combos:
                        rule_fs = torch.ones(n, dtype=dtype, device=dev)
                        for i, j in enumerate(combo):
                            xi = X_t[:, i]
                            ci = c[i, j]
                            si = s[i, j].abs() + eps
                            if self.mf_type == 'gaussian':
                                mf_val = torch.exp(-0.5 * ((xi - ci) / si) ** 2)
                            elif self.mf_type == 'bell':
                                mf_val = 1.0 / (1.0 + ((xi - ci) / si).abs() ** 4 + eps)
                            else:
                                mf_val = torch.exp(-0.5 * ((xi - ci) / si) ** 2)
                            rule_fs = rule_fs * mf_val
                        fs_list.append(rule_fs)
                    fs = torch.stack(fs_list, dim=1)  # (n, n_rules)

                total = fs.sum(dim=1, keepdim=True)
                fn = fs / (total.clamp(min=1e-10))  # (n, n_rules)

                # LSE consequents (torch.linalg.lstsq)
                n_rules = fn.shape[1]
                X_aug = torch.cat([torch.ones(n, 1, dtype=dtype, device=dev), X_t], dim=1)
                A_blocks = []
                for r in range(n_rules):
                    A_blocks.append(fn[:, r:r+1] * X_aug)
                A = torch.cat(A_blocks, dim=1)  # (n, n_rules*(n_inputs+1))
                ATA = A.T @ A + self.alpha * torch.eye(A.shape[1], dtype=dtype, device=dev)
                ATy = A.T @ yt
                try:
                    theta = torch.linalg.solve(ATA, ATy)
                except Exception:
                    theta = torch.linalg.lstsq(A, yt.unsqueeze(1)).solution.squeeze()
                cons = theta.reshape(n_rules, self.n_inputs + 1)

                # Inference
                rule_out = X_aug @ cons.T  # (n, n_rules)
                y_pred = (fn * rule_out).sum(dim=1)
                return y_pred, cons, fn

            optimizer = torch.optim.LBFGS(
                [centers_p, spreads_p],
                max_iter=min(self.max_iter, 100),
                tolerance_grad=1e-7,
                tolerance_change=1e-9,
                line_search_fn='strong_wolfe'
            )

            best_loss = [float('inf')]
            best_params = [self._pack(), self.consequent.copy()]

            def closure():
                optimizer.zero_grad()
                y_pred, _, _ = _torch_forward(centers_p, spreads_p, Xt)
                loss = ((yt - y_pred) ** 2).mean()
                loss.backward()
                return loss

            for _ in range(max(1, self.patience // 10)):
                loss_val = optimizer.step(closure)
                if loss_val is not None and float(loss_val) < best_loss[0]:
                    best_loss[0] = float(loss_val)
                    best_params[0] = np.concatenate([
                        centers_p.detach().cpu().numpy().ravel(),
                        spreads_p.detach().cpu().numpy().ravel()
                    ])

            # En iyi parametreleri geri yukle
            self.centers, self.spreads = self._unpack(best_params[0], X_norm.shape[1])
            _, fn_best = self._firing_strengths(X_norm)
            self.consequent = self._lse_consequents(X_norm, y_train, fn_best)
            logger.debug(f"[ANFIS-GPU] torch LBFGS tamamlandi, loss={best_loss[0]:.6f}, device={dev}")
            return True

        except Exception as e:
            logger.debug(f"[ANFIS-GPU] PyTorch LBFGS hatasi (scipy'a geciyor): {e}")
            return False

    def predict(self, X):
        X_norm = self.scaler.transform(X)
        _, fn = self._firing_strengths(X_norm)
        return self._infer(X_norm, fn)


class OutlierDetector:
    """Detect and remove outliers from training data based on model residuals."""

    def __init__(self, iqr_multiplier: float = 3.0, zscore_threshold: float = 3.0):
        self.iqr_multiplier = iqr_multiplier
        self.zscore_threshold = zscore_threshold

    def detect(self, residuals: np.ndarray) -> np.ndarray:
        """Return boolean mask of NON-outlier samples (True = keep)."""
        res = np.asarray(residuals).ravel()
        # IQR method
        q1, q3 = np.percentile(res, 25), np.percentile(res, 75)
        iqr = q3 - q1
        iqr_mask = (res >= q1 - self.iqr_multiplier * iqr) & \
                   (res <= q3 + self.iqr_multiplier * iqr)
        # Z-score method
        mu, sigma = np.mean(res), np.std(res) + 1e-10
        z_mask = np.abs(res - mu) / sigma <= self.zscore_threshold
        # Keep only samples that pass BOTH
        keep_mask = iqr_mask & z_mask
        # Safety: always keep at least 90% of samples (small unique-nucleus datasets)
        if keep_mask.sum() < max(5, int(0.90 * len(res))):
            keep_mask = iqr_mask  # relax to IQR-only
        return keep_mask


# ============================================================================
# ANFIS PARALLEL TRAINER V2
# ============================================================================

class ANFISParallelTrainerV2:
    """
    ANFIS Parallel Trainer V2

    Features:
    - Trains ANFIS models in parallel
    - Supports multiple configurations
    - Multi-dataset training
    - Progress monitoring
    """

    def __init__(self,
                 datasets_dir: str = None,
                 output_dir: str = None,
                 n_workers: int = None,
                 gpu_enabled: bool = False,
                 use_config_manager: bool = True,
                 use_adaptive_strategy: bool = False,
                 use_performance_analyzer: bool = True,
                 save_datasets: bool = True):
        """
        Initialize ANFIS Parallel Trainer

        Args:
            datasets_dir: Directory containing datasets from PFAZ 1
            output_dir: Output directory for trained ANFIS models (PFAZ 3 output)
            n_workers: Number of parallel workers (None = auto)
            gpu_enabled: Enable PyTorch GPU for L-BFGS premise optimization
            use_config_manager: Use ANFIS config manager for 8 FIS configurations
            use_adaptive_strategy: Use adaptive learning strategy (3-stage)
            use_performance_analyzer: Use performance analyzer for detailed metrics
            save_datasets: Save train/val/test datasets in .mat, .csv, .xlsx formats
        """
        self.output_dir = Path(output_dir) if output_dir else Path('trained_anfis_models')
        self.output_dir.mkdir(parents=True, exist_ok=True)

        self.datasets_dir = Path(datasets_dir) if datasets_dir else None
        self.gpu_enabled = gpu_enabled

        # Determine number of workers
        if n_workers is None:
            import multiprocessing
            self.n_workers = max(2, multiprocessing.cpu_count() - 2)
        else:
            self.n_workers = n_workers

        # New features
        self.use_config_manager = use_config_manager
        self.use_adaptive_strategy = use_adaptive_strategy
        self.use_performance_analyzer = use_performance_analyzer
        self.save_datasets = save_datasets

        # Storage
        self.training_results = []
        self.failed_jobs = []

        # Thread-safe kernel usage tracker
        import threading
        self.kernel_usage_tracker = {}
        self.kernel_tracker_lock = threading.Lock()

        logger.info("=" * 80)
        logger.info("ANFIS PARALLEL TRAINER V2 INITIALIZED")
        logger.info("=" * 80)
        logger.info(f"Datasets directory: {self.datasets_dir}")
        logger.info(f"Output directory: {self.output_dir}")
        logger.info(f"Workers: {self.n_workers}")
        logger.info(f"Config Manager: {self.use_config_manager}")
        logger.info(f"Adaptive Strategy: {self.use_adaptive_strategy}")
        logger.info(f"Performance Analyzer: {self.use_performance_analyzer}")
        logger.info(f"Save Datasets: {self.save_datasets}")
        logger.info("=" * 80)

    def _get_feature_set_from_name(self, dataset_name: str) -> List[str]:
        """Feature set kolonlarini dataset adindan belirle.

        Yeni format: {TARGET}_{SIZE}_{SCENARIO}_{FEATURE_CODE}_{SCALING}_{SAMPLING}[_NoAnomaly]
        Dataset CSV'i zaten sadece ilgili feature kolonlarini iceriyor.
        None dondurerek adaptive selection kullanilir.

        NOT: substring matching YAPILMAZ - 'AZN' kodu 'AZNNP' icinde de bulunur,
        yanlis esleme olusturur.
        """
        logger.debug(f"Adaptive feature selection for: {dataset_name}")
        return None

    def _load_split_csv(self, dataset_path: Path, split: str) -> pd.DataFrame:
        """
        Pre-split CSV'yi yukle.
        Yeni format: headerless (baslık YOK) — sutun adlari metadata.json'dan.
        Eski format: baslik satirli (geriye donuk uyumluluk).
        """
        csv_file = dataset_path / f"{split}.csv"
        if not csv_file.exists():
            raise FileNotFoundError(f"Split file not found: {csv_file}")

        # metadata.json'dan sutun adlarini al
        col_names = None
        meta_file = dataset_path / 'metadata.json'
        if meta_file.exists():
            try:
                import json as _json
                with open(meta_file, encoding='utf-8') as _f:
                    meta = _json.load(_f)
                feat = meta.get('feature_names') or meta.get('feature_columns', [])
                tgt  = meta.get('target_names')  or meta.get('target_columns',  [])
                if feat and tgt:
                    col_names = list(feat) + list(tgt)
            except Exception:
                pass

        if col_names:
            # Yeni headerless format
            return pd.read_csv(csv_file, header=None, names=col_names)
        else:
            # Eski baslikli format (geriye donuk uyumluluk)
            return pd.read_csv(csv_file)

    def _clean_df(self, df: pd.DataFrame) -> pd.DataFrame:
        """Basic string/NaN cleaning on a dataframe."""
        df = df.replace(['unknown', 'Unknown', 'UNKNOWN'], np.nan)
        for col in df.columns:
            if df[col].dtype == 'object':
                df[col] = df[col].astype(str).str.replace('\u2212', '-', regex=False)
                df[col] = df[col].replace(['nan', '', 'NaN'], np.nan)
        return df

    def _detect_target_cols(self, df: pd.DataFrame, dataset_name: str) -> List[str]:
        """Identify target column(s) from dataframe.
        New dataset format: {TARGET}_{SIZE}_{SCENARIO}_{FEATURE_CODE}_{SCALING}_{SAMPLING}
        Actual column names: 'MAGNETIC MOMENT [?]', 'QUADRUPOLE MOMENT [Q]', 'Beta_2'
        Also supports legacy short names: MM, Q, Beta_2
        """
        # Canonical target name from dataset name (handles Beta_2 two-part prefix)
        name = dataset_name
        if name.startswith('Beta_2_'):
            target_key = 'Beta_2'
        elif name.startswith('MM_QM_'):
            target_key = 'MM_QM'
        elif name.startswith('MM_'):
            target_key = 'MM'
        elif name.startswith('QM_'):
            target_key = 'QM'
        else:
            target_key = None

        # Candidate columns per target (long and short forms)
        candidates_map = {
            'MM':    [c for c in df.columns if 'MAGNETIC MOMENT' in c or c == 'MM'],
            'QM':    [c for c in df.columns if 'QUADRUPOLE MOMENT' in c or c in ('Q', 'QM')],
            'Beta_2': [c for c in df.columns if c == 'Beta_2'],
            'MM_QM': [c for c in df.columns if 'MAGNETIC MOMENT' in c or c == 'MM'] +
                     [c for c in df.columns if 'QUADRUPOLE MOMENT' in c or c in ('Q', 'QM')],
        }

        if target_key and target_key in candidates_map:
            target_cols = [c for c in candidates_map[target_key] if c in df.columns]
            if target_cols:
                return target_cols

        # Fallback: try all known target columns
        fallback = []
        for col in df.columns:
            if 'MAGNETIC MOMENT' in col or 'QUADRUPOLE MOMENT' in col or col in ('MM', 'Q', 'QM', 'Beta_2'):
                fallback.append(col)
        return fallback

    def load_dataset(self, dataset_path: Path):
        """Load pre-split train/val/test CSVs from dataset directory.
        Returns:
            X_train, y_train, X_val, y_val, X_test, y_test,
            nucleus_names  (dict with keys 'train', 'val', 'test' → list[str]),
            feature_cols   (list[str])
        """
        dataset_name = dataset_path.name

        train_df = self._clean_df(self._load_split_csv(dataset_path, 'train'))
        val_df   = self._clean_df(self._load_split_csv(dataset_path, 'val'))
        test_df  = self._clean_df(self._load_split_csv(dataset_path, 'test'))

        logger.info(f"Loaded splits: train={len(train_df)}, val={len(val_df)}, test={len(test_df)}")
        logger.info(f"Columns: {list(train_df.columns)}")

        # Identify target columns
        target_cols = self._detect_target_cols(train_df, dataset_name)
        if not target_cols:
            logger.error(f"Available columns: {list(train_df.columns)}")
            raise ValueError(f"No target columns found in {dataset_path}")
        logger.info(f"Target columns: {target_cols}")

        # Adaptive feature selection: all non-target, non-NUCLEUS columns
        feature_cols = [col for col in train_df.columns if col not in target_cols and col != 'NUCLEUS']

        def prepare_split(df):
            all_cols = feature_cols + target_cols
            for col in all_cols:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors='coerce')
            df_clean = df[all_cols].dropna()
            X = df_clean[feature_cols].values.astype(np.float32)
            y = df_clean[target_cols].values.astype(np.float32)
            # Nucleus names — re-align with cleaned rows
            if 'NUCLEUS' in df.columns:
                nucleus = df.loc[df_clean.index, 'NUCLEUS'].astype(str).tolist() \
                          if hasattr(df_clean, 'index') else df['NUCLEUS'].astype(str).tolist()
            else:
                nucleus = [f"sample_{i}" for i in range(len(X))]
            return X, y, nucleus

        X_train, y_train, nuc_train = prepare_split(train_df)
        X_val,   y_val,   nuc_val   = prepare_split(val_df)
        X_test,  y_test,  nuc_test  = prepare_split(test_df)

        logger.info(f"After cleaning: train={len(X_train)}, val={len(X_val)}, test={len(X_test)}")

        if len(X_train) == 0:
            raise ValueError(f"No valid training data in {dataset_path}")

        nucleus_names = {'train': nuc_train, 'val': nuc_val, 'test': nuc_test}
        return X_train, y_train, X_val, y_val, X_test, y_test, nucleus_names, feature_cols

    def calculate_metrics(self, y_true: np.ndarray, y_pred: np.ndarray) -> Dict:
        """Calculate regression metrics"""

        if y_true.ndim > 1:
            y_true = y_true.flatten()
        if y_pred.ndim > 1:
            y_pred = y_pred.flatten()

        metrics = {
            'r2': float(r2_score(y_true, y_pred)),
            'rmse': float(np.sqrt(mean_squared_error(y_true, y_pred))),
            'mae': float(mean_absolute_error(y_true, y_pred))
        }

        return metrics

    def _build_single_anfis(self, X_tr, n_inputs: int, config: Dict) -> TakagiSugenoANFIS:
        """Instantiate a TakagiSugenoANFIS with adaptive n_mfs for one output."""
        method = config.get('method', 'grid')
        requested_mfs = config.get('n_mfs', 3)
        safe_mfs = _adaptive_n_mfs(n_inputs, len(X_tr), requested_mfs) if method == 'grid' \
                   else config.get('n_mfs', 6)
        return TakagiSugenoANFIS(
            n_inputs=n_inputs,
            n_mfs=safe_mfs,
            mf_type=config.get('mf_type', 'gaussian'),
            method=method,
            max_iter=config.get('max_iter', 300),
            patience=30,
            alpha=1e-2,
            use_gradient=config.get('use_gradient', True),
            gpu_enabled=getattr(self, 'gpu_enabled', False),
        )

    def _train_anfis_on_data(self, X_tr, y_tr, X_vl, y_vl, config: Dict):
        """Train ANFIS. Handles multi-output by training one model per output.
        Returns either a single TakagiSugenoANFIS or a list of them (multi-output).
        """
        n_inputs = X_tr.shape[1]
        y_tr = np.asarray(y_tr)
        y_vl = np.asarray(y_vl) if y_vl is not None else None

        if y_tr.ndim == 1 or y_tr.shape[1] == 1:
            # Single output
            y_tr_1d = y_tr.ravel()
            y_vl_1d = y_vl.ravel() if y_vl is not None else None
            model = self._build_single_anfis(X_tr, n_inputs, config)
            model.fit(X_tr, y_tr_1d, X_val=X_vl, y_val=y_vl_1d)
            return model
        else:
            # Multi-output: train one ANFIS per column
            n_outputs = y_tr.shape[1]
            models = []
            for col in range(n_outputs):
                y_col = y_tr[:, col]
                y_vl_col = y_vl[:, col] if y_vl is not None else None
                m = self._build_single_anfis(X_tr, n_inputs, config)
                m.fit(X_tr, y_col, X_val=X_vl, y_val=y_vl_col)
                models.append(m)
            return models   # list of models, one per output

    def train_single_anfis(self, job: ANFISTrainingJob) -> ANFISTrainingResult:
        """Train single ANFIS model with outlier detection and metadata logging."""
        try:
            import joblib
            start_time = time.time()

            # ---- Load pre-split data -------------------------------------------
            X_train, y_train, X_val, y_val, X_test, y_test, nucleus_names, feature_cols = \
                self.load_dataset(job.dataset_path)

            n_inputs  = X_train.shape[1]
            mf_type   = job.config.get('mf_type', 'gaussian')
            method    = job.config.get('method', 'grid')
            n_mfs_req = job.config.get('n_mfs', 3)
            safe_mfs  = _adaptive_n_mfs(n_inputs, len(X_train), n_mfs_req) if method == 'grid' else n_mfs_req
            n_rules   = safe_mfs ** n_inputs if method == 'grid' else safe_mfs

            logger.info(f"Training ANFIS: {job.job_id} | "
                        f"inputs={n_inputs} | MF={mf_type} | method={method} | "
                        f"n_mfs={safe_mfs} | n_rules={n_rules} | "
                        f"train_samples={len(X_train)}")

            # ---- Helper: predict with single or multi-output model -------------
            def model_predict(model, X):
                if isinstance(model, list):
                    # Multi-output: stack predictions column-wise
                    preds = np.column_stack([m.predict(X) for m in model])
                    return preds
                return model.predict(X)

            # ---- Scale features (ANFIS is sensitive to unscaled inputs) --------
            from sklearn.preprocessing import StandardScaler as _SS
            self._input_scaler = _SS()
            X_train = self._input_scaler.fit_transform(X_train)
            X_val   = self._input_scaler.transform(X_val)
            X_test  = self._input_scaler.transform(X_test)

            # ---- Train on full training set ------------------------------------
            anfis = self._train_anfis_on_data(X_train, y_train, X_val, y_val, job.config)
            y_train_pred_full = model_predict(anfis, X_train)

            # ---- Outlier detection on training residuals ----------------------
            # Use first output column (or only column) for outlier detection
            y_ref = y_train.ravel() if y_train.ndim == 1 else y_train[:, 0]
            pred_ref = y_train_pred_full.ravel() if y_train_pred_full.ndim == 1 \
                       else y_train_pred_full[:, 0]
            residuals = y_ref - pred_ref
            detector = OutlierDetector(iqr_multiplier=3.0, zscore_threshold=3.0)
            keep_mask = detector.detect(residuals)
            n_outliers = int((~keep_mask).sum())

            best_anfis = anfis
            outlier_cleaned = False
            outlier_nuclei = []

            if n_outliers > 0 and keep_mask.sum() >= max(5, int(0.9 * len(X_train))):
                logger.info(f"  Outliers detected: {n_outliers} / {len(X_train)} samples — retraining without them")
                outlier_nuclei = [nucleus_names['train'][i] for i in range(len(keep_mask)) if not keep_mask[i]]
                X_clean = X_train[keep_mask]
                y_clean = y_train[keep_mask] if y_train.ndim == 1 else y_train[keep_mask, :]
                y_vl_clean = y_val[keep_mask] if len(y_val) == len(keep_mask) else y_val
                anfis_clean = self._train_anfis_on_data(X_clean, y_clean, X_val, y_val, job.config)

                # Compare val performance (first output)
                y_val_ref = y_val.ravel() if y_val.ndim == 1 else y_val[:, 0]
                pred_orig  = model_predict(anfis, X_val)
                pred_clean = model_predict(anfis_clean, X_val)
                r2_orig  = float(r2_score(y_val_ref,
                                          pred_orig.ravel() if pred_orig.ndim == 1 else pred_orig[:, 0]))
                r2_clean = float(r2_score(y_val_ref,
                                          pred_clean.ravel() if pred_clean.ndim == 1 else pred_clean[:, 0]))

                if r2_clean > r2_orig:
                    logger.info(f"  Clean model better: val_R2 {r2_orig:.4f} -> {r2_clean:.4f}")
                    best_anfis = anfis_clean
                    outlier_cleaned = True
                else:
                    logger.info(f"  Original model kept: val_R2 {r2_orig:.4f} >= clean {r2_clean:.4f}")

            # ---- Final evaluation ---------------------------------------------
            y_train_pred = model_predict(best_anfis, X_train)
            y_val_pred   = model_predict(best_anfis, X_val)
            y_test_pred  = model_predict(best_anfis, X_test)

            train_metrics = self.calculate_metrics(y_train, y_train_pred)
            val_metrics   = self.calculate_metrics(y_val,   y_val_pred)
            test_metrics  = self.calculate_metrics(y_test,  y_test_pred)

            # ---- Divergence detection -----------------------------------------
            DIVERGENCE_R2_THRESHOLD = -2.0
            val_r2_value = val_metrics.get('r2', 0.0)
            if isinstance(val_r2_value, float) and val_r2_value < DIVERGENCE_R2_THRESHOLD:
                logger.warning(f"[DIVERGED] {job.job_id} | val_R2={val_r2_value:.4f} < {DIVERGENCE_R2_THRESHOLD} — skipping save")
                return ANFISTrainingResult(
                    job_id=job.job_id,
                    config_id=job.config['id'],
                    dataset_name=job.dataset_name,
                    success=False,
                    error_message=f"ANFIS diverged: val_R2={val_r2_value:.4f}"
                )

            # ---- Quality filter -----------------------------------------------
            if isinstance(val_r2_value, float) and val_r2_value < R2_MIN_SAVE_THRESHOLD:
                logger.warning(
                    f"[POOR] {job.job_id} | val_R2={val_r2_value:.4f} < {R2_MIN_SAVE_THRESHOLD} "
                    f"(Poor/Failed category) — skipping save"
                )
                return ANFISTrainingResult(
                    job_id=job.job_id,
                    config_id=job.config['id'],
                    dataset_name=job.dataset_name,
                    success=False,
                    metrics={'train': train_metrics, 'val': val_metrics, 'test': test_metrics},
                    error_message=f"Poor/Failed: val_R2={val_r2_value:.4f} < {R2_MIN_SAVE_THRESHOLD}",
                    training_mode=job.training_mode,
                )

            # ---- Build metadata -----------------------------------------------
            training_meta = {
                'dataset_name':              job.dataset_name,
                'config_id':                 job.config['id'],
                'mf_type':                   mf_type,
                'method':                    method,
                'n_inputs':                  n_inputs,
                'feature_cols':              feature_cols,
                'n_mfs_per_input':           int(safe_mfs),
                'n_rules':                   int(n_rules),
                'n_train':                   int(len(X_train)),
                'n_val':                     int(len(X_val)),
                'n_test':                    int(len(X_test)),
                'n_outliers_detected':       n_outliers,
                'outlier_cleaning_applied':  outlier_cleaned,
                'nucleus_train':             nucleus_names['train'],
                'nucleus_val':               nucleus_names['val'],
                'nucleus_test':              nucleus_names['test'],
                'nucleus_outliers_removed':  outlier_nuclei,
                'training_mode':             job.training_mode,
            }

            metrics = {
                'train': train_metrics,
                'val':   val_metrics,
                'test':  test_metrics,
                'training_meta': training_meta,
            }

            # ---- Save model ---------------------------------------------------
            job.output_dir.mkdir(parents=True, exist_ok=True)
            model_path = job.output_dir / f"model_{job.config['id']}.pkl"
            joblib.dump(best_anfis, model_path)

            # Save metrics JSON
            metrics_file = job.output_dir / f"metrics_{job.config['id']}.json"
            with open(metrics_file, 'w') as f:
                json.dump(metrics, f, indent=2)

            # ---- Workspace + FIS kaydi (ANFISModelSaver) --------------------
            # Egitim girdisi CSV; CIKTI olarak workspace ve FIS dosyalari kaydedilir.
            # workspace: FIS yapisi + egitim/dogrulama hatalari + metrikler (.mat)
            # fis      : Sadece FIS yapisi (.mat)
            try:
                from pfaz_modules.pfaz03_anfis_training.anfis_model_saver import ANFISModelSaver

                # FIS veri yapisi: Python ANFIS nesnesinden cikar
                fis_data = {
                    'n_inputs':    int(best_anfis.n_inputs),
                    'n_mfs':       int(best_anfis.n_mfs),
                    'n_rules':     int(best_anfis.n_rules),
                    'mf_type':     best_anfis.mf_type,
                    'method':      best_anfis.method,
                    'centers':     best_anfis.centers if best_anfis.centers is not None else np.array([]),
                    'spreads':     best_anfis.spreads if best_anfis.spreads is not None else np.array([]),
                    'consequent':  best_anfis.consequent if best_anfis.consequent is not None else np.array([]),
                    'scaler_mean': best_anfis.scaler.mean_ if hasattr(best_anfis.scaler, 'mean_') else np.array([]),
                    'scaler_std':  best_anfis.scaler.scale_ if hasattr(best_anfis.scaler, 'scale_') else np.array([]),
                }

                model_saver = ANFISModelSaver(base_dir=str(job.output_dir))
                saver_result = model_saver.save_anfis_model(
                    model_data={
                        'fis': fis_data,
                        'training_error':   getattr(best_anfis, 'train_errors_',   []),
                        'validation_error': getattr(best_anfis, 'val_errors_',     []),
                        'metrics': metrics,
                        'outliers': None,
                    },
                    config_name=job.config['id'],
                    dataset_info={
                        'dataset_name': job.dataset_name,
                        'n_train': int(len(X_train)),
                        'n_val':   int(len(X_val)),
                        'n_test':  int(len(X_test)),
                        'n_inputs': int(n_inputs),
                        'n_rules':  int(n_rules),
                        'mf_type':  mf_type,
                        'method':   method,
                    },
                    output_dir=str(job.output_dir)
                )
                _files = saver_result.get('files', {})
                logger.info(f"  [OK] Workspace: {Path(_files.get('workspace', '?')).name}  "
                            f"FIS: {Path(_files.get('fis', '?')).name}")
            except Exception as e_saver:
                logger.warning(f"  [WARNING] ANFISModelSaver basarisiz (devam): {e_saver}")

            # ---- Track kernel usage -------------------------------------------
            if self.save_datasets:
                self.save_training_datasets(
                    X_train, y_train, X_val, y_val, X_test, y_test,
                    dataset_name=job.dataset_name,
                    config_id=job.config['id']
                )
                self.track_kernel_usage(
                    dataset_name=job.dataset_name,
                    config_id=job.config['id'],
                    kernel_info={
                        'n_train':    len(X_train),
                        'n_val':      len(X_val),
                        'n_test':     len(X_test),
                        'n_features': n_inputs,
                        'n_rules':    n_rules,
                        'mf_type':    mf_type,
                        'method':     method,
                    }
                )

            training_time = time.time() - start_time
            logger.info(f"[SUCCESS] {job.job_id} | "
                        f"val_R2={val_metrics['r2']:.4f} | "
                        f"test_R2={test_metrics['r2']:.4f} | "
                        f"outliers={n_outliers} | {training_time:.1f}s")

            return ANFISTrainingResult(
                job_id=job.job_id,
                config_id=job.config['id'],
                dataset_name=job.dataset_name,
                success=True,
                metrics=metrics,
                model_path=model_path,
                training_time=training_time,
                training_mode=job.training_mode,
            )

        except Exception as e:
            logger.error(f"[ERROR] {job.job_id} | Error: {str(e)}", exc_info=False)
            return ANFISTrainingResult(
                job_id=job.job_id,
                config_id=job.config['id'],
                dataset_name=job.dataset_name,
                success=False,
                error_message=str(e),
                training_mode=getattr(job, 'training_mode', 'standard'),
            )

    def create_anfis_configs(self, n_configs: int = 8) -> List[Dict]:
        """Create ANFIS training configurations.

        Each config specifies:
          - id          : unique string id
          - mf_type     : 'gaussian' | 'bell' | 'triangle' | 'trapezoid'
          - method      : 'grid' | 'subclust'
          - n_mfs       : membership functions per input (grid only)
          - max_iter    : gradient optimization iterations
          - use_gradient: whether to run L-BFGS-B premise refinement
        """
        all_configs = [
            # --- Best performers first (trapezoid and bell 2MF, empirically validated) ---
            {'id': 'CFG_Grid_2MF_Trap',   'mf_type': 'trapezoid',   'method': 'grid',     'n_mfs': 2, 'max_iter': 300, 'use_gradient': True},
            {'id': 'CFG_Grid_2MF_Bell',   'mf_type': 'bell',        'method': 'grid',     'n_mfs': 2, 'max_iter': 300, 'use_gradient': True},
            {'id': 'CFG_Grid_2MF_Gauss',  'mf_type': 'gaussian',    'method': 'grid',     'n_mfs': 2, 'max_iter': 300, 'use_gradient': True},
            {'id': 'CFG_Grid_2MF_Tri',    'mf_type': 'triangle',    'method': 'grid',     'n_mfs': 2, 'max_iter': 300, 'use_gradient': True},
            # --- 3MF configs ---
            {'id': 'CFG_Grid_3MF_Bell',   'mf_type': 'bell',        'method': 'grid',     'n_mfs': 3, 'max_iter': 400, 'use_gradient': True},
            {'id': 'CFG_Grid_3MF_Gauss',  'mf_type': 'gaussian',    'method': 'grid',     'n_mfs': 3, 'max_iter': 400, 'use_gradient': True},
            # --- SubClust configs ---
            {'id': 'CFG_SubClust_5',      'mf_type': 'gaussian',    'method': 'subclust', 'n_mfs': 5, 'max_iter': 300, 'use_gradient': True},
            {'id': 'CFG_SubClust_8',      'mf_type': 'gaussian',    'method': 'subclust', 'n_mfs': 8, 'max_iter': 300, 'use_gradient': True},
        ]
        return all_configs[:n_configs]

    def discover_datasets(self) -> List[Path]:
        """Discover dataset directories (excluding reports and metadata)"""

        if self.datasets_dir is None or not self.datasets_dir.exists():
            raise ValueError(f"Datasets directory not found: {self.datasets_dir}")

        dataset_paths = []

        # Directories to exclude (reports, metadata, logs, etc.)
        EXCLUDE_DIRS = {
            'quality_reports',
            'validation_reports',
            'logs',
            'reports',
            'metadata',
            'temp',
            'cache',
            '__pycache__'
        }

        for subdir in self.datasets_dir.iterdir():
            if subdir.is_dir():
                # Skip excluded directories
                if subdir.name in EXCLUDE_DIRS:
                    logger.info(f"  Skipping non-dataset directory: {subdir.name}")
                    continue

                has_data = (
                    list(subdir.glob('*.csv')) or
                    list(subdir.glob('*.xlsx')) or
                    list(subdir.glob('*.tsv'))
                )

                if has_data:
                    dataset_paths.append(subdir)
                    logger.info(f"  Found dataset: {subdir.name}")

        return dataset_paths

    def _run_parallel_wave(self, jobs: List[ANFISTrainingJob], start_time: float,
                           wave_name: str = 'WAVE') -> List[ANFISTrainingResult]:
        """Run a batch of ANFIS training jobs in parallel and return all results."""
        if not jobs:
            return []
        results = []
        with ThreadPoolExecutor(max_workers=self.n_workers) as executor:
            future_to_job = {executor.submit(self.train_single_anfis, job): job for job in jobs}
            completed = 0
            for future in as_completed(future_to_job):
                job = future_to_job[future]
                try:
                    result = future.result()
                    results.append(result)
                    completed += 1
                    if completed % 5 == 0:
                        elapsed = time.time() - start_time
                        logger.info(f"[{wave_name}] {completed}/{len(jobs)} | {elapsed/60:.1f}m")
                except Exception as e:
                    logger.error(f"[{wave_name}] Job failed: {job.job_id} | {str(e)}")
                    results.append(ANFISTrainingResult(
                        job_id=job.job_id, config_id=job.config['id'],
                        dataset_name=job.dataset_name, success=False,
                        error_message=str(e), training_mode=getattr(job, 'training_mode', 'standard'),
                    ))
        return results

    def train_all_anfis_parallel(self, n_configs: int = 10) -> Dict:
        """
        Main entry point for training all ANFIS models.

        3-Phase soft strategy:
          Wave 1 (PILOT)   : 4x 2MF configs run for all datasets
          Wave 2 (ADVANCED): 4x 3MF+SubClust configs; if best pilot val_R2 < 0.3,
                             max_iter reduced to 100 (soft — no elimination)

        Args:
            n_configs: Number of configurations to use (max 8)
        """
        logger.info("\n" + "=" * 80)
        logger.info("ANFIS PARALLEL TRAINING (3-PHASE SOFT STRATEGY)")
        logger.info("=" * 80)

        configs = self.create_anfis_configs(n_configs)
        logger.info(f"Using {len(configs)} ANFIS configurations")

        dataset_paths = self.discover_datasets()
        logger.info(f"Found {len(dataset_paths)} datasets")

        # Split configs into pilot (2MF) and advanced (3MF + SubClust)
        PILOT_IDS = {'CFG_Grid_2MF_Trap', 'CFG_Grid_2MF_Bell',
                     'CFG_Grid_2MF_Gauss', 'CFG_Grid_2MF_Tri'}
        pilot_configs   = [c for c in configs if c['id'] in PILOT_IDS]
        advanced_configs = [c for c in configs if c['id'] not in PILOT_IDS]
        if not pilot_configs:
            pilot_configs, advanced_configs = configs, []

        start_time = time.time()
        results: List[ANFISTrainingResult] = []
        skipped_existing = 0

        # ---- Wave 1: Pilot -----------------------------------------------
        pilot_jobs = []
        for dp in dataset_paths:
            for cfg in pilot_configs:
                out_dir = self.output_dir / dp.name / cfg['id']
                if (out_dir / f"model_{cfg['id']}.pkl").exists():
                    skipped_existing += 1
                    continue
                pilot_jobs.append(ANFISTrainingJob(
                    job_id=f"{dp.name}_ANFIS_{cfg['id']}",
                    config=cfg, dataset_path=dp,
                    dataset_name=dp.name, output_dir=out_dir,
                    training_mode='pilot',
                ))

        logger.info(f"Wave 1 PILOT: {len(pilot_jobs)} jobs ({skipped_existing} skipped)")
        pilot_results = self._run_parallel_wave(pilot_jobs, start_time, 'PILOT')
        results.extend(pilot_results)

        # Compute best pilot val_R2 per dataset
        pilot_best_r2: Dict[str, float] = {}
        for r in pilot_results:
            if r.success and r.metrics:
                vr2 = r.metrics.get('val', {}).get('r2', -999.0)
                ds  = r.dataset_name
                if vr2 > pilot_best_r2.get(ds, -999.0):
                    pilot_best_r2[ds] = vr2

        # ---- Wave 2: Advanced --------------------------------------------
        advanced_jobs = []
        for dp in dataset_paths:
            best_r2 = pilot_best_r2.get(dp.name, 0.0)
            for cfg in advanced_configs:
                out_dir = self.output_dir / dp.name / cfg['id']
                if (out_dir / f"model_{cfg['id']}.pkl").exists():
                    skipped_existing += 1
                    continue
                cfg_copy = dict(cfg)
                if best_r2 < 0.3:
                    cfg_copy['max_iter'] = 100
                    mode = 'advanced_reduced'
                else:
                    mode = 'advanced'
                advanced_jobs.append(ANFISTrainingJob(
                    job_id=f"{dp.name}_ANFIS_{cfg['id']}",
                    config=cfg_copy, dataset_path=dp,
                    dataset_name=dp.name, output_dir=out_dir,
                    training_mode=mode,
                ))

        logger.info(f"Wave 2 ADVANCED: {len(advanced_jobs)} jobs")
        advanced_results = self._run_parallel_wave(advanced_jobs, start_time, 'ADVANCED')
        results.extend(advanced_results)

        if skipped_existing:
            logger.info(f"[RESUME] {skipped_existing} ANFIS jobs skipped (model file exists)")

        total_time = time.time() - start_time
        failed = [r for r in results if not r.success]
        successful = len(results) - len(failed)

        logger.info("\n" + "=" * 80)
        logger.info("ANFIS PARALLEL TRAINING COMPLETED")
        logger.info("=" * 80)
        logger.info(f"Total jobs: {len(results)}")
        logger.info(f"Successful: {successful}")
        logger.info(f"Failed: {len(failed)}")
        logger.info(f"Total time: {total_time/60:.2f} minutes")
        logger.info("=" * 80)

        self.training_results = results
        self.failed_jobs = failed

        # Save summary
        self.save_summary_report()
        self.generate_comparison_excel()

        # ---- ANFISPerformanceAnalyzer: config karsilastirma raporu ----
        try:
            from pfaz_modules.pfaz03_anfis_training.anfis_performance_analyzer import ANFISPerformanceAnalyzer
            _perf_dir = self.output_dir / 'performance_analysis'
            _perf = ANFISPerformanceAnalyzer(
                trained_models_dir=str(self.output_dir),
                output_dir=str(_perf_dir)
            )
            _perf.load_training_results()
            _perf.generate_excel_report()
            logger.info("[OK] ANFISPerformanceAnalyzer: Excel raporu olusturuldu -> performance_analysis/")
        except Exception as _e:
            logger.warning(f"[WARNING] ANFISPerformanceAnalyzer basarisiz (devam): {_e}")

        # ---- ANFISVisualizer: target bazli karsilastirma grafigi ----
        try:
            from pfaz_modules.pfaz03_anfis_training.anfis_visualizer import ANFISVisualizer
            _results_by_target: dict = {}
            for _r in self.training_results:
                if not _r.success or not _r.metrics:
                    continue
                _ds = _r.dataset_name
                if   _ds.startswith('Beta_2'): _tgt = 'Beta_2'
                elif _ds.startswith('MM_QM'):  _tgt = 'MM_QM'
                elif _ds.startswith('MM'):     _tgt = 'MM'
                elif _ds.startswith('QM'):     _tgt = 'QM'
                else:                          _tgt = _ds.split('_')[0]
                _tm = _r.metrics.get('test', {})
                _results_by_target.setdefault(_tgt, []).append({
                    'metrics': {
                        'test': {
                            'R2':   _tm.get('r2',   0.0),
                            'RMSE': _tm.get('rmse', 0.0),
                            'MAE':  _tm.get('mae',  0.0),
                        }
                    },
                    'training_time': _r.training_time or 0.0,
                })
            if _results_by_target:
                _viz = ANFISVisualizer(output_dir=str(self.output_dir / 'anfis_visualizations'))
                _viz.plot_target_comparison(_results_by_target)
                logger.info("[OK] ANFISVisualizer: target karsilastirma grafigi -> anfis_visualizations/")
        except Exception as _e:
            logger.warning(f"[WARNING] ANFISVisualizer basarisiz (devam): {_e}")

        # ---- ANFISRobustnessTester: en iyi 3 sonuc uzerinde iteratif outlier test ----
        try:
            from pfaz_modules.pfaz03_anfis_training.anfis_robustness_tester import ANFISRobustnessTester
            import pandas as _pd_rt
            # Basarili sonuclari test R2'ye gore sirala, en iyi 3 al
            _rob_candidates = sorted(
                [_r for _r in self.training_results if _r.success and _r.metrics],
                key=lambda _r: _r.metrics.get('test', {}).get('r2', -999),
                reverse=True
            )[:3]
            if _rob_candidates:
                _rob_tester = ANFISRobustnessTester(
                    output_dir=str(self.output_dir / 'robustness_analysis'),
                    max_iterations=3,  # hizli calistir (orijinal 5 yerine)
                )
                for _rc in _rob_candidates:
                    try:
                        # Dataset CSV'lerini bul
                        _ds_dir = None
                        for _cand in [
                            self.output_dir.parent / 'generated_datasets' / _rc.dataset_name,
                            self.output_dir.parent.parent / 'outputs' / 'generated_datasets' / _rc.dataset_name,
                        ]:
                            if _cand.exists():
                                _ds_dir = _cand
                                break
                        if _ds_dir is None:
                            continue
                        import pandas as _pds_r
                        import json as _jsn_r
                        _meta_f = _ds_dir / 'metadata.json'
                        _col_names_r = None
                        if _meta_f.exists():
                            with open(_meta_f) as _mfr:
                                _meta_r = _jsn_r.load(_mfr)
                            _fc = _meta_r.get('feature_names') or _meta_r.get('feature_columns', [])
                            _tc = _meta_r.get('target_names')  or _meta_r.get('target_columns',  [])
                            if _fc and _tc:
                                _col_names_r = list(_fc) + list(_tc)
                        _rkw = {'header': None, 'names': _col_names_r} if _col_names_r else {}
                        _df_tr = _pds_r.read_csv(_ds_dir / 'train.csv', **_rkw)
                        _df_va = _pds_r.read_csv(_ds_dir / 'val.csv',   **_rkw)
                        _df_te = _pds_r.read_csv(_ds_dir / 'test.csv',  **_rkw)
                        # Target sutununu belirle
                        _tgt_col = next(
                            (c for c in (_df_tr.columns[-1:])
                             if str(c) not in {'NUCLEUS'}),
                            _df_tr.columns[-1]
                        )
                        _feat_cols_r = [c for c in _df_tr.columns
                                        if c not in {'NUCLEUS', _tgt_col}]
                        import numpy as _np_r
                        _X_tr = _df_tr[_feat_cols_r].fillna(0).values.astype(_np_r.float32)
                        _y_tr = _df_tr[_tgt_col].fillna(0).values.astype(_np_r.float32)
                        _X_va = _df_va[_feat_cols_r].fillna(0).values.astype(_np_r.float32)
                        _y_va = _df_va[_tgt_col].fillna(0).values.astype(_np_r.float32)
                        _X_te = _df_te[_feat_cols_r].fillna(0).values.astype(_np_r.float32)
                        _y_te = _df_te[_tgt_col].fillna(0).values.astype(_np_r.float32)
                        _rob_tester.test_robustness(
                            _X_tr, _y_tr, _X_va, _y_va, _X_te, _y_te,
                            config_name=f'{_rc.dataset_name}_{_rc.config_id}'
                        )
                    except Exception as _rce:
                        logger.warning(f"  [WARNING] RobustnessTester {_rc.dataset_name}: {_rce}")
                _rob_tester.generate_excel_report('ANFIS_Robustness_Report.xlsx')
                logger.info("[OK] ANFISRobustnessTester: iteratif test raporu -> robustness_analysis/")
            else:
                logger.info("  [INFO] ANFISRobustnessTester: basarili sonuc yok — atlanıyor")
        except Exception as _e:
            logger.warning(f"[WARNING] ANFISRobustnessTester basarisiz (devam): {_e}")

        # ---- ANFISAdaptiveStrategy / PatternTracker: egitim örüntü analizi ----
        try:
            from pfaz_modules.pfaz03_anfis_training.anfis_adaptive_strategy import PatternTracker
            import json as _jsn_pt
            _pt = PatternTracker()
            for _r in self.training_results:
                if not _r.success or not _r.metrics:
                    continue
                _ds = _r.dataset_name
                if   _ds.startswith('Beta_2'): _tgt_pt = 'Beta_2'
                elif _ds.startswith('MM_QM'):  _tgt_pt = 'MM_QM'
                elif _ds.startswith('MM'):     _tgt_pt = 'MM'
                elif _ds.startswith('QM'):     _tgt_pt = 'QM'
                else:                          _tgt_pt = _ds.split('_')[0]
                _tm_pt = _r.metrics.get('test', {})
                _pt.record_result(
                    config_info={
                        'model':    'ANFIS',
                        'target':   _tgt_pt,
                        'features': '_'.join(_ds.split('_')[3:5]) if len(_ds.split('_')) >= 5 else _ds,
                        'scaling':  _ds.split('_')[5] if len(_ds.split('_')) > 5 else 'Unknown',
                    },
                    metrics={'R2': _tm_pt.get('r2', 0.0), 'RMSE': _tm_pt.get('rmse', 0.0)}
                )
            _pt_analysis = _pt.analyze_patterns()
            # JSON olarak kaydet
            _pt_dir = self.output_dir / 'adaptive_pattern_analysis'
            _pt_dir.mkdir(parents=True, exist_ok=True)
            with open(_pt_dir / 'pattern_analysis.json', 'w', encoding='utf-8') as _ptf:
                # defaultdict → normal dict dönüşümü
                _jsn_pt.dump(
                    {k: [[n, s] for n, s in v] for k, v in _pt_analysis.items()},
                    _ptf, indent=2, ensure_ascii=False
                )
            # Excel olarak da kaydet
            import pandas as _pds_pt
            _pt_rows = []
            for _cat, _sorted in _pt_analysis.items():
                for _name, _st in _sorted:
                    _pt_rows.append({
                        'Kategori': _cat, 'Ad': _name,
                        'Ort_R2': round(_st['mean'], 4),
                        'Std_R2': round(_st['std'],  4),
                        'Sayı':   _st['count'],
                    })
            if _pt_rows:
                _pt_df = _pds_pt.DataFrame(_pt_rows)
                _pt_df.to_excel(_pt_dir / 'pattern_analysis.xlsx', index=False)
            logger.info("[OK] ANFISAdaptiveStrategy/PatternTracker: örüntü analizi -> adaptive_pattern_analysis/")
        except Exception as _e:
            logger.warning(f"[WARNING] ANFISAdaptiveStrategy basarisiz (devam): {_e}")

        # ANFISDatasetSelector: deactivated — ANFIS 3-phase strategy covers all datasets
        logger.info("[INFO] ANFISDatasetSelector: deactivated (3-phase pilot strategy active)")

        # ---- NuclearPatternAnalyzer: ANFIS sonrasi nuklear desen analizi ----
        try:
            from pfaz_modules.pfaz12_advanced_analytics.nuclear_pattern_analyzer import NuclearPatternAnalyzer
            from pathlib import Path as _Path3
            _npa3_candidates = [
                'aaa2.txt',
                str(self.output_dir / 'aaa2.txt'),
            ]
            _enriched3 = list(_Path3('outputs').glob('**/aaa2_enriched*.csv')) + \
                         list(_Path3('outputs').glob('**/aaa2_enriched*.xlsx'))
            _npa3_data = None
            for _c3 in _npa3_candidates:
                if _Path3(_c3).exists():
                    _npa3_data = _c3
                    break
            if _npa3_data is None and _enriched3:
                _npa3_data = str(_enriched3[0])
            if _npa3_data:
                _npa3 = NuclearPatternAnalyzer(
                    data_path=_npa3_data,
                    output_dir=str(self.output_dir / 'nuclear_patterns')
                )
                _npa3.run_all()
                logger.info("[OK] NuclearPatternAnalyzer: Excel + grafikler -> nuclear_patterns/")
            else:
                logger.warning("[WARNING] NuclearPatternAnalyzer: veri dosyası bulunamadı (aaa2.txt yok)")
        except Exception as _npa3_e:
            logger.warning(f"[WARNING] NuclearPatternAnalyzer basarisiz (devam): {_npa3_e}")

        return {
            'status': 'completed',
            'total_jobs': len(results),
            'successful': successful,
            'failed': len(failed),
            'results': results
        }

    def save_summary_report(self):
        """Save summary report as JSON and Excel."""

        report_file = self.output_dir / 'anfis_training_summary.json'

        summary = {
            'total_jobs': len(self.training_results),
            'successful': len([r for r in self.training_results if r.success]),
            'failed': len(self.failed_jobs),
            'results': []
        }

        flat_rows = []   # for Excel

        for result in self.training_results:
            result_dict = {
                'job_id': result.job_id,
                'config_id': result.config_id,
                'dataset_name': result.dataset_name,
                'success': result.success,
                'training_time': result.training_time
            }

            if result.success:
                result_dict['metrics'] = result.metrics
                result_dict['model_path'] = str(result.model_path)

                meta = result.metrics.get('training_meta', {})
                train_m = result.metrics.get('train', {})
                val_m   = result.metrics.get('val',   {})
                test_m  = result.metrics.get('test',  {})

                # Parse target from dataset_name
                ds = result.dataset_name
                if ds.startswith('Beta_2_'):
                    target = 'Beta_2'
                elif ds.startswith('MM_QM_'):
                    target = 'MM_QM'
                elif ds.startswith('MM_'):
                    target = 'MM'
                elif ds.startswith('QM_'):
                    target = 'QM'
                else:
                    target = ds.split('_')[0] if '_' in ds else ds

                parts = ds.replace('Beta_2_', 'Beta2_').replace('MM_QM_', 'MMQM_').split('_')

                def r2_category(r2):
                    if r2 is None or (isinstance(r2, float) and r2 != r2):
                        return 'N/A'
                    if r2 < 0:       return 'Failed (<0)'
                    if r2 < 0.5:    return 'Poor (0-0.5)'
                    if r2 < 0.7:    return 'Low (0.5-0.7)'
                    if r2 < 0.85:   return 'Medium (0.7-0.85)'
                    if r2 < 0.95:   return 'Good (0.85-0.95)'
                    return 'Excellent (>=0.95)'

                val_r2  = val_m.get('r2')
                test_r2 = test_m.get('r2')

                flat_rows.append({
                    'Job_ID':           result.job_id,
                    'Dataset_Name':     result.dataset_name,
                    'Target':           target,
                    'Nucleus_Count':    parts[1] if len(parts) > 1 else '',
                    'Scenario':         meta.get('dataset_name', ds).split('_')[2] if '_' in ds else '',
                    'Feature_Code':     parts[3] if len(parts) > 3 else '',
                    'Config_ID':        result.config_id,
                    'MF_Type':          meta.get('mf_type', ''),
                    'Method':           meta.get('method', ''),
                    'N_Inputs':         meta.get('n_inputs', ''),
                    'N_MFs':            meta.get('n_mfs_per_input', ''),
                    'N_Rules':          meta.get('n_rules', ''),
                    'Feature_Cols':     ', '.join(meta.get('feature_cols', [])),
                    'N_Train':          meta.get('n_train', ''),
                    'N_Val':            meta.get('n_val', ''),
                    'N_Test':           meta.get('n_test', ''),
                    'N_Outliers':       meta.get('n_outliers_detected', 0),
                    'Outlier_Cleaned':  meta.get('outlier_cleaning_applied', False),
                    'Train_R2':         train_m.get('r2'),
                    'Train_RMSE':       train_m.get('rmse'),
                    'Train_MAE':        train_m.get('mae'),
                    'Val_R2':           val_r2,
                    'Val_RMSE':         val_m.get('rmse'),
                    'Val_MAE':          val_m.get('mae'),
                    'Test_R2':          test_r2,
                    'Test_RMSE':        test_m.get('rmse'),
                    'Test_MAE':         test_m.get('mae'),
                    'Val_R2_Category':  r2_category(val_r2),
                    'Test_R2_Category': r2_category(test_r2),
                    'PKL_Saved':        True,
                    'Status_Note':      'OK',
                    'Training_Mode':    meta.get('training_mode', result.training_mode),
                    'Training_Time_s':  result.training_time,
                    'Nucleus_Train':    ', '.join(meta.get('nucleus_train', [])),
                    'Nucleus_Val':      ', '.join(meta.get('nucleus_val', [])),
                    'Nucleus_Test':     ', '.join(meta.get('nucleus_test', [])),
                    'Nucleus_Outliers_Removed': ', '.join(meta.get('nucleus_outliers_removed', [])),
                })
            else:
                result_dict['error'] = result.error_message
                # Include actual scores when available (Poor/Failed filter passes metrics)
                _fm = result.metrics or {}
                _fval = _fm.get('val', {})
                _ftrain = _fm.get('train', {})
                _ftest  = _fm.get('test', {})
                _fval_r2  = _fval.get('r2')
                _ftest_r2 = _ftest.get('r2')
                _err = result.error_message or ''
                if 'Poor/Failed' in _err:
                    _status_note = 'POOR_R2_FILTER'
                elif 'diverged' in _err.lower():
                    _status_note = 'DIVERGED'
                elif _err:
                    _status_note = f'ERROR: {_err[:80]}'
                else:
                    _status_note = 'FAILED'
                flat_rows.append({
                    'Job_ID':           result.job_id,
                    'Dataset_Name':     result.dataset_name,
                    'Config_ID':        result.config_id,
                    'Train_R2':         _ftrain.get('r2'),
                    'Train_RMSE':       _ftrain.get('rmse'),
                    'Val_R2':           _fval_r2,
                    'Val_RMSE':         _fval.get('rmse'),
                    'Test_R2':          _ftest_r2,
                    'Test_RMSE':        _ftest.get('rmse'),
                    'Val_R2_Category':  r2_category(_fval_r2) if _fval_r2 is not None else 'ERROR',
                    'Test_R2_Category': r2_category(_ftest_r2) if _ftest_r2 is not None else 'ERROR',
                    'PKL_Saved':        False,
                    'Status_Note':      _status_note,
                    'Training_Mode':    result.training_mode,
                    'Error':            result.error_message,
                })

            summary['results'].append(result_dict)

        # Save JSON
        with open(report_file, 'w', encoding='utf-8') as f:
            json.dump(summary, f, indent=2, ensure_ascii=False)
        logger.info(f"Summary JSON saved: {report_file}")

        # Save Excel
        try:
            df_all = pd.DataFrame(flat_rows)
            excel_path = self.output_dir / 'anfis_training_results.xlsx'
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                # ---- Sheet 1: All Results
                df_all.to_excel(writer, sheet_name='All_Results', index=False)

                # ---- Sheet 2: Best per Dataset (best val_R2 per dataset)
                if 'Val_R2' in df_all.columns and df_all['Val_R2'].notna().any():
                    df_success = df_all[df_all['Val_R2'].notna()].copy()
                    df_best = df_success.sort_values('Val_R2', ascending=False) \
                                        .drop_duplicates(subset=['Dataset_Name'], keep='first')
                    df_best.to_excel(writer, sheet_name='Best_Per_Dataset', index=False)

                # ---- Sheet 3: Nucleus Tracking
                nuc_rows = []
                for row in flat_rows:
                    ds  = row.get('Dataset_Name', '')
                    cfg = row.get('Config_ID', '')
                    for split, col in [('train', 'Nucleus_Train'),
                                       ('val',   'Nucleus_Val'),
                                       ('test',  'Nucleus_Test')]:
                        for nuc in [n.strip() for n in row.get(col, '').split(',') if n.strip()]:
                            nuc_rows.append({
                                'Dataset_Name': ds,
                                'Config_ID':    cfg,
                                'Split':        split,
                                'Nucleus':      nuc,
                                'Role':         'used',
                            })
                    for nuc in [n.strip() for n in row.get('Nucleus_Outliers_Removed', '').split(',') if n.strip()]:
                        nuc_rows.append({
                            'Dataset_Name': ds,
                            'Config_ID':    cfg,
                            'Split':        'train',
                            'Nucleus':      nuc,
                            'Role':         'outlier_removed',
                        })
                if nuc_rows:
                    pd.DataFrame(nuc_rows).to_excel(writer, sheet_name='Nucleus_Tracking', index=False)

                # ---- Sheet 4: Outlier_Frequency (per-nucleus outlier statistics)
                # Counts how many times each nucleus was detected as outlier and
                # how many times the clean model was actually adopted.
                outlier_freq: Dict[str, Dict] = {}
                for row in flat_rows:
                    ds = row.get('Dataset_Name', '')
                    outlier_cleaned = row.get('Outlier_Cleaned', False)
                    for nuc in [n.strip() for n in row.get('Nucleus_Outliers_Removed', '').split(',') if n.strip()]:
                        if nuc not in outlier_freq:
                            outlier_freq[nuc] = {
                                'Nucleus': nuc,
                                'Times_Outlier_Detected': 0,
                                'Times_Actually_Removed': 0,
                                'Datasets_Affected': [],
                            }
                        outlier_freq[nuc]['Times_Outlier_Detected'] += 1
                        if outlier_cleaned:
                            outlier_freq[nuc]['Times_Actually_Removed'] += 1
                        if ds not in outlier_freq[nuc]['Datasets_Affected']:
                            outlier_freq[nuc]['Datasets_Affected'].append(ds)
                if outlier_freq:
                    _of_rows = []
                    for rec in sorted(outlier_freq.values(),
                                      key=lambda x: x['Times_Outlier_Detected'], reverse=True):
                        _of_rows.append({
                            'Nucleus':                rec['Nucleus'],
                            'Times_Outlier_Detected': rec['Times_Outlier_Detected'],
                            'Times_Actually_Removed': rec['Times_Actually_Removed'],
                            'Removal_Rate_%':         round(
                                100 * rec['Times_Actually_Removed'] / max(1, rec['Times_Outlier_Detected']), 1),
                            'Datasets_Affected_Count': len(rec['Datasets_Affected']),
                            'Datasets_Affected':      ', '.join(rec['Datasets_Affected'][:10]),
                        })
                    pd.DataFrame(_of_rows).to_excel(writer, sheet_name='Outlier_Frequency', index=False)

                # ---- Sheet 6: R2 Category Summary
                if 'Val_R2_Category' in df_all.columns and 'Target' in df_all.columns:
                    pivot = df_all.groupby(['Target', 'Val_R2_Category']).size() \
                                  .reset_index(name='Count')
                    pivot.to_excel(writer, sheet_name='R2_Category_Summary', index=False)

            logger.info(f"Summary Excel saved: {excel_path}")
            self._apply_excel_formatting(excel_path)
        except Exception as e:
            logger.warning(f"Excel save failed: {e}")

    def _apply_excel_formatting(self, excel_path: Path):
        """Apply auto-filter, freeze panes, and conditional R2 coloring."""
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill, Font
            from openpyxl.utils import get_column_letter

            wb = load_workbook(excel_path)

            # Color scale for R2 columns
            r2_colors = {
                'Failed (<0)':      'FFB3B3',  # light red
                'Poor (0-0.5)':     'FFD9B3',  # light orange
                'Low (0.5-0.7)':    'FFFF99',  # light yellow
                'Medium (0.7-0.85)':'C6EFCE',  # light green
                'Good (0.85-0.95)': '70AD47',  # green
                'Excellent (>=0.95)':'375623',  # dark green
            }

            for sheet_name in ['All_Results', 'Best_Per_Dataset']:
                if sheet_name not in wb.sheetnames:
                    continue
                ws = wb[sheet_name]
                # Auto-filter on header row
                ws.auto_filter.ref = ws.dimensions
                # Freeze top row
                ws.freeze_panes = 'A2'
                # Auto-width
                for col in ws.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 40)
                # Color Val_R2_Category and Test_R2_Category columns
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        hdr = ws.cell(1, cell.column).value or ''
                        if 'Category' in str(hdr) and cell.value in r2_colors:
                            cell.fill = PatternFill('solid', fgColor=r2_colors[cell.value])

            for sheet_name in ['Nucleus_Tracking', 'Outlier_Frequency', 'R2_Category_Summary']:
                if sheet_name not in wb.sheetnames:
                    continue
                ws = wb[sheet_name]
                ws.auto_filter.ref = ws.dimensions
                ws.freeze_panes = 'A2'
                for col in ws.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 40)

            wb.save(excel_path)
        except Exception as e:
            logger.debug(f"Excel formatting skipped: {e}")


    def generate_comparison_excel(self, ai_results_dir: str = None):
        """
        Build AI vs ANFIS comparison Excel.
        Reads PFAZ2 results from outputs/trained_models/ and PFAZ3 results
        from this trainer's training_results.

        Output: anfis_vs_ai_comparison.xlsx  (in output_dir)
        Sheets:
          - Comparison      : per dataset, best AI vs best ANFIS, R2 diff, winner
          - ANFIS_Wins      : datasets where ANFIS > best AI (or AI failed)
          - AI_Failed       : datasets where best AI R2 < threshold
          - By_Target_Pivot : R2 summary grouped by target
          - By_Config_Pivot : best ANFIS config per dataset
        """
        try:
            logger.info("Generating AI vs ANFIS comparison Excel...")

            # ---------------------------------------------------------------- #
            # 1.  Collect ANFIS results (best val_R2 per dataset from this run)
            # ---------------------------------------------------------------- #
            anfis_rows = {}
            for result in self.training_results:
                if not result.success:
                    continue
                ds    = result.dataset_name
                val_r2 = result.metrics.get('val', {}).get('r2', None)
                if val_r2 is None:
                    continue
                meta = result.metrics.get('training_meta', {})
                row = {
                    'Dataset_Name':    ds,
                    'ANFIS_Config':    result.config_id,
                    'ANFIS_MF_Type':   meta.get('mf_type', ''),
                    'ANFIS_Method':    meta.get('method', ''),
                    'ANFIS_N_Rules':   meta.get('n_rules', ''),
                    'ANFIS_Val_R2':    result.metrics.get('val', {}).get('r2'),
                    'ANFIS_Test_R2':   result.metrics.get('test', {}).get('r2'),
                    'ANFIS_Val_RMSE':  result.metrics.get('val', {}).get('rmse'),
                    'ANFIS_Test_RMSE': result.metrics.get('test', {}).get('rmse'),
                    'ANFIS_N_Outliers': meta.get('n_outliers_detected', 0),
                    'ANFIS_Outlier_Cleaned': meta.get('outlier_cleaning_applied', False),
                }
                # Keep best val_R2 per dataset
                if ds not in anfis_rows or (val_r2 > (anfis_rows[ds].get('ANFIS_Val_R2') or -999)):
                    anfis_rows[ds] = row

            # ---------------------------------------------------------------- #
            # 2.  Collect AI results from PFAZ2 trained_models directory
            # ---------------------------------------------------------------- #
            if ai_results_dir is None:
                # Try relative to output_dir (usually outputs/anfis_models → outputs/trained_models)
                ai_results_dir = str(self.output_dir.parent / 'trained_models')

            ai_rows = {}
            ai_dir = Path(ai_results_dir)
            if ai_dir.exists():
                for ds_dir in ai_dir.iterdir():
                    if not ds_dir.is_dir():
                        continue
                    ds = ds_dir.name
                    best_val_r2   = None
                    best_model    = None
                    best_test_r2  = None
                    best_val_rmse = None
                    for model_dir in ds_dir.iterdir():
                        if not model_dir.is_dir():
                            continue
                        model_type = model_dir.name
                        for cfg_dir in model_dir.iterdir():
                            if not cfg_dir.is_dir():
                                continue
                            for mf in cfg_dir.glob('metrics_*.json'):
                                try:
                                    with open(mf, encoding='utf-8') as f:
                                        m = json.load(f)
                                    val_r2  = m.get('val', {}).get('r2')
                                    test_r2 = m.get('test', {}).get('r2')
                                    val_rmse= m.get('val', {}).get('rmse')
                                    if val_r2 is not None:
                                        if best_val_r2 is None or val_r2 > best_val_r2:
                                            best_val_r2   = val_r2
                                            best_test_r2  = test_r2
                                            best_val_rmse = val_rmse
                                            best_model    = f"{model_type}/{cfg_dir.name}"
                                except Exception:
                                    pass
                    if best_val_r2 is not None:
                        ai_rows[ds] = {
                            'Dataset_Name':  ds,
                            'AI_Best_Model': best_model,
                            'AI_Val_R2':     best_val_r2,
                            'AI_Test_R2':    best_test_r2,
                            'AI_Val_RMSE':   best_val_rmse,
                        }
            else:
                logger.warning(f"AI results dir not found: {ai_dir}")

            # ---------------------------------------------------------------- #
            # 3.  Merge into comparison table
            # ---------------------------------------------------------------- #
            all_datasets = set(anfis_rows.keys()) | set(ai_rows.keys())

            def r2_cat(v):
                if v is None: return 'N/A'
                if v < 0:     return 'Failed (<0)'
                if v < 0.5:   return 'Poor (0-0.5)'
                if v < 0.7:   return 'Low (0.5-0.7)'
                if v < 0.85:  return 'Medium (0.7-0.85)'
                if v < 0.95:  return 'Good (0.85-0.95)'
                return 'Excellent (>=0.95)'

            def parse_target(ds):
                if ds.startswith('Beta_2_'): return 'Beta_2'
                if ds.startswith('MM_QM_'):  return 'MM_QM'
                if ds.startswith('MM_'):     return 'MM'
                if ds.startswith('QM_'):     return 'QM'
                return ds.split('_')[0]

            comparison = []
            for ds in sorted(all_datasets):
                ai  = ai_rows.get(ds, {})
                anf = anfis_rows.get(ds, {})

                ai_r2   = ai.get('AI_Val_R2')
                anf_r2  = anf.get('ANFIS_Val_R2')

                # Determine winner
                if ai_r2 is not None and anf_r2 is not None:
                    diff = anf_r2 - ai_r2
                    if diff > 0.02:
                        winner = 'ANFIS'
                    elif diff < -0.02:
                        winner = 'AI'
                    else:
                        winner = 'Tie'
                elif anf_r2 is not None:
                    winner = 'ANFIS_Only'
                elif ai_r2 is not None:
                    winner = 'AI_Only'
                else:
                    winner = 'N/A'

                parts = ds.replace('Beta_2_', 'Beta2_').replace('MM_QM_', 'MMQM_').split('_')

                row = {
                    'Dataset_Name':       ds,
                    'Target':             parse_target(ds),
                    'Nucleus_Count':      parts[1] if len(parts) > 1 else '',
                    'Scenario':           parts[2] if len(parts) > 2 else '',
                    'Feature_Code':       parts[3] if len(parts) > 3 else '',
                    'NoAnomaly':          'Yes' if 'NoAnomaly' in ds else 'No',
                    # AI columns
                    'AI_Best_Model':      ai.get('AI_Best_Model', ''),
                    'AI_Val_R2':          ai_r2,
                    'AI_Test_R2':         ai.get('AI_Test_R2'),
                    'AI_Val_RMSE':        ai.get('AI_Val_RMSE'),
                    'AI_R2_Category':     r2_cat(ai_r2),
                    # ANFIS columns
                    'ANFIS_Config':       anf.get('ANFIS_Config', ''),
                    'ANFIS_MF_Type':      anf.get('ANFIS_MF_Type', ''),
                    'ANFIS_Method':       anf.get('ANFIS_Method', ''),
                    'ANFIS_N_Rules':      anf.get('ANFIS_N_Rules', ''),
                    'ANFIS_Val_R2':       anf_r2,
                    'ANFIS_Test_R2':      anf.get('ANFIS_Test_R2'),
                    'ANFIS_Val_RMSE':     anf.get('ANFIS_Val_RMSE'),
                    'ANFIS_R2_Category':  r2_cat(anf_r2),
                    'ANFIS_N_Outliers':   anf.get('ANFIS_N_Outliers', ''),
                    'ANFIS_Cleaned':      anf.get('ANFIS_Outlier_Cleaned', ''),
                    # Delta
                    'Delta_Val_R2':       round(anf_r2 - ai_r2, 4) if (ai_r2 is not None and anf_r2 is not None) else None,
                    'Winner':             winner,
                }
                comparison.append(row)

            df_cmp = pd.DataFrame(comparison)

            # ---------------------------------------------------------------- #
            # 4.  Build derived sheets
            # ---------------------------------------------------------------- #
            df_ai_failed = df_cmp[
                df_cmp['AI_R2_Category'].isin(['Failed (<0)', 'Poor (0-0.5)', 'N/A'])
            ].copy() if not df_cmp.empty else pd.DataFrame()

            df_anfis_wins = df_cmp[df_cmp['Winner'] == 'ANFIS'].copy() \
                if not df_cmp.empty else pd.DataFrame()

            # By-Target pivot
            if not df_cmp.empty and 'Target' in df_cmp.columns:
                pivot_target = df_cmp.groupby('Target').agg(
                    N_Datasets=('Dataset_Name', 'count'),
                    AI_Mean_Val_R2=('AI_Val_R2', 'mean'),
                    ANFIS_Mean_Val_R2=('ANFIS_Val_R2', 'mean'),
                    ANFIS_Wins=('Winner', lambda x: (x == 'ANFIS').sum()),
                    AI_Wins=('Winner', lambda x: (x == 'AI').sum()),
                    Ties=('Winner', lambda x: (x == 'Tie').sum()),
                ).reset_index()
            else:
                pivot_target = pd.DataFrame()

            # By-Config pivot
            if not df_cmp.empty and 'ANFIS_Config' in df_cmp.columns:
                pivot_config = df_cmp.groupby('ANFIS_Config').agg(
                    N_Datasets=('Dataset_Name', 'count'),
                    Mean_Val_R2=('ANFIS_Val_R2', 'mean'),
                    Max_Val_R2=('ANFIS_Val_R2', 'max'),
                    N_Wins=('Winner', lambda x: (x == 'ANFIS').sum()),
                ).reset_index().sort_values('Mean_Val_R2', ascending=False)
            else:
                pivot_config = pd.DataFrame()

            # ---------------------------------------------------------------- #
            # 5.  Write Excel
            # ---------------------------------------------------------------- #
            excel_path = self.output_dir / 'anfis_vs_ai_comparison.xlsx'
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                df_cmp.to_excel(writer,         sheet_name='Comparison',      index=False)
                df_anfis_wins.to_excel(writer,  sheet_name='ANFIS_Wins',      index=False)
                df_ai_failed.to_excel(writer,   sheet_name='AI_Failed',       index=False)
                pivot_target.to_excel(writer,   sheet_name='By_Target_Pivot', index=False)
                pivot_config.to_excel(writer,   sheet_name='By_Config_Pivot', index=False)

            # Apply formatting
            self._apply_comparison_formatting(excel_path)
            logger.info(f"Comparison Excel saved: {excel_path}")

        except Exception as e:
            logger.error(f"Comparison Excel generation failed: {e}", exc_info=True)

    def _apply_comparison_formatting(self, excel_path: Path):
        """Apply auto-filter, freeze panes, conditional coloring to comparison Excel."""
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter

            WINNER_COLORS = {
                'ANFIS':      '70AD47',   # green
                'AI':         '4472C4',   # blue
                'Tie':        'FFD966',   # yellow
                'ANFIS_Only': 'A9D18E',   # light green
                'AI_Only':    '9DC3E6',   # light blue
                'N/A':        'D9D9D9',   # grey
            }
            R2_COLORS = {
                'Failed (<0)':       'FFB3B3',
                'Poor (0-0.5)':      'FFD9B3',
                'Low (0.5-0.7)':     'FFFF99',
                'Medium (0.7-0.85)': 'C6EFCE',
                'Good (0.85-0.95)':  '70AD47',
                'Excellent (>=0.95)':'375623',
                'N/A':               'D9D9D9',
            }

            wb = load_workbook(excel_path)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                ws.auto_filter.ref = ws.dimensions
                ws.freeze_panes = 'A2'

                # Auto column width
                for col in ws.columns:
                    max_len = max((len(str(c.value or '')) for c in col), default=8)
                    ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 45)

                # Conditional cell coloring
                headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        hdr = ws.cell(1, cell.column).value or ''
                        val = cell.value
                        if hdr == 'Winner' and val in WINNER_COLORS:
                            cell.fill = PatternFill('solid', fgColor=WINNER_COLORS[val])
                            cell.font = Font(bold=True)
                        elif 'Category' in str(hdr) and val in R2_COLORS:
                            cell.fill = PatternFill('solid', fgColor=R2_COLORS[val])
                        elif hdr == 'Delta_Val_R2' and val is not None:
                            try:
                                fv = float(val)
                                if fv > 0.05:
                                    cell.fill = PatternFill('solid', fgColor='70AD47')
                                elif fv < -0.05:
                                    cell.fill = PatternFill('solid', fgColor='FF6666')
                            except (TypeError, ValueError):
                                pass

            wb.save(excel_path)
        except Exception as e:
            logger.debug(f"Comparison formatting skipped: {e}")

    def save_training_datasets(self, X_train, y_train, X_val, y_val, X_test, y_test, dataset_name: str, config_id: str):
        """
        Save train/val/test datasets in multiple formats (.mat, .csv, .xlsx)

        Args:
            X_train, y_train: Training data
            X_val, y_val: Validation data
            X_test, y_test: Test data
            dataset_name: Name of the dataset
            config_id: Configuration ID
        """
        if not self.save_datasets:
            return

        try:
            # Create dataset save directory
            save_dir = self.output_dir / 'training_datasets' / f'{dataset_name}_{config_id}'
            save_dir.mkdir(parents=True, exist_ok=True)

            # Save as CSV
            for split_name, (X, y) in [('train', (X_train, y_train)),
                                        ('val', (X_val, y_val)),
                                        ('test', (X_test, y_test))]:
                df = pd.DataFrame(X)

                # Handle multi-target y (shape: n_samples, n_targets)
                y_arr = np.asarray(y)
                if y_arr.ndim == 1:
                    df['target'] = y_arr
                else:
                    # Multi-target: add each target as a separate column
                    for i in range(y_arr.shape[1]):
                        df[f'target_{i}'] = y_arr[:, i]

                csv_path = save_dir / f'{split_name}.csv'
                df.to_csv(csv_path, index=False, header=False)
                logger.info(f"  Saved {split_name}.csv (headerless)")

            # Save as Excel
            try:
                with pd.ExcelWriter(save_dir / 'all_splits.xlsx') as writer:
                    for split_name, (X, y) in [('train', (X_train, y_train)),
                                                ('val', (X_val, y_val)),
                                                ('test', (X_test, y_test))]:
                        df = pd.DataFrame(X)

                        # Handle multi-target y (shape: n_samples, n_targets)
                        y_arr = np.asarray(y)
                        if y_arr.ndim == 1:
                            df['target'] = y_arr
                        else:
                            # Multi-target: add each target as a separate column
                            for i in range(y_arr.shape[1]):
                                df[f'target_{i}'] = y_arr[:, i]

                        df.to_excel(writer, sheet_name=split_name, index=False)
                logger.info(f"  Saved all_splits.xlsx")
            except Exception as e:
                logger.warning(f"  Could not save Excel: {e}")

            # Not: .mat kaydi kaldirildi — egitim CSV ile yapilir.

            logger.info(f"[DATASET SAVED] {dataset_name}_{config_id}")

        except Exception as e:
            logger.error(f"[ERROR] Could not save datasets: {e}")

    def track_kernel_usage(self, dataset_name: str, config_id: str, kernel_info: Dict):
        """
        Track which kernels (nuclei) are used in which training (thread-safe)

        Args:
            dataset_name: Name of the dataset
            config_id: Configuration ID
            kernel_info: Dictionary with kernel information (e.g., nucleus A, Z, N)
        """
        tracking_key = f'{dataset_name}_{config_id}'

        # Thread-safe update
        with self.kernel_tracker_lock:
            self.kernel_usage_tracker[tracking_key] = {
                'dataset': dataset_name,
                'config': config_id,
                'timestamp': datetime.now().isoformat(),
                'kernel_info': kernel_info
            }

            # Save tracker to file
            tracker_file = self.output_dir / 'kernel_usage_tracker.json'
            with open(tracker_file, 'w') as f:
                json.dump(self.kernel_usage_tracker, f, indent=2)

        logger.info(f"[KERNEL TRACKER] Updated for {tracking_key}")

    def generate_kernel_usage_report(self) -> Dict:
        """
        Generate a comprehensive report of kernel usage across all trainings

        Returns:
            Dictionary with kernel usage statistics
        """
        report = {
            'total_trainings': len(self.kernel_usage_tracker),
            'trainings': self.kernel_usage_tracker,
            'generated_at': datetime.now().isoformat()
        }

        # Save report
        report_file = self.output_dir / 'kernel_usage_report.json'
        with open(report_file, 'w') as f:
            json.dump(report, f, indent=2)

        logger.info(f"[KERNEL REPORT] Generated: {report_file}")
        logger.info(f"  Total trainings tracked: {report['total_trainings']}")

        return report


# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    """Main execution for testing"""

    print("\n" + "=" * 80)
    print("PFAZ 3: ANFIS PARALLEL TRAINER V2 - TEST")
    print("=" * 80)

    # Initialize trainer
    trainer = ANFISParallelTrainerV2(
        output_dir='test_anfis_models',
        n_workers=2,
        use_config_manager=True,
        use_adaptive_strategy=False,
        use_performance_analyzer=True,
        save_datasets=True
    )

    print("\n[SUCCESS] ANFIS Parallel Trainer V2 initialized with all features!")


if __name__ == "__main__":
    main()
