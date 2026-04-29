"""
AutoML Retraining Loop
======================
AI ve ANFIS modellerini Good/Medium/Poor kategorilerine gore secip optimize eder.

Workflow:
  1. PFAZ2 cikti dizinindeki metrics_*.json dosyalarini tara
  2. Kategorile: Poor (R2<0.70), Medium (0.70-0.90), Good (0.90-0.95)
  3. Her kategoriden n_per_category (default 25) secilir
  4. Her aday icin AutoMLOptimizer calistir (Optuna TPE)
  5. Test seti uzerinde tahmin karsilastirmasi yap
  6. ANFIS optimizer'i ayni kategorilerden 25'er dataset ile calistir
  7. Deneysel verilerle (aaa2.txt) karsilastirma ekle
  8. automl_improvement_report.xlsx ve automl_retraining_log.json olustur
"""

from __future__ import annotations

import json
import logging
import time
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Optional imports
# ---------------------------------------------------------------------------
try:
    from .automl_optimizer import AutoMLOptimizer, OPTUNA_AVAILABLE
except ImportError:
    try:
        from automl_optimizer import AutoMLOptimizer, OPTUNA_AVAILABLE
    except ImportError:
        AutoMLOptimizer = None
        OPTUNA_AVAILABLE = False

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _load_metrics_files(models_dir: Path) -> List[Dict]:
    """PFAZ2 metrics_*.json dosyalarını toplar ve listeler."""
    records = []
    for mf in models_dir.rglob('metrics_*.json'):
        try:
            with open(mf, 'r') as f:
                data = json.load(f)
            # Flat list veya tek obje olabilir
            if isinstance(data, list):
                for item in data:
                    item.setdefault('_source_file', str(mf))
                    records.append(item)
            elif isinstance(data, dict):
                data.setdefault('_source_file', str(mf))
                records.append(data)
        except Exception as e:
            logger.debug(f"metrics okuma hatası {mf}: {e}")
    return records


def _infer_model_type_from_path(path_str: str) -> Optional[str]:
    """Dosya yolundan model tipini tahmin eder."""
    p = path_str.lower()
    for mt in ('rf', 'xgb', 'xgboost', 'gbm', 'lgb', 'lightgbm', 'cb', 'catboost', 'svr', 'dnn', 'mlp'):
        if mt in p:
            # Normalise
            return {'xgboost': 'xgb', 'lightgbm': 'lgb', 'catboost': 'cb',
                    'mlp': 'dnn'}.get(mt, mt)
    return None


def _load_test_data(datasets_dir: Path, dataset_name: str, target: str):
    """
    train+val+test verilerini yukler.
    Returns: (X_train, y_train, X_val, y_val, X_test, y_test, nucleus_ids) veya None
    """
    MM_QM_COLS = ['MM', 'QM']
    ID_COLS = ['NUCLEUS', 'Z', 'N', 'A']

    for candidate in datasets_dir.rglob(f'*{dataset_name}*'):
        if not candidate.is_dir():
            continue
        train_csv = candidate / 'train.csv'
        val_csv   = candidate / 'val.csv'
        test_csv  = candidate / 'test.csv'
        if not (train_csv.exists() and val_csv.exists() and test_csv.exists()):
            continue
        try:
            df_train = pd.read_csv(train_csv)
            df_val   = pd.read_csv(val_csv)
            df_test  = pd.read_csv(test_csv)

            feature_cols = [c for c in df_train.columns
                            if c not in (['MM', 'QM', 'Beta_2', 'MM_QM', 'NUCLEUS',
                                          'A', 'Z', 'N', 'SPIN', 'PARITY'])]
            if target == 'MM_QM':
                target_cols = [c for c in MM_QM_COLS if c in df_train.columns]
            else:
                target_cols = [target] if target in df_train.columns else []

            if not feature_cols or not target_cols:
                continue

            # Nucleus IDs from test set
            nucleus_ids = []
            for ic in ID_COLS:
                if ic in df_test.columns:
                    nucleus_ids = df_test[ic].tolist()
                    break

            X_train = df_train[feature_cols].values
            y_train = df_train[target_cols].values
            X_val   = df_val[feature_cols].values
            y_val   = df_val[target_cols].values
            X_test  = df_test[feature_cols].values
            y_test  = df_test[target_cols].values

            if target != 'MM_QM':
                y_train = y_train.ravel()
                y_val   = y_val.ravel()
                y_test  = y_test.ravel()

            return X_train, y_train, X_val, y_val, X_test, y_test, nucleus_ids
        except Exception as e:
            logger.debug(f"test data yuklenemedi {candidate}: {e}")

    return None


def _load_split_data(datasets_dir: Path, dataset_name: str, target: str):
    """
    train.csv / val.csv çiftini yükler.
    Returns: (X_train, y_train, X_val, y_val) veya None
    """
    MM_QM_COLS = ['MM', 'QM']

    for candidate in datasets_dir.rglob(f'*{dataset_name}*'):
        train_csv = candidate / 'train.csv' if candidate.is_dir() else None
        val_csv   = candidate / 'val.csv'   if candidate.is_dir() else None
        if train_csv and train_csv.exists() and val_csv and val_csv.exists():
            try:
                df_train = pd.read_csv(train_csv)
                df_val   = pd.read_csv(val_csv)

                feature_cols = [c for c in df_train.columns
                                if c not in (['MM', 'QM', 'Beta_2', 'MM_QM', 'NUCLEUS',
                                              'A', 'Z', 'N', 'SPIN', 'PARITY'])]

                if target == 'MM_QM':
                    target_cols = [c for c in MM_QM_COLS if c in df_train.columns]
                else:
                    target_cols = [target] if target in df_train.columns else []

                if not feature_cols or not target_cols:
                    continue

                X_train = df_train[feature_cols].values
                y_train = df_train[target_cols].values
                X_val   = df_val[feature_cols].values
                y_val   = df_val[target_cols].values

                if target != 'MM_QM':
                    y_train = y_train.ravel()
                    y_val   = y_val.ravel()

                return X_train, y_train, X_val, y_val
            except Exception as e:
                logger.debug(f"CSV yükleme hatası {candidate}: {e}")

    # Fallback: tam yol araması
    for csv_dir in datasets_dir.rglob('train.csv'):
        parent = csv_dir.parent
        if dataset_name.replace('_', '') in parent.name.replace('_', '').lower():
            try:
                df_train = pd.read_csv(csv_dir)
                df_val   = pd.read_csv(parent / 'val.csv')
                feature_cols = [c for c in df_train.columns
                                if c not in ['MM', 'QM', 'Beta_2', 'MM_QM', 'NUCLEUS',
                                             'A', 'Z', 'N', 'SPIN', 'PARITY']]
                if target == 'MM_QM':
                    target_cols = [c for c in MM_QM_COLS if c in df_train.columns]
                else:
                    target_cols = [target] if target in df_train.columns else []

                if not feature_cols or not target_cols:
                    continue

                X_train = df_train[feature_cols].values
                y_train = df_train[target_cols].values
                X_val   = df_val[feature_cols].values
                y_val   = df_val[target_cols].values

                if target != 'MM_QM':
                    y_train = y_train.ravel()
                    y_val   = y_val.ravel()

                return X_train, y_train, X_val, y_val
            except Exception:
                pass

    return None


# ---------------------------------------------------------------------------
# Main class
# ---------------------------------------------------------------------------

class AutoMLRetrainingLoop:
    """
    AI ve ANFIS modellerini Good/Medium/Poor kategorilerine gore secip optimize eder.
    Her kategoriden n_per_category dataset alinir; test seti ve deneysel veri
    ile karsilastirmali Excel raporu uretilir.

    Parameters
    ----------
    models_dir      : PFAZ2 cikti dizini (trained_models)
    datasets_dir    : PFAZ1 cikti dizini (generated_datasets)
    output_dir      : PFAZ13 cikti dizini
    n_trials        : Her kombinasyon icin Optuna trial sayisi
    model_types     : Denenecek model tipleri (None = kayit dosyasindaki tipi kullan)
    n_per_category  : Her kategoriden secilecek max dataset sayisi (varsayilan 25)
    aaa2_txt_path   : Deneysel verilerin bulundugu aaa2.txt yolu (opsiyonel)
    anfis_models_dir: PFAZ3 cikti dizini - ANFIS modelleri icin (opsiyonel)
    """

    # R2 kategori sinirlari
    POOR_MAX      = 0.70
    MEDIUM_MAX    = 0.90
    GOOD_MAX      = 0.95
    EXCELLENT_MIN = 0.95  # R2 >= 0.95 — Excellent, en dusuk skorlular secilir

    def __init__(
        self,
        models_dir: str,
        datasets_dir: str,
        output_dir: str,
        r2_threshold: float = 0.80,   # geri uyumluluk icin tutuldu
        n_trials: int = 30,
        model_types: Optional[List[str]] = None,
        max_retrain: int = 0,
        n_per_category: int = 25,
        aaa2_txt_path: Optional[str] = None,
        anfis_models_dir: Optional[str] = None,
        gpu_enabled: bool = False,
    ):
        self.models_dir      = Path(models_dir)
        self.datasets_dir    = Path(datasets_dir)
        self.output_dir      = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

        self.r2_threshold    = r2_threshold
        self.n_trials        = n_trials
        self.model_types     = model_types or ['rf', 'xgb', 'lgb']
        self.max_retrain     = max_retrain
        self.n_per_category  = n_per_category
        self.anfis_models_dir = Path(anfis_models_dir) if anfis_models_dir else None
        self.gpu_enabled     = gpu_enabled

        # aaa2_txt_path: explicit veya config.json'dan otomatik
        if aaa2_txt_path:
            self.aaa2_txt_path = aaa2_txt_path
        else:
            self.aaa2_txt_path = self._auto_detect_aaa2_path()

        self._improvement_records: List[Dict] = []
        self._anfis_records: List[Dict] = []
        self._prediction_records: List[Dict] = []

    def _auto_detect_aaa2_path(self) -> Optional[str]:
        """config.json'dan veya proje kokundan aaa2.txt yolunu otomatik belirler."""
        # 1) config.json arama — proje koku veya calisiyor dizin
        for cfg_cand in [
            Path('config.json'),
            self.models_dir.parent / 'config.json',
            self.models_dir.parent.parent / 'config.json',
        ]:
            if cfg_cand.exists():
                try:
                    with open(cfg_cand, 'r', encoding='utf-8') as _f:
                        cfg = json.load(_f)
                    data_dir = cfg.get('paths', {}).get('data_dir', 'data')
                    input_file = cfg.get('data', {}).get('input_file', 'aaa2.txt')
                    candidate = Path(data_dir) / input_file
                    if candidate.exists():
                        logger.info(f"[AutoMLRetrain] aaa2_txt_path config'den alindi: {candidate}")
                        return str(candidate)
                    # config.json'un yanindaki dizinden dene
                    candidate2 = cfg_cand.parent / data_dir / input_file
                    if candidate2.exists():
                        logger.info(f"[AutoMLRetrain] aaa2_txt_path config'den alindi: {candidate2}")
                        return str(candidate2)
                except Exception:
                    pass

        # 2) Standart fallback konumlari
        for fallback in [
            Path('data/aaa2.txt'),
            self.models_dir.parent / 'data' / 'aaa2.txt',
            self.models_dir.parent.parent / 'data' / 'aaa2.txt',
        ]:
            if fallback.exists():
                logger.info(f"[AutoMLRetrain] aaa2_txt_path otomatik bulundu: {fallback}")
                return str(fallback)

        logger.info("[AutoMLRetrain] aaa2_txt_path bulunamadi — deneysel karsilastirma atlanacak")
        return None

    # ------------------------------------------------------------------
    # Step 1: Scan PFAZ2 metrics — kategorik secim
    # ------------------------------------------------------------------

    def _extract_val_r2(self, rec: Dict) -> Optional[float]:
        """metrics dict'inden val_r2 cikartir."""
        val_r2 = rec.get('val_r2') or rec.get('Val_R2') or rec.get('val_R2')
        if val_r2 is None:
            for key in ('metrics', 'val_metrics', 'validation', 'val'):
                sub = rec.get(key, {})
                if isinstance(sub, dict):
                    val_r2 = sub.get('r2') or sub.get('R2')
                    if val_r2 is not None:
                        break
        if val_r2 is None:
            return None
        try:
            v = float(val_r2)
            return v if not (v != v) else None  # NaN check
        except (TypeError, ValueError):
            return None

    def find_categorized_candidates(self) -> Dict[str, List[Dict]]:
        """
        Metrics dosyalarini tara, R2 degerine gore kategorile ve
        her kategoriden n_per_category en dusuk skorlu adayi sec.

        Kategoriler:
          poor      : R2 < 0.70
          medium    : 0.70 <= R2 < 0.90
          good      : 0.90 <= R2 < 0.95
          excellent : R2 >= 0.95  (en dusuk 25 aday — marjinal iyilesme aramaak icin)
        """
        all_records = _load_metrics_files(self.models_dir)
        poor, medium, good, excellent = [], [], [], []

        for rec in all_records:
            val_r2 = self._extract_val_r2(rec)
            if val_r2 is None:
                continue

            target     = rec.get('target') or rec.get('Target') or ''
            dataset    = rec.get('dataset_name') or rec.get('dataset') or rec.get('Dataset') or ''
            model_type = (rec.get('model_type') or rec.get('model') or
                          _infer_model_type_from_path(rec.get('_source_file', '')))

            if not target or not dataset or not model_type:
                continue

            entry = {
                'target':        target,
                'dataset':       dataset,
                'model_type':    model_type,
                'before_val_r2': val_r2,
                '_source_file':  rec.get('_source_file', ''),
            }

            if val_r2 < self.POOR_MAX:
                poor.append(entry)
            elif val_r2 < self.MEDIUM_MAX:
                medium.append(entry)
            elif val_r2 < self.GOOD_MAX:
                good.append(entry)
            else:
                excellent.append(entry)  # R2 >= 0.95

        # Her kategoriden en dusuk R2'li n_per_category adayi sec
        poor      = sorted(poor,      key=lambda x: x['before_val_r2'])[:self.n_per_category]
        medium    = sorted(medium,    key=lambda x: x['before_val_r2'])[:self.n_per_category]
        good      = sorted(good,      key=lambda x: x['before_val_r2'])[:self.n_per_category]
        excellent = sorted(excellent, key=lambda x: x['before_val_r2'])[:self.n_per_category]

        logger.info(f"[AutoMLRetrain] {len(all_records)} kayit tarandı.")
        logger.info(f"  Poor      (R2<0.70)   : {len(poor)} secildi")
        logger.info(f"  Medium    (0.70-0.90) : {len(medium)} secildi")
        logger.info(f"  Good      (0.90-0.95) : {len(good)} secildi")
        logger.info(f"  Excellent (R2>=0.95)  : {len(excellent)} secildi")
        return {'poor': poor, 'medium': medium, 'good': good, 'excellent': excellent}

    def find_low_scoring_candidates(self) -> List[Dict]:
        """Geri uyumluluk icin — categorized'dan flat liste doner."""
        cats = self.find_categorized_candidates()
        return cats['poor'] + cats['medium'] + cats['good'] + cats['excellent']

    # ------------------------------------------------------------------
    # Step 2: Retrain a single candidate
    # ------------------------------------------------------------------

    def _retrain_one(self, cand: Dict, category: str = '') -> Optional[Dict]:
        """
        Tek bir (target, dataset, model_type) kombinasyonunu optimize eder.
        Test seti tahminleri de uretilir (deneysel karsilastirma icin).
        """
        if not OPTUNA_AVAILABLE or AutoMLOptimizer is None:
            logger.warning("[AutoMLRetrain] Optuna/AutoMLOptimizer mevcut degil — atlaniyor")
            return None

        target     = cand['target']
        dataset    = cand['dataset']
        model_type = cand['model_type']
        before_r2  = cand['before_val_r2']

        logger.info(f"[AutoMLRetrain] [{category.upper() or 'ALL'}] "
                    f"{target} | {dataset} | {model_type} (before R2={before_r2:.4f})")

        # Tum split'leri (train/val/test) yukle
        full_data = _load_test_data(self.datasets_dir, dataset, target)
        if full_data is None:
            # Fallback: sadece train+val
            split = _load_split_data(self.datasets_dir, dataset, target)
            if split is None:
                logger.warning(f"[AutoMLRetrain] Veri bulunamadı: {dataset}/{target} — atlaniyor")
                return None
            X_train, y_train, X_val, y_val = split
            X_test, y_test, nucleus_ids = None, None, []
        else:
            X_train, y_train, X_val, y_val, X_test, y_test, nucleus_ids = full_data

        # Hangi model tiplerini dene
        best_r2         = before_r2
        best_params     = {}
        best_model_type = model_type
        best_study      = None

        model_candidates = self.model_types if self.model_types else [model_type]

        for mt in model_candidates:
            try:
                opt = AutoMLOptimizer(X_train, y_train, X_val, y_val,
                                     model_type=mt, gpu_enabled=self.gpu_enabled)
                t0 = time.time()
                study = opt.optimize(n_trials=self.n_trials)
                elapsed = time.time() - t0

                after_r2 = study.best_value
                params   = study.best_params

                logger.info(f"  [{mt}] after Val R2={after_r2:.4f}  "
                            f"(delta={after_r2 - before_r2:+.4f})  elapsed={elapsed:.1f}s")

                if after_r2 > best_r2:
                    best_r2         = after_r2
                    best_params     = params
                    best_model_type = mt
                    best_study      = study

            except Exception as e:
                logger.warning(f"  [{mt}] optimize hatasi: {e}")

        improvement = best_r2 - before_r2

        # Test seti tahmini — en iyi parametrelerle yeniden egit ve test'e uygula
        test_preds_after = None
        test_r2_after    = None
        if X_test is not None and best_params and best_study is not None:
            try:
                best_opt = AutoMLOptimizer(X_train, y_train, X_val, y_val, model_type=best_model_type)
                best_model_obj = best_opt._build_model(best_params)
                best_model_obj.fit(X_train, y_train)
                test_preds_after = best_model_obj.predict(X_test).tolist()
                from sklearn.metrics import r2_score
                test_r2_after = float(r2_score(y_test, test_preds_after))
                logger.info(f"  [TEST] after Test R2={test_r2_after:.4f}")
            except Exception as te:
                logger.debug(f"  Test tahmini hatasi: {te}")

        # Tahmin kayitlarini depola
        if X_test is not None and test_preds_after is not None:
            y_test_list = y_test.tolist() if hasattr(y_test, 'tolist') else list(y_test)
            for i, nid in enumerate(nucleus_ids or range(len(y_test_list))):
                exp_val = y_test_list[i] if i < len(y_test_list) else None
                pred_after = test_preds_after[i] if i < len(test_preds_after) else None
                self._prediction_records.append({
                    'nucleus':        str(nid),
                    'target':         target,
                    'dataset':        dataset,
                    'category':       category,
                    'model':          best_model_type,
                    'experimental':   round(float(exp_val), 6) if exp_val is not None else None,
                    'pred_after':     round(float(pred_after), 6) if pred_after is not None else None,
                    'error_after':    round(float(abs(exp_val - pred_after)), 6)
                                      if (exp_val is not None and pred_after is not None) else None,
                })

        record = {
            'target':           target,
            'dataset':          dataset,
            'category':         category,
            'original_model':   model_type,
            'best_model':       best_model_type,
            'before_val_r2':    round(float(before_r2), 6),
            'after_val_r2':     round(float(best_r2), 6),
            'improvement':      round(float(improvement), 6),
            'improved':         improvement > 0.001,
            'after_test_r2':    round(test_r2_after, 6) if test_r2_after is not None else None,
            'n_trials':         self.n_trials,
            'best_params':      best_params,
            'timestamp':        time.strftime('%Y-%m-%d %H:%M:%S'),
        }
        return record

    # ------------------------------------------------------------------
    # Step 3: ANFIS optimization for multiple datasets
    # ------------------------------------------------------------------

    def _run_anfis_for_category(self, candidates: List[Dict], category: str) -> List[Dict]:
        """Bir kategori icin ANFIS optimizasyonunu calistir."""
        anfis_results = []
        try:
            from pfaz_modules.pfaz13_automl.automl_anfis_optimizer import AutoMLANFISOptimizer
        except ImportError:
            try:
                from automl_anfis_optimizer import AutoMLANFISOptimizer
            except ImportError:
                logger.warning("[AutoMLANFIS] AutoMLANFISOptimizer import edilemedi — atlaniyor")
                return []

        anfis_dir = self.output_dir / 'anfis_optimization' / category
        anfis_dir.mkdir(parents=True, exist_ok=True)

        for cand in candidates:
            target  = cand['target']
            dataset = cand['dataset']
            before_r2 = cand['before_val_r2']

            full_data = _load_test_data(self.datasets_dir, dataset, target)
            if full_data is None:
                split = _load_split_data(self.datasets_dir, dataset, target)
                if split is None:
                    logger.warning(f"[AutoMLANFIS] Veri yok: {dataset}/{target}")
                    continue
                X_tr, y_tr, X_val, y_val = split
                X_test, y_test, nucleus_ids = None, None, []
            else:
                X_tr, y_tr, X_val, y_val, X_test, y_test, nucleus_ids = full_data

            try:
                _anfis_opt = AutoMLANFISOptimizer(
                    n_trials=min(20, self.n_trials),
                    timeout=300,
                    output_dir=str(anfis_dir),
                    use_logging=False,
                    use_matlab=False,
                )
                y_tr_1d  = y_tr.ravel()  if y_tr.ndim  > 1 else y_tr
                y_val_1d = y_val.ravel() if y_val.ndim > 1 else y_val
                result = _anfis_opt.optimize(
                    X_train=X_tr, y_train=y_tr_1d,
                    X_val=X_val,  y_val=y_val_1d,
                    dataset_name=dataset,
                    objective='r2',
                )
                after_r2 = result.get('best_score', before_r2)
                best_cfg  = result.get('best_config', {})

                # Test seti tahmini
                test_r2_after = None
                if X_test is not None:
                    try:
                        from pfaz_modules.pfaz03_anfis_training.anfis_core import TakagiSugenoANFIS
                        from sklearn.preprocessing import StandardScaler
                        scaler = StandardScaler()
                        X_sc = scaler.fit_transform(X_tr)
                        mf_type = best_cfg.get('mf_type', 'gaussian')
                        n_mfs   = best_cfg.get('n_mfs', 2)
                        anfis = TakagiSugenoANFIS(n_inputs=X_sc.shape[1],
                                                   n_mfs=n_mfs, mf_type=mf_type)
                        anfis.fit(X_sc, y_tr_1d)
                        Xt_sc = scaler.transform(X_test)
                        pred_test = anfis.predict(Xt_sc)
                        from sklearn.metrics import r2_score
                        y_t_1d = y_test.ravel() if y_test.ndim > 1 else y_test
                        test_r2_after = float(r2_score(y_t_1d, pred_test))
                    except Exception as te:
                        logger.debug(f"[AutoMLANFIS] Test tahmini hatasi: {te}")

                rec = {
                    'target':        target,
                    'dataset':       dataset,
                    'category':      category,
                    'before_val_r2': round(float(before_r2), 6),
                    'after_val_r2':  round(float(after_r2), 6),
                    'improvement':   round(float(after_r2 - before_r2), 6),
                    'improved':      (after_r2 - before_r2) > 0.001,
                    'after_test_r2': round(test_r2_after, 6) if test_r2_after is not None else None,
                    'best_config':   str(best_cfg),
                    'n_trials':      min(20, self.n_trials),
                    'timestamp':     time.strftime('%Y-%m-%d %H:%M:%S'),
                }
                anfis_results.append(rec)
                logger.info(f"  [ANFIS/{category}] {dataset}/{target}: "
                            f"R2 {before_r2:.4f} -> {after_r2:.4f}")
            except Exception as ae:
                logger.warning(f"[AutoMLANFIS] {dataset}/{target} hatasi: {ae}")

        return anfis_results

    # ------------------------------------------------------------------
    # Step 4: Run all candidates (kategorik)
    # ------------------------------------------------------------------

    def run(self) -> Dict:
        """
        Kategorik secim (Poor/Medium/Good) ile AI + ANFIS optimize et,
        test seti ve deneysel veri karsilastirmali rapor uret.
        """
        categorized = self.find_categorized_candidates()
        all_candidates = (categorized['poor'] + categorized['medium'] +
                          categorized['good'] + categorized.get('excellent', []))

        if not all_candidates:
            logger.info("[AutoMLRetrain] Optimize edilecek aday yok.")
            self._save_empty_report()
            return {'status': 'no_candidates'}

        logger.info(f"[AutoMLRetrain] Toplam: {len(all_candidates)} AI adayi "
                    f"(poor={len(categorized['poor'])}, "
                    f"medium={len(categorized['medium'])}, "
                    f"good={len(categorized['good'])}, "
                    f"excellent={len(categorized.get('excellent', []))})")

        # --- AI Optimizasyonu ---
        for category, cands in categorized.items():
            logger.info(f"\n[AutoMLRetrain] === {category.upper()} kategorisi "
                        f"({len(cands)} dataset) ===")
            for cand in cands:
                record = self._retrain_one(cand, category=category)
                if record:
                    self._improvement_records.append(record)

        # --- AI özet log ---
        n_ai_improved = sum(1 for r in self._improvement_records if r.get('improved'))
        logger.info(f"\n[AutoMLRetrain] AI tamamlandi: {len(self._improvement_records)} "
                    f"kombinasyon, {n_ai_improved} iyilesti")

        # --- ANFIS Optimizasyonu (her kategori icin) ---
        logger.info("\n[AutoMLRetrain] === ANFIS Optimizasyonu basliyor ===")
        for category, cands in categorized.items():
            if cands:
                anfis_recs = self._run_anfis_for_category(cands, category)
                self._anfis_records.extend(anfis_recs)

        n_anfis_improved = sum(1 for r in self._anfis_records if r.get('improved'))
        logger.info(f"[AutoMLRetrain] ANFIS tamamlandi: {len(self._anfis_records)} "
                    f"kombinasyon, {n_anfis_improved} iyilesti")

        # --- AutoMLFeatureEngineer (en iyi iyilesen icin) ---
        fe_result = None
        try:
            from pfaz_modules.pfaz13_automl.automl_feature_engineer import AutoMLFeatureEngineer
            _best_rec = max(self._improvement_records,
                            key=lambda r: r.get('improvement', 0.0),
                            default=None)
            if _best_rec:
                _split = _load_split_data(self.datasets_dir,
                                           _best_rec['dataset'], _best_rec['target'])
                if _split is not None:
                    _X_tr, _y_tr, _X_val, _y_val = _split
                    _n_feats = _X_tr.shape[1]
                    _feat_names = [f'f{i}' for i in range(_n_feats)]
                    _fe = AutoMLFeatureEngineer(
                        polynomial_degree=2, include_transforms=True,
                        include_interactions=True,
                        target_n_features=min(40, max(10, _n_feats * 3)),
                        selection_methods=['rfe', 'lasso', 'mutual_info'],
                        output_dir=str(self.output_dir / 'feature_engineering')
                    )
                    _y_1d = _y_tr.ravel() if _y_tr.ndim > 1 else _y_tr
                    _X_eng, _selected_names = _fe.fit_transform(
                        _X_tr, _y_1d, _feat_names, X_val=_X_val)
                    fe_result = {
                        'dataset':           _best_rec['dataset'],
                        'target':            _best_rec['target'],
                        'original_features': _n_feats,
                        'selected_features': len(_selected_names),
                    }
                    logger.info(f"[AutoMLFeatEng] {_n_feats} -> {len(_selected_names)} ozellik")
        except Exception as _fe_e:
            logger.warning(f"[WARNING] AutoMLFeatureEngineer basarisiz: {_fe_e}")

        # --- JSON + Excel ---
        self._save_json_log()
        excel_path = self._save_excel_report()

        # --- AutoMLVisualizer ---
        try:
            from pfaz_modules.pfaz13_automl.automl_visualizer import AutoMLVisualizer
            _aml_xlsx = self.output_dir / 'automl_improvement_report.xlsx'
            if _aml_xlsx.exists():
                _aml_viz = AutoMLVisualizer(
                    excel_path=str(_aml_xlsx),
                    output_dir=str(self.output_dir / 'automl_visualizations')
                )
                _aml_viz.generate_all_plots()
                logger.info("[OK] AutoMLVisualizer -> automl_visualizations/")
        except ImportError:
            pass
        except Exception as _av_e:
            logger.warning(f"[WARNING] AutoMLVisualizer basarisiz: {_av_e}")

        return {
            'status':            'completed',
            'candidates_found':  len(all_candidates),
            'ai_retrained':      len(self._improvement_records),
            'ai_improved':       n_ai_improved,
            'anfis_retrained':   len(self._anfis_records),
            'anfis_improved':    n_anfis_improved,
            'excel_report':      str(excel_path) if excel_path else None,
            'records':           self._improvement_records,
            'anfis_records':     self._anfis_records,
            'feature_engineering': fe_result,
        }

    # ------------------------------------------------------------------
    # Reporting
    # ------------------------------------------------------------------

    def _save_json_log(self):
        log_path = self.output_dir / 'automl_retraining_log.json'
        try:
            combined = {
                'ai_records':    self._improvement_records,
                'anfis_records': self._anfis_records,
            }
            with open(log_path, 'w') as f:
                json.dump(combined, f, indent=2)
            logger.info(f"[AutoMLRetrain] JSON log: {log_path}")
        except Exception as e:
            logger.warning(f"[AutoMLRetrain] JSON kayit hatasi: {e}")

    def _save_empty_report(self):
        """Aday yoksa boş rapor yaz."""
        log_path = self.output_dir / 'automl_retraining_log.json'
        try:
            with open(log_path, 'w') as f:
                json.dump([], f)
        except Exception:
            pass

    def _save_excel_report(self) -> Optional[Path]:
        """
        automl_improvement_report.xlsx — 6 sheet:
          1. AI_Summary        : AI per-combination ozet (kategori sutunlu)
          2. ANFIS_Summary     : ANFIS per-combination ozet
          3. Predictions_AI    : Test seti uzerinde tahmin vs deneysel
          4. Predictions_ANFIS : ANFIS test tahminleri (varsa)
          5. Best_Params       : En iyi hiperparametreler
          6. Overview          : Ozet istatistikler
        """
        if not OPENPYXL_AVAILABLE:
            logger.warning("[AutoMLRetrain] openpyxl yok — Excel raporu atlandi")
            return None
        if not self._improvement_records and not self._anfis_records:
            return None

        out_path = self.output_dir / 'automl_improvement_report.xlsx'

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

            GREEN_FILL  = PatternFill('solid', fgColor='C6EFCE')
            RED_FILL    = PatternFill('solid', fgColor='FFC7CE')
            YELLOW_FILL = PatternFill('solid', fgColor='FFEB9C')
            BLUE_FILL   = PatternFill('solid', fgColor='9DC3E6')
            HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
            HEADER_FONT = Font(bold=True, color='FFFFFF')
            CAT_FILLS   = {
                'poor':   PatternFill('solid', fgColor='FFD7D7'),
                'medium': PatternFill('solid', fgColor='FFF2CC'),
                'good':   PatternFill('solid', fgColor='D9EAD3'),
            }

            def _write_header(ws, cols):
                for ci, hdr in enumerate(cols, 1):
                    c = ws.cell(row=1, column=ci, value=hdr)
                    c.font = HEADER_FONT; c.fill = HEADER_FILL
                    c.alignment = Alignment(horizontal='center')

            def _auto_width(ws, max_w=50):
                for col in ws.columns:
                    ml = max((len(str(c.value or '')) for c in col), default=8)
                    ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 3, max_w)

            wb = Workbook()

            # ==================================================================
            # Sheet 1: AI_Summary
            # ==================================================================
            ws1 = wb.active
            ws1.title = 'AI_Summary'
            ai_cols = [
                'Kategori', 'Hedef', 'Dataset', 'Orijinal Model', 'En Iyi Model',
                'Once Val R2', 'Sonra Val R2', 'Iyilesme', 'Test R2 (After)',
                'Iyilesti?', 'Deneme Sayisi', 'Zaman',
            ]
            _write_header(ws1, ai_cols)
            for ri, rec in enumerate(self._improvement_records, 2):
                cat = rec.get('category', '')
                vals = [
                    cat.upper(),
                    rec.get('target'),          rec.get('dataset'),
                    rec.get('original_model'),  rec.get('best_model'),
                    rec.get('before_val_r2'),   rec.get('after_val_r2'),
                    rec.get('improvement'),     rec.get('after_test_r2'),
                    'EVET' if rec.get('improved') else 'HAYIR',
                    rec.get('n_trials'),        rec.get('timestamp'),
                ]
                for ci, val in enumerate(vals, 1):
                    ws1.cell(row=ri, column=ci, value=val)
                row_fill = CAT_FILLS.get(cat, GREEN_FILL if rec.get('improved') else RED_FILL)
                for ci in range(1, len(ai_cols) + 1):
                    ws1.cell(row=ri, column=ci).fill = row_fill
            _auto_width(ws1)

            # ==================================================================
            # Sheet 2: ANFIS_Summary
            # ==================================================================
            ws2 = wb.create_sheet('ANFIS_Summary')
            anfis_cols = [
                'Kategori', 'Hedef', 'Dataset',
                'Once Val R2', 'Sonra Val R2', 'Iyilesme', 'Test R2 (After)',
                'Iyilesti?', 'En Iyi Config', 'Deneme Sayisi', 'Zaman',
            ]
            _write_header(ws2, anfis_cols)
            for ri, rec in enumerate(self._anfis_records, 2):
                cat = rec.get('category', '')
                vals = [
                    cat.upper(),
                    rec.get('target'),          rec.get('dataset'),
                    rec.get('before_val_r2'),   rec.get('after_val_r2'),
                    rec.get('improvement'),     rec.get('after_test_r2'),
                    'EVET' if rec.get('improved') else 'HAYIR',
                    rec.get('best_config'),     rec.get('n_trials'),
                    rec.get('timestamp'),
                ]
                for ci, val in enumerate(vals, 1):
                    ws2.cell(row=ri, column=ci, value=val)
                row_fill = CAT_FILLS.get(cat, GREEN_FILL if rec.get('improved') else RED_FILL)
                for ci in range(1, len(anfis_cols) + 1):
                    ws2.cell(row=ri, column=ci).fill = row_fill
            _auto_width(ws2)

            # ==================================================================
            # Sheet 3: Predictions_AI — tahmin vs deneysel
            # ==================================================================
            ws3 = wb.create_sheet('Predictions_AI')
            pred_cols = [
                'Cekirdek', 'Hedef', 'Dataset', 'Kategori', 'Model',
                'Deneysel Deger', 'Tahmin (Sonra)', 'Mutlak Hata (Sonra)',
            ]
            _write_header(ws3, pred_cols)
            for ri, prec in enumerate(self._prediction_records, 2):
                vals = [
                    prec.get('nucleus'),    prec.get('target'),
                    prec.get('dataset'),    prec.get('category'),
                    prec.get('model'),      prec.get('experimental'),
                    prec.get('pred_after'), prec.get('error_after'),
                ]
                for ci, val in enumerate(vals, 1):
                    ws3.cell(row=ri, column=ci, value=val)
                # Hata rengine gore renk
                err = prec.get('error_after')
                if err is not None:
                    fill = GREEN_FILL if err < 0.1 else (YELLOW_FILL if err < 0.3 else RED_FILL)
                    for ci in range(1, len(pred_cols) + 1):
                        ws3.cell(row=ri, column=ci).fill = fill
            _auto_width(ws3)

            # ==================================================================
            # Sheet 4: Predictions_ANFIS (gecici — ileride doldurulabilir)
            # ==================================================================
            ws4 = wb.create_sheet('Predictions_ANFIS')
            _write_header(ws4, ['Hedef', 'Dataset', 'Kategori', 'Before R2', 'After R2', 'Test R2', 'Iyilesti?'])
            for ri, rec in enumerate(self._anfis_records, 2):
                vals = [
                    rec.get('target'), rec.get('dataset'), rec.get('category'),
                    rec.get('before_val_r2'), rec.get('after_val_r2'),
                    rec.get('after_test_r2'),
                    'EVET' if rec.get('improved') else 'HAYIR',
                ]
                for ci, val in enumerate(vals, 1):
                    ws4.cell(row=ri, column=ci, value=val)
                fill = GREEN_FILL if rec.get('improved') else RED_FILL
                for ci in range(1, 8):
                    ws4.cell(row=ri, column=ci).fill = fill
            _auto_width(ws4)

            # ==================================================================
            # Sheet 5: Best_Params
            # ==================================================================
            ws5 = wb.create_sheet('Best_Params')
            _write_header(ws5, ['Kategori', 'Hedef', 'Dataset', 'Model', 'Parametre', 'Deger'])
            ri5 = 2
            for rec in self._improvement_records:
                for pname, pval in (rec.get('best_params') or {}).items():
                    ws5.cell(row=ri5, column=1, value=rec.get('category', '').upper())
                    ws5.cell(row=ri5, column=2, value=rec.get('target'))
                    ws5.cell(row=ri5, column=3, value=rec.get('dataset'))
                    ws5.cell(row=ri5, column=4, value=rec.get('best_model'))
                    ws5.cell(row=ri5, column=5, value=pname)
                    ws5.cell(row=ri5, column=6, value=str(pval))
                    ri5 += 1
            _auto_width(ws5)

            # ==================================================================
            # Sheet 6: Overview
            # ==================================================================
            ws6 = wb.create_sheet('Overview')

            def _ov_row(ws, label, val, ri):
                ws.cell(row=ri, column=1, value=label)
                ws.cell(row=ri, column=2, value=val)

            _write_header(ws6, ['Metrik', 'Deger'])
            ri6 = 2
            ai_recs = self._improvement_records
            an_recs = self._anfis_records

            # Kategori bazli AI ozet
            for cat in ('poor', 'medium', 'good'):
                cat_recs = [r for r in ai_recs if r.get('category') == cat]
                if not cat_recs:
                    continue
                n_imp = sum(1 for r in cat_recs if r.get('improved'))
                avg_imp = float(np.mean([r['improvement'] for r in cat_recs]))
                _ov_row(ws6, f'AI [{cat.upper()}] toplam', len(cat_recs), ri6); ri6 += 1
                _ov_row(ws6, f'AI [{cat.upper()}] iyilesti', n_imp, ri6); ri6 += 1
                _ov_row(ws6, f'AI [{cat.upper()}] ort. iyilesme', round(avg_imp, 4), ri6); ri6 += 1
                ws6.cell(row=ri6-3, column=1).fill = CAT_FILLS[cat]

            # Genel AI
            _ov_row(ws6, '--- AI Genel ---', '', ri6); ri6 += 1
            _ov_row(ws6, 'AI Toplam optimize',  len(ai_recs), ri6); ri6 += 1
            _ov_row(ws6, 'AI Iyilesen',          sum(1 for r in ai_recs if r.get('improved')), ri6); ri6 += 1
            if ai_recs:
                _ov_row(ws6, 'AI Max iyilesme',
                        round(float(max(r['improvement'] for r in ai_recs)), 4), ri6); ri6 += 1
                test_r2s = [r['after_test_r2'] for r in ai_recs if r.get('after_test_r2') is not None]
                if test_r2s:
                    _ov_row(ws6, 'AI Ort. Test R2 (sonra)',
                            round(float(np.mean(test_r2s)), 4), ri6); ri6 += 1

            # ANFIS ozet
            _ov_row(ws6, '--- ANFIS ---', '', ri6); ri6 += 1
            _ov_row(ws6, 'ANFIS Toplam optimize', len(an_recs), ri6); ri6 += 1
            _ov_row(ws6, 'ANFIS Iyilesen', sum(1 for r in an_recs if r.get('improved')), ri6); ri6 += 1
            if an_recs:
                a_test = [r['after_test_r2'] for r in an_recs if r.get('after_test_r2') is not None]
                if a_test:
                    _ov_row(ws6, 'ANFIS Ort. Test R2 (sonra)',
                            round(float(np.mean(a_test)), 4), ri6); ri6 += 1

            # Deneysel karsilastirma ozeti
            _ov_row(ws6, '--- Deneysel Karsilastirma ---', '', ri6); ri6 += 1
            if self._prediction_records:
                errs = [r['error_after'] for r in self._prediction_records
                        if r.get('error_after') is not None]
                if errs:
                    _ov_row(ws6, 'Ort. mutlak hata (test)',  round(float(np.mean(errs)), 4), ri6); ri6 += 1
                    _ov_row(ws6, 'Max mutlak hata (test)',   round(float(max(errs)), 4), ri6); ri6 += 1
                    _ov_row(ws6, 'Hata < 0.10 olan cekirdek', sum(1 for e in errs if e < 0.10), ri6); ri6 += 1

            # Parametreler
            _ov_row(ws6, '--- Parametreler ---', '', ri6); ri6 += 1
            _ov_row(ws6, 'n_per_category', self.n_per_category, ri6); ri6 += 1
            _ov_row(ws6, 'n_trials',       self.n_trials,       ri6); ri6 += 1
            _ov_row(ws6, 'Kategoriler',    'Poor<0.70 / Medium 0.70-0.90 / Good 0.90-0.95', ri6)

            ws6.column_dimensions['A'].width = 42
            ws6.column_dimensions['B'].width = 22

            wb.save(out_path)
            logger.info(f"[AutoMLRetrain] Excel raporu kaydedildi: {out_path}")
            return out_path

        except Exception as e:
            logger.warning(f"[AutoMLRetrain] Excel kayit hatasi: {e}")
            return None

    # ------------------------------------------------------------------
    # Convenience: get records for PFAZ6 sheet
    # ------------------------------------------------------------------

    def get_improvement_records(self) -> List[Dict]:
        """PFAZ6 entegrasyonu için iyileşme kayıtlarını döner."""
        return self._improvement_records
