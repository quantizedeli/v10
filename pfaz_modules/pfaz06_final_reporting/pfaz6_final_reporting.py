"""
PFAZ 6: Final Reporting & Thesis Integration
Tum sonuclarin toplanmasi ve thesis-ready raporlar

v2.0 - Kapsamli yeniden yazim:
  - ANFIS training_meta alanlari (mf_type, n_rules, n_inputs vb.)
  - Unknown predictions toplama
  - Dataset adından target/size/scenario/feature_set parse
  - Feature abbreviation tablosu (PFAZ_DEVELOPMENT_NOTES todo)
  - Anomali vs NoAnomaly karsilastirma (PFAZ_DEVELOPMENT_NOTES todo)
  - Best models per target + best feature set per target
  - openpyxl: conditional formatting, freeze panes, autofilter, renk kodlamasi
  - Overview/Dashboard sheet
"""

import pandas as pd
import numpy as np
from pathlib import Path
import json
import logging
from datetime import datetime

logging.basicConfig(level=logging.INFO, format='%(message)s')
logger = logging.getLogger(__name__)

# =============================================================================
# FEATURE ABBREVIATION TABLE (PFAZ1 DEVELOPMENT NOTES'tan)
# =============================================================================
FEATURE_ABBREV = {
    'A':    ('A',                'Kutle numarasi'),
    'Z':    ('Z',                'Proton sayisi'),
    'N':    ('N',                'Notron sayisi'),
    'S':    ('SPIN',             'Nukleus spini'),
    'PAR':  ('PARITY',           'Parite (+1/-1)'),
    'MC':   ('magic_character',  'Kabuk yapisi skoru (0-1)'),
    'BEPA': ('BE_per_A',         'Baglanma enerjisi / nukleon'),
    'B2E':  ('Beta_2_estimated', 'Hesaplanan deformasyon beta_2'),
    'ZMD':  ('Z_magic_dist',     'Z magic sayiya uzaklik'),
    'NMD':  ('N_magic_dist',     'N magic sayiya uzaklik'),
    'BEA':  ('BE_asymmetry',     'Asimetri baglanma enerjisi'),
    'ZV':   ('Z_valence',        'Valans proton sayisi'),
    'NV':   ('N_valence',        'Valans notron sayisi'),
    'ZSG':  ('Z_shell_gap',      'Proton shell gap enerjisi (MeV)'),
    'NSG':  ('N_shell_gap',      'Notron shell gap enerjisi (MeV)'),
    'BEP':  ('BE_pairing',       'Pairing enerjisi'),
    'SPHI': ('spherical_index',  'Kuresellik indeksi (0-1)'),
    'CP':   ('Q0_intrinsic',     'Kolektif/intrinsik kuadrupol'),
    'PF':   ('P_FACTOR',         'P-faktoru'),
    'BET':  ('BE_total',         'Toplam baglanma enerjisi'),
    'SN':   ('S_n_approx',       'Notron ayrilma enerjisi (yaklasik)'),
    'SP':   ('S_p_approx',       'Proton ayrilma enerjisi (yaklasik)'),
    'NN':   ('Nn',               'Valans notron sayisi (aaa2.txt ham sutunu)'),
    'NP':   ('Np',               'Valans proton sayisi (aaa2.txt ham sutunu)'),
}

# Bilinen feature set kombinasyonlari (kisaltma -> acik list)
FEATURE_SET_EXPAND = {
    'AZS':       ['A', 'Z', 'S'],
    'AZMC':      ['A', 'Z', 'MC'],
    'AZBEPA':    ['A', 'Z', 'BEPA'],
    'ASMC':      ['A', 'S', 'MC'],
    'AMCBEPA':   ['A', 'MC', 'BEPA'],
    'NNPMC':     ['NN', 'NP', 'MC'],
    'AZSMC':     ['A', 'Z', 'S', 'MC'],
    'AZSBEPA':   ['A', 'Z', 'S', 'BEPA'],
    'AZMCBEPA':  ['A', 'Z', 'MC', 'BEPA'],
    'AZSB2E':    ['A', 'Z', 'S', 'B2E'],
    'AZNNP':     ['A', 'Z', 'NN', 'NP'],
    'AZSMCBEPA': ['A', 'Z', 'S', 'MC', 'BEPA'],
    'AZSMCB2E':  ['A', 'Z', 'S', 'MC', 'B2E'],
    'AZSNNNP':   ['A', 'Z', 'S', 'NN', 'NP'],
    'AZB2E':     ['A', 'Z', 'B2E'],
    'ZB2EMC':    ['Z', 'B2E', 'MC'],
    'B2EMCBEA':  ['B2E', 'MC', 'BEA'],
    'AZB2EMC':   ['A', 'Z', 'B2E', 'MC'],
    'ZB2EMCS':   ['Z', 'B2E', 'MC', 'S'],
    'AZB2EBEA':  ['A', 'Z', 'B2E', 'BEA'],
    'ZNNPMC':    ['Z', 'NN', 'NP', 'MC'],
    'AZB2EMCS':  ['A', 'Z', 'B2E', 'MC', 'S'],
    'AZB2EMCBEA':['A', 'Z', 'B2E', 'MC', 'BEA'],
    'AZNNPMC':   ['A', 'Z', 'NN', 'NP', 'MC'],
    'AZN':       ['A', 'Z', 'N'],
    'MCZMNM':    ['MC', 'ZMD', 'NMD'],
    'AZVNV':     ['A', 'ZV', 'NV'],
    'ZMNMBEA':   ['ZMD', 'NMD', 'BEA'],
    'MCZMNMZV':  ['MC', 'ZMD', 'NMD', 'ZV'],
    'MCZMNMBEA': ['MC', 'ZMD', 'NMD', 'BEA'],
    'AMCZMNM':   ['A', 'MC', 'ZMD', 'NMD'],
    'ZVNVZMNM':  ['ZV', 'NV', 'ZMD', 'NMD'],
    'ZNNPMC':    ['Z', 'NN', 'NP', 'MC'],
    'MCZMNMZVNV':['MC', 'ZMD', 'NMD', 'ZV', 'NV'],
    'AMCZMNMBEA':['A', 'MC', 'ZMD', 'NMD', 'BEA'],
}

# openpyxl renk sabitleri
CLR_HEADER    = 'FF2C3E50'   # koyu lacivert
CLR_TITLE     = 'FF3498DB'   # mavi
CLR_EXCEL_BG  = 'FFDCE6F1'  # acik mavi
CLR_EXCEL_ALT = 'FFF0F5FB'  # daha acik mavi
CLR_EXCELLENT = 'FF00B050'   # koyu yesil
CLR_GOOD      = 'FF92D050'   # yesil
CLR_MEDIUM    = 'FFFFC000'   # sari
CLR_POOR      = 'FFFF0000'   # kirmizi
CLR_WHITE     = 'FFFFFFFF'


def _parse_dataset_name(name: str) -> dict:
    """
    Dataset adini parcala: TARGET_SIZE_SCENARIO_FEATURESET[_NoAnomaly]
    Ornek: MM_150_S70_AZSMC_NoScaling_Random_NoAnomaly
    Donus: {'target': 'MM', 'size': '150', 'scenario': 'S70',
             'feature_set': 'AZSMC', 'no_anomaly': True,
             'n_inputs': 4}
    """
    result = {'target': '', 'size': '', 'scenario': '',
              'feature_set': '', 'no_anomaly': False, 'n_inputs': 0,
              'feature_names': ''}

    # NoAnomaly kontrolü
    if name.endswith('_NoAnomaly'):
        result['no_anomaly'] = True
        name = name[:-len('_NoAnomaly')]

    # Iki-parcali target tespiti
    if name.startswith('Beta_2_'):
        result['target'] = 'Beta_2'
        rest = name[len('Beta_2_'):]
    elif name.startswith('MM_QM_'):
        result['target'] = 'MM_QM'
        rest = name[len('MM_QM_'):]
    elif name.startswith('MM_'):
        result['target'] = 'MM'
        rest = name[len('MM_'):]
    elif name.startswith('QM_'):
        result['target'] = 'QM'
        rest = name[len('QM_'):]
    else:
        parts = name.split('_')
        result['target'] = parts[0] if parts else name
        rest = '_'.join(parts[1:])

    # rest: SIZE_SCENARIO_FEATURESET_NoScaling_Random
    parts = rest.split('_')
    if parts:
        result['size'] = parts[0]   # 75/100/150/200/ALL
    if len(parts) > 1:
        result['scenario'] = parts[1]  # S70/S80
    if len(parts) > 2:
        result['feature_set'] = parts[2]  # AZSMC vb.
        fs = parts[2]
        result['n_inputs'] = len(FEATURE_SET_EXPAND.get(fs, []))
        # Feature isimlerini genislet
        expanded = FEATURE_SET_EXPAND.get(fs, [])
        result['feature_names'] = ', '.join(
            FEATURE_ABBREV.get(ab, (ab, ''))[0] for ab in expanded
        )

    return result


def _write_df_chunked(df: 'pd.DataFrame', writer, base_sheet_name: str,
                      max_rows: int = 50_000) -> int:
    """
    DataFrame'i ExcelWriter'a yazar; max_rows satırı geçerse birden fazla sheet'e böler.

    Sheet adı  : base_sheet_name           (< max_rows)
                 base_sheet_name_2         (2. parça)
                 base_sheet_name_3         (3. parça)  ...

    Returns: yazılan sheet sayısı
    """
    if len(df) <= max_rows:
        df.to_excel(writer, sheet_name=base_sheet_name[:31], index=False)
        return 1

    n_chunks = 0
    for i, start in enumerate(range(0, len(df), max_rows)):
        chunk = df.iloc[start: start + max_rows]
        suffix = '' if i == 0 else f'_{i + 1}'
        sheet = (base_sheet_name + suffix)[:31]
        chunk.to_excel(writer, sheet_name=sheet, index=False)
        n_chunks += 1
    return n_chunks


def _format_worksheet_inline(ws, n_data_rows: int = 0, skip_cells: bool = False):
    """
    Tek bir openpyxl worksheet'ini formatla (ExcelWriter context icinde cagrilir).
    skip_cells=True: sadece baslik, freeze, autofilter (buyuk sheet'ler icin)
    """
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.formatting.rule import ColorScaleRule
        from openpyxl.utils import get_column_letter
    except ImportError:
        return

    header_fill  = PatternFill('solid', fgColor=CLR_HEADER)
    header_font  = Font(bold=True, color=CLR_WHITE, size=11)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    alt_fill     = PatternFill('solid', fgColor=CLR_EXCEL_ALT)
    thin         = Side(style='thin', color='FFB0B0B0')
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    max_col = ws.max_column or 1
    max_row = ws.max_row   or 1

    # Baslik satiri
    for cell in ws[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align

    if not skip_cells and max_row <= 8000:
        # Alternatif satir rengi + border
        for row_idx in range(2, max_row + 1):
            fill = alt_fill if row_idx % 2 == 0 else None
            for cell in ws[row_idx]:
                if fill:
                    cell.fill = fill
                cell.border = border

        # Sutun genislikleri (ilk 300 satir ornekle)
        for col in ws.iter_cols(max_row=min(max_row, 300)):
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 55)
    else:
        # Buyuk sheet: sadece baslik uzunluguna gore genislik
        for col_idx in range(1, max_col + 1):
            hdr = ws.cell(row=1, column=col_idx).value
            col_letter = get_column_letter(col_idx)
            w = len(str(hdr)) + 4 if hdr else 12
            ws.column_dimensions[col_letter].width = min(max(w, 10), 40)

    # R2 sutunlarina renk olcegi (range bazli — hizli)
    for col_idx in range(1, max_col + 1):
        hdr = ws.cell(row=1, column=col_idx).value
        if hdr and isinstance(hdr, str) and ('R2' in hdr or '_r2' in hdr.lower()):
            col_letter = get_column_letter(col_idx)
            if max_row > 1:
                ws.conditional_formatting.add(
                    f'{col_letter}2:{col_letter}{max_row}',
                    ColorScaleRule(
                        start_type='min', start_color='FFFF4444',
                        mid_type='num',   mid_value=0.7, mid_color='FFFFFF00',
                        end_type='max',   end_color='FF00AA00'
                    )
                )

    # Freeze panes
    ws.freeze_panes = 'A2'

    # AutoFilter
    if ws.title not in ('Overview', 'Feature_Abbreviations'):
        ws.auto_filter.ref = ws.dimensions


def _apply_openpyxl_formatting(excel_file: Path):
    """Artik kullanilmiyor — formatlama inline yapiliyor. Geriye donuk uyumluluk."""
    logger.info(f"  [INFO] Inline formatlama kullaniliyor, post-process atlandi")


class FinalReportingPipeline:
    """Thesis icin final raporlama sistemi v2.0"""

    def __init__(self, reports_dir=None, output_dir='outputs/reports',
                 ai_models_dir=None, anfis_models_dir=None,
                 use_excel_charts: bool = True, use_latex_generator: bool = True):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

        base = self.output_dir.parent  # outputs/
        self.ai_models_dir    = Path(ai_models_dir)    if ai_models_dir    else base / 'trained_models'
        self.anfis_models_dir = Path(anfis_models_dir) if anfis_models_dir else base / 'anfis_models'
        self.cross_model_dir  = base / 'cross_model_analysis'
        self.unknown_dir      = base / 'unknown_predictions'
        self.validation_dir   = self.ai_models_dir / 'model_validation'
        self.datasets_dir     = base / 'generated_datasets'
        self.aaa2_txt_path    = None  # Set externally if isotope chain analysis needed
        self.pfaz9_output_dir = None  # Set externally for MC summary sheet
        self.pfaz13_output_dir = None  # Set externally for AutoML improvements sheet

        self.all_results = {}
        self.use_excel_charts    = use_excel_charts
        self.use_latex_generator = use_latex_generator

        logger.info(f"AI Models Dir : {self.ai_models_dir}")
        logger.info(f"ANFIS Dir     : {self.anfis_models_dir}")
        logger.info(f"Output Dir    : {self.output_dir}")

    # =========================================================================
    # DATA COLLECTION
    # =========================================================================

    def collect_all_results(self):
        """Tum fazlardan sonuclari topla"""
        logger.info("\n" + "="*80)
        logger.info("TUM SONUCLAR TOPLANIYOR")
        logger.info("="*80)

        logger.info("\n1. AI Model Sonuclari...")
        self._collect_ai_results()

        logger.info("2. ANFIS Sonuclari...")
        self._collect_anfis_results()

        logger.info("3. Saglamlik (CV) Sonuclari...")
        self._collect_robustness_results()

        logger.info("4. Cross-Model Analiz...")
        self._collect_crossmodel_results()

        logger.info("5. Unknown Predictions...")
        self._collect_unknown_predictions()

        n_ai    = len(self.all_results.get('ai_rows', []))
        n_anfis = len(self.all_results.get('anfis_results', []))
        n_unk   = len(self.all_results.get('unknown_rows', []))
        logger.info(f"\n[OK] AI konfigurasyonlari: {n_ai}")
        logger.info(f"[OK] ANFIS kayitlari      : {n_anfis}")
        logger.info(f"[OK] Unknown predictions  : {n_unk}")
        return self.all_results

    def _collect_ai_results(self):
        """
        AI model sonuclarini gercek dizin yapisindan topla.
        Yapi: {ai_models_dir}/{dataset_name}/{model_type}/{config_id}/metrics_{config_id}.json
        """
        self.all_results['ai_models'] = {}
        self.all_results['ai_rows']   = []

        if not self.ai_models_dir.exists():
            logger.warning(f"  [WARN] AI models dir bulunamadi: {self.ai_models_dir}")
            return

        count = 0
        skip_dirs = {'model_validation', 'seed_reports', 'hyperparameter_tuning'}
        for dataset_dir in sorted(self.ai_models_dir.iterdir()):
            if not dataset_dir.is_dir() or dataset_dir.name in skip_dirs:
                continue
            dataset_name = dataset_dir.name
            parsed = _parse_dataset_name(dataset_name)

            for model_type_dir in sorted(dataset_dir.iterdir()):
                if not model_type_dir.is_dir():
                    continue
                model_type = model_type_dir.name

                for config_dir in sorted(model_type_dir.iterdir()):
                    if not config_dir.is_dir():
                        continue
                    config_id = config_dir.name

                    metrics_file = config_dir / f"metrics_{config_id}.json"
                    if not metrics_file.exists():
                        continue

                    try:
                        with open(metrics_file, encoding='utf-8') as f:
                            m = json.load(f)
                    except Exception:
                        continue

                    # Skip diverged DNN models (val_R2 < -2 flag)
                    if m.get('diverged', False):
                        logger.debug(f"  [SKIP] Diverged DNN: {config_dir}")
                        continue

                    # Skip models with extreme R2 values (corruption guard)
                    train_m = m.get('train', {})
                    val_m   = m.get('val',   {})
                    test_m  = m.get('test',  {})
                    val_r2 = val_m.get('r2', 0.0)
                    if isinstance(val_r2, (int, float)) and not np.isnan(val_r2) and val_r2 < -10:
                        logger.warning(
                            f"  [SKIP] Extreme val_R2={val_r2:.1f} in {config_dir} — "
                            f"excluded from reporting to preserve chart scales"
                        )
                        continue

                    # CV sonuclari
                    cv_file = config_dir / f"cv_results_{config_id}.json"
                    cv_r2_mean = np.nan
                    cv_r2_std  = np.nan
                    if cv_file.exists():
                        try:
                            with open(cv_file, encoding='utf-8') as f:
                                cv = json.load(f)
                            cv_data = cv.get('cv_results', {})
                            cv_r2_mean = cv_data.get('r2_test_mean', np.nan)
                            cv_r2_std  = cv_data.get('r2_test_std',  np.nan)
                        except Exception:
                            pass

                    row = {
                        'Dataset':       dataset_name,
                        'Target':        parsed['target'],
                        'Size':          parsed['size'],
                        'Scenario':      parsed['scenario'],
                        'Feature_Set':   parsed['feature_set'],
                        'N_Inputs':      parsed['n_inputs'],
                        'Feature_Names': parsed['feature_names'],
                        'NoAnomaly':     parsed['no_anomaly'],
                        'Model_Type':    model_type,
                        'Config_ID':     config_id,
                        'Train_R2':      train_m.get('r2',   np.nan),
                        'Train_RMSE':    train_m.get('rmse', np.nan),
                        'Train_MAE':     train_m.get('mae',  np.nan),
                        'Val_R2':        val_m.get('r2',     np.nan),
                        'Val_RMSE':      val_m.get('rmse',   np.nan),
                        'Val_MAE':       val_m.get('mae',    np.nan),
                        'Test_R2':       test_m.get('r2',    np.nan),
                        'Test_RMSE':     test_m.get('rmse',  np.nan),
                        'Test_MAE':      test_m.get('mae',   np.nan),
                        'CV_R2_Mean':    cv_r2_mean,
                        'CV_R2_Std':     cv_r2_std,
                        'Training_Time_s': m.get('training_time', np.nan),
                    }
                    self.all_results['ai_rows'].append(row)
                    self.all_results['ai_models'][f"{dataset_name}__{model_type}__{config_id}"] = row
                    count += 1

        logger.info(f"  [OK] {count} AI model kaydedildi")

    def _collect_anfis_results(self):
        """
        ANFIS sonuclarini gercek dizin yapisindan topla.
        training_meta alanlarini (mf_type, method, n_inputs, n_rules vb.) da ekle.
        Yapi: {anfis_models_dir}/{dataset_name}/{config_id}/metrics_{config_id}.json
        """
        self.all_results['anfis_results'] = []

        if not self.anfis_models_dir.exists():
            logger.warning(f"  [WARN] ANFIS dir bulunamadi: {self.anfis_models_dir}")
            return

        count = 0
        for dataset_dir in sorted(self.anfis_models_dir.iterdir()):
            if not dataset_dir.is_dir() or dataset_dir.name in ('training_datasets',):
                continue
            dataset_name = dataset_dir.name
            parsed = _parse_dataset_name(dataset_name)

            for config_dir in sorted(dataset_dir.iterdir()):
                if not config_dir.is_dir():
                    continue
                config_id = config_dir.name

                metrics_file = config_dir / f"metrics_{config_id}.json"
                if not metrics_file.exists():
                    continue

                try:
                    with open(metrics_file, encoding='utf-8') as f:
                        m = json.load(f)
                except Exception:
                    continue

                train_m = m.get('train', {})
                val_m   = m.get('val',   {})
                test_m  = m.get('test',  {})
                meta    = m.get('training_meta', {})

                row = {
                    'Dataset':          dataset_name,
                    'Target':           parsed['target'],
                    'Size':             parsed['size'],
                    'Scenario':         parsed['scenario'],
                    'Feature_Set':      parsed['feature_set'],
                    'N_Inputs':         meta.get('n_inputs',       parsed['n_inputs']),
                    'Feature_Names':    ', '.join(meta.get('feature_cols', [])) or parsed['feature_names'],
                    'NoAnomaly':        parsed['no_anomaly'],
                    'Config_ID':        config_id,
                    'MF_Type':          meta.get('mf_type',            ''),
                    'Method':           meta.get('method',              ''),
                    'N_MFs_Per_Input':  meta.get('n_mfs_per_input',    np.nan),
                    'N_Rules':          meta.get('n_rules',            np.nan),
                    'N_Train':          meta.get('n_train',            np.nan),
                    'N_Val':            meta.get('n_val',              np.nan),
                    'N_Test':           meta.get('n_test',             np.nan),
                    'N_Outliers':       meta.get('n_outliers_detected', np.nan),
                    'Outlier_Cleaning': meta.get('outlier_cleaning_applied', False),
                    'Train_R2':         train_m.get('r2',   np.nan),
                    'Train_RMSE':       train_m.get('rmse', np.nan),
                    'Train_MAE':        train_m.get('mae',  np.nan),
                    'Val_R2':           val_m.get('r2',     np.nan),
                    'Val_RMSE':         val_m.get('rmse',   np.nan),
                    'Val_MAE':          val_m.get('mae',    np.nan),
                    'Test_R2':          test_m.get('r2',    np.nan),
                    'Test_RMSE':        test_m.get('rmse',  np.nan),
                    'Test_MAE':         test_m.get('mae',   np.nan),
                }
                self.all_results['anfis_results'].append(row)
                count += 1

        logger.info(f"  [OK] {count} ANFIS kaydi")

    def _collect_robustness_results(self):
        """Saglamlik / Cross-Validation sonuclarini topla"""
        self.all_results['robustness'] = []
        if not self.validation_dir.exists():
            logger.info(f"  [INFO] model_validation dizini yok: {self.validation_dir}")
            return
        count = 0
        for model_dir in sorted(self.validation_dir.iterdir()):
            if not model_dir.is_dir():
                continue
            cv_file = model_dir / 'cv_summary.json'
            if not cv_file.exists():
                continue
            try:
                with open(cv_file, encoding='utf-8') as f:
                    cv = json.load(f)
                row = {'Model': model_dir.name}
                row.update(cv)
                self.all_results['robustness'].append(row)
                count += 1
            except Exception:
                pass
        logger.info(f"  [OK] {count} saglamlik sonucu")

    def _collect_crossmodel_results(self):
        """Cross-model analiz sonuclarini topla"""
        json_file = self.cross_model_dir / 'cross_model_analysis_summary.json'
        if json_file.exists():
            try:
                with open(json_file, encoding='utf-8') as f:
                    self.all_results['cross_model'] = json.load(f)
                logger.info(f"  [OK] Cross-model analiz yuklendi")
            except Exception as e:
                logger.warning(f"  [WARN] Cross-model analiz yuklenemedi: {e}")
        else:
            logger.info(f"  [INFO] Cross-model JSON yok: {json_file}")

    def _collect_unknown_predictions(self):
        """Unknown nuclei tahmin sonuclarini topla"""
        self.all_results['unknown_rows'] = []

        # PFAZ4 ciktisi: Unknown_Nuclei_Results.xlsx
        xlsx_file = self.unknown_dir / 'Unknown_Nuclei_Results.xlsx'
        if not xlsx_file.exists():
            logger.info(f"  [INFO] Unknown predictions dosyasi yok: {xlsx_file}")
            return

        # Lock dosyasi varsa (Excel'de acik), okuma atlaniyor
        lock_file = xlsx_file.parent / f"~${xlsx_file.name}"
        if lock_file.exists():
            logger.warning(f"  [WARN] Unknown predictions dosyasi Excel'de acik (lock mevcut), atlaniyor.")
            return

        try:
            df = pd.read_excel(xlsx_file, sheet_name=None)
            # All_Results ya da ilk sheet
            sheet_name = 'All_Results' if 'All_Results' in df else list(df.keys())[0]
            self.all_results['unknown_rows'] = df[sheet_name].to_dict('records')
            logger.info(f"  [OK] Unknown predictions yuklendi: {len(self.all_results['unknown_rows'])} kayit")
            self.all_results['unknown_sheets'] = {k: v for k, v in df.items()}
        except Exception as e:
            logger.warning(f"  [WARN] Unknown predictions yuklenemedi: {e}")

    # =========================================================================
    # EXCEL GENERATION
    # =========================================================================

    def generate_thesis_tables(self):
        """Tez icin KAPSAMLI Excel tablolari — gercek verilerle + openpyxl formatlama"""
        logger.info("\n" + "="*80)
        logger.info("TEZ TABLOLARI OLUSTURULUYOR")
        logger.info("="*80)

        base_file = self.output_dir / 'THESIS_COMPLETE_RESULTS.xlsx'
        lock_file = self.output_dir / '~$THESIS_COMPLETE_RESULTS.xlsx'
        if lock_file.exists():
            # Dosya Excel'de acik - zaman damgali isim kullan
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            excel_file = self.output_dir / f'THESIS_COMPLETE_RESULTS_{ts}.xlsx'
            logger.warning(f"  [WARN] Ana dosya kilitle, yeni dosya: {excel_file.name}")
        else:
            excel_file = base_file

        ai_rows    = self.all_results.get('ai_rows',      [])
        anfis_rows = self.all_results.get('anfis_results', [])
        ai_df      = pd.DataFrame(ai_rows)    if ai_rows    else pd.DataFrame()
        anfis_df   = pd.DataFrame(anfis_rows) if anfis_rows else pd.DataFrame()

        # Remove diverged/extreme R2 rows to prevent Excel chart scale corruption
        if not ai_df.empty:
            _R2_FLOOR = -10.0
            for _col in ['Val_R2', 'Test_R2', 'Train_R2']:
                if _col in ai_df.columns:
                    ai_df = ai_df[pd.to_numeric(ai_df[_col], errors='coerce').fillna(0) >= _R2_FLOOR]
            logger.info(f"  [FILTER] AI models after R2 floor filter ({_R2_FLOOR}): {len(ai_df)} rows")

        sheets_written = 0

        def _fmt(sheet_name, large=False):
            """ExcelWriter context icinde inline formatlama"""
            if sheet_name in writer.sheets:
                _format_worksheet_inline(writer.sheets[sheet_name], skip_cells=large)

        def _safe_write(fn, *args, **kwargs):
            nonlocal sheets_written
            try:
                fn(*args, **kwargs)
                sheets_written += 1
            except Exception as e:
                logger.error(f"  [ERROR] Sheet yazma hatasi ({fn.__name__}): {e}")

        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:

            # SHEET 1: Overview / Dashboard
            _safe_write(self._write_overview_sheet, writer, ai_df, anfis_df)
            _fmt('Overview')

            # SHEET 2: All AI Models (buyuk — 50k satır/sheet limiti ile bölünür)
            if not ai_df.empty:
                try:
                    n_sheets = _write_df_chunked(ai_df, writer, 'All_AI_Models', max_rows=50_000)
                    _fmt('All_AI_Models', large=True)
                    logger.info(f"  [OK] All_AI_Models ({len(ai_df)} satir, {n_sheets} sheet)")
                    sheets_written += n_sheets
                except Exception as e:
                    logger.error(f"  [ERROR] All_AI_Models: {e}")

            # SHEET 3-5: Model-tipi bazli (buyuk — 50k satır/sheet limiti ile)
            if not ai_df.empty:
                for mtype in sorted(ai_df['Model_Type'].dropna().unique()):
                    try:
                        sub = ai_df[ai_df['Model_Type'] == mtype].sort_values('Val_R2', ascending=False)
                        base_sheet = f'{mtype}_Models'
                        n_s = _write_df_chunked(sub, writer, base_sheet, max_rows=50_000)
                        _fmt(base_sheet[:31], large=len(sub) > 8000)
                        logger.info(f"  [OK] {base_sheet} ({len(sub)} satir, {n_s} sheet)")
                        sheets_written += n_s
                    except Exception as e:
                        logger.error(f"  [ERROR] {mtype}_Models: {e}")

            # SHEET: AI Dataset Summary
            if not ai_df.empty:
                try:
                    grp_cols = [c for c in ['Dataset', 'Target', 'Size', 'Scenario',
                                             'Feature_Set', 'N_Inputs', 'Feature_Names',
                                             'NoAnomaly', 'Model_Type'] if c in ai_df.columns]
                    summary_ai = ai_df.groupby(grp_cols).agg(
                        Best_Val_R2 =('Val_R2',  'max'),
                        Mean_Val_R2 =('Val_R2',  'mean'),
                        Best_Test_R2=('Test_R2', 'max'),
                        Mean_Test_R2=('Test_R2', 'mean'),
                        Mean_RMSE   =('Test_RMSE','mean'),
                        N_Configs   =('Config_ID','count')
                    ).reset_index()
                    summary_ai.to_excel(writer, sheet_name='AI_Dataset_Summary', index=False)
                    _fmt('AI_Dataset_Summary')
                    logger.info(f"  [OK] AI_Dataset_Summary ({len(summary_ai)} satir)")
                    sheets_written += 1
                except Exception as e:
                    logger.error(f"  [ERROR] AI_Dataset_Summary: {e}")

            # SHEET: All ANFIS Models
            if not anfis_df.empty:
                try:
                    anfis_df.to_excel(writer, sheet_name='All_ANFIS_Models', index=False)
                    _fmt('All_ANFIS_Models')
                    logger.info(f"  [OK] All_ANFIS_Models ({len(anfis_df)} satir)")
                    sheets_written += 1
                except Exception as e:
                    logger.error(f"  [ERROR] All_ANFIS_Models: {e}")

                try:
                    grp_a = [c for c in ['Dataset', 'Target', 'Size', 'Scenario',
                                          'Feature_Set', 'N_Inputs', 'Feature_Names'] if c in anfis_df.columns]
                    summary_anfis = anfis_df.groupby(grp_a).agg(
                        Best_Val_R2 =('Val_R2',  'max'),
                        Mean_Val_R2 =('Val_R2',  'mean'),
                        Best_Test_R2=('Test_R2', 'max'),
                        Mean_Test_R2=('Test_R2', 'mean'),
                        Mean_N_Rules=('N_Rules', 'mean'),
                        N_Configs   =('Config_ID','count')
                    ).reset_index()
                    summary_anfis.to_excel(writer, sheet_name='ANFIS_Dataset_Summary', index=False)
                    _fmt('ANFIS_Dataset_Summary')
                    logger.info(f"  [OK] ANFIS_Dataset_Summary ({len(summary_anfis)} satir)")
                    sheets_written += 1
                except Exception as e:
                    logger.error(f"  [ERROR] ANFIS_Dataset_Summary: {e}")

                try:
                    grp_c = [c for c in ['Config_ID', 'MF_Type', 'Method', 'N_MFs_Per_Input'] if c in anfis_df.columns]
                    anfis_cfg = anfis_df.groupby(grp_c).agg(
                        Mean_Val_R2 =('Val_R2',  'mean'),
                        Best_Val_R2 =('Val_R2',  'max'),
                        Mean_Test_R2=('Test_R2', 'mean'),
                        Best_Test_R2=('Test_R2', 'max'),
                        Mean_N_Rules=('N_Rules', 'mean'),
                        N_Datasets  =('Dataset', 'count')
                    ).reset_index().sort_values('Mean_Test_R2', ascending=False)
                    anfis_cfg.to_excel(writer, sheet_name='ANFIS_Config_Comparison', index=False)
                    _fmt('ANFIS_Config_Comparison')
                    logger.info(f"  [OK] ANFIS_Config_Comparison ({len(anfis_cfg)} satir)")
                    sheets_written += 1
                except Exception as e:
                    logger.error(f"  [ERROR] ANFIS_Config_Comparison: {e}")

            # SHEET: AI vs ANFIS Comparison
            _safe_write(self._write_ai_anfis_comparison, writer, ai_df, anfis_df)
            _fmt('AI_vs_ANFIS_Comparison')

            # SHEET: Best Models Per Target
            _safe_write(self._write_best_models_per_target, writer, ai_df, anfis_df)
            _fmt('Best_Models_Per_Target')

            # SHEET: Best Feature Set Per Target
            _safe_write(self._write_best_feature_set_per_target, writer, ai_df, anfis_df)
            _fmt('Best_FeatureSet_Per_Target')

            # SHEET: Anomaly vs NoAnomaly
            _safe_write(self._write_anomaly_comparison, writer, ai_df, anfis_df)
            _fmt('Anomaly_vs_NoAnomaly')

            # SHEET: Anomaly Explained (per-nucleus why selected)
            _safe_write(self._write_anomaly_explained, writer)
            _fmt('Anomaly_Explained')

            # SHEET: Target Statistics
            _safe_write(self._write_target_statistics, writer, ai_df, anfis_df)
            _fmt('Target_Statistics')

            # SHEET: Robustness CV
            robustness = self.all_results.get('robustness', [])
            if robustness:
                try:
                    pd.DataFrame(robustness).to_excel(writer, sheet_name='Robustness_CV', index=False)
                    _fmt('Robustness_CV')
                    logger.info(f"  [OK] Robustness_CV ({len(robustness)} satir)")
                    sheets_written += 1
                except Exception as e:
                    logger.error(f"  [ERROR] Robustness_CV: {e}")

            # SHEET: Cross-Model Summary
            cm_data = self.all_results.get('cross_model', {})
            if cm_data and isinstance(cm_data, dict):
                try:
                    rows_cm = [{'Key': k, 'Value': str(v)} for k, v in cm_data.items()]
                    if rows_cm:
                        pd.DataFrame(rows_cm).to_excel(writer, sheet_name='CrossModel_Summary', index=False)
                        _fmt('CrossModel_Summary')
                        logger.info(f"  [OK] CrossModel_Summary")
                        sheets_written += 1
                except Exception as e:
                    logger.error(f"  [ERROR] CrossModel_Summary: {e}")

            # SHEET: Unknown Predictions
            unknown_rows = self.all_results.get('unknown_rows', [])
            if unknown_rows:
                try:
                    pd.DataFrame(unknown_rows).to_excel(writer, sheet_name='Unknown_Predictions', index=False)
                    _fmt('Unknown_Predictions')
                    logger.info(f"  [OK] Unknown_Predictions ({len(unknown_rows)} satir)")
                    sheets_written += 1
                except Exception as e:
                    logger.error(f"  [ERROR] Unknown_Predictions: {e}")

            # SHEET: Overall Statistics
            _safe_write(self._write_overall_statistics, writer, ai_df, anfis_df)
            _fmt('Overall_Statistics')

            # SHEET: Isotope Chain Sudden Changes
            _safe_write(self._write_isotope_chain_analysis, writer)
            _fmt('IsoChain_SuddenChanges')

            # SHEET: Monte Carlo + Robustness Summary
            _safe_write(self._write_monte_carlo_robustness_summary, writer, ai_df)
            _fmt('MC_Robustness_Summary')

            # SHEET: AutoML Improvements (PFAZ13 retraining before/after)
            _safe_write(self._write_automl_improvements, writer)
            _fmt('AutoML_Improvements')

            # SHEET: Band Analizi (NuclearMomentBandAnalyzer ozeti — PFAZ12)
            _safe_write(self._write_band_analysis_sheet, writer)
            _fmt('Band_Analizi')

            # SHEET: Feature Abbreviations
            _safe_write(self._write_feature_abbreviations, writer, ai_df)
            _fmt('Feature_Abbreviations')

        # openpyxl post-processing: formatlama
        _apply_openpyxl_formatting(excel_file)

        logger.info(f"\n[OK] Excel: {excel_file}")
        logger.info(f"  Toplam sheet: {sheets_written}")
        return excel_file

    # -------------------------------------------------------------------------
    # Excel Sheet yazma yardimcilari
    # -------------------------------------------------------------------------

    def _write_overview_sheet(self, writer, ai_df, anfis_df):
        """Overview / Dashboard sheet"""
        now = datetime.now().strftime('%Y-%m-%d %H:%M')
        rows = [
            {'Bilgi': 'Rapor Tarihi',          'Deger': now},
            {'Bilgi': 'Proje',                 'Deger': 'Nukleer Fizik AI - Nuc Data v1'},
            {'Bilgi': '',                       'Deger': ''},
            {'Bilgi': '--- AI MODELLER ---',   'Deger': ''},
            {'Bilgi': 'AI Toplam Konfigurasyon','Deger': len(ai_df)},
            {'Bilgi': 'AI Dataset Sayisi',     'Deger': ai_df['Dataset'].nunique() if not ai_df.empty else 0},
            {'Bilgi': 'AI Model Tipleri',       'Deger': ', '.join(sorted(ai_df['Model_Type'].unique())) if not ai_df.empty else '-'},
            {'Bilgi': 'AI En Iyi Val R2',       'Deger': round(float(ai_df['Val_R2'].max()),  4) if not ai_df.empty else '-'},
            {'Bilgi': 'AI En Iyi Test R2',      'Deger': round(float(ai_df['Test_R2'].max()), 4) if not ai_df.empty else '-'},
            {'Bilgi': 'AI Ort Val R2',          'Deger': round(float(ai_df['Val_R2'].mean()), 4) if not ai_df.empty else '-'},
            {'Bilgi': '',                       'Deger': ''},
            {'Bilgi': '--- ANFIS MODELLER ---', 'Deger': ''},
            {'Bilgi': 'ANFIS Toplam Kayit',     'Deger': len(anfis_df)},
            {'Bilgi': 'ANFIS Dataset Sayisi',   'Deger': anfis_df['Dataset'].nunique() if not anfis_df.empty else 0},
            {'Bilgi': 'ANFIS En Iyi Val R2',    'Deger': round(float(anfis_df['Val_R2'].max()),  4) if not anfis_df.empty else '-'},
            {'Bilgi': 'ANFIS En Iyi Test R2',   'Deger': round(float(anfis_df['Test_R2'].max()), 4) if not anfis_df.empty else '-'},
            {'Bilgi': 'ANFIS Ort Val R2',       'Deger': round(float(anfis_df['Val_R2'].mean()), 4) if not anfis_df.empty else '-'},
        ]

        def _best_row_info(df, tgt, r2_col, label_cols):
            sub = df[df['Target'] == tgt]
            valid = sub[r2_col].dropna()
            if valid.empty:
                return '-'
            idx = valid.idxmax()
            best = valid[idx]
            labels = '/'.join(str(sub.at[idx, c]) for c in label_cols if c in sub.columns)
            return f"{best:.4f} ({labels})"

        # Hedef bazli en iyi AI R2
        if not ai_df.empty:
            rows.append({'Bilgi': '', 'Deger': ''})
            rows.append({'Bilgi': '--- HEDEF BAZLI EN IYI AI VAL R2 ---', 'Deger': ''})
            for tgt in sorted(ai_df['Target'].dropna().unique()):
                info = _best_row_info(ai_df, tgt, 'Val_R2', ['Model_Type', 'Feature_Set'])
                rows.append({'Bilgi': f'{tgt} - En Iyi Val R2', 'Deger': info})

        if not anfis_df.empty:
            rows.append({'Bilgi': '', 'Deger': ''})
            rows.append({'Bilgi': '--- HEDEF BAZLI EN IYI ANFIS VAL R2 ---', 'Deger': ''})
            for tgt in sorted(anfis_df['Target'].dropna().unique()):
                info = _best_row_info(anfis_df, tgt, 'Val_R2', ['Config_ID', 'Feature_Set'])
                rows.append({'Bilgi': f'{tgt} - En Iyi Val R2', 'Deger': info})

        pd.DataFrame(rows).to_excel(writer, sheet_name='Overview', index=False)
        logger.info(f"  [OK] Overview sheet")

    def _write_ai_anfis_comparison(self, writer, ai_df, anfis_df):
        """AI vs ANFIS dataset-bazli karsilastirma"""
        data = []
        datasets = set()
        if not ai_df.empty:
            datasets.update(ai_df['Dataset'].unique())
        if not anfis_df.empty:
            datasets.update(anfis_df['Dataset'].unique())

        for ds in sorted(datasets):
            parsed = _parse_dataset_name(ds)
            row = {
                'Dataset':     ds,
                'Target':      parsed['target'],
                'Size':        parsed['size'],
                'Scenario':    parsed['scenario'],
                'Feature_Set': parsed['feature_set'],
                'N_Inputs':    parsed['n_inputs'],
                'NoAnomaly':   parsed['no_anomaly'],
            }

            def _safe_idxmax_val(df_sub, r2_col, label_col):
                """idxmax NaN donmemesi icin dropna ile guvenli lookup"""
                valid = df_sub[r2_col].dropna()
                if valid.empty:
                    return ''
                idx = valid.idxmax()
                return df_sub.at[idx, label_col] if label_col in df_sub.columns else ''

            if not ai_df.empty:
                ai_sub = ai_df[ai_df['Dataset'] == ds]
                row['AI_N_Configs']    = len(ai_sub)
                row['AI_Best_Val_R2']  = ai_sub['Val_R2'].max()  if len(ai_sub) else np.nan
                row['AI_Mean_Val_R2']  = ai_sub['Val_R2'].mean() if len(ai_sub) else np.nan
                row['AI_Best_Test_R2'] = ai_sub['Test_R2'].max() if len(ai_sub) else np.nan
                row['AI_Best_CV_R2']   = ai_sub['CV_R2_Mean'].max() if ('CV_R2_Mean' in ai_sub.columns and len(ai_sub)) else np.nan
                row['AI_Best_Model']   = _safe_idxmax_val(ai_sub, 'Val_R2', 'Model_Type')
            else:
                for k in ('AI_N_Configs','AI_Best_Val_R2','AI_Mean_Val_R2','AI_Best_Test_R2','AI_Best_CV_R2','AI_Best_Model'):
                    row[k] = np.nan

            if not anfis_df.empty:
                an_sub = anfis_df[anfis_df['Dataset'] == ds]
                row['ANFIS_N_Configs']    = len(an_sub)
                row['ANFIS_Best_Val_R2']  = an_sub['Val_R2'].max()  if len(an_sub) else np.nan
                row['ANFIS_Mean_Val_R2']  = an_sub['Val_R2'].mean() if len(an_sub) else np.nan
                row['ANFIS_Best_Test_R2'] = an_sub['Test_R2'].max() if len(an_sub) else np.nan
                row['ANFIS_Best_Config']  = _safe_idxmax_val(an_sub, 'Val_R2', 'Config_ID')
                row['ANFIS_Best_N_Rules'] = (
                    an_sub.at[an_sub['Val_R2'].dropna().idxmax(), 'N_Rules']
                    if not an_sub['Val_R2'].dropna().empty else np.nan
                )
            else:
                for k in ('ANFIS_N_Configs','ANFIS_Best_Val_R2','ANFIS_Mean_Val_R2','ANFIS_Best_Test_R2','ANFIS_Best_Config','ANFIS_Best_N_Rules'):
                    row[k] = np.nan

            # Winner
            ai_r2   = row.get('AI_Best_Val_R2',    np.nan)
            anfis_r2 = row.get('ANFIS_Best_Val_R2', np.nan)
            if pd.isna(ai_r2) and pd.isna(anfis_r2):
                row['Winner'] = '-'
                row['Delta_R2'] = np.nan
            elif pd.isna(anfis_r2):
                row['Winner'] = 'AI'
                row['Delta_R2'] = np.nan
            elif pd.isna(ai_r2):
                row['Winner'] = 'ANFIS'
                row['Delta_R2'] = np.nan
            else:
                delta = float(anfis_r2) - float(ai_r2)
                row['Delta_R2'] = round(delta, 4)
                row['Winner'] = 'ANFIS' if delta > 0.001 else ('AI' if delta < -0.001 else 'Tie')

            data.append(row)

        df = pd.DataFrame(data)
        df.to_excel(writer, sheet_name='AI_vs_ANFIS_Comparison', index=False)
        logger.info(f"  [OK] AI_vs_ANFIS_Comparison ({len(df)} dataset)")

    def _write_best_models_per_target(self, writer, ai_df, anfis_df):
        """Hedef bazinda Top-N model siralama"""
        rows = []
        all_targets = set()
        if not ai_df.empty:
            all_targets.update(ai_df['Target'].dropna().unique())
        if not anfis_df.empty:
            all_targets.update(anfis_df['Target'].dropna().unique())

        def _safe_round(val):
            try:
                v = float(val)
                return round(v, 4) if not np.isnan(v) else np.nan
            except Exception:
                return np.nan

        for tgt in sorted(all_targets):
            rank = 1
            if not ai_df.empty:
                sub = ai_df[ai_df['Target'] == tgt]
                sub = sub[sub['Val_R2'].notna()].nlargest(10, 'Val_R2')
                for _, r in sub.iterrows():
                    rows.append({
                        'Target':      tgt,
                        'Rank':        rank,
                        'Model_Cat':   'AI',
                        'Model_Type':  r.get('Model_Type', ''),
                        'Config_ID':   r.get('Config_ID',  ''),
                        'Dataset':     r.get('Dataset',    ''),
                        'Feature_Set': r.get('Feature_Set',''),
                        'N_Inputs':    r.get('N_Inputs',   ''),
                        'Size':        r.get('Size',       ''),
                        'Scenario':    r.get('Scenario',   ''),
                        'Val_R2':      _safe_round(r.get('Val_R2')),
                        'Test_R2':     _safe_round(r.get('Test_R2')),
                        'Val_RMSE':    _safe_round(r.get('Val_RMSE')),
                        'Test_RMSE':   _safe_round(r.get('Test_RMSE')),
                    })
                    rank += 1

            if not anfis_df.empty:
                sub = anfis_df[anfis_df['Target'] == tgt]
                sub = sub[sub['Val_R2'].notna()].nlargest(5, 'Val_R2')
                for _, r in sub.iterrows():
                    rows.append({
                        'Target':      tgt,
                        'Rank':        rank,
                        'Model_Cat':   'ANFIS',
                        'Model_Type':  str(r.get('MF_Type','')) + '_' + str(r.get('Method','')),
                        'Config_ID':   r.get('Config_ID',  ''),
                        'Dataset':     r.get('Dataset',    ''),
                        'Feature_Set': r.get('Feature_Set',''),
                        'N_Inputs':    r.get('N_Inputs',   ''),
                        'Size':        r.get('Size',       ''),
                        'Scenario':    r.get('Scenario',   ''),
                        'Val_R2':      _safe_round(r.get('Val_R2')),
                        'Test_R2':     _safe_round(r.get('Test_R2')),
                        'Val_RMSE':    _safe_round(r.get('Val_RMSE')),
                        'Test_RMSE':   _safe_round(r.get('Test_RMSE')),
                    })
                    rank += 1

        if rows:
            pd.DataFrame(rows).to_excel(writer, sheet_name='Best_Models_Per_Target', index=False)
            logger.info(f"  [OK] Best_Models_Per_Target ({len(rows)} satir)")

    def _write_best_feature_set_per_target(self, writer, ai_df, anfis_df):
        """
        Hangi feature setinin hangi hedef icin en iyi sonucu verdigi ozet tablosu.
        (PFAZ_DEVELOPMENT_NOTES todo)
        """
        rows = []
        all_targets = set()
        if not ai_df.empty:
            all_targets.update(ai_df['Target'].dropna().unique())
        if not anfis_df.empty:
            all_targets.update(anfis_df['Target'].dropna().unique())

        for tgt in sorted(all_targets):
            # AI: feature_set bazli en iyi val_r2
            if not ai_df.empty:
                sub = ai_df[ai_df['Target'] == tgt]
                if not sub.empty and 'Feature_Set' in sub.columns:
                    fs_grp = sub.groupby('Feature_Set').agg(
                        Best_Val_R2 =('Val_R2',  'max'),
                        Mean_Val_R2 =('Val_R2',  'mean'),
                        Best_Test_R2=('Test_R2', 'max'),
                        Mean_Test_R2=('Test_R2', 'mean'),
                        N_Configs   =('Config_ID','count'),
                    ).reset_index().sort_values('Best_Val_R2', ascending=False)
                    for rank, (_, r) in enumerate(fs_grp.iterrows(), 1):
                        fs = r['Feature_Set']
                        rows.append({
                            'Target':        tgt,
                            'Model_Cat':     'AI',
                            'Rank':          rank,
                            'Feature_Set':   fs,
                            'Feature_Names': FEATURE_SET_EXPAND.get(fs, []),
                            'N_Inputs':      len(FEATURE_SET_EXPAND.get(fs, [])),
                            'Best_Val_R2':   round(float(r['Best_Val_R2']),  4),
                            'Mean_Val_R2':   round(float(r['Mean_Val_R2']),  4),
                            'Best_Test_R2':  round(float(r['Best_Test_R2']), 4),
                            'Mean_Test_R2':  round(float(r['Mean_Test_R2']), 4),
                            'N_Configs':     int(r['N_Configs']),
                        })

            # ANFIS: feature_set bazli en iyi val_r2
            if not anfis_df.empty:
                sub = anfis_df[anfis_df['Target'] == tgt]
                if not sub.empty and 'Feature_Set' in sub.columns:
                    fs_grp = sub.groupby('Feature_Set').agg(
                        Best_Val_R2 =('Val_R2',  'max'),
                        Mean_Val_R2 =('Val_R2',  'mean'),
                        Best_Test_R2=('Test_R2', 'max'),
                        Mean_Test_R2=('Test_R2', 'mean'),
                        N_Configs   =('Config_ID','count'),
                    ).reset_index().sort_values('Best_Val_R2', ascending=False)
                    for rank, (_, r) in enumerate(fs_grp.iterrows(), 1):
                        fs = r['Feature_Set']
                        rows.append({
                            'Target':        tgt,
                            'Model_Cat':     'ANFIS',
                            'Rank':          rank,
                            'Feature_Set':   fs,
                            'Feature_Names': FEATURE_SET_EXPAND.get(fs, []),
                            'N_Inputs':      len(FEATURE_SET_EXPAND.get(fs, [])),
                            'Best_Val_R2':   round(float(r['Best_Val_R2']),  4),
                            'Mean_Val_R2':   round(float(r['Mean_Val_R2']),  4),
                            'Best_Test_R2':  round(float(r['Best_Test_R2']), 4),
                            'Mean_Test_R2':  round(float(r['Mean_Test_R2']), 4),
                            'N_Configs':     int(r['N_Configs']),
                        })

        if rows:
            df = pd.DataFrame(rows)
            # Feature_Names listesini stringe cevir
            df['Feature_Names'] = df['Feature_Names'].apply(
                lambda x: ', '.join(x) if isinstance(x, list) else str(x))
            df.to_excel(writer, sheet_name='Best_FeatureSet_Per_Target', index=False)
            logger.info(f"  [OK] Best_FeatureSet_Per_Target ({len(df)} satir)")

    def _write_anomaly_comparison(self, writer, ai_df, anfis_df):
        """
        Anomali vs NoAnomaly dataset karsilastirmasi.
        (PFAZ_DEVELOPMENT_NOTES todo)
        """
        rows = []

        def _compare(df, model_cat):
            if df.empty or 'NoAnomaly' not in df.columns:
                return
            for ds_base in df['Dataset'].str.replace('_NoAnomaly', '', regex=False).unique():
                normal = df[df['Dataset'] == ds_base]
                noanom = df[df['Dataset'] == ds_base + '_NoAnomaly']
                if normal.empty or noanom.empty:
                    continue
                parsed = _parse_dataset_name(ds_base)
                rows.append({
                    'Dataset_Base':       ds_base,
                    'Target':             parsed['target'],
                    'Size':               parsed['size'],
                    'Scenario':           parsed['scenario'],
                    'Feature_Set':        parsed['feature_set'],
                    'Model_Cat':          model_cat,
                    'Normal_Best_Val_R2': round(float(normal['Val_R2'].max()),  4),
                    'NoAnomaly_Best_Val_R2': round(float(noanom['Val_R2'].max()), 4),
                    'Delta_R2': round(float(noanom['Val_R2'].max()) - float(normal['Val_R2'].max()), 4),
                    'Normal_Best_Test_R2': round(float(normal['Test_R2'].max()),  4),
                    'NoAnomaly_Best_Test_R2': round(float(noanom['Test_R2'].max()), 4),
                    'Delta_Test_R2': round(float(noanom['Test_R2'].max()) - float(normal['Test_R2'].max()), 4),
                    'Winner': ('NoAnomaly' if noanom['Val_R2'].max() > normal['Val_R2'].max() + 0.001
                               else ('Normal' if normal['Val_R2'].max() > noanom['Val_R2'].max() + 0.001
                                     else 'Tie')),
                })

        _compare(ai_df,    'AI')
        _compare(anfis_df, 'ANFIS')

        if rows:
            pd.DataFrame(rows).to_excel(writer, sheet_name='Anomaly_vs_NoAnomaly', index=False)
            logger.info(f"  [OK] Anomaly_vs_NoAnomaly ({len(rows)} satir)")
        else:
            logger.info(f"  [INFO] Anomaly vs NoAnomaly karsilastirma: veri bulunamadi")

    def _write_anomaly_explained(self, writer):
        """
        Her anomali cekirdegi icin neden secildigini aciklayan rapor.
        Kaynak: datasets/ alt dizinindeki anomaly_explanation_report.json dosyalari.
        Sutunlar: Target, NUCLEUS, A, Z, N, SPIN, PARITY, n_violations,
                  worst_column, worst_iqr_ratio, triggered_columns (string)
        """
        import json, glob as _glob

        rows = []

        # Search for anomaly_explanation_report.json files under datasets dir
        search_roots = []
        datasets_dir = Path(self.datasets_dir) if hasattr(self, 'datasets_dir') and self.datasets_dir else None
        if datasets_dir and datasets_dir.exists():
            search_roots.append(datasets_dir)
        # Also search relative paths
        for rel in ['datasets', '../datasets', 'output/datasets']:
            p = Path(rel)
            if p.exists():
                search_roots.append(p)

        report_files = []
        for root in search_roots:
            report_files.extend(root.rglob('anomaly_explanation_report.json'))
        # Deduplicate
        seen = set()
        unique_files = []
        for f in report_files:
            key = str(f.resolve())
            if key not in seen:
                seen.add(key)
                unique_files.append(f)

        for report_file in unique_files:
            try:
                with open(report_file, 'r') as f:
                    data = json.load(f)
                for target, tdata in data.items():
                    for rec in tdata.get('nuclei', []):
                        # Flatten triggered_columns to a readable string
                        triggered_str = '; '.join(
                            f"{t['column']}={t['value']:.3g}({t['direction']},×{t['iqr_ratio']:.1f}IQR)"
                            for t in rec.get('triggered_columns', [])[:5]
                        )
                        rows.append({
                            'Target':          target,
                            'NUCLEUS':         rec.get('NUCLEUS', ''),
                            'A':               rec.get('A', ''),
                            'Z':               rec.get('Z', ''),
                            'N':               rec.get('N', ''),
                            'SPIN':            rec.get('SPIN', ''),
                            'PARITY':          rec.get('PARITY', ''),
                            'N_Violations':    rec.get('n_violations', 0),
                            'Worst_Column':    rec.get('worst_column', ''),
                            'Worst_IQR_Ratio': rec.get('worst_iqr_ratio', 0.0),
                            'Anomaly_Reason':  triggered_str,
                        })
            except Exception as e:
                logger.warning(f"  [WARNING] anomaly_explained: could not read {report_file}: {e}")

        if rows:
            out_df = pd.DataFrame(rows).sort_values(['Target', 'Worst_IQR_Ratio'], ascending=[True, False])
            out_df.to_excel(writer, sheet_name='Anomaly_Explained', index=False)
            logger.info(f"  [OK] Anomaly_Explained ({len(rows)} satirlik anomali kaydı)")
        else:
            logger.info(f"  [INFO] Anomaly_Explained: anomaly_explanation_report.json bulunamadi")

    def _write_target_statistics(self, writer, ai_df, anfis_df):
        """Hedef bazli istatistikler"""
        rows = []

        for cat, df in [('AI', ai_df), ('ANFIS', anfis_df)]:
            if df.empty or 'Target' not in df.columns:
                continue
            r2_col = 'Val_R2'
            for tgt in sorted(df['Target'].dropna().unique()):
                sub = df[df['Target'] == tgt][r2_col].dropna()
                if sub.empty:
                    continue
                rows.append({
                    'Model_Cat':       cat,
                    'Target':          tgt,
                    'N_Records':       len(sub),
                    'Mean_Val_R2':     round(float(sub.mean()), 4),
                    'Std_Val_R2':      round(float(sub.std()),  4),
                    'Max_Val_R2':      round(float(sub.max()),  4),
                    'Min_Val_R2':      round(float(sub.min()),  4),
                    'Median_Val_R2':   round(float(sub.median()), 4),
                    'Pct_R2_gt_07':    round(float((sub > 0.7).mean() * 100), 1),
                    'Pct_R2_gt_08':    round(float((sub > 0.8).mean() * 100), 1),
                    'Pct_R2_gt_09':    round(float((sub > 0.9).mean() * 100), 1),
                })

        if rows:
            pd.DataFrame(rows).to_excel(writer, sheet_name='Target_Statistics', index=False)
            logger.info(f"  [OK] Target_Statistics ({len(rows)} satir)")

    def _write_overall_statistics(self, writer, ai_df, anfis_df):
        """Genel istatistikler"""
        data = [
            {'Metrik': 'Toplam AI Konfigurasyonu', 'Deger': len(ai_df)},
            {'Metrik': 'Toplam ANFIS Kaydi',        'Deger': len(anfis_df)},
            {'Metrik': 'AI Dataset Sayisi',    'Deger': ai_df['Dataset'].nunique()    if not ai_df.empty    else 0},
            {'Metrik': 'ANFIS Dataset Sayisi', 'Deger': anfis_df['Dataset'].nunique() if not anfis_df.empty else 0},
        ]
        if not ai_df.empty:
            data += [
                {'Metrik': 'AI En Iyi Val R2',   'Deger': round(float(ai_df['Val_R2'].max()),   4)},
                {'Metrik': 'AI En Iyi Test R2',  'Deger': round(float(ai_df['Test_R2'].max()),  4)},
                {'Metrik': 'AI Ort Val R2',      'Deger': round(float(ai_df['Val_R2'].mean()),  4)},
                {'Metrik': 'AI Model Tipleri',   'Deger': ', '.join(sorted(ai_df['Model_Type'].unique()))},
            ]
        if not anfis_df.empty:
            data += [
                {'Metrik': 'ANFIS En Iyi Val R2',  'Deger': round(float(anfis_df['Val_R2'].max()),   4)},
                {'Metrik': 'ANFIS En Iyi Test R2', 'Deger': round(float(anfis_df['Test_R2'].max()),  4)},
                {'Metrik': 'ANFIS Ort Val R2',     'Deger': round(float(anfis_df['Val_R2'].mean()),  4)},
                {'Metrik': 'ANFIS Konfigurasyonlar', 'Deger': ', '.join(sorted(anfis_df['Config_ID'].unique()))},
            ]
        pd.DataFrame(data).to_excel(writer, sheet_name='Overall_Statistics', index=False)
        logger.info(f"  [OK] Overall_Statistics")

    def _write_isotope_chain_analysis(self, writer):
        """
        İzotop zinciri ani değer değişimleri + magic number korelasyonu.
        Kaynak: aaa2.txt (self.aaa2_txt_path) veya otomatik arama.
        """
        try:
            from pfaz_modules.pfaz05_cross_model.isotope_chain_analyzer import IsotopeChainAnalyzer
        except ImportError:
            try:
                import sys
                sys.path.insert(0, str(Path(__file__).parent.parent.parent))
                from pfaz_modules.pfaz05_cross_model.isotope_chain_analyzer import IsotopeChainAnalyzer
            except ImportError:
                logger.warning("  [INFO] IsotopeChainAnalyzer import edilemedi — IsoChain_SuddenChanges atlandı")
                return

        # Find aaa2.txt
        aaa2_path = self.aaa2_txt_path
        if aaa2_path is None:
            for candidate in ['aaa2.txt', '../aaa2.txt', '../../aaa2.txt']:
                p = Path(candidate)
                if p.exists():
                    aaa2_path = str(p)
                    break

        if aaa2_path is None or not Path(aaa2_path).exists():
            logger.info("  [INFO] aaa2.txt bulunamadi — IsoChain_SuddenChanges atlandı")
            return

        try:
            analyzer = IsotopeChainAnalyzer(aaa2_txt_path=aaa2_path,
                                            output_dir=str(self.output_dir / 'isotope_chain'))
            analyzer.run_full_analysis()

            # Summary sheet
            if analyzer.summary:
                summary_rows = []
                for target, s in analyzer.summary.items():
                    summary_rows.append({
                        'Target':                 target,
                        'Total_Transitions':      s['total_transitions'],
                        'Sudden_Changes':         s['n_sudden'],
                        'Sudden_Pct':             s['sudden_pct'],
                        'Magic_N_Count':          s['magic_N_in_sudden'],
                        'Magic_N_Pct':            s['magic_N_pct'],
                        'Magic_Z_Count':          s['magic_Z_in_sudden'],
                        'Magic_Z_Pct':            s['magic_Z_pct'],
                        'Mean_Delta_at_Magic':    s['mean_delta_at_magic'],
                        'Mean_Delta_at_NonMagic': s['mean_delta_at_nonmagic'],
                        'Magic_Amplification':    s['magic_amplification'],
                    })
                pd.DataFrame(summary_rows).to_excel(writer, sheet_name='IsoChain_SuddenChanges', index=False)

            # Detailed sudden changes (all targets combined)
            flat = analyzer.get_sudden_changes_flat()
            if not flat.empty:
                flat.to_excel(writer, sheet_name='IsoChain_Detail', index=False)

            n_sudden = sum(len(v) for v in analyzer.sudden_changes.values())
            logger.info(f"  [OK] IsoChain_SuddenChanges ({n_sudden} ani degisim)")

        except Exception as e:
            logger.warning(f"  [WARNING] IsoChain analizi hatasi: {e}")

    def _write_monte_carlo_robustness_summary(self, writer, ai_df):
        """
        Monte Carlo belirsizlik özeti (PFAZ9) + robustness (CV) özeti (PFAZ2).

        MC özeti: PFAZ9 output'unda AAA2_Complete_{target}.xlsx dosyalarından
                  Uncertainty sheet'ini okur → per-target mean std / CI özeti.
        Robustness özeti: ai_df'deki CV_R2 sütunlarından hesaplanır.
        """
        rows_mc  = []
        rows_rob = []

        # ---- Monte Carlo from PFAZ9 outputs -----------------------------------
        pfaz9_dir = self.pfaz9_output_dir
        if pfaz9_dir is None:
            # Guess from outputs tree
            for candidate in [
                self.output_dir.parent / 'pfaz9_output',
                self.output_dir.parent / 'aaa2_control_group',
                self.output_dir.parent.parent / 'outputs' / 'pfaz9',
            ]:
                if candidate.exists():
                    pfaz9_dir = candidate
                    break

        if pfaz9_dir and Path(pfaz9_dir).exists():
            pfaz9_dir = Path(pfaz9_dir)
            for target in ['MM', 'QM']:
                excel_path = pfaz9_dir / f'AAA2_Complete_{target}.xlsx'
                if not excel_path.exists():
                    continue
                try:
                    unc_df = pd.read_excel(excel_path, sheet_name='Uncertainty')
                    if unc_df.empty:
                        continue
                    rows_mc.append({
                        'Target':             target,
                        'N_Nuclei':           len(unc_df),
                        'MC_Mean_Pred':       round(float(unc_df['Mean_Prediction'].mean()), 6) if 'Mean_Prediction' in unc_df.columns else np.nan,
                        'MC_Mean_Std':        round(float(unc_df['Std_Prediction'].mean()), 6)  if 'Std_Prediction'  in unc_df.columns else np.nan,
                        'MC_Max_Std':         round(float(unc_df['Std_Prediction'].max()),  6)  if 'Std_Prediction'  in unc_df.columns else np.nan,
                        'MC_Mean_CV':         round(float(unc_df['CV'].mean()), 6)              if 'CV'              in unc_df.columns else np.nan,
                        'MC_CI_Width_Mean':   round(float((unc_df['CI_Upper'] - unc_df['CI_Lower']).mean()), 6)
                                              if 'CI_Upper' in unc_df.columns and 'CI_Lower' in unc_df.columns else np.nan,
                        'High_Uncertainty_N': int((unc_df['CV'] > 0.3).sum()) if 'CV' in unc_df.columns else np.nan,
                    })
                    logger.info(f"  [OK] MC summary for {target}: {len(unc_df)} nuclei")
                except Exception as e:
                    logger.warning(f"  [WARNING] MC read error {target}: {e}")
        else:
            logger.info("  [INFO] PFAZ9 output dir bulunamadi — MC özeti atlandı")

        # ---- Robustness from CV results in ai_df ------------------------------
        cv_col = 'CV_R2' if 'CV_R2' in ai_df.columns else None
        if cv_col is None:
            # Try to find a CV column
            for c in ai_df.columns:
                if 'cv' in c.lower() and 'r2' in c.lower():
                    cv_col = c
                    break

        if cv_col and not ai_df.empty and 'Target' in ai_df.columns:
            for target in sorted(ai_df['Target'].dropna().unique()):
                sub = ai_df[ai_df['Target'] == target][cv_col].dropna()
                if sub.empty:
                    continue
                model_type_col = 'Model_Type' if 'Model_Type' in ai_df.columns else None
                best_model = ''
                if model_type_col:
                    idx = ai_df[ai_df['Target'] == target][cv_col].idxmax() if not sub.empty else None
                    if idx is not None:
                        best_model = str(ai_df.loc[idx, model_type_col])

                rows_rob.append({
                    'Target':             target,
                    'N_Models_with_CV':   int(len(sub)),
                    'CV_R2_Mean':         round(float(sub.mean()), 6),
                    'CV_R2_Std':          round(float(sub.std()),  6),
                    'CV_R2_Max':          round(float(sub.max()),  6),
                    'CV_R2_Min':          round(float(sub.min()),  6),
                    'Best_CV_Model':      best_model,
                    'Robust_N':           int((sub >= 0.7).sum()),
                    'Robust_Pct':         round((sub >= 0.7).sum() / max(1, len(sub)) * 100, 2),
                })

        # Write sheets
        if rows_mc:
            pd.DataFrame(rows_mc).to_excel(writer, sheet_name='MC_Robustness_Summary', index=False)
            logger.info(f"  [OK] MC_Robustness_Summary — MC: {len(rows_mc)} targets")
        elif rows_rob:
            pd.DataFrame(rows_rob).to_excel(writer, sheet_name='MC_Robustness_Summary', index=False)
            logger.info(f"  [OK] MC_Robustness_Summary — Rob: {len(rows_rob)} targets")
        else:
            # Write empty placeholder
            pd.DataFrame([{'Info': 'PFAZ9 MC verisi henüz mevcut değil. PFAZ9 tamamlandıktan sonra yeniden çalıştırın.'}]).to_excel(
                writer, sheet_name='MC_Robustness_Summary', index=False)
            logger.info("  [INFO] MC_Robustness_Summary: veri bulunamadi — placeholder yazıldı")

        # Also write robustness detail if available
        if rows_rob:
            try:
                pd.DataFrame(rows_rob).to_excel(writer, sheet_name='Robustness_CV_Summary', index=False)
                logger.info(f"  [OK] Robustness_CV_Summary ({len(rows_rob)} hedef)")
            except Exception:
                pass

    def _write_automl_improvements(self, writer):
        """
        PFAZ13 AutoML Retraining Loop sonuçlarını üç alt-bölümde yazar:
          1. Summary   — her kombinasyon için önceki/sonraki Val R² ve iyileşme
          2. Best Params — en iyi hiperparametreler (long format)
          3. Overview  — aggregate istatistikler

        Kaynak: pfaz13_output_dir/automl_retraining_log.json
                VEYA pfaz13_output_dir/automl_improvement_report.xlsx (fall-back)
        """
        import json
        from pathlib import Path

        # --- find records ---------------------------------------------------
        records = []

        pfaz13_dir = self.pfaz13_output_dir
        if pfaz13_dir is None:
            # guess from output tree
            for candidate in [
                self.output_dir.parent / 'automl_results',
                self.output_dir.parent.parent / 'automl_results',
            ]:
                if candidate.exists():
                    pfaz13_dir = candidate
                    break

        if pfaz13_dir:
            log_json = Path(pfaz13_dir) / 'automl_retraining_log.json'
            if log_json.exists():
                try:
                    with open(log_json) as f:
                        records = json.load(f)
                except Exception as e:
                    logger.warning(f"  [WARN] automl_retraining_log.json okuma hatası: {e}")

        if not records:
            # Write a placeholder so the sheet still appears
            pd.DataFrame([{
                'Info': 'PFAZ13 AutoML Retraining Loop henüz çalıştırılmadı veya aday bulunamadı.',
            }]).to_excel(writer, sheet_name='AutoML_Improvements', index=False)
            logger.info("  [INFO] AutoML_Improvements: veri yok, yer tutucu yazıldı")
            return

        # ---- Summary -------------------------------------------------------
        summary_rows = []
        for r in records:
            summary_rows.append({
                'Target':           r.get('target', ''),
                'Dataset':          r.get('dataset', ''),
                'Original Model':   r.get('original_model', ''),
                'Best Model':       r.get('best_model', ''),
                'Before Val R²':    r.get('before_val_r2'),
                'After Val R²':     r.get('after_val_r2'),
                'Improvement (ΔR²)': r.get('improvement'),
                'Improved?':        'YES' if r.get('improved') else 'NO',
                'N Trials':         r.get('n_trials'),
                'Timestamp':        r.get('timestamp', ''),
            })
        pd.DataFrame(summary_rows).to_excel(
            writer, sheet_name='AutoML_Improvements', index=False)
        logger.info(f"  [OK] AutoML_Improvements — Summary ({len(summary_rows)} satır)")

        # ---- Best Params (long) -------------------------------------------
        param_rows = []
        for r in records:
            for pname, pval in (r.get('best_params') or {}).items():
                param_rows.append({
                    'Target':    r.get('target', ''),
                    'Dataset':   r.get('dataset', ''),
                    'Model':     r.get('best_model', ''),
                    'Parameter': pname,
                    'Value':     str(pval),
                })
        if param_rows:
            pd.DataFrame(param_rows).to_excel(
                writer, sheet_name='AutoML_BestParams', index=False)
            logger.info(f"  [OK] AutoML_BestParams ({len(param_rows)} satır)")

        # ---- Overview stats -----------------------------------------------
        import numpy as np
        improvements = [r.get('improvement', 0.0) for r in records]
        improved_only = [x for x, r in zip(improvements, records) if r.get('improved')]
        n_improved    = sum(1 for r in records if r.get('improved'))

        overview_rows = [
            {'Metrik': 'Toplam yeniden eğitilen kombinasyon', 'Değer': len(records)},
            {'Metrik': 'İyileşen kombinasyon sayısı',         'Değer': n_improved},
            {'Metrik': 'İyileşmeyen kombinasyon sayısı',      'Değer': len(records) - n_improved},
            {'Metrik': 'Ortalama ΔR² (tümü)',                 'Değer': round(float(np.mean(improvements)), 4)},
            {'Metrik': 'Ortalama ΔR² (sadece iyileşenler)',  'Değer': round(float(np.mean(improved_only)), 4) if improved_only else 0.0},
            {'Metrik': 'Maksimum ΔR²',                        'Değer': round(float(max(improvements)), 4)},
            {'Metrik': 'Minimum ΔR²',                         'Değer': round(float(min(improvements)), 4)},
        ]
        pd.DataFrame(overview_rows).to_excel(
            writer, sheet_name='AutoML_Overview', index=False)
        logger.info(f"  [OK] AutoML_Overview")

    def _write_feature_abbreviations(self, writer, ai_df):
        """
        Feature kisaltma tablosu + feature set aciklamalari.
        (PFAZ_DEVELOPMENT_NOTES todo)
        """
        # 1. Feature kisaltma tablosu
        abbrev_rows = [
            {'Kisaltma': k, 'Sutun_Adi': v[0], 'Aciklama': v[1]}
            for k, v in FEATURE_ABBREV.items()
        ]
        df_abbrev = pd.DataFrame(abbrev_rows)

        # 2. Feature set aciklamalari
        set_rows = []
        for fs_code, members in FEATURE_SET_EXPAND.items():
            member_cols = [FEATURE_ABBREV.get(m, (m, ''))[0] for m in members]
            set_rows.append({
                'Feature_Set_Kodu': fs_code,
                'N_Giris':          len(members),
                'Giriş_Kisaltmalari': ', '.join(members),
                'Giriş_Sutun_Adlari': ', '.join(member_cols),
            })
        df_sets = pd.DataFrame(set_rows)

        # Hedef bazli hangi feature setleri kullanilmis
        used_rows = []
        if not ai_df.empty and 'Feature_Set' in ai_df.columns:
            for tgt in sorted(ai_df['Target'].dropna().unique()):
                sub = ai_df[ai_df['Target'] == tgt]
                for fs in sorted(sub['Feature_Set'].dropna().unique()):
                    best_r2 = sub[sub['Feature_Set'] == fs]['Val_R2'].max()
                    used_rows.append({
                        'Target':      tgt,
                        'Feature_Set': fs,
                        'N_Giris':     len(FEATURE_SET_EXPAND.get(fs, [])),
                        'Best_Val_R2': round(float(best_r2), 4),
                    })
        df_used = pd.DataFrame(used_rows)

        # Excele yaz — 3 tablo ayni sheet'e alt alta
        startrow = 0
        df_abbrev.to_excel(writer, sheet_name='Feature_Abbreviations', index=False, startrow=startrow)
        startrow += len(df_abbrev) + 3

        df_sets.to_excel(writer, sheet_name='Feature_Abbreviations', index=False, startrow=startrow)
        startrow += len(df_sets) + 3

        if not df_used.empty:
            df_used.to_excel(writer, sheet_name='Feature_Abbreviations', index=False, startrow=startrow)

        logger.info(f"  [OK] Feature_Abbreviations ({len(abbrev_rows)} kisaltma, {len(set_rows)} set)")

    def _write_band_analysis_sheet(self, writer):
        """
        PFAZ12 NuclearMomentBandAnalyzer ciktisinden ozet sheet.
        Kaynak: outputs/advanced_analytics/band_analysis/nuclear_band_analysis.xlsx
        Sheets: Bant_Ozeti + Korelasyon
        """
        import pathlib

        # Locate band analysis Excel produced by PFAZ12
        candidates = []
        base = pathlib.Path(self.output_dir).parent
        for search_root in [base, base.parent]:
            candidates += list(search_root.rglob('nuclear_band_analysis.xlsx'))
        if not candidates:
            logger.info("  [INFO] Band_Analizi: nuclear_band_analysis.xlsx bulunamadi — sheet atlandi")
            pd.DataFrame([{'Durum': 'PFAZ12 band analizi henuz calistirilmadi'}]).to_excel(
                writer, sheet_name='Band_Analizi', index=False)
            return

        band_excel = candidates[0]
        try:
            xl_band = pd.ExcelFile(band_excel)
        except Exception as e:
            logger.warning(f"  [WARN] Band_Analizi Excel acilamadi: {e}")
            return

        startrow = 0

        # -- Bant Ozeti tablosu --
        if 'Bant_Ozeti' in xl_band.sheet_names:
            try:
                df_bo = pd.read_excel(xl_band, sheet_name='Bant_Ozeti')
                header_df = pd.DataFrame([['=== MOMENT BANT OZETI ===', '', '', '', '', '']],
                                          columns=df_bo.columns[:6].tolist() + ([''] * max(0, 6 - len(df_bo.columns))))
                label_row = pd.DataFrame([{'Bilgi': '--- MOMENT BANT OZETI ---'}])
                label_row.to_excel(writer, sheet_name='Band_Analizi', index=False, startrow=startrow)
                startrow += 2
                df_bo.to_excel(writer, sheet_name='Band_Analizi', index=False, startrow=startrow)
                startrow += len(df_bo) + 3
                logger.info(f"  [OK] Band_Analizi/Bant_Ozeti ({len(df_bo)} bant)")
            except Exception as e:
                logger.warning(f"  [WARN] Bant_Ozeti: {e}")

        # -- Korelasyon tablosu (top 20) --
        if 'Korelasyon' in xl_band.sheet_names:
            try:
                df_cor = pd.read_excel(xl_band, sheet_name='Korelasyon')
                # Pick spearman column
                spear_col = next((c for c in df_cor.columns if 'spearman' in c.lower()), None)
                if spear_col:
                    df_cor_top = df_cor.nlargest(20, spear_col)
                else:
                    df_cor_top = df_cor.head(20)
                label_row = pd.DataFrame([{'Bilgi': '--- OZELLIK KORELASYONLARI (Top-20 Spearman |r|) ---'}])
                label_row.to_excel(writer, sheet_name='Band_Analizi', index=False, startrow=startrow)
                startrow += 2
                df_cor_top.to_excel(writer, sheet_name='Band_Analizi', index=False, startrow=startrow)
                startrow += len(df_cor_top) + 3
                logger.info(f"  [OK] Band_Analizi/Korelasyon ({len(df_cor_top)} satir)")
            except Exception as e:
                logger.warning(f"  [WARN] Korelasyon: {e}")

    # =========================================================================
    # JSON / CHARTS / LATEX
    # =========================================================================

    def generate_summary_json(self):
        """JSON ozet dosyasi — gercek verilerle zenginlestirilmis"""
        ai_rows   = self.all_results.get('ai_rows',      [])
        an_rows   = self.all_results.get('anfis_results', [])
        ai_df = pd.DataFrame(ai_rows) if ai_rows else pd.DataFrame()
        an_df = pd.DataFrame(an_rows) if an_rows else pd.DataFrame()

        def _safe_loc(df_sub, r2_col, label_cols):
            """dropna ile guvenli idxmax + loc"""
            valid = df_sub[r2_col].dropna()
            if valid.empty:
                return None
            idx = valid.idxmax()
            row = df_sub.loc[idx]
            return row

        # Hedef bazli en iyi AI/ANFIS
        per_target = {}
        for tgt in ['MM', 'QM']:
            per_target[tgt] = {}
            if not ai_df.empty and 'Target' in ai_df.columns:
                sub = ai_df[ai_df['Target'] == tgt]
                best = _safe_loc(sub, 'Val_R2', [])
                if best is not None:
                    per_target[tgt]['AI_best_val_r2']      = round(float(best.get('Val_R2',  0) or 0), 4)
                    per_target[tgt]['AI_best_test_r2']     = round(float(best.get('Test_R2', 0) or 0), 4)
                    per_target[tgt]['AI_best_model']       = str(best.get('Model_Type', ''))
                    per_target[tgt]['AI_best_feature_set'] = str(best.get('Feature_Set', ''))
            if not an_df.empty and 'Target' in an_df.columns:
                sub = an_df[an_df['Target'] == tgt]
                best = _safe_loc(sub, 'Val_R2', [])
                if best is not None:
                    per_target[tgt]['ANFIS_best_val_r2']      = round(float(best.get('Val_R2',  0) or 0), 4)
                    per_target[tgt]['ANFIS_best_test_r2']     = round(float(best.get('Test_R2', 0) or 0), 4)
                    per_target[tgt]['ANFIS_best_config']      = str(best.get('Config_ID', ''))
                    per_target[tgt]['ANFIS_best_feature_set'] = str(best.get('Feature_Set', ''))

        summary = {
            'timestamp':            datetime.now().isoformat(),
            'total_ai_configs':     len(ai_rows),
            'total_anfis_records':  len(an_rows),
            'ai_datasets':          sorted(ai_df['Dataset'].unique().tolist())   if not ai_df.empty else [],
            'anfis_datasets':       sorted(an_df['Dataset'].unique().tolist())   if not an_df.empty else [],
            'targets':              ['MM', 'QM'],
            'phases_completed':     ['PFAZ1', 'PFAZ2', 'PFAZ3', 'PFAZ4', 'PFAZ5', 'PFAZ6'],
            'ai_best_val_r2':       float(ai_df['Val_R2'].max())  if not ai_df.empty else None,
            'ai_best_test_r2':      float(ai_df['Test_R2'].max()) if not ai_df.empty else None,
            'anfis_best_val_r2':    float(an_df['Val_R2'].max())  if not an_df.empty else None,
            'anfis_best_test_r2':   float(an_df['Test_R2'].max()) if not an_df.empty else None,
            'ai_model_types':       sorted(ai_df['Model_Type'].unique().tolist()) if not ai_df.empty else [],
            'per_target_best':      per_target,
            'feature_abbreviations': FEATURE_ABBREV,
            'feature_sets':         {k: v for k, v in FEATURE_SET_EXPAND.items()},
        }

        json_file = self.output_dir / 'final_summary.json'
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(summary, f, indent=2, ensure_ascii=False)

        logger.info(f"\n[OK] JSON: {json_file}")
        return json_file

    def generate_excel_charts(self):
        """Excel grafikleri"""
        if not self.use_excel_charts:
            logger.info("[SKIP] Excel charts disabled")
            return None
        try:
            from pfaz_modules.pfaz06_final_reporting.excel_charts import ExcelChartGenerator
            logger.info("\n[EXCEL CHARTS] Generating charts...")
            chart_gen = ExcelChartGenerator(output_dir=str(self.output_dir))
            charts = chart_gen.generate_all_charts(self.all_results)
            logger.info(f"[OK] Generated {len(charts)} Excel charts")
            return charts
        except Exception as e:
            logger.error(f"[ERROR] Excel charts generation failed: {e}")
            return None

    def generate_latex_report(self):
        """LaTeX raporu"""
        if not self.use_latex_generator:
            logger.info("[SKIP] LaTeX generator disabled")
            return None
        try:
            from pfaz_modules.pfaz06_final_reporting.latex_generator import LaTeXReportGenerator
            logger.info("\n[LATEX] Generating LaTeX report...")
            latex_gen = LaTeXReportGenerator(output_dir=str(self.output_dir))
            latex_file = latex_gen.generate_report(self.all_results)
            logger.info(f"[OK] LaTeX report: {latex_file}")
            return latex_file
        except Exception as e:
            logger.error(f"[ERROR] LaTeX generation failed: {e}")
            return None

    # =========================================================================
    # MAIN PIPELINE
    # =========================================================================

    def run_complete_pipeline(self):
        """Tam pipeline calistir"""
        start = datetime.now()

        logger.info("\n" + "="*80)
        logger.info("PFAZ 6: FINAL REPORTING & THESIS v2.0")
        logger.info("="*80)

        self.collect_all_results()
        excel_file   = self.generate_thesis_tables()
        json_file    = self.generate_summary_json()
        excel_charts = self.generate_excel_charts()
        latex_file   = self.generate_latex_report()

        # ---- ComprehensiveExcelReporter: ANFIS config detay raporu ----
        comprehensive_excel = None
        try:
            from pfaz_modules.pfaz06_final_reporting.comprehensive_excel_reporter import ComprehensiveExcelReporter
            _anfis_rows = self.all_results.get('anfis_results', [])
            if _anfis_rows:
                import pandas as _pd_cr
                _anfis_df_cr = _pd_cr.DataFrame(_anfis_rows)
                # Rename columns to match ComprehensiveExcelReporter expected schema
                _col_map = {
                    'Target': 'target', 'Config_ID': 'config_id',
                    'Val_R2': 'val_r2', 'Test_R2': 'test_r2',
                    'Val_RMSE': 'val_rmse', 'Test_RMSE': 'test_rmse',
                    'Dataset': 'dataset_name',
                }
                _anfis_df_cr = _anfis_df_cr.rename(columns={k: v for k, v in _col_map.items() if k in _anfis_df_cr.columns})
                _reporter = ComprehensiveExcelReporter(output_dir=str(self.output_dir))
                comprehensive_excel = str(_reporter.generate_full_report(_anfis_df_cr))
                logger.info(f"[OK] ComprehensiveExcelReporter: {comprehensive_excel}")
        except Exception as _e:
            logger.warning(f"[WARNING] ComprehensiveExcelReporter basarisiz (devam): {_e}")

        # ---- BootstrapConfidenceIntervals: Val R2 skorlari icin %95 CI ----
        bootstrap_result = None
        try:
            import numpy as _np_bs
            from pfaz_modules.pfaz12_advanced_analytics.bootstrap_confidence_intervals import BootstrapConfidenceIntervals
            _ai_rows = self.all_results.get('ai_rows', [])
            if _ai_rows:
                import pandas as _pd_bs
                _ai_df_bs = _pd_bs.DataFrame(_ai_rows)
                _r2_col = 'Val_R2' if 'Val_R2' in _ai_df_bs.columns else None
                if _r2_col:
                    _r2_vals = _ai_df_bs[_r2_col].dropna().values.astype(float)
                    if len(_r2_vals) >= 10:
                        _bci = BootstrapConfidenceIntervals(
                            n_bootstrap=5000,
                            confidence_level=0.95,
                            random_state=42,
                            output_dir=str(self.output_dir / 'bootstrap_ci')
                        )
                        bootstrap_result = _bci.bootstrap_statistic(_r2_vals, statistic=_np_bs.mean, method='percentile')
                        logger.info(
                            f"[OK] BootstrapCI (AI Val R2): "
                            f"mean={bootstrap_result['point_estimate']:.4f}  "
                            f"95% CI=[{bootstrap_result['ci_lower']:.4f}, {bootstrap_result['ci_upper']:.4f}]"
                        )
                        # Save CI to JSON
                        import json as _json_bs
                        _ci_path = self.output_dir / 'bootstrap_ci' / 'ai_val_r2_bootstrap_ci.json'
                        _ci_path.parent.mkdir(parents=True, exist_ok=True)
                        _ci_save = {k: float(v) if hasattr(v, 'item') else v
                                    for k, v in bootstrap_result.items()
                                    if k != 'bootstrap_distribution'}
                        with open(_ci_path, 'w') as _f_ci:
                            _json_bs.dump(_ci_save, _f_ci, indent=2)
        except Exception as _e:
            logger.warning(f"[WARNING] BootstrapConfidenceIntervals basarisiz (devam): {_e}")

        # ---- BootstrapConfidenceIntervals (ANFIS Val R2) ----
        bootstrap_anfis_result = None
        try:
            import numpy as _np_bsa
            from pfaz_modules.pfaz12_advanced_analytics.bootstrap_confidence_intervals import BootstrapConfidenceIntervals
            _anfis_rows_bs = self.all_results.get('anfis_results', [])
            if _anfis_rows_bs:
                import pandas as _pd_bsa
                _an_df_bsa = _pd_bsa.DataFrame(_anfis_rows_bs)
                _r2_col_a = 'Val_R2' if 'Val_R2' in _an_df_bsa.columns else None
                if _r2_col_a:
                    _r2_vals_a = _an_df_bsa[_r2_col_a].dropna().values.astype(float)
                    if len(_r2_vals_a) >= 10:
                        _bci_a = BootstrapConfidenceIntervals(
                            n_bootstrap=5000, confidence_level=0.95, random_state=42,
                            output_dir=str(self.output_dir / 'bootstrap_ci')
                        )
                        bootstrap_anfis_result = _bci_a.bootstrap_statistic(
                            _r2_vals_a, statistic=_np_bsa.mean, method='percentile')
                        logger.info(
                            f"[OK] BootstrapCI (ANFIS Val R2): "
                            f"mean={bootstrap_anfis_result['point_estimate']:.4f}  "
                            f"95% CI=[{bootstrap_anfis_result['ci_lower']:.4f}, "
                            f"{bootstrap_anfis_result['ci_upper']:.4f}]"
                        )
                        import json as _json_bsa
                        _ci_a_path = self.output_dir / 'bootstrap_ci' / 'anfis_val_r2_bootstrap_ci.json'
                        _ci_a_path.parent.mkdir(parents=True, exist_ok=True)
                        _ci_a_save = {k: float(v) if hasattr(v, 'item') else v
                                      for k, v in bootstrap_anfis_result.items()
                                      if k != 'bootstrap_distribution'}
                        with open(_ci_a_path, 'w') as _f_cia:
                            _json_bsa.dump(_ci_a_save, _f_cia, indent=2)
        except Exception as _e_bsa:
            logger.warning(f"[WARNING] BootstrapCI (ANFIS) basarisiz (devam): {_e_bsa}")

        # ---- AdvancedSensitivityAnalysis (PFAZ12): tornado diyagramı ile duyarlılık ----
        sensitivity_result = None
        try:
            from pfaz_modules.pfaz12_advanced_analytics.advanced_sensitivity_analysis import AdvancedSensitivityAnalysis
            import joblib as _jl_sa
            import pandas as _pds_sa
            import numpy as _np_sa
            import json as _jsn_sa
            # En iyi RF/XGB modelini bul
            _sa_model = None
            _sa_feat_names = None
            _sa_X_test = None
            _models_root_sa = None
            for _cand_sa in [
                self.output_dir.parent / 'trained_models',
                self.output_dir.parent.parent / 'outputs' / 'trained_models',
            ]:
                if _cand_sa.exists():
                    _models_root_sa = _cand_sa
                    break
            if _models_root_sa is not None:
                _best_r2_sa = -999.0
                _best_pkl_sa = None
                for _pkl_sa in list(_models_root_sa.rglob('model_*.pkl'))[:200]:
                    _pstr_sa = str(_pkl_sa).lower()
                    if not any(_t in _pstr_sa for _t in ('rf', 'xgb', 'gbm')):
                        continue
                    _mf_sa = _pkl_sa.parent / f'metrics_{_pkl_sa.parent.name}.json'
                    if not _mf_sa.exists():
                        continue
                    with open(_mf_sa) as _mff_sa:
                        _met_sa = _jsn_sa.load(_mff_sa)
                    _vr2_sa = _met_sa.get('val', {}).get('r2', -999.0)
                    if _vr2_sa > _best_r2_sa:
                        _best_r2_sa = _vr2_sa
                        _best_pkl_sa = _pkl_sa
                if _best_pkl_sa is not None:
                    _sa_model = _jl_sa.load(_best_pkl_sa)
                    # Dataset CSV'sini bul
                    _ds_name_sa = _best_pkl_sa.parts[-4] if len(_best_pkl_sa.parts) >= 4 else ''
                    _ds_dir_sa = _models_root_sa.parent / 'generated_datasets' / _ds_name_sa
                    if _ds_dir_sa.exists():
                        _meta_f_sa = _ds_dir_sa / 'metadata.json'
                        _col_n_sa = None
                        if _meta_f_sa.exists():
                            with open(_meta_f_sa) as _mf3:
                                _meta_sa = _jsn_sa.load(_mf3)
                            _fc_sa = _meta_sa.get('feature_names') or _meta_sa.get('feature_columns', [])
                            _tc_sa = _meta_sa.get('target_names')  or _meta_sa.get('target_columns',  [])
                            if _fc_sa and _tc_sa:
                                _col_n_sa = list(_fc_sa) + list(_tc_sa)
                        _rkw_sa = {'header': None, 'names': _col_n_sa} if _col_n_sa else {}
                        _df_test_sa = _pds_sa.read_csv(_ds_dir_sa / 'test.csv', **_rkw_sa)
                        _tgt_c_sa = _df_test_sa.columns[-1]
                        _sa_feat_names = [c for c in _df_test_sa.columns if c not in {'NUCLEUS', _tgt_c_sa}]
                        _sa_X_test = _df_test_sa[_sa_feat_names].fillna(0).values.astype(_np_sa.float32)
            if _sa_model is not None and _sa_feat_names is not None and _sa_X_test is not None:
                # Baseline: ortalama özellik değerleri; aralık: ±1 std
                _means_sa = _sa_X_test.mean(axis=0)
                _stds_sa  = _sa_X_test.std(axis=0).clip(min=1e-6)
                _baseline_sa = {_sa_feat_names[i]: float(_means_sa[i]) for i in range(len(_sa_feat_names))}
                _ranges_sa   = {
                    _sa_feat_names[i]: (float(_means_sa[i] - _stds_sa[i]), float(_means_sa[i] + _stds_sa[i]))
                    for i in range(len(_sa_feat_names))
                }
                # model_func: params dict -> tek skalerli tahmin
                def _model_func_sa(params):
                    _row = _np_sa.array([params[f] for f in _sa_feat_names], dtype=_np_sa.float32).reshape(1, -1)
                    return float(_sa_model.predict(_row)[0])
                _sa = AdvancedSensitivityAnalysis(
                    output_dir=str(self.output_dir / 'sensitivity_analysis'),
                    random_state=42
                )
                sensitivity_result = _sa.tornado_analysis(_model_func_sa, _baseline_sa, _ranges_sa)
                _sa.plot_tornado_diagram(sensitivity_result)
                _sa.export_to_excel('sensitivity_analysis.xlsx')
                logger.info(f"[OK] AdvancedSensitivityAnalysis: tornado diyagramı + Excel -> sensitivity_analysis/")
            else:
                logger.info("  [INFO] AdvancedSensitivityAnalysis: uygun model bulunamadı — atlanıyor")
        except Exception as _e:
            logger.warning(f"[WARNING] AdvancedSensitivityAnalysis basarisiz (devam): {_e}")

        duration = (datetime.now() - start).total_seconds()

        logger.info("\n" + "="*80)
        logger.info("[SUCCESS] PFAZ 6 TAMAMLANDI")
        logger.info("="*80)
        logger.info(f"Sure   : {duration:.1f} saniye")
        logger.info(f"Excel  : {excel_file}")
        logger.info(f"JSON   : {json_file}")
        if latex_file:
            logger.info(f"LaTeX  : {latex_file}")

        return {
            'results':               self.all_results,
            'excel_file':            str(excel_file),
            'json_file':             str(json_file),
            'excel_charts':          excel_charts,
            'latex_file':            str(latex_file) if latex_file else None,
            'comprehensive_excel':   comprehensive_excel,
            'bootstrap_ci':          bootstrap_result,
            'bootstrap_ci_anfis':    bootstrap_anfis_result,
            'sensitivity_analysis':  sensitivity_result,
            'duration':              duration
        }


def main():
    pipeline = FinalReportingPipeline()
    results = pipeline.run_complete_pipeline()
    return results


if __name__ == "__main__":
    main()
