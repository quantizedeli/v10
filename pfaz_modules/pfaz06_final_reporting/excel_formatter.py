"""
Advanced Excel Formatter
Conditional formatting, data validation, hyperlinks

11. modül - reporting/excel_formatter.py
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, IconSetRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.hyperlink import Hyperlink
from pathlib import Path
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class AdvancedExcelFormatter:
    """Excel için gelişmiş formatlama"""
    
    def __init__(self):
        self.colors = {
            'header': 'FF4472C4',
            'even_row': 'FFD9E1F2',
            'odd_row': 'FFFFFFFF',
            'good': 'FF92D050',
            'medium': 'FFFFF2CC',
            'poor': 'FFFFC7CE'
        }
        
        logger.info("Advanced Excel Formatter başlatıldı")
    
    def format_workbook(self, excel_file):
        """
        Complete workbook formatting
        
        Args:
            excel_file: Excel file path
        """
        
        excel_file = Path(excel_file)
        wb = load_workbook(excel_file)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Apply formatting
            self._format_headers(ws)
            self._apply_alternating_rows(ws)
            self._auto_fit_columns(ws)
            self._add_borders(ws)
            
        wb.save(excel_file)
        logger.info(f"✓ Workbook formatted: {excel_file}")
    
    def add_conditional_formatting(self, excel_file, sheet_name, range_str, rule_type='colorscale'):
        """
        Add conditional formatting
        
        Args:
            rule_type: 'colorscale', 'databar', 'iconset'
        """
        
        wb = load_workbook(excel_file)
        ws = wb[sheet_name]
        
        if rule_type == 'colorscale':
            rule = ColorScaleRule(
                start_type='min', start_color='FFF8696B',
                mid_type='percentile', mid_value=50, mid_color='FFFFEB84',
                end_type='max', end_color='FF63BE7B'
            )
        elif rule_type == 'databar':
            rule = DataBarRule(
                start_type='min', start_value=0,
                end_type='max', end_value=1,
                color='FF638EC6'
            )
        elif rule_type == 'iconset':
            rule = IconSetRule('3TrafficLights1', 'num', [0, 33, 67])
        else:
            return
        
        ws.conditional_formatting.add(range_str, rule)
        wb.save(excel_file)
        
        logger.info(f"✓ Conditional formatting added: {sheet_name}!{range_str}")
    
    def add_data_validation(self, excel_file, sheet_name, range_str, validation_list):
        """Add dropdown list validation"""
        
        wb = load_workbook(excel_file)
        ws = wb[sheet_name]
        
        dv = DataValidation(type="list", formula1=f'"{",".join(validation_list)}"')
        ws.add_data_validation(dv)
        dv.add(range_str)
        
        wb.save(excel_file)
        logger.info(f"✓ Data validation added: {sheet_name}!{range_str}")
    
    def _format_headers(self, ws):
        """Format header row"""
        header_font = Font(bold=True, color='FFFFFFFF', size=11)
        header_fill = PatternFill(start_color=self.colors['header'], 
                                  end_color=self.colors['header'], 
                                  fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
    
    def _apply_alternating_rows(self, ws):
        """Apply alternating row colors"""
        even_fill = PatternFill(start_color=self.colors['even_row'],
                               end_color=self.colors['even_row'],
                               fill_type='solid')
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if row_idx % 2 == 0:
                for cell in row:
                    cell.fill = even_fill
    
    def _auto_fit_columns(self, ws):
        """Auto-fit column widths"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _add_borders(self, ws):
        """Add borders to all cells"""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border


if __name__ == "__main__":
    print("✓ Excel Formatter modülü hazır - reporting/excel_formatter.py")

# ==================== EKLEME BAŞI ====================
class ExcelFormatter:
    """Wrapper for AdvancedExcelFormatter"""
    def __init__(self):
        self.formatter = AdvancedExcelFormatter()
    
    def create_comprehensive_report(self, results, output_path):
        return self.formatter.create_comprehensive_report(results, output_path)
# ==================== EKLEME SON ====================