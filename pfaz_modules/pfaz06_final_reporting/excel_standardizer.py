"""
Excel Standardizer
==================
Tüm PFAZ modülleri için standart Excel formatlama aracı.
AdvancedExcelFormatter yerine geçen genel amaçlı yardımcı.

Özellikler:
- Bold, renkli başlıklar (mavi header)
- Hücre uzunluğuna göre autosize
- Koşullu biçimlendirme (R² renk skalası, data bar, trafik ışığı)
- Pivot tablo oluşturma (openpyxl)
- İyi/Orta/Kötü renk sınıflandırması
- Her PFAZ modülünden `from ... import ExcelStandardizer` ile kullanılabilir

Kullanım:
    from pfaz_modules.pfaz06_final_reporting.excel_standardizer import ExcelStandardizer
    with ExcelStandardizer("rapor.xlsx") as es:
        es.write_sheet("Özet", df_ozet, conditional_cols=["R2", "RMSE"])
        es.write_pivot("Pivot", df, rows="Target", cols="Model", values="R2")
"""

import logging
from pathlib import Path
from typing import Dict, List, Optional, Union
from datetime import datetime

import pandas as pd
import numpy as np

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# xlsxwriter ile çalışır (birincil engine); yoksa openpyxl fallback
# ---------------------------------------------------------------------------
try:
    import xlsxwriter
    _XW_AVAILABLE = True
except ImportError:
    _XW_AVAILABLE = False

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import (
        ColorScaleRule, DataBarRule, CellIsRule, FormulaRule
    )
    from openpyxl.worksheet.table import Table, TableStyleInfo
    _OPX_AVAILABLE = True
except ImportError:
    _OPX_AVAILABLE = False


# ---------------------------------------------------------------------------
# Renk sabitleri
# ---------------------------------------------------------------------------
class _C:
    HEADER_BG   = "4472C4"   # Mavi başlık
    HEADER_FONT = "FFFFFF"   # Beyaz yazı
    TITLE_BG    = "D9E1F2"   # Açık mavi başlık
    GOOD_BG     = "C6EFCE"   # Yeşil (iyi)
    GOOD_FONT   = "006100"
    WARN_BG     = "FFEB9C"   # Sarı (orta)
    WARN_FONT   = "9C6500"
    BAD_BG      = "FFC7CE"   # Kırmızı (kötü)
    BAD_FONT    = "9C0006"
    ALT_ROW     = "F2F2F2"   # Alternatif satır grisi
    BORDER      = "BFBFBF"


# ===========================================================================
# Ana Sınıf — Context Manager
# ===========================================================================
class ExcelStandardizer:
    """
    Standart Excel formatlama aracı.

    with ExcelStandardizer("dosya.xlsx") as es:
        es.write_sheet("Sayfa1", df)
        es.write_pivot("Pivot", df, rows="Target", cols="Model", values="R2")
    """

    def __init__(self, filepath: Union[str, Path]):
        self.filepath = Path(filepath)
        self.filepath.parent.mkdir(parents=True, exist_ok=True)
        self._writer: Optional[pd.ExcelWriter] = None
        self._sheets_written: List[str] = []

    # -----------------------------------------------------------------------
    # Context manager
    # -----------------------------------------------------------------------
    def __enter__(self):
        engine = "xlsxwriter" if _XW_AVAILABLE else ("openpyxl" if _OPX_AVAILABLE else None)
        if engine is None:
            raise ImportError("xlsxwriter veya openpyxl kurulmalı: pip install xlsxwriter")
        self._writer = pd.ExcelWriter(str(self.filepath), engine=engine)
        self._engine = engine
        return self

    def __exit__(self, *args):
        if self._writer:
            self._writer.close()
        logger.info(f"[ExcelStandardizer] Kaydedildi: {self.filepath.name} "
                    f"({len(self._sheets_written)} sayfa)")

    # -----------------------------------------------------------------------
    # Temel sayfa yazma
    # -----------------------------------------------------------------------
    def write_sheet(
        self,
        sheet_name: str,
        df: pd.DataFrame,
        *,
        conditional_cols: Optional[List[str]] = None,
        r2_cols: Optional[List[str]] = None,
        freeze_header: bool = True,
        max_col_width: int = 50,
        index: bool = False,
    ) -> None:
        """
        DataFrame'i standart biçimlendirilmiş sayfaya yazar.

        Args:
            sheet_name: Excel sayfa adı
            df: Yazılacak veri
            conditional_cols: Koşullu biçimlendirme uygulanacak sütunlar
                               (ör. R², RMSE — iyi=yeşil, kötü=kırmızı)
            r2_cols: R² sütunları (0-1 skalasına göre renk)
            freeze_header: İlk satırı dondur
            max_col_width: Maksimum sütun genişliği (karakter)
            index: Satır indeksini yaz
        """
        if self._writer is None:
            raise RuntimeError("ExcelStandardizer context manager içinde kullanılmalı")

        sheet_name = sheet_name[:31]  # Excel 31 karakter sınırı
        df.to_excel(self._writer, sheet_name=sheet_name, index=index)
        self._sheets_written.append(sheet_name)

        if self._engine == "xlsxwriter":
            self._format_xlsxwriter(
                sheet_name, df,
                conditional_cols=conditional_cols or [],
                r2_cols=r2_cols or [],
                freeze_header=freeze_header,
                max_col_width=max_col_width,
                index=index,
            )
        else:
            self._format_openpyxl(
                sheet_name, df,
                conditional_cols=conditional_cols or [],
                r2_cols=r2_cols or [],
                freeze_header=freeze_header,
                max_col_width=max_col_width,
                index=index,
            )

    # -----------------------------------------------------------------------
    # xlsxwriter formatlama
    # -----------------------------------------------------------------------
    def _format_xlsxwriter(self, sheet_name, df, *, conditional_cols,
                           r2_cols, freeze_header, max_col_width, index):
        wb = self._writer.book
        ws = self._writer.sheets[sheet_name]
        n_rows, n_cols = df.shape
        col_offset = 1 if index else 0

        # --- Format tanımları ---
        fmt_header = wb.add_format({
            'bold': True, 'bg_color': f'#{_C.HEADER_BG}',
            'font_color': f'#{_C.HEADER_FONT}', 'align': 'center',
            'valign': 'vcenter', 'border': 1, 'text_wrap': True,
        })
        fmt_good   = wb.add_format({'bg_color': f'#{_C.GOOD_BG}',  'font_color': f'#{_C.GOOD_FONT}'})
        fmt_warn   = wb.add_format({'bg_color': f'#{_C.WARN_BG}',  'font_color': f'#{_C.WARN_FONT}'})
        fmt_bad    = wb.add_format({'bg_color': f'#{_C.BAD_BG}',   'font_color': f'#{_C.BAD_FONT}'})
        fmt_num    = wb.add_format({'num_format': '0.0000', 'align': 'center'})
        fmt_int    = wb.add_format({'num_format': '0', 'align': 'center'})

        # --- Başlık satırını yeniden yaz (formatla) ---
        for col_idx, col_name in enumerate(df.columns):
            ws.write(0, col_idx + col_offset, str(col_name), fmt_header)

        # --- Autosize ---
        for col_idx, col_name in enumerate(df.columns):
            series = df.iloc[:, col_idx].astype(str)
            col_width = min(max(series.str.len().max(), len(str(col_name))) + 2, max_col_width)
            ws.set_column(col_idx + col_offset, col_idx + col_offset, col_width)

        # --- Başlık satırını dondur ---
        if freeze_header:
            ws.freeze_panes(1, 0)

        # --- Koşullu biçimlendirme (R² sütunları) ---
        for col_name in r2_cols:
            if col_name in df.columns:
                col_idx = list(df.columns).index(col_name) + col_offset
                col_letter = self._col_letter(col_idx)
                # Renk skalası: kırmızı(0) → sarı(0.7) → yeşil(1)
                ws.conditional_format(
                    f'{col_letter}2:{col_letter}{n_rows + 1}',
                    {'type': '3_color_scale',
                     'min_color': f'#{_C.BAD_BG}',
                     'mid_color': f'#{_C.WARN_BG}',
                     'max_color': f'#{_C.GOOD_BG}',
                     'min_value': 0, 'mid_value': 0.7, 'max_value': 1.0,
                     'min_type': 'num', 'mid_type': 'num', 'max_type': 'num'}
                )

        # --- Koşullu biçimlendirme (genel sütunlar) ---
        for col_name in conditional_cols:
            if col_name in df.columns and col_name not in r2_cols:
                col_idx = list(df.columns).index(col_name) + col_offset
                col_letter = self._col_letter(col_idx)
                ws.conditional_format(
                    f'{col_letter}2:{col_letter}{n_rows + 1}',
                    {'type': 'data_bar',
                     'bar_color': f'#{_C.HEADER_BG}',
                     'bar_only': False}
                )

    # -----------------------------------------------------------------------
    # openpyxl formatlama (fallback)
    # -----------------------------------------------------------------------
    def _format_openpyxl(self, sheet_name, df, *, conditional_cols,
                         r2_cols, freeze_header, max_col_width, index):
        if not _OPX_AVAILABLE:
            return
        ws = self._writer.sheets[sheet_name]
        n_rows, n_cols = df.shape
        col_offset = 1 if index else 0

        header_fill = PatternFill("solid", fgColor=_C.HEADER_BG)
        header_font = Font(bold=True, color=_C.HEADER_FONT)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin", color=_C.BORDER)
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # --- Başlık ---
        for col_idx, col_name in enumerate(df.columns):
            cell = ws.cell(row=1, column=col_idx + 1 + col_offset)
            cell.value = str(col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border

        # --- Autosize ---
        for col_idx, col_name in enumerate(df.columns):
            series = df.iloc[:, col_idx].astype(str)
            col_width = min(max(series.str.len().max(), len(str(col_name))) + 2, max_col_width)
            col_letter = get_column_letter(col_idx + 1 + col_offset)
            ws.column_dimensions[col_letter].width = col_width

        # --- Dondur ---
        if freeze_header:
            ws.freeze_panes = ws.cell(row=2, column=1)

        # --- Koşullu biçimlendirme (R² renk skalası) ---
        for col_name in r2_cols:
            if col_name in df.columns:
                col_idx = list(df.columns).index(col_name)
                col_letter = get_column_letter(col_idx + 1 + col_offset)
                cell_range = f"{col_letter}2:{col_letter}{n_rows + 1}"
                rule = ColorScaleRule(
                    start_type="num", start_value=0,   start_color=_C.BAD_BG,
                    mid_type="num",   mid_value=0.7,   mid_color=_C.WARN_BG,
                    end_type="num",   end_value=1.0,   end_color=_C.GOOD_BG,
                )
                ws.conditional_formatting.add(cell_range, rule)

        # --- Koşullu biçimlendirme (data bar — genel) ---
        for col_name in conditional_cols:
            if col_name in df.columns and col_name not in r2_cols:
                col_idx = list(df.columns).index(col_name)
                col_letter = get_column_letter(col_idx + 1 + col_offset)
                cell_range = f"{col_letter}2:{col_letter}{n_rows + 1}"
                rule = DataBarRule(start_type="min", start_value=0,
                                   end_type="max",   end_value=100,
                                   color=_C.HEADER_BG)
                ws.conditional_formatting.add(cell_range, rule)

    # -----------------------------------------------------------------------
    # Pivot tablo
    # -----------------------------------------------------------------------
    def write_pivot(
        self,
        sheet_name: str,
        df: pd.DataFrame,
        rows: Union[str, List[str]],
        cols: Union[str, List[str]],
        values: str,
        aggfunc: str = "mean",
        *,
        r2_scale: bool = True,
    ) -> None:
        """
        Pivot tablo oluştur ve standart formatlı sayfaya yaz.

        Args:
            sheet_name: Sayfa adı
            df: Kaynak veri
            rows: Satır(lar)
            cols: Sütun(lar)
            values: Değer sütunu
            aggfunc: Toparlama fonksiyonu ('mean', 'max', 'min', 'count')
            r2_scale: R² renk skalası uygula
        """
        try:
            pivot = pd.pivot_table(
                df, values=values, index=rows, columns=cols,
                aggfunc=aggfunc, fill_value=np.nan
            ).round(4)
            # Çok seviyeli sütunları düzleştir
            if isinstance(pivot.columns, pd.MultiIndex):
                pivot.columns = ["_".join(str(c) for c in col) for col in pivot.columns]
            pivot = pivot.reset_index()
            r2 = [c for c in pivot.columns if r2_scale and "r2" in str(c).lower()]
            self.write_sheet(sheet_name, pivot, r2_cols=r2, index=False)
        except Exception as e:
            logger.warning(f"[ExcelStandardizer] Pivot hata ({sheet_name}): {e}")

    # -----------------------------------------------------------------------
    # Renk sınıflandırma yardımcısı (sütun bazlı trafik ışığı)
    # -----------------------------------------------------------------------
    @staticmethod
    def classify_r2(r2: float) -> str:
        """R² değerine göre sınıf döndürür."""
        if r2 >= 0.85:
            return "Excellent"
        elif r2 >= 0.70:
            return "Good"
        elif r2 >= 0.50:
            return "Moderate"
        else:
            return "Poor"

    # -----------------------------------------------------------------------
    # Yardımcı
    # -----------------------------------------------------------------------
    @staticmethod
    def _col_letter(col_idx: int) -> str:
        """0-tabanlı sütun indeksinden Excel harfi üret (A, B, …, AA, …)"""
        result = ""
        n = col_idx + 1
        while n:
            n, remainder = divmod(n - 1, 26)
            result = chr(65 + remainder) + result
        return result

    # -----------------------------------------------------------------------
    # Tek satırlık hızlı yazma (context manager olmadan)
    # -----------------------------------------------------------------------
    @classmethod
    def save(
        cls,
        filepath: Union[str, Path],
        sheets: Dict[str, pd.DataFrame],
        *,
        r2_cols_map: Optional[Dict[str, List[str]]] = None,
        conditional_map: Optional[Dict[str, List[str]]] = None,
    ) -> Path:
        """
        Birden fazla sayfayı tek çağrıda kaydet.

        Args:
            filepath: Çıktı dosyası
            sheets: {"Sayfa Adı": dataframe} sözlüğü
            r2_cols_map: {"Sayfa Adı": ["R2", "Val_R2"]} — R² renk skalası
            conditional_map: {"Sayfa Adı": ["RMSE"]} — data bar
        Returns:
            Path: Kaydedilen dosya yolu
        """
        r2_cols_map     = r2_cols_map     or {}
        conditional_map = conditional_map or {}
        fp = Path(filepath)
        with cls(fp) as es:
            for sname, df in sheets.items():
                es.write_sheet(
                    sname, df,
                    r2_cols=r2_cols_map.get(sname, []),
                    conditional_cols=conditional_map.get(sname, []),
                )
        return fp


# ===========================================================================
# Bağımsız yardımcı fonksiyonlar (import ederek doğrudan kullanılabilir)
# ===========================================================================

def autosize_and_header(ws, df: pd.DataFrame, *, max_width: int = 50,
                        freeze: bool = True) -> None:
    """
    openpyxl worksheet'e başlık + autosize uygular.
    Var olan bir worksheet'i sonradan biçimlendirmek için.
    """
    if not _OPX_AVAILABLE:
        return
    header_fill = PatternFill("solid", fgColor=_C.HEADER_BG)
    header_font = Font(bold=True, color=_C.HEADER_FONT)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill  = header_fill
        cell.font  = header_font
        cell.alignment = center
        if not cell.value:
            cell.value = str(col_name)
        series = df.iloc[:, col_idx - 1].astype(str)
        width  = min(max(series.str.len().max(), len(str(col_name))) + 2, max_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    if freeze:
        ws.freeze_panes = ws.cell(row=2, column=1)


def add_r2_color_scale(ws, col_letter: str, n_rows: int) -> None:
    """
    openpyxl worksheet'e R² renk skalası ekler.
    Örnek: add_r2_color_scale(ws, "D", 100)
    """
    if not _OPX_AVAILABLE:
        return
    cell_range = f"{col_letter}2:{col_letter}{n_rows + 1}"
    rule = ColorScaleRule(
        start_type="num", start_value=0,   start_color=_C.BAD_BG,
        mid_type="num",   mid_value=0.7,   mid_color=_C.WARN_BG,
        end_type="num",   end_value=1.0,   end_color=_C.GOOD_BG,
    )
    ws.conditional_formatting.add(cell_range, rule)


def color_cell(cell, r2: float) -> None:
    """
    Tek hücreye R² değerine göre renk uygular.
    openpyxl Cell nesnesi beklenir.
    """
    if not _OPX_AVAILABLE:
        return
    if r2 >= 0.85:
        bg, fg = _C.GOOD_BG, _C.GOOD_FONT
    elif r2 >= 0.70:
        bg, fg = _C.WARN_BG, _C.WARN_FONT
    else:
        bg, fg = _C.BAD_BG, _C.BAD_FONT
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(color=fg, bold=True)
