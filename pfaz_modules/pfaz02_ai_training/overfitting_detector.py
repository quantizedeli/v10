# -*- coding: utf-8 -*-
"""
PFAZ 2: Overfitting Detector
=============================

Train/Val gap analysis ve overfitting detection sistemi

Features:
- Epoch-by-epoch overfitting detection
- Train/Val performance gap analysis
- Visual charts (loss curves, gap trends)
- Excel comprehensive reports
- Overfitting remedies suggestions

Author: Nuclear Physics AI Training Pipeline
Version: 1.0.0
Date: 2025-10-15
"""

import numpy as np
import pandas as pd
from pathlib import Path
import json
import logging
from typing import Dict, List, Tuple, Optional
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

try:
    import matplotlib
    matplotlib.use('Agg')  # Non-interactive backend
    import matplotlib.pyplot as plt
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    matplotlib = None
    plt = None
    MATPLOTLIB_AVAILABLE = False
    logging.warning("Matplotlib not available - charts will be skipped")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.chart import LineChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    Workbook = None
    Font = None
    PatternFill = None
    Alignment = None
    LineChart = None
    Reference = None
    OPENPYXL_AVAILABLE = False
    logging.warning("openpyxl not available")

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


# ============================================================================
# OVERFITTING METRICS
# ============================================================================

class OverfittingMetrics:
    """Calculate various overfitting metrics"""
    
    @staticmethod
    def calculate_gap(train_metric: float, val_metric: float, metric_type: str = 'r2') -> float:
        """
        Calculate train-val gap
        
        Args:
            train_metric: Training metric value
            val_metric: Validation metric value
            metric_type: 'r2' (higher is better) or 'loss' (lower is better)
        
        Returns:
            Gap value (positive means overfitting)
        """
        if metric_type in ['r2', 'accuracy']:
            # Higher is better - gap = train - val
            return train_metric - val_metric
        else:
            # Lower is better - gap = val - train
            return val_metric - train_metric
    
    @staticmethod
    def calculate_gap_percentage(train_metric: float, val_metric: float, metric_type: str = 'r2') -> float:
        """Calculate percentage gap"""
        gap = OverfittingMetrics.calculate_gap(train_metric, val_metric, metric_type)
        
        if metric_type in ['r2', 'accuracy']:
            if train_metric > 0:
                return (gap / train_metric) * 100
        else:
            if train_metric > 0:
                return (gap / train_metric) * 100
        
        return 0.0
    
    @staticmethod
    def detect_overfitting_severity(gap_percentage: float) -> str:
        """
        Classify overfitting severity
        
        Returns:
            'none', 'mild', 'moderate', 'severe', 'extreme'
        """
        if gap_percentage < 5:
            return 'none'
        elif gap_percentage < 10:
            return 'mild'
        elif gap_percentage < 20:
            return 'moderate'
        elif gap_percentage < 40:
            return 'severe'
        else:
            return 'extreme'


# ============================================================================
# TRAINING HISTORY ANALYZER
# ============================================================================

class TrainingHistoryAnalyzer:
    """Analyze training history for overfitting patterns"""
    
    def __init__(self, history: Dict):
        """
        Args:
            history: Training history dict with keys like:
                     'train_loss', 'val_loss', 'train_r2', 'val_r2', etc.
        """
        self.history = history
        self.epochs = len(history.get('train_loss', []))
        
    def detect_divergence_point(self) -> Optional[int]:
        """
        Detect epoch where train and val start diverging significantly
        
        Returns:
            Epoch number or None
        """
        if 'train_loss' not in self.history or 'val_loss' not in self.history:
            return None
        
        train_losses = np.array(self.history['train_loss'])
        val_losses = np.array(self.history['val_loss'])
        
        if len(train_losses) < 5:
            return None
        
        # Calculate gap over time
        gaps = val_losses - train_losses
        
        # Use moving average to smooth
        window = min(5, len(gaps) // 3)
        if window < 2:
            return None
        
        gaps_smooth = np.convolve(gaps, np.ones(window)/window, mode='valid')
        
        # Find where gap starts increasing significantly
        gap_diff = np.diff(gaps_smooth)
        
        # Threshold: gap increase > 0.1 * initial gap
        threshold = 0.1 * abs(gaps_smooth[0]) if gaps_smooth[0] != 0 else 0.01
        
        for i, diff in enumerate(gap_diff):
            if diff > threshold:
                return i + window
        
        return None
    
    def calculate_final_gap(self, metric: str = 'r2') -> Dict:
        """Calculate final train-val gap"""
        train_key = f'train_{metric}'
        val_key = f'val_{metric}'
        
        if train_key not in self.history or val_key not in self.history:
            return {}
        
        train_value = self.history[train_key][-1]
        val_value = self.history[val_key][-1]
        
        gap = OverfittingMetrics.calculate_gap(train_value, val_value, metric)
        gap_pct = OverfittingMetrics.calculate_gap_percentage(train_value, val_value, metric)
        severity = OverfittingMetrics.detect_overfitting_severity(gap_pct)
        
        return {
            'train': train_value,
            'val': val_value,
            'gap': gap,
            'gap_percentage': gap_pct,
            'severity': severity
        }
    
    def get_best_epoch(self, metric: str = 'val_r2', mode: str = 'max') -> int:
        """Find best epoch based on validation metric"""
        key = f'val_{metric}' if not metric.startswith('val_') else metric
        
        if key not in self.history:
            return 0
        
        values = self.history[key]
        
        if mode == 'max':
            return int(np.argmax(values))
        else:
            return int(np.argmin(values))


# ============================================================================
# OVERFITTING DETECTOR
# ============================================================================

class OverfittingDetector:
    """
    Main overfitting detection system
    
    Features:
    - Load training histories
    - Detect overfitting patterns
    - Generate visual charts
    - Create Excel reports
    - Suggest remedies
    """
    
    def __init__(self, output_dir: str = 'overfitting_analysis'):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        self.analyses = []
        
        logger.info("=" * 80)
        logger.info("OVERFITTING DETECTOR INITIALIZED")
        logger.info("=" * 80)
        logger.info(f"Output directory: {self.output_dir}")
        logger.info("=" * 80)
    
    def load_training_results(self, results_dir: Path) -> List[Dict]:
        """
        Load training results from directory
        
        Expected structure:
        results_dir/
            dataset_name/
                model_type/
                    config_id/
                        metrics_*.json  (contains history)
        """
        results = []
        
        # Find all metrics files
        metrics_files = list(results_dir.glob('**/metrics_*.json'))
        
        logger.info(f"Found {len(metrics_files)} training result files")
        
        for metrics_file in metrics_files:
            try:
                with open(metrics_file, 'r', encoding='utf-8') as f:
                    metrics = json.load(f)
                
                # Extract identifiers from path
                config_id = metrics_file.parent.name
                model_type = metrics_file.parent.parent.name
                dataset_name = metrics_file.parent.parent.parent.name
                
                result = {
                    'dataset_name': dataset_name,
                    'model_type': model_type,
                    'config_id': config_id,
                    'metrics': metrics,
                    'file_path': metrics_file
                }
                
                results.append(result)
            
            except Exception as e:
                logger.error(f"Failed to load {metrics_file}: {e}")
        
        logger.info(f"Successfully loaded {len(results)} training results")
        return results
    
    def analyze_single_training(self, result: Dict) -> Dict:
        """Analyze single training result for overfitting"""
        
        metrics = result['metrics']
        
        analysis = {
            'dataset_name': result['dataset_name'],
            'model_type': result['model_type'],
            'config_id': result['config_id'],
            'has_overfitting': False,
            'severity': 'none',
            'gaps': {},
            'divergence_epoch': None,
            'best_epoch': None,
            'suggestions': []
        }
        
        # Check if history exists
        if 'history' in metrics:
            history = metrics['history']
            analyzer = TrainingHistoryAnalyzer(history)
            
            # Analyze R2 gap
            if 'train_r2' in history and 'val_r2' in history:
                r2_gap = analyzer.calculate_final_gap('r2')
                analysis['gaps']['r2'] = r2_gap
                
                if r2_gap.get('severity') not in ['none', 'mild']:
                    analysis['has_overfitting'] = True
                    analysis['severity'] = r2_gap['severity']
            
            # Analyze loss gap
            if 'train_loss' in history and 'val_loss' in history:
                loss_gap = analyzer.calculate_final_gap('loss')
                analysis['gaps']['loss'] = loss_gap
            
            # Detect divergence point
            divergence = analyzer.detect_divergence_point()
            if divergence:
                analysis['divergence_epoch'] = divergence
            
            # Find best epoch
            if 'val_r2' in history:
                analysis['best_epoch'] = analyzer.get_best_epoch('r2', 'max')
            elif 'val_loss' in history:
                analysis['best_epoch'] = analyzer.get_best_epoch('loss', 'min')
        
        # Generate suggestions
        analysis['suggestions'] = self._generate_suggestions(analysis)
        
        return analysis
    
    def analyze_all_trainings(self, results: List[Dict]) -> List[Dict]:
        """Analyze all training results"""
        
        logger.info("\n" + "=" * 80)
        logger.info("ANALYZING ALL TRAININGS FOR OVERFITTING")
        logger.info("=" * 80)
        
        analyses = []
        
        for result in results:
            analysis = self.analyze_single_training(result)
            analyses.append(analysis)
            
            if analysis['has_overfitting']:
                logger.warning(
                    f"[WARNING]  {result['dataset_name']}/{result['model_type']}/{result['config_id']}: "
                    f"Overfitting detected ({analysis['severity']})"
                )
        
        self.analyses = analyses
        
        # Summary statistics
        total = len(analyses)
        overfitted = len([a for a in analyses if a['has_overfitting']])
        
        logger.info("\n" + "=" * 80)
        logger.info("OVERFITTING ANALYSIS SUMMARY")
        logger.info("=" * 80)
        logger.info(f"Total trainings analyzed: {total}")
        logger.info(f"Overfitting detected: {overfitted} ({overfitted/total*100:.1f}%)")
        logger.info("=" * 80 + "\n")
        
        return analyses
    
    def _generate_suggestions(self, analysis: Dict) -> List[str]:
        """Generate suggestions to reduce overfitting"""
        
        suggestions = []
        
        if not analysis['has_overfitting']:
            return ["[SUCCESS] No significant overfitting detected"]
        
        severity = analysis['severity']
        
        if severity == 'extreme':
            suggestions.append("[ERROR] CRITICAL: Extreme overfitting detected!")
            suggestions.append("-> Reduce model complexity significantly")
            suggestions.append("-> Increase dropout rates")
            suggestions.append("-> Use stronger regularization (L1/L2)")
            suggestions.append("-> Consider using more training data")
            suggestions.append("-> Try early stopping at epoch {}".format(analysis.get('divergence_epoch', 'N/A')))
        
        elif severity == 'severe':
            suggestions.append("🟠 Severe overfitting detected")
            suggestions.append("-> Increase dropout rates (try +0.1 to +0.2)")
            suggestions.append("-> Add L2 regularization")
            suggestions.append("-> Reduce model layers or units")
            suggestions.append("-> Use data augmentation if possible")
        
        elif severity == 'moderate':
            suggestions.append("🟡 Moderate overfitting detected")
            suggestions.append("-> Slightly increase dropout")
            suggestions.append("-> Consider early stopping")
            suggestions.append("-> Try batch normalization")
        
        elif severity == 'mild':
            suggestions.append("[OK] Mild overfitting - generally acceptable")
            suggestions.append("-> Monitor in production")
            suggestions.append("-> Consider ensemble methods")
        
        return suggestions
    
    def generate_charts(self, top_n: int = 20):
        """Generate overfitting visualization charts"""
        
        if not MATPLOTLIB_AVAILABLE:
            logger.warning("Matplotlib not available - skipping charts")
            return
        
        logger.info("Generating overfitting charts...")
        
        charts_dir = self.output_dir / 'charts'
        charts_dir.mkdir(exist_ok=True)
        
        # Sort by severity
        severity_order = {'extreme': 5, 'severe': 4, 'moderate': 3, 'mild': 2, 'none': 1}
        sorted_analyses = sorted(
            [a for a in self.analyses if a['has_overfitting']],
            key=lambda x: severity_order.get(x['severity'], 0),
            reverse=True
        )[:top_n]
        
        for analysis in sorted_analyses:
            try:
                self._create_single_chart(analysis, charts_dir)
            except Exception as e:
                logger.error(f"Failed to create chart: {e}")
        
        # Create summary chart
        self._create_summary_chart(charts_dir)
        
        logger.info(f"[SUCCESS] Charts saved to {charts_dir}")
    
    def _create_single_chart(self, analysis: Dict, output_dir: Path):
        """Create chart for single training"""
        
        # This would require loading full history
        # Placeholder for now
        pass
    
    def _create_summary_chart(self, output_dir: Path):
        """Create summary overfitting chart"""
        
        if not self.analyses:
            return
        
        # Count by severity
        severity_counts = {
            'none': 0,
            'mild': 0,
            'moderate': 0,
            'severe': 0,
            'extreme': 0
        }
        
        for analysis in self.analyses:
            severity = analysis['severity']
            severity_counts[severity] += 1
        
        # Create bar chart
        fig, ax = plt.subplots(figsize=(10, 6))
        
        severities = list(severity_counts.keys())
        counts = list(severity_counts.values())
        colors = ['green', 'lightgreen', 'yellow', 'orange', 'red']
        
        bars = ax.bar(severities, counts, color=colors, alpha=0.7, edgecolor='black')
        
        # Add value labels on bars
        for bar, count in zip(bars, counts):
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2., height,
                   f'{count}',
                   ha='center', va='bottom', fontsize=12, fontweight='bold')
        
        ax.set_xlabel('Overfitting Severity', fontsize=12, fontweight='bold')
        ax.set_ylabel('Number of Trainings', fontsize=12, fontweight='bold')
        ax.set_title('Overfitting Severity Distribution', fontsize=14, fontweight='bold')
        ax.grid(axis='y', alpha=0.3)
        
        plt.tight_layout()
        plt.savefig(output_dir / 'overfitting_summary.png', dpi=150, bbox_inches='tight')
        plt.close()
        
        logger.info("[SUCCESS] Summary chart created")
    
    def generate_excel_report(self, filename: str = 'Overfitting_Analysis.xlsx'):
        """Generate comprehensive Excel report"""
        
        if not self.analyses:
            logger.warning("No analyses to export")
            return
        
        excel_path = self.output_dir / filename
        
        logger.info(f"Generating Excel report: {excel_path}")
        
        if OPENPYXL_AVAILABLE:
            self._generate_formatted_excel(excel_path)
        else:
            # Fallback to simple Excel
            self._generate_simple_excel(excel_path)
        
        logger.info(f"[SUCCESS] Excel report saved: {excel_path}")
    
    def _generate_simple_excel(self, excel_path: Path):
        """Generate simple Excel without formatting"""
        
        # Create dataframe
        data = []
        
        for analysis in self.analyses:
            row = {
                'Dataset': analysis['dataset_name'],
                'Model_Type': analysis['model_type'],
                'Config_ID': analysis['config_id'],
                'Has_Overfitting': analysis['has_overfitting'],
                'Severity': analysis['severity'],
                'Divergence_Epoch': analysis.get('divergence_epoch', 'N/A'),
                'Best_Epoch': analysis.get('best_epoch', 'N/A')
            }
            
            # Add gap metrics
            if 'r2' in analysis['gaps']:
                r2_gap = analysis['gaps']['r2']
                row['R2_Train'] = r2_gap.get('train', 'N/A')
                row['R2_Val'] = r2_gap.get('val', 'N/A')
                row['R2_Gap'] = r2_gap.get('gap', 'N/A')
                row['R2_Gap_Pct'] = r2_gap.get('gap_percentage', 'N/A')
            
            # Add suggestions
            row['Suggestions'] = ' | '.join(analysis['suggestions'])
            
            data.append(row)
        
        df = pd.DataFrame(data)
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Overfitting_Analysis', index=False)
            
            # Summary statistics
            summary_data = {
                'Metric': [
                    'Total Trainings',
                    'With Overfitting',
                    'None Severity',
                    'Mild Severity',
                    'Moderate Severity',
                    'Severe Severity',
                    'Extreme Severity'
                ],
                'Count': [
                    len(self.analyses),
                    len([a for a in self.analyses if a['has_overfitting']]),
                    len([a for a in self.analyses if a['severity'] == 'none']),
                    len([a for a in self.analyses if a['severity'] == 'mild']),
                    len([a for a in self.analyses if a['severity'] == 'moderate']),
                    len([a for a in self.analyses if a['severity'] == 'severe']),
                    len([a for a in self.analyses if a['severity'] == 'extreme'])
                ]
            }
            
            pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', index=False)
    
    def _generate_formatted_excel(self, excel_path: Path):
        """Generate formatted Excel with styling"""
        
        wb = Workbook()
        wb.remove(wb.active)
        
        # Sheet 1: Detailed Analysis
        ws1 = wb.create_sheet('Overfitting_Analysis')
        
        headers = ['Dataset', 'Model', 'Config', 'Severity', 'R2_Train', 'R2_Val', 
                   'R2_Gap%', 'Best_Epoch', 'Suggestions']
        ws1.append(headers)
        
        # Style header
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in ws1[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Add data with conditional formatting
        for analysis in self.analyses:
            r2_gap = analysis['gaps'].get('r2', {})
            
            row = [
                analysis['dataset_name'],
                analysis['model_type'],
                analysis['config_id'],
                analysis['severity'],
                r2_gap.get('train', 'N/A'),
                r2_gap.get('val', 'N/A'),
                r2_gap.get('gap_percentage', 'N/A'),
                analysis.get('best_epoch', 'N/A'),
                ' | '.join(analysis['suggestions'][:2])  # First 2 suggestions
            ]
            
            ws1.append(row)
            
            # Color code by severity
            current_row = ws1.max_row
            severity_cell = ws1.cell(current_row, 4)
            
            if analysis['severity'] == 'extreme':
                severity_cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                severity_cell.font = Font(color='FFFFFF', bold=True)
            elif analysis['severity'] == 'severe':
                severity_cell.fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
            elif analysis['severity'] == 'moderate':
                severity_cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            elif analysis['severity'] == 'mild':
                severity_cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        
        # Auto-adjust columns
        for column in ws1.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception as e:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws1.column_dimensions[column_letter].width = adjusted_width
        
        # Sheet 2: Summary
        ws2 = wb.create_sheet('Summary')
        ws2.append(['Metric', 'Count'])
        
        summary_stats = [
            ('Total Trainings', len(self.analyses)),
            ('With Overfitting', len([a for a in self.analyses if a['has_overfitting']])),
            ('None Severity', len([a for a in self.analyses if a['severity'] == 'none'])),
            ('Mild Severity', len([a for a in self.analyses if a['severity'] == 'mild'])),
            ('Moderate Severity', len([a for a in self.analyses if a['severity'] == 'moderate'])),
            ('Severe Severity', len([a for a in self.analyses if a['severity'] == 'severe'])),
            ('Extreme Severity', len([a for a in self.analyses if a['severity'] == 'extreme']))
        ]
        
        for metric, count in summary_stats:
            ws2.append([metric, count])
        
        # Style summary header
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        wb.save(excel_path)


# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    """Main execution for testing"""
    
    print("\n" + "=" * 80)
    print("PFAZ 2: OVERFITTING DETECTOR - TEST")
    print("=" * 80)
    
    # Initialize detector
    detector = OverfittingDetector(output_dir='test_overfitting_analysis')
    
    # Create mock training results for testing
    print("\nCreating mock training results...")
    test_results_dir = Path('test_training_results')
    test_results_dir.mkdir(exist_ok=True)
    
    # Create sample result
    sample_history = {
        'train_loss': [0.5, 0.3, 0.2, 0.15, 0.12, 0.10, 0.09, 0.08],
        'val_loss': [0.55, 0.35, 0.25, 0.22, 0.24, 0.26, 0.28, 0.30],
        'train_r2': [0.5, 0.7, 0.8, 0.85, 0.88, 0.90, 0.91, 0.92],
        'val_r2': [0.45, 0.65, 0.75, 0.78, 0.76, 0.74, 0.72, 0.70]
    }
    
    sample_result_dir = test_results_dir / 'MM_75nuclei' / 'DNN' / 'TRAIN_001'
    sample_result_dir.mkdir(parents=True, exist_ok=True)
    
    with open(sample_result_dir / 'metrics_TRAIN_001.json', 'w', encoding='utf-8') as f:
        json.dump({'history': sample_history}, f)
    
    # Load results
    print("\nLoading training results...")
    results = detector.load_training_results(test_results_dir)
    
    # Analyze
    print("\nAnalyzing for overfitting...")
    analyses = detector.analyze_all_trainings(results)
    
    # Generate charts
    print("\nGenerating charts...")
    detector.generate_charts()
    
    # Generate Excel report
    print("\nGenerating Excel report...")
    detector.generate_excel_report()
    
    print("\n[SUCCESS] OVERFITTING DETECTION TEST COMPLETED!")
    print(f"Output directory: {detector.output_dir}")


if __name__ == "__main__":
    main()
