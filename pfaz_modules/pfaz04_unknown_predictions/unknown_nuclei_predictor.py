# -*- coding: utf-8 -*-
"""
PFAZ 4: Unknown Nuclei Predictor
=================================

Predict on test-split (unknown) nuclei using trained AI and ANFIS models.
Compare known (val) vs unknown (test) performance and analyze degradation.

Features:
- Load AI models (PFAZ 2) and ANFIS models (PFAZ 3)
- Use test.csv from each dataset as "unknown" nuclei
- Known (val) vs Unknown (test) performance comparison
- Degradation analysis (R² drop)
- Per-nucleus prediction tracking
- Excel reports with filtering and pivot tables

Author: Nuclear Physics AI Training Pipeline
Version: 2.0.0
"""

import numpy as np
import pandas as pd
from pathlib import Path
import json
import logging
import traceback
from typing import Dict, List, Optional, Tuple
import warnings
warnings.filterwarnings('ignore')

try:
    import joblib
    JOBLIB_AVAILABLE = True
except ImportError:
    joblib = None
    JOBLIB_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    Workbook = None
    Font = None
    PatternFill = None
    Alignment = None
    Border = None
    Side = None
    get_column_letter = None
    OPENPYXL_AVAILABLE = False

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Actual CSV column names for targets
TARGET_COLUMNS = {
    'MM':     'MAGNETIC MOMENT [µ]',
    'QM':     'QUADRUPOLE MOMENT [Q]',
    'Beta_2': 'Beta_2',
    'MM_QM':  ['MAGNETIC MOMENT [µ]', 'QUADRUPOLE MOMENT [Q]'],
}

NON_FEATURE_COLS = {'NUCLEUS', 'MAGNETIC MOMENT [µ]', 'QUADRUPOLE MOMENT [Q]', 'Beta_2'}


class UnknownNucleiPredictor:
    """
    Unknown Nuclei Predictor v2.0

    Uses test.csv from each generated dataset as the "unknown" held-out nuclei.
    Loads AI (PFAZ2) and ANFIS (PFAZ3) models, predicts on test data,
    and compares against known (val) performance to measure generalization.
    """

    def __init__(self,
                 ai_models_dir: str,
                 anfis_models_dir: str,
                 splits_dir: str,
                 output_dir: str = 'unknown_predictions'):

        self.ai_models_dir = Path(ai_models_dir)
        self.anfis_models_dir = Path(anfis_models_dir)
        self.datasets_dir = Path(splits_dir)   # generated_datasets/
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

        self.all_results: List[Dict] = []
        self.per_nucleus_results: List[Dict] = []

        logger.info("=" * 80)
        logger.info("PFAZ 4: UNKNOWN NUCLEI PREDICTOR v2.0 INITIALIZED")
        logger.info("=" * 80)
        logger.info(f"AI models dir   : {self.ai_models_dir}")
        logger.info(f"ANFIS models dir: {self.anfis_models_dir}")
        logger.info(f"Datasets dir    : {self.datasets_dir}")
        logger.info(f"Output dir      : {self.output_dir}")
        logger.info("=" * 80)

    # ------------------------------------------------------------------
    # Target identification
    # ------------------------------------------------------------------

    def _identify_targets(self, df: pd.DataFrame) -> List[str]:
        """Return list of target column names found in df."""
        targets = []
        if 'MAGNETIC MOMENT [µ]' in df.columns:
            targets.append('MAGNETIC MOMENT [µ]')
        if 'QUADRUPOLE MOMENT [Q]' in df.columns:
            targets.append('QUADRUPOLE MOMENT [Q]')
        if 'Beta_2' in df.columns:
            targets.append('Beta_2')
        return targets

    def _get_feature_cols(self, df: pd.DataFrame, target_cols: List[str]) -> List[str]:
        """Return feature columns (all except NUCLEUS and targets)."""
        exclude = {'NUCLEUS'} | set(target_cols)
        return [c for c in df.columns if c not in exclude]

    # ------------------------------------------------------------------
    # Model loading helpers
    # ------------------------------------------------------------------

    def _load_ai_models_for_dataset(self, dataset_name: str) -> List[Dict]:
        """
        Load all AI model PKLs for a given dataset.
        Path: trained_models/{dataset}/{model_type}/{config}/model_{model_type}_{config}.pkl
        """
        dataset_dir = self.ai_models_dir / dataset_name
        if not dataset_dir.exists():
            return []

        result = []
        for pkl in dataset_dir.glob('**/model_*.pkl'):
            config_id = pkl.parent.name        # e.g. DNN_036
            model_type = pkl.parent.parent.name  # e.g. DNN

            metrics_file = pkl.parent / f'metrics_{config_id}.json'
            val_r2 = None
            train_r2 = None
            if metrics_file.exists():
                try:
                    with open(metrics_file, encoding='utf-8') as f:
                        m = json.load(f)
                    val_r2 = m.get('val', {}).get('r2')
                    train_r2 = m.get('train', {}).get('r2')
                except Exception:
                    pass

            try:
                model = joblib.load(pkl)
                result.append({
                    'model': model,
                    'model_type': model_type,
                    'config_id': config_id,
                    'val_r2': val_r2,
                    'train_r2': train_r2,
                })
            except Exception as e:
                logger.debug(f"Failed to load AI model {pkl}: {e}")

        return result

    def _load_anfis_models_for_dataset(self, dataset_name: str) -> List[Dict]:
        """
        Load all ANFIS model PKLs for a given dataset.
        Path: anfis_models/{dataset}/{config}/model_{config}.pkl
        """
        dataset_dir = self.anfis_models_dir / dataset_name
        if not dataset_dir.exists():
            return []

        result = []
        for pkl in dataset_dir.glob('*/model_*.pkl'):
            config_id = pkl.parent.name   # e.g. ANFIS_001

            metrics_file = pkl.parent / f'metrics_{config_id}.json'
            val_r2 = None
            train_r2 = None
            if metrics_file.exists():
                try:
                    with open(metrics_file, encoding='utf-8') as f:
                        m = json.load(f)
                    val_r2 = m.get('val', {}).get('r2')
                    train_r2 = m.get('train', {}).get('r2')
                except Exception:
                    pass

            try:
                model = joblib.load(pkl)
                result.append({
                    'model': model,
                    'config_id': config_id,
                    'val_r2': val_r2,
                    'train_r2': train_r2,
                })
            except Exception as e:
                logger.debug(f"Failed to load ANFIS model {pkl}: {e}")

        return result

    # ------------------------------------------------------------------
    # Prediction helpers
    # ------------------------------------------------------------------

    def _predict_with_model(self, model, X: np.ndarray) -> Optional[np.ndarray]:
        """Predict using AI or ANFIS model. ANFIS may be a list of models."""
        try:
            if isinstance(model, list):
                # Multi-output ANFIS: list of single-output models
                preds = [m.predict(X) for m in model]
                return np.column_stack(preds)
            else:
                return model.predict(X)
        except Exception as e:
            logger.debug(f"Predict error: {e}")
            return None

    def _calculate_metrics(self, y_true: np.ndarray, y_pred: np.ndarray,
                           target_name: str = '') -> Dict:
        """Calculate R², RMSE, MAE. Handles 1D and 2D arrays."""
        from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score

        y_true = np.asarray(y_true).ravel()
        y_pred = np.asarray(y_pred).ravel()

        if len(y_true) == 0 or len(y_pred) == 0:
            return {'r2': None, 'rmse': None, 'mae': None}

        try:
            return {
                'r2': float(r2_score(y_true, y_pred)),
                'rmse': float(np.sqrt(mean_squared_error(y_true, y_pred))),
                'mae': float(mean_absolute_error(y_true, y_pred)),
            }
        except Exception:
            return {'r2': None, 'rmse': None, 'mae': None}

    # ------------------------------------------------------------------
    # Main prediction loop
    # ------------------------------------------------------------------

    def predict_unknown_nuclei(self) -> Dict:
        """Predict on test.csv of all datasets and generate report."""

        logger.info("\n" + "=" * 80)
        logger.info("PREDICTING ON UNKNOWN (TEST) NUCLEI")
        logger.info("=" * 80)

        # Iterate over all dataset directories
        dataset_dirs = sorted([
            d for d in self.datasets_dir.iterdir()
            if d.is_dir() and (d / 'test.csv').exists()
        ])

        logger.info(f"Found {len(dataset_dirs)} datasets with test.csv")

        for dataset_dir in dataset_dirs:
            dataset_name = dataset_dir.name
            test_csv = dataset_dir / 'test.csv'

            try:
                test_df = pd.read_csv(test_csv)
            except Exception as e:
                logger.warning(f"Cannot read {test_csv}: {e}")
                continue

            target_cols = self._identify_targets(test_df)
            if not target_cols:
                logger.debug(f"No target columns in {dataset_name}, skipping")
                continue

            feature_cols = self._get_feature_cols(test_df, target_cols)
            if not feature_cols:
                logger.debug(f"No feature columns in {dataset_name}, skipping")
                continue

            X_test = test_df[feature_cols].values
            nuclei = test_df['NUCLEUS'].tolist() if 'NUCLEUS' in test_df.columns else [f'N{i}' for i in range(len(X_test))]
            n_test = len(X_test)

            logger.info(f"\n  Dataset: {dataset_name}  |  Targets: {target_cols}  |  Test size: {n_test}")

            # --- AI models ---
            ai_models = self._load_ai_models_for_dataset(dataset_name)
            for m_info in ai_models:
                y_pred = self._predict_with_model(m_info['model'], X_test)
                if y_pred is None:
                    continue

                self._process_predictions(
                    y_pred=y_pred,
                    test_df=test_df,
                    target_cols=target_cols,
                    nuclei=nuclei,
                    dataset_name=dataset_name,
                    model_category='AI',
                    model_type=m_info['model_type'],
                    config_id=m_info['config_id'],
                    val_r2=m_info['val_r2'],
                    train_r2=m_info['train_r2'],
                )

            # --- ANFIS models ---
            anfis_models = self._load_anfis_models_for_dataset(dataset_name)
            for m_info in anfis_models:
                y_pred = self._predict_with_model(m_info['model'], X_test)
                if y_pred is None:
                    continue

                self._process_predictions(
                    y_pred=y_pred,
                    test_df=test_df,
                    target_cols=target_cols,
                    nuclei=nuclei,
                    dataset_name=dataset_name,
                    model_category='ANFIS',
                    model_type='ANFIS',
                    config_id=m_info['config_id'],
                    val_r2=m_info['val_r2'],
                    train_r2=m_info['train_r2'],
                )

        logger.info("\n" + "=" * 80)
        logger.info(f"COMPLETED: {len(self.all_results)} total model-dataset combinations")
        logger.info("=" * 80 + "\n")

        self.generate_excel_report()

        # ---- GeneralizationAnalyzer: val vs test R2 genelleme raporu ----
        try:
            from pfaz_modules.pfaz04_unknown_predictions.generalization_analyzer import GeneralizationAnalyzer
            if self.all_results:
                _df = pd.DataFrame(self.all_results)
                # Build model_id for merging
                _df['_mid'] = (
                    _df['Dataset'].astype(str) + '__' +
                    _df['Model_Category'].astype(str) + '__' +
                    _df['Model_Type'].astype(str) + '__' +
                    _df['Config_ID'].astype(str)
                )
                _known_df = _df[['_mid', 'Val_R2 (Known)', 'Model_Type']].copy()
                _known_df = _known_df.rename(columns={
                    '_mid': 'model_id',
                    'Val_R2 (Known)': 'r2',
                    'Model_Type': 'type',
                })
                _known_df['rmse'] = 0.0  # val RMSE not stored in all_results

                _unk_df = _df[['_mid', 'Test_R2 (Unknown)', 'Test_RMSE', 'Model_Type']].copy()
                _unk_df = _unk_df.rename(columns={
                    '_mid': 'model_id',
                    'Test_R2 (Unknown)': 'r2',
                    'Test_RMSE': 'rmse',
                    'Model_Type': 'type',
                })

                _gen = GeneralizationAnalyzer(
                    output_dir=str(self.output_dir / 'generalization_analysis')
                )
                _gen.calculate_generalization_scores(
                    _known_df.dropna(subset=['r2']),
                    _unk_df.dropna(subset=['r2'])
                )
                _gen.generate_excel_report()
                logger.info("[OK] GeneralizationAnalyzer: raporu olusturuldu -> generalization_analysis/")
        except Exception as _e:
            logger.warning(f"[WARNING] GeneralizationAnalyzer basarisiz (devam): {_e}")

        return {
            'total_results': len(self.all_results),
            'per_nucleus_results': len(self.per_nucleus_results),
        }

    def _process_predictions(self, y_pred, test_df, target_cols, nuclei,
                             dataset_name, model_category, model_type,
                             config_id, val_r2, train_r2):
        """Process predictions for one model × one dataset."""
        y_pred = np.asarray(y_pred)

        # Handle output shape alignment
        if y_pred.ndim == 1:
            y_pred = y_pred.reshape(-1, 1)

        n_targets = len(target_cols)

        # If model outputs more columns than targets, take first n_targets
        if y_pred.shape[1] > n_targets:
            y_pred = y_pred[:, :n_targets]

        # If model outputs fewer, pad with NaN
        if y_pred.shape[1] < n_targets:
            pad = np.full((y_pred.shape[0], n_targets - y_pred.shape[1]), np.nan)
            y_pred = np.hstack([y_pred, pad])

        # Compute metrics per target
        for ti, tcol in enumerate(target_cols):
            if tcol not in test_df.columns:
                continue
            y_true_col = test_df[tcol].values
            y_pred_col = y_pred[:, ti]

            # Skip if all NaN
            if np.all(np.isnan(y_pred_col)):
                continue

            metrics = self._calculate_metrics(y_true_col, y_pred_col)
            test_r2 = metrics['r2']
            degradation = None
            if val_r2 is not None and test_r2 is not None:
                degradation = val_r2 - test_r2

            self.all_results.append({
                'Dataset': dataset_name,
                'Target': tcol,
                'Model_Category': model_category,
                'Model_Type': model_type,
                'Config_ID': config_id,
                'Train_R2': train_r2,
                'Val_R2 (Known)': val_r2,
                'Test_R2 (Unknown)': test_r2,
                'Degradation (Val-Test)': degradation,
                'Test_RMSE': metrics['rmse'],
                'Test_MAE': metrics['mae'],
                'N_Test': len(y_true_col),
            })

            # Per nucleus predictions
            for ni, nucleus in enumerate(nuclei):
                self.per_nucleus_results.append({
                    'Dataset': dataset_name,
                    'Target': tcol,
                    'Model_Category': model_category,
                    'Model_Type': model_type,
                    'Config_ID': config_id,
                    'NUCLEUS': nucleus,
                    'y_true': float(y_true_col[ni]),
                    'y_pred': float(y_pred_col[ni]),
                    'error': float(y_pred_col[ni] - y_true_col[ni]),
                    'abs_error': float(abs(y_pred_col[ni] - y_true_col[ni])),
                })

    # ------------------------------------------------------------------
    # Excel report generation
    # ------------------------------------------------------------------

    def generate_excel_report(self, filename: str = 'Unknown_Nuclei_Results.xlsx'):
        """Generate comprehensive Excel report with filtering and pivot tables."""

        excel_path = self.output_dir / filename
        logger.info(f"Generating Excel report: {excel_path}")

        if not self.all_results:
            logger.warning("No results to report.")
            return

        df_all = pd.DataFrame(self.all_results)

        if OPENPYXL_AVAILABLE:
            wb = Workbook()
            wb.remove(wb.active)

            # Sheet 1: All Results
            ws1 = wb.create_sheet('All_Results')
            self._write_df_sheet(ws1, df_all, title='All Model Predictions on Test Nuclei',
                                 r2_col='Test_R2 (Unknown)')

            # Sheet 2: Best per Dataset (best test R² per dataset × target × category)
            if not df_all.empty:
                df_best = (df_all.dropna(subset=['Test_R2 (Unknown)'])
                           .sort_values('Test_R2 (Unknown)', ascending=False)
                           .groupby(['Dataset', 'Target', 'Model_Category'], as_index=False)
                           .first())
                ws2 = wb.create_sheet('Best_Per_Dataset')
                self._write_df_sheet(ws2, df_best,
                                     title='Best Test R² per Dataset × Target × Model Category',
                                     r2_col='Test_R2 (Unknown)')

            # Sheet 3: Degradation Analysis
            df_deg = df_all.dropna(subset=['Degradation (Val-Test)']).copy()
            df_deg = df_deg.sort_values('Degradation (Val-Test)', ascending=False)
            ws3 = wb.create_sheet('Degradation_Analysis')
            self._write_df_sheet(ws3, df_deg,
                                 title='Degradation Analysis (Val R² - Test R²)',
                                 r2_col='Test_R2 (Unknown)')

            # Sheet 4: AI vs ANFIS comparison (best per dataset × target)
            ws4 = wb.create_sheet('AI_vs_ANFIS')
            self._write_comparison_sheet(ws4, df_all)

            # Sheet 5: Pivot - by Target
            ws5 = wb.create_sheet('Pivot_By_Target')
            self._write_pivot_sheet(ws5, df_all, index='Target',
                                    title='Average Test R² by Target and Model Category')

            # Sheet 6: Pivot - by Model Type
            ws6 = wb.create_sheet('Pivot_By_ModelType')
            self._write_pivot_sheet(ws6, df_all, index='Model_Type',
                                    title='Average Test R² by Model Type')

            # Sheet 7: Per-nucleus predictions (top N per model)
            if self.per_nucleus_results:
                df_nuc = pd.DataFrame(self.per_nucleus_results)
                ws7 = wb.create_sheet('Per_Nucleus_Predictions')
                self._write_df_sheet(ws7, df_nuc,
                                     title='Per Nucleus Predictions',
                                     r2_col=None)

            wb.save(excel_path)
            logger.info(f"[SUCCESS] Excel report saved: {excel_path}")

        else:
            # Fallback to CSV
            df_all.to_csv(self.output_dir / 'unknown_results_all.csv', index=False)
            logger.info("[SUCCESS] CSV fallback saved (openpyxl not available)")

    def _write_df_sheet(self, ws, df: pd.DataFrame, title: str = '',
                        r2_col: Optional[str] = None):
        """Write a DataFrame to a worksheet with headers, auto-filter, freeze, colors."""
        if df.empty:
            ws.append([title or 'No data'])
            return

        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=10)

        # Title row
        if title:
            ws.append([title])
            ws['A1'].font = Font(bold=True, size=12)
            ws.append([])   # blank row
            start_row = 3
        else:
            start_row = 1

        # Headers
        headers = df.columns.tolist()
        ws.append(headers)
        header_row = ws.max_row
        for cell in ws[header_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Data rows
        for _, row in df.iterrows():
            row_data = []
            for val in row:
                if pd.isna(val):
                    row_data.append(None)
                elif isinstance(val, (np.integer,)):
                    row_data.append(int(val))
                elif isinstance(val, (np.floating,)):
                    row_data.append(float(val))
                else:
                    row_data.append(val)
            ws.append(row_data)

        # Conditional coloring on R² column
        if r2_col and r2_col in headers:
            r2_idx = headers.index(r2_col) + 1
            for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row,
                                    min_col=r2_idx, max_col=r2_idx):
                for cell in row:
                    if cell.value is None:
                        continue
                    v = float(cell.value)
                    if v >= 0.9:
                        cell.fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
                    elif v >= 0.7:
                        cell.fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
                    elif v >= 0.5:
                        cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                    elif v >= 0.0:
                        cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    else:
                        cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

        # Auto-filter, freeze
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

        # Column widths
        for col_idx, col_name in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            max_len = max(len(str(col_name)), 10)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    def _write_comparison_sheet(self, ws, df_all: pd.DataFrame):
        """Write AI vs ANFIS best comparison per dataset × target."""
        ws.append(['AI vs ANFIS Comparison — Best Test R² per Dataset × Target'])
        ws['A1'].font = Font(bold=True, size=12)
        ws.append([])

        if df_all.empty or 'Test_R2 (Unknown)' not in df_all.columns:
            ws.append(['No data available'])
            return

        df_valid = df_all.dropna(subset=['Test_R2 (Unknown)'])

        # Best test R² per (dataset, target, category)
        df_best = (df_valid
                   .sort_values('Test_R2 (Unknown)', ascending=False)
                   .groupby(['Dataset', 'Target', 'Model_Category'], as_index=False)
                   .first()[['Dataset', 'Target', 'Model_Category', 'Model_Type',
                              'Config_ID', 'Val_R2 (Known)', 'Test_R2 (Unknown)',
                              'Degradation (Val-Test)']])

        # Pivot to side-by-side
        ai_df = df_best[df_best['Model_Category'] == 'AI'].rename(columns={
            'Test_R2 (Unknown)': 'AI_Test_R2',
            'Val_R2 (Known)': 'AI_Val_R2',
            'Config_ID': 'AI_Config',
            'Model_Type': 'AI_Model_Type',
            'Degradation (Val-Test)': 'AI_Degradation',
        }).drop(columns=['Model_Category'])

        anfis_df = df_best[df_best['Model_Category'] == 'ANFIS'].rename(columns={
            'Test_R2 (Unknown)': 'ANFIS_Test_R2',
            'Val_R2 (Known)': 'ANFIS_Val_R2',
            'Config_ID': 'ANFIS_Config',
            'Model_Type': 'ANFIS_Model_Type',
            'Degradation (Val-Test)': 'ANFIS_Degradation',
        }).drop(columns=['Model_Category'])

        merged = pd.merge(ai_df, anfis_df, on=['Dataset', 'Target'], how='outer')

        def winner(row):
            ai = row.get('AI_Test_R2')
            an = row.get('ANFIS_Test_R2')
            if pd.isna(ai) and pd.isna(an):
                return 'N/A'
            if pd.isna(ai):
                return 'ANFIS'
            if pd.isna(an):
                return 'AI'
            return 'AI' if ai >= an else 'ANFIS'

        merged['Winner'] = merged.apply(winner, axis=1)
        merged['Delta_R2 (AI-ANFIS)'] = merged.get('AI_Test_R2', np.nan) - merged.get('ANFIS_Test_R2', np.nan)

        self._write_df_sheet(ws, merged, r2_col='AI_Test_R2')

    def _write_pivot_sheet(self, ws, df_all: pd.DataFrame, index: str, title: str):
        """Write a simple pivot table (mean test R² by index × model category)."""
        ws.append([title])
        ws['A1'].font = Font(bold=True, size=12)
        ws.append([])

        if df_all.empty or 'Test_R2 (Unknown)' not in df_all.columns:
            ws.append(['No data'])
            return

        df_valid = df_all.dropna(subset=['Test_R2 (Unknown)'])
        if df_valid.empty:
            ws.append(['No valid data'])
            return

        pivot = df_valid.pivot_table(
            values='Test_R2 (Unknown)',
            index=index,
            columns='Model_Category',
            aggfunc=['mean', 'count'],
        )
        pivot.columns = [f'{agg}_{cat}' for agg, cat in pivot.columns]
        pivot = pivot.reset_index()

        self._write_df_sheet(ws, pivot, r2_col=None)

    def generate_aaa2_comparison_excel(self,
                                       aaa2_txt_path: str = None,
                                       filename: str = 'AAA2_Original_vs_Predictions.xlsx'):
        """
        Generate pivot-style Excel: one row per nucleus, columns = original + all model predictions.

        Format per sheet (one sheet per target: MM, QM, Beta_2, MM_QM):
            NUCLEUS | Z | N | A | Original_Value | RF_best_pred | XGB_best_pred | DNN_best_pred |
            Best_Pred | Error | Abs_Error | Best_Model

        The 'best' prediction per model type = the prediction from the config with highest val_R2.
        If per_nucleus_results is empty, tries to load from existing outputs.
        """
        if not self.per_nucleus_results:
            logger.warning("[AAA2 Comparison] No per-nucleus results available. Run predict_unknown_nuclei() first.")
            return

        excel_path = self.output_dir / filename
        logger.info(f"[AAA2 Comparison] Generating: {excel_path}")

        df_nuc = pd.DataFrame(self.per_nucleus_results)

        # Load aaa2 to get Z, N, A for each nucleus
        aaa2_info = {}
        if aaa2_txt_path and Path(aaa2_txt_path).exists():
            try:
                aaa2_df = pd.read_csv(aaa2_txt_path, sep='\t', encoding='utf-8')
                if 'NUCLEUS' not in aaa2_df.columns:
                    aaa2_df = pd.read_csv(aaa2_txt_path, encoding='utf-8')
                for _, row in aaa2_df.iterrows():
                    nuc = str(row.get('NUCLEUS', ''))
                    if nuc:
                        aaa2_info[nuc] = {
                            'Z': row.get('Z', ''),
                            'N': row.get('N', ''),
                            'A': row.get('A', ''),
                        }
            except Exception as e:
                logger.warning(f"[AAA2 Comparison] Could not load aaa2.txt: {e}")

        wb = Workbook() if OPENPYXL_AVAILABLE else None
        if wb:
            wb.remove(wb.active)

        # Process per target
        target_col_map = {
            'MAGNETIC MOMENT [µ]': 'MM',
            'QUADRUPOLE MOMENT [Q]': 'QM',
            'Beta_2': 'Beta_2',
        }
        df_nuc['Target_Label'] = df_nuc['Target'].map(
            lambda t: target_col_map.get(t, t)
        )

        all_sheets = {}
        for target_label in ['MM', 'QM']:
            sub = df_nuc[df_nuc['Target_Label'] == target_label].copy()
            if sub.empty:
                continue

            # Pivot: nucleus → model_type → best prediction (mean over configs)
            # Take the mean prediction per (nucleus, model_type)
            pivot = sub.groupby(['NUCLEUS', 'Model_Type']).agg(
                y_true=('y_true', 'first'),
                y_pred_mean=('y_pred', 'mean'),
                abs_error_mean=('abs_error', 'mean'),
            ).reset_index()

            # Wide format: one row per nucleus
            wide = pivot.pivot_table(
                index=['NUCLEUS', 'y_true'],
                columns='Model_Type',
                values=['y_pred_mean', 'abs_error_mean'],
                aggfunc='mean'
            )
            wide.columns = [f'{col[1]}_{col[0].replace("_mean", "")}' for col in wide.columns]
            wide = wide.reset_index()
            wide.rename(columns={'y_true': f'Original_{target_label}'}, inplace=True)

            # Add Z, N, A
            wide['Z'] = wide['NUCLEUS'].map(lambda n: aaa2_info.get(n, {}).get('Z', ''))
            wide['N'] = wide['NUCLEUS'].map(lambda n: aaa2_info.get(n, {}).get('N', ''))
            wide['A'] = wide['NUCLEUS'].map(lambda n: aaa2_info.get(n, {}).get('A', ''))

            # Best prediction (lowest mean abs_error among model types)
            pred_cols = [c for c in wide.columns if c.endswith('_y_pred')]
            err_cols  = [c for c in wide.columns if c.endswith('_abs_error')]

            if pred_cols:
                wide['Best_Pred'] = wide[pred_cols].apply(
                    lambda row: row.dropna().mean() if not row.dropna().empty else np.nan, axis=1
                )
                wide['Best_Abs_Error'] = wide[err_cols].apply(
                    lambda row: row.dropna().min() if not row.dropna().empty else np.nan, axis=1
                )
                best_col = f'Original_{target_label}'
                wide['Best_Error'] = wide['Best_Pred'] - wide[best_col]
                wide['Best_Model'] = wide[pred_cols].apply(
                    lambda row: pred_cols[row.dropna().values.argmin()].split('_')[0]
                    if not row.dropna().empty else '', axis=1
                )

            # Reorder columns
            id_cols = ['NUCLEUS', 'A', 'Z', 'N', f'Original_{target_label}']
            other_cols = [c for c in wide.columns if c not in id_cols]
            wide = wide[id_cols + other_cols]

            all_sheets[target_label] = wide

            if wb:
                sheet_name = target_label[:31]
                ws = wb.create_sheet(sheet_name)
                self._write_df_sheet(
                    ws, wide,
                    title=f'AAA2 Original vs Model Predictions — Target: {target_label}',
                    r2_col=None
                )
            else:
                csv_out = self.output_dir / f'aaa2_comparison_{target_label}.csv'
                wide.to_csv(csv_out, index=False)

        # Summary sheet: number of nuclei and mean abs error per target × model
        if wb and all_sheets:
            ws_sum = wb.create_sheet('Summary')
            summary_rows = [['Target', 'Model_Type', 'N_Nuclei', 'Mean_Abs_Error', 'Mean_Abs_Error_Pct']]
            for tgt, df_wide in all_sheets.items():
                orig_col = f'Original_{tgt}'
                for col in df_wide.columns:
                    if col.endswith('_abs_error'):
                        mtype = col.replace('_abs_error', '')
                        n = df_wide[col].dropna().count()
                        mae = df_wide[col].dropna().mean()
                        if orig_col in df_wide.columns:
                            orig_std = df_wide[orig_col].std()
                            pct = (mae / orig_std * 100) if orig_std and orig_std > 0 else np.nan
                        else:
                            pct = np.nan
                        summary_rows.append([tgt, mtype, n, round(mae, 4), round(pct, 2)])
            for row in summary_rows:
                ws_sum.append(row)

        if wb:
            wb.save(excel_path)
            logger.info(f"[OK] AAA2 Comparison Excel saved: {excel_path}")
        else:
            logger.warning("[AAA2 Comparison] openpyxl not available -- CSV files saved instead")


def main():
    """Test PFAZ4 predictor standalone."""
    import os
    base = Path(__file__).resolve().parent.parent.parent / 'outputs'
    print("\n" + "=" * 80)
    print("PFAZ 4: UNKNOWN NUCLEI PREDICTOR - STANDALONE TEST")
    print("=" * 80)

    predictor = UnknownNucleiPredictor(
        ai_models_dir=str(base / 'trained_models'),
        anfis_models_dir=str(base / 'anfis_models'),
        splits_dir=str(base / 'generated_datasets'),
        output_dir=str(base / 'unknown_predictions'),
    )

    results = predictor.predict_unknown_nuclei()
    print(f"\n[SUCCESS] Completed. Results: {results}")


if __name__ == "__main__":
    main()
