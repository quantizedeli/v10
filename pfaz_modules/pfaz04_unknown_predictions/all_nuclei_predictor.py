# -*- coding: utf-8 -*-
"""
PFAZ 2: All Nuclei Predictor
=============================

Her trained model ile ALL nuclei dataset'ler üzerinde tahmin yapıp
comprehensive Excel raporu oluşturur

Features:
- Load ALL datasets (MM, QM, MM_QM, Beta_2)
- Load all trained models
- Predict on ALL nuclei
- Calculate Deltas (Prediction - Experimental)
- Generate multi-sheet Excel report
- Best model identification per nucleus

Excel Format:
- Sheets: MM_ALL_267nuclei, QM_ALL_219nuclei, MM_QM_ALL_219nuclei, Beta_2_ALL_267nuclei
- Columns: NUCLEUS | A | Z | N | Experimental_Value | Model1_Pred | Model1_Delta | ...

Author: Nuclear Physics AI Training Pipeline
Version: 1.0.0
Date: 2025-10-15
"""

import numpy as np
import pandas as pd
from pathlib import Path
import json
import logging
from typing import Dict, List, Tuple, Optional
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logging.warning("openpyxl not available - Excel formatting will be limited")

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


# ============================================================================
# MODEL LOADER
# ============================================================================

class ModelLoader:
    """Load trained models from directory"""
    
    def __init__(self, models_dir: Path):
        self.models_dir = Path(models_dir)
        self.loaded_models = {}
    
    def load_all_models(self, dataset_name: str) -> Dict:
        """
        Load all trained models for a specific dataset
        
        Args:
            dataset_name: Dataset name (e.g., 'MM_ALL_267nuclei')
        
        Returns:
            Dictionary of {model_id: model_object}
        """
        import joblib
        
        dataset_dir = self.models_dir / dataset_name
        
        if not dataset_dir.exists():
            logger.warning(f"Model directory not found: {dataset_dir}")
            return {}
        
        models = {}
        
        # Find all model files (.pkl)
        model_files = list(dataset_dir.glob('**/*.pkl'))
        
        logger.info(f"Loading models from {dataset_dir}")
        logger.info(f"Found {len(model_files)} model files")
        
        for model_file in model_files:
            try:
                # Extract model identifier from path
                # Example: MM_ALL_267nuclei/RF/TRAIN_001/model_RF_TRAIN_001.pkl
                model_type = model_file.parent.parent.name
                config_id = model_file.parent.name
                model_id = f"{model_type}_{config_id}"
                
                # Load model
                model = joblib.load(model_file)
                models[model_id] = {
                    'model': model,
                    'model_type': model_type,
                    'config_id': config_id,
                    'file_path': model_file
                }
                
                logger.debug(f"[SUCCESS] Loaded: {model_id}")
            
            except Exception as e:
                logger.error(f"[ERROR] Failed to load {model_file}: {e}")
        
        logger.info(f"[SUCCESS] Successfully loaded {len(models)} models for {dataset_name}")
        
        return models


# ============================================================================
# DATASET LOADER
# ============================================================================

class AllNucleiDatasetLoader:
    """Load ALL nuclei datasets"""
    
    def __init__(self, datasets_dir: Path):
        self.datasets_dir = Path(datasets_dir)
    
    def load_all_datasets(self) -> Dict[str, pd.DataFrame]:
        """
        Load all ALL nuclei datasets
        
        Returns:
            Dictionary of {dataset_name: dataframe}
        """
        datasets = {}
        
        # Target patterns to look for
        target_patterns = [
            'MM_ALL_*nuclei',
            'QM_ALL_*nuclei',
            'MM_QM_ALL_*nuclei',
            'Beta_2_ALL_*nuclei'
        ]
        
        for pattern in target_patterns:
            matching_dirs = list(self.datasets_dir.glob(pattern))
            
            for dataset_dir in matching_dirs:
                dataset_name = dataset_dir.name
                
                # Find data file
                csv_files = list(dataset_dir.glob('*.csv'))
                
                if not csv_files:
                    logger.warning(f"No CSV file found in {dataset_dir}")
                    continue
                
                data_file = csv_files[0]
                
                try:
                    df = pd.read_csv(data_file)
                    datasets[dataset_name] = df
                    logger.info(f"[SUCCESS] Loaded: {dataset_name} ({len(df)} nuclei)")
                
                except Exception as e:
                    logger.error(f"[ERROR] Failed to load {dataset_name}: {e}")
        
        if not datasets:
            logger.warning("No ALL nuclei datasets found!")
        else:
            logger.info(f"[SUCCESS] Total datasets loaded: {len(datasets)}")
        
        return datasets


# ============================================================================
# PREDICTOR
# ============================================================================

class AllNucleiPredictor:
    """
    Predict on ALL nuclei with all trained models
    
    Workflow:
    1. Load ALL datasets
    2. Load all trained models per dataset
    3. Make predictions
    4. Calculate metrics and deltas
    5. Identify best models
    """
    
    def __init__(self,
                 datasets_dir: str,
                 models_dir: str,
                 output_dir: str = 'all_nuclei_predictions'):
        """
        Initialize All Nuclei Predictor
        
        Args:
            datasets_dir: Directory containing ALL nuclei datasets
            models_dir: Directory containing trained models
            output_dir: Output directory for predictions
        """
        self.datasets_dir = Path(datasets_dir)
        self.models_dir = Path(models_dir)
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        self.dataset_loader = AllNucleiDatasetLoader(self.datasets_dir)
        self.model_loader = ModelLoader(self.models_dir)
        
        self.predictions = {}  # {dataset_name: predictions_df}
        
        logger.info("=" * 80)
        logger.info("ALL NUCLEI PREDICTOR INITIALIZED")
        logger.info("=" * 80)
        logger.info(f"Datasets directory: {self.datasets_dir}")
        logger.info(f"Models directory: {self.models_dir}")
        logger.info(f"Output directory: {self.output_dir}")
        logger.info("=" * 80)
    
    def identify_target_columns(self, dataset_name: str, df: pd.DataFrame) -> List[str]:
        """Identify target columns from dataset name and dataframe"""
        
        if 'MM_QM' in dataset_name:
            return ['MM', 'Q']
        elif 'Beta_2' in dataset_name:
            return ['Beta_2']
        elif 'MM' in dataset_name:
            return ['MM']
        elif 'QM' in dataset_name:
            return ['Q']
        else:
            # Try to infer from columns
            target_cols = []
            for col in ['MM', 'Q', 'Beta_2']:
                if col in df.columns:
                    target_cols.append(col)
            return target_cols
    
    def prepare_features(self, df: pd.DataFrame, target_cols: List[str]) -> np.ndarray:
        """Prepare feature matrix"""
        
        # Exclude target columns and NUCLEUS
        feature_cols = [col for col in df.columns 
                       if col not in target_cols and col != 'NUCLEUS']
        
        X = df[feature_cols].values
        return X
    
    def predict_all_nuclei(self):
        """
        Main prediction workflow
        
        1. Load datasets
        2. For each dataset:
           - Load models
           - Make predictions
           - Calculate deltas
           - Find best models
        """
        logger.info("\n" + "=" * 80)
        logger.info("STARTING ALL NUCLEI PREDICTIONS")
        logger.info("=" * 80)
        
        # Load datasets
        datasets = self.dataset_loader.load_all_datasets()
        
        if not datasets:
            logger.error("No datasets loaded! Aborting.")
            return
        
        # Process each dataset
        for dataset_name, df in datasets.items():
            logger.info(f"\n{'='*80}")
            logger.info(f"PROCESSING: {dataset_name}")
            logger.info(f"{'='*80}")
            
            # Identify targets
            target_cols = self.identify_target_columns(dataset_name, df)
            logger.info(f"Target columns: {target_cols}")
            
            # Prepare features
            X = self.prepare_features(df, target_cols)
            logger.info(f"Features shape: {X.shape}")
            
            # Get experimental values
            y_true = df[target_cols].values
            if y_true.ndim == 1:
                y_true = y_true.reshape(-1, 1)
            
            # Load models for this dataset
            models = self.model_loader.load_all_models(dataset_name)
            
            if not models:
                logger.warning(f"No models found for {dataset_name}, skipping")
                continue
            
            logger.info(f"Loaded {len(models)} models")
            
            # Create results dataframe
            results_df = df[['NUCLEUS', 'A', 'Z', 'N']].copy()
            
            # Add experimental values
            for i, target_col in enumerate(target_cols):
                results_df[f'Experimental_{target_col}'] = y_true[:, i]
            
            # Make predictions with each model
            model_errors = {}  # Track errors per model
            
            for model_id, model_info in models.items():
                try:
                    model = model_info['model']
                    
                    # Predict
                    y_pred = model.predict(X)
                    
                    # Handle dimensions
                    if y_pred.ndim == 1:
                        y_pred = y_pred.reshape(-1, 1)
                    
                    # Add predictions and deltas for each target
                    for i, target_col in enumerate(target_cols):
                        pred_col = f'{model_id}_Pred_{target_col}'
                        delta_col = f'{model_id}_Delta_{target_col}'
                        error_col = f'{model_id}_Error_{target_col}'
                        
                        results_df[pred_col] = y_pred[:, i]
                        results_df[delta_col] = y_pred[:, i] - y_true[:, i]
                        results_df[error_col] = np.abs(y_pred[:, i] - y_true[:, i])
                    
                    # Calculate overall error for this model
                    overall_error = np.mean([
                        np.abs(y_pred[:, i] - y_true[:, i]).mean()
                        for i in range(len(target_cols))
                    ])
                    model_errors[model_id] = overall_error
                    
                    logger.info(f"  [SUCCESS] {model_id}: MAE = {overall_error:.4f}")
                
                except Exception as e:
                    logger.error(f"  [ERROR] {model_id}: Prediction failed - {e}")
            
            # Find best model for each nucleus
            if model_errors:
                # For each nucleus, find model with minimum error
                best_models = []
                best_preds = {target: [] for target in target_cols}
                best_deltas = {target: [] for target in target_cols}
                
                for idx in range(len(results_df)):
                    # Find model with minimum error for this nucleus
                    nucleus_errors = {}
                    
                    for model_id in model_errors.keys():
                        # Average error across all targets for this nucleus
                        avg_error = np.mean([
                            results_df.loc[idx, f'{model_id}_Error_{target}']
                            for target in target_cols
                        ])
                        nucleus_errors[model_id] = avg_error
                    
                    # Best model for this nucleus
                    best_model_id = min(nucleus_errors, key=nucleus_errors.get)
                    best_models.append(best_model_id)
                    
                    # Best predictions and deltas
                    for target in target_cols:
                        best_preds[target].append(
                            results_df.loc[idx, f'{best_model_id}_Pred_{target}']
                        )
                        best_deltas[target].append(
                            results_df.loc[idx, f'{best_model_id}_Delta_{target}']
                        )
                
                # Add best model columns
                results_df['Best_Model'] = best_models
                
                for target in target_cols:
                    results_df[f'Best_Pred_{target}'] = best_preds[target]
                    results_df[f'Best_Delta_{target}'] = best_deltas[target]
            
            # Store results
            self.predictions[dataset_name] = results_df
            
            logger.info(f"[SUCCESS] Predictions completed for {dataset_name}")
            logger.info(f"   Results shape: {results_df.shape}")
        
        logger.info("\n" + "=" * 80)
        logger.info("ALL PREDICTIONS COMPLETED")
        logger.info("=" * 80)
    
    def generate_excel_report(self, filename: str = 'All_Nuclei_Predictions.xlsx'):
        """
        Generate comprehensive Excel report
        
        Format:
        - One sheet per dataset
        - Formatted columns
        - Color coding for best models
        """
        if not self.predictions:
            logger.error("No predictions to export!")
            return
        
        excel_path = self.output_dir / filename
        
        logger.info(f"\n{'='*80}")
        logger.info(f"GENERATING EXCEL REPORT: {excel_path}")
        logger.info(f"{'='*80}")
        
        if OPENPYXL_AVAILABLE:
            self._generate_formatted_excel(excel_path)
        else:
            # Fallback to simple Excel
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                for dataset_name, df in self.predictions.items():
                    sheet_name = dataset_name[:31]  # Excel sheet name limit
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            logger.info(f"[SUCCESS] Excel report saved: {excel_path}")
    
    def _generate_formatted_excel(self, excel_path: Path):
        """Generate formatted Excel with styling"""
        
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        for dataset_name, df in self.predictions.items():
            sheet_name = dataset_name[:31]
            ws = wb.create_sheet(title=sheet_name)
            
            logger.info(f"  Creating sheet: {sheet_name}")
            
            # Write header
            headers = df.columns.tolist()
            ws.append(headers)
            
            # Style header
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Write data
            for _, row in df.iterrows():
                ws.append(row.tolist())
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze panes (freeze first row and first 4 columns)
            ws.freeze_panes = 'E2'
        
        # Save
        wb.save(excel_path)
        logger.info(f"[SUCCESS] Formatted Excel report saved: {excel_path}")
    
    def save_csv_predictions(self):
        """Save predictions as CSV files (one per dataset)"""
        
        for dataset_name, df in self.predictions.items():
            csv_path = self.output_dir / f"{dataset_name}_predictions.csv"
            df.to_csv(csv_path, index=False)
            logger.info(f"[SUCCESS] CSV saved: {csv_path}")
    
    def generate_summary_statistics(self) -> pd.DataFrame:
        """Generate summary statistics across all datasets"""
        
        summary_data = []
        
        for dataset_name, df in self.predictions.items():
            # Find experimental and prediction columns
            exp_cols = [col for col in df.columns if col.startswith('Experimental_')]
            best_pred_cols = [col for col in df.columns if col.startswith('Best_Pred_')]
            best_delta_cols = [col for col in df.columns if col.startswith('Best_Delta_')]
            
            # Calculate statistics
            for exp_col, pred_col, delta_col in zip(exp_cols, best_pred_cols, best_delta_cols):
                target = exp_col.replace('Experimental_', '')
                
                exp_values = df[exp_col].values
                pred_values = df[pred_col].values
                delta_values = df[delta_col].values
                
                summary_data.append({
                    'Dataset': dataset_name,
                    'Target': target,
                    'N_Nuclei': len(df),
                    'MAE': np.abs(delta_values).mean(),
                    'RMSE': np.sqrt((delta_values**2).mean()),
                    'Mean_Delta': delta_values.mean(),
                    'Std_Delta': delta_values.std(),
                    'Max_Error': np.abs(delta_values).max()
                })
        
        summary_df = pd.DataFrame(summary_data)
        
        # Save
        summary_path = self.output_dir / 'prediction_summary_statistics.csv'
        summary_df.to_csv(summary_path, index=False)
        logger.info(f"[SUCCESS] Summary statistics saved: {summary_path}")
        
        return summary_df


# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    """Main execution for testing"""
    
    print("\n" + "=" * 80)
    print("PFAZ 2: ALL NUCLEI PREDICTOR - TEST")
    print("=" * 80)
    
    # Paths (adjust as needed)
    DATASETS_DIR = "generated_datasets"
    MODELS_DIR = "trained_models"
    OUTPUT_DIR = "all_nuclei_predictions"
    
    # Initialize predictor
    predictor = AllNucleiPredictor(
        datasets_dir=DATASETS_DIR,
        models_dir=MODELS_DIR,
        output_dir=OUTPUT_DIR
    )
    
    # Run predictions
    print("\nRunning predictions on ALL nuclei...")
    predictor.predict_all_nuclei()
    
    # Generate Excel report
    print("\nGenerating Excel report...")
    predictor.generate_excel_report()
    
    # Save CSV files
    print("\nSaving CSV files...")
    predictor.save_csv_predictions()
    
    # Generate summary statistics
    print("\nGenerating summary statistics...")
    summary = predictor.generate_summary_statistics()
    print("\nSummary Statistics:")
    print(summary.to_string(index=False))
    
    print("\n[SUCCESS] ALL NUCLEI PREDICTION COMPLETED!")


if __name__ == "__main__":
    main()
