"""
Single Nucleus Predictor
========================
Kullanicinin verdigi tek cekirdek veya liste icin
tum egitilmis AI ve ANFIS modellerinden tahmin alir.

========================================================
GIRIS FORMATI
========================================================

1. DICT (tek cekirdek):
   {'Z': 26, 'N': 30, 'SPIN': 0.0, 'PARITY': 1}
   A otomatik hesaplanir (A = Z + N).
   Ek ozellikler (magic_character, BE_per_A, vb.) varsa
   kullanilir; yoksa TheoreticalCalculationsManager ile
   otomatik turetilir.

2. PRED_INPUT.TXT (onerilen format):
   # Her satir bir cekirdek
   # Sutunlar: Z  N  SPIN  PARITY
   # Z, N: tam sayi; SPIN: yarim tam sayi (0, 0.5, 1, 1.5, ...)
   # PARITY: +1 veya -1
   # Ornek: Fe-56 (Z=26, N=30, I=0, pi=+1)
   26  30  0.0  1

3. AAA2.TXT formatli dosya (ayni format):
   Satir yapisi: Z  N  SPIN  PARITY  [MM]  [QM]  [...]
   Ilk 4 sutun ozellik olarak alinir; geri kalanlari yok sayilir.

4. CSV dosyasi (Z, N, SPIN, PARITY sutunlu):
   Z,N,SPIN,PARITY
   26,30,0.0,1

5. EXCEL (.xlsx) - ayni sutun yapisi ile

Not: Giris olarak yalnizca Z ve N yeterlidir.
     Diger ozellikler TheoreticalCalculationsManager ile
     otomatik hesaplanir.

========================================================
CIKTI
========================================================
Her hedef (MM, QM, Beta_2, MM_QM) icin:
  - Hedef bazinda en iyi 25 modelin tahminleri
  - Consensus (ortalama) degeri
Excel raporu + JSON ozeti + Grafik PNG'leri

========================================================
"""

from __future__ import annotations

import json
import logging
from pathlib import Path
from typing import Any, Dict, List, Optional, Union
import warnings
warnings.filterwarnings('ignore')

import numpy as np
import pandas as pd

logger = logging.getLogger(__name__)

ALL_TARGETS = ['MM', 'QM']

# Hedef kolonu gercek adlarina map
TARGET_COL_MAP = {
    'MM':     'MAGNETIC MOMENT [u]',
    'QM':     'QUADRUPOLE MOMENT [Q]',
    'Beta_2': 'Beta_2',
    'MM_QM':  ['MAGNETIC MOMENT [u]', 'QUADRUPOLE MOMENT [Q]'],
}

# Minimum giris sutunlari — bunlar yoksa enrichment yapilamaz
REQUIRED_COLUMNS = ['Z', 'N']


class SingleNucleusPredictor:
    """
    Egitilmis tum modellerden tek cekirdek / kucuk liste tahmini yapar.

    Parameters
    ----------
    ai_models_dir    : PFAZ2 ciktisi (trained_models/ veya benzer)
    anfis_models_dir : PFAZ3 ciktisi (anfis_models/)
    splits_dir       : PFAZ1 ciktisi (generated_datasets/) -- feature_cols icin
    output_dir       : Tahmin ciktilarinin kaydedilecegi yer
    targets          : Tahmin edilecek hedefler (None = hepsi)
    top_n_models     : Her hedef icin en iyi N model kullanilsin (0 = hepsi)
    """

    def __init__(
        self,
        ai_models_dir: str,
        anfis_models_dir: str,
        splits_dir: str,
        output_dir: str = 'single_nucleus_predictions',
        targets: Optional[List[str]] = None,
        top_n_models: int = 25,
    ):
        self.ai_dir     = Path(ai_models_dir)
        self.anfis_dir  = Path(anfis_models_dir)
        self.ds_dir     = Path(splits_dir)
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

        self.targets      = targets or ALL_TARGETS
        self.top_n_models = top_n_models

        # Cache: {(target, model_id): model_obj}
        self._model_cache: Dict = {}
        # Feature columns per target: {target: [col1, col2, ...]}
        self._feature_cols: Dict[str, List[str]] = {}

    # ------------------------------------------------------------------
    # PUBLIC API
    # ------------------------------------------------------------------

    def predict_from_dict(
        self,
        nucleus_data: Dict[str, Any],
        nucleus_label: str = 'input_nucleus',
    ) -> Dict:
        """
        Tek cekirdek sozlugunden tahmin.

        Ornek:
            {'Z': 26, 'N': 30, 'SPIN': 0.0, 'PARITY': 1}
        """
        df = pd.DataFrame([nucleus_data])
        return self.predict_from_dataframe(df, nucleus_labels=[nucleus_label])

    def predict_from_file(self, file_path: str) -> Dict:
        """
        Dosyadan tahmin:
          - .xlsx / .xls  : Excel
          - .csv          : CSV
          - diger         : pred_input.txt / aaa2.txt tipi (bosluk ayrimli)
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Dosya bulunamadi: {path}")

        if path.suffix.lower() in ('.xlsx', '.xls'):
            df = pd.read_excel(path)
        elif path.suffix.lower() == '.csv':
            df = pd.read_csv(path)
        else:
            df = _parse_nucleus_txt(path)

        labels = None
        if 'NUCLEUS' in df.columns:
            labels = df['NUCLEUS'].astype(str).tolist()
        elif 'Z' in df.columns and 'N' in df.columns:
            labels = [f"Z{int(r.get('Z', 0))}-N{int(r.get('N', 0))}"
                      for _, r in df.iterrows()]

        return self.predict_from_dataframe(df, nucleus_labels=labels)

    def predict_from_dataframe(
        self,
        df: pd.DataFrame,
        nucleus_labels: Optional[List[str]] = None,
    ) -> Dict:
        """
        DataFrame'den tahmin — her satir bir cekirdek.

        Returns:
            {
              'predictions': [
                  {'nucleus': ...,
                   'MM': {'model_A': v1, 'consensus': v_avg, 'n_models': N},
                   ...},
                  ...
              ],
              'summary': {target: {'mean':..., 'std':..., 'n_nuclei':...}},
              'excel_path': ...,
              'json_path':  ...,
              'plot_paths': [...],
            }
        """
        n_nuclei = len(df)
        if nucleus_labels is None:
            nucleus_labels = [f"nucleus_{i+1}" for i in range(n_nuclei)]

        logger.info(f"[SinglePredictor] {n_nuclei} cekirdek icin tahmin basliyor...")
        logger.info(f"  Hedefler: {self.targets}")
        logger.info(f"  Top-N models: {self.top_n_models}")

        # Feature zenginlestirme
        df = self._enrich_features(df)

        all_predictions = [{'nucleus': nucleus_labels[i]} for i in range(n_nuclei)]

        for target in self.targets:
            logger.info(f"\n  [Target: {target}]")
            try:
                target_preds = self._predict_target(df, target)

                # MM_QM: model_id+'_MM' / model_id+'_QM' anahtarlari gelir
                # Bunlari iki ayri alt-hedef olarak depola
                is_mmqm = (target == 'MM_QM')
                if is_mmqm:
                    sub_keys = ('MM_QM_MM', 'MM_QM_QM')
                    sub_preds: Dict[str, Dict] = {sk: {} for sk in sub_keys}
                    for model_key, arr in target_preds.items():
                        if model_key.endswith('_MM'):
                            sub_preds['MM_QM_MM'][model_key] = arr
                        elif model_key.endswith('_QM'):
                            sub_preds['MM_QM_QM'][model_key] = arr

                    for sub_key, sub_dict in sub_preds.items():
                        for i in range(n_nuclei):
                            vals = {mid: float(arr[i])
                                    for mid, arr in sub_dict.items()
                                    if i < len(arr) and not np.isnan(arr[i])}
                            vs = list(vals.values())
                            vals['consensus'] = float(np.mean(vs)) if vs else float('nan')
                            vals['n_models']  = len(vs)
                            all_predictions[i][sub_key] = vals
                    # Toplam MM_QM consensus (ortalama MM ve QM)
                    for i in range(n_nuclei):
                        mm_c = all_predictions[i].get('MM_QM_MM', {}).get('consensus', float('nan'))
                        qm_c = all_predictions[i].get('MM_QM_QM', {}).get('consensus', float('nan'))
                        all_predictions[i]['MM_QM'] = {
                            'MM_consensus': mm_c,
                            'QM_consensus': qm_c,
                            'consensus':    float(np.nanmean([mm_c, qm_c])),
                            'n_models':     all_predictions[i].get('MM_QM_MM', {}).get('n_models', 0),
                        }
                else:
                    for i in range(n_nuclei):
                        preds_for_nuc = {
                            model_id: float(preds[i])
                            for model_id, preds in target_preds.items()
                            if i < len(preds)
                        }
                        vals = [v for v in preds_for_nuc.values() if not np.isnan(v)]
                        preds_for_nuc['consensus'] = float(np.mean(vals)) if vals else float('nan')
                        preds_for_nuc['n_models']  = len(vals)
                        all_predictions[i][target] = preds_for_nuc
            except Exception as e:
                logger.warning(f"  [WARN] {target} tahmin hatasi: {e}")
                for i in range(n_nuclei):
                    all_predictions[i][target] = {'error': str(e)}

        # Summary
        summary = {}
        for target in self.targets:
            cvals = []
            for pred in all_predictions:
                c = pred.get(target, {}).get('consensus', float('nan'))
                if isinstance(c, float) and not np.isnan(c):
                    cvals.append(c)
            summary[target] = {
                'mean':     float(np.mean(cvals)) if cvals else None,
                'std':      float(np.std(cvals))  if len(cvals) > 1 else None,
                'n_nuclei': len(cvals),
            }

        # Kaydet
        excel_path = self._save_excel(all_predictions, nucleus_labels)
        json_path  = self._save_json(all_predictions, summary)
        plot_paths = self._generate_visualizations(all_predictions, nucleus_labels, summary)

        logger.info(f"\n[SinglePredictor] Tahmin tamamlandi.")
        logger.info(f"  Excel : {excel_path}")
        logger.info(f"  JSON  : {json_path}")
        logger.info(f"  Grafik: {len(plot_paths)} dosya")

        return {
            'predictions': all_predictions,
            'summary':     summary,
            'excel_path':  str(excel_path) if excel_path else None,
            'json_path':   str(json_path)  if json_path  else None,
            'plot_paths':  plot_paths,
        }

    # ------------------------------------------------------------------
    # Feature Enrichment
    # ------------------------------------------------------------------

    def _enrich_features(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Giris DataFrame'ini teorik hesaplama ozelliklerle zenginlestir.
        TheoreticalCalculationsManager kullanarak magic_character, BE_per_A,
        Z_magic_dist, vb. tureter.
        Basarisiz olursa orijinal DataFrame dondurulur (0-fill ile devam edilir).
        """
        # A = Z + N hesapla (yoksa)
        df = df.copy()
        if 'A' not in df.columns and 'Z' in df.columns and 'N' in df.columns:
            df['A'] = df['Z'].astype(int) + df['N'].astype(int)

        # Temel sayisal dogrulama
        for col in REQUIRED_COLUMNS:
            if col not in df.columns:
                logger.warning(f"[ENRICH] Gerekli sutun eksik: {col}. Zenginlestirme atlanacak.")
                return df

        try:
            import sys
            project_root = Path(__file__).resolve().parents[3]
            if str(project_root) not in sys.path:
                sys.path.insert(0, str(project_root))

            from physics_modules.theoretical_calculations_manager import (
                TheoreticalCalculationsManager
            )

            logger.info("[ENRICH] TheoreticalCalculationsManager ile ozellikler hesaplaniyor...")
            mgr = TheoreticalCalculationsManager(enable_all=True)

            enriched = mgr.calculate_all_theoretical_properties(df, save_report=False)
            n_new = len(enriched.columns) - len(df.columns)
            logger.info(f"[ENRICH] {n_new} yeni ozellik eklendi. Toplam: {len(enriched.columns)}")
            return enriched

        except Exception as e:
            logger.warning(
                f"[ENRICH] Teorik zenginlestirme basarisiz ({e}). "
                f"Eksik ozellikler 0 ile doldurulacak."
            )
            return df

    # ------------------------------------------------------------------
    # Predict one target
    # ------------------------------------------------------------------

    def _predict_target(self, df: pd.DataFrame, target: str) -> Dict[str, np.ndarray]:
        """
        Her model kendi feature listesiyle X olusturur.
        MM_QM modelleri 2-sutun cikti verir: {model_id+'_MM': arr, model_id+'_QM': arr}
        """
        is_mmqm = (target == 'MM_QM')
        records = self._collect_model_records(target)
        if not records:
            logger.warning(f"  [WARN] {target}: hicbir model bulunamadi")
            return {}

        results: Dict[str, np.ndarray] = {}

        for rec in records:
            model_id  = rec['model_id']
            model_obj = rec['model']
            feat_cols = rec['feature_names']

            # Her model icin kendi feature setinden X olustur
            X = self._build_X(df, feat_cols, model_id)
            if X is None:
                continue

            try:
                raw = model_obj.predict(X)
                arr = np.asarray(raw)

                if is_mmqm:
                    # MM_QM: 2D cikti (n, 2) veya 1D (n,) — her iki durumu isle
                    if arr.ndim == 2 and arr.shape[1] >= 2:
                        results[model_id + '_MM'] = arr[:, 0]
                        results[model_id + '_QM'] = arr[:, 1]
                    elif arr.ndim == 1:
                        # Tek cikti varsa MM olarak kaydet
                        results[model_id + '_MM'] = arr
                else:
                    results[model_id] = arr.ravel()

                logger.debug(f"    [{rec['source']}] {model_id}: OK")
            except Exception as e:
                logger.debug(f"    [SKIP] {model_id}: {e}")

        logger.info(f"  => {target}: {len(results)} tahmin serisi elde edildi")
        return results

    def _build_X(self, df: pd.DataFrame, feat_cols: List[str],
                 model_id: str) -> Optional[np.ndarray]:
        """Zenginlestirilmis df'den model-ozel feature matrisi olustur."""
        X = pd.DataFrame(index=df.index)
        missing = []
        for col in feat_cols:
            if col in df.columns:
                X[col] = pd.to_numeric(df[col], errors='coerce').fillna(0.0)
            else:
                X[col] = 0.0
                missing.append(col)
        if missing:
            logger.debug(f"    [FILL-0] {model_id}: {len(missing)} eksik "
                         f"ozellik 0 ile dolduruldu: {missing[:3]}")
        return X[feat_cols].values.astype(np.float32)

    # ------------------------------------------------------------------
    # Model + feature-list collection
    # ------------------------------------------------------------------

    def _collect_model_records(self, target: str) -> List[Dict]:
        """
        AI ve ANFIS modellerini Val_R2'ye gore siralar.
        Her kayit: {model_id, model, feature_names, source, val_r2}
        Her model kendi metadata.json'undaki feature_names'i tasir.
        """
        import joblib

        records: List[Dict] = []

        # Metadata cache: dataset_name -> feature_names
        meta_cache: Dict[str, List[str]] = {}

        def _load_meta(dataset_dir_name: str) -> List[str]:
            if dataset_dir_name in meta_cache:
                return meta_cache[dataset_dir_name]
            # generated_datasets/ altinda metadata.json
            meta_path = self.ds_dir / dataset_dir_name / 'metadata.json'
            if not meta_path.exists():
                meta_cache[dataset_dir_name] = []
                return []
            try:
                with open(meta_path, encoding='utf-8') as f:
                    m = json.load(f)
                feats = m.get('feature_names', [])
                meta_cache[dataset_dir_name] = feats
                return feats
            except Exception:
                meta_cache[dataset_dir_name] = []
                return []

        # ---- AI modeller ------------------------------------------------
        if self.ai_dir.exists():
            for mf in self.ai_dir.rglob('metrics_*.json'):
                try:
                    with open(mf, encoding='utf-8') as f:
                        m = json.load(f)
                    mtarget = m.get('target') or m.get('Target', '')
                    if mtarget != target:
                        continue
                    val_r2 = float(m.get('val', {}).get('r2')
                                   or m.get('val_r2')
                                   or m.get('Val_R2', -999.0) or -999.0)
                    if val_r2 < -5.0:
                        continue
                    # Dataset dizini: metrics_*.json -> config/ -> model_type/ -> dataset/
                    dataset_dir_name = mf.parents[2].name
                    feat_cols = _load_meta(dataset_dir_name)
                    if not feat_cols:
                        continue
                    for ext in ('*.pkl', '*.joblib', '*.h5', '*.keras'):
                        for mp in mf.parent.glob(ext):
                            records.append({
                                'val_r2':       val_r2,
                                'model_path':   mp,
                                'feature_names': feat_cols,
                                'source':       'AI',
                                'dataset':      dataset_dir_name,
                            })
                except Exception:
                    continue

        # ---- ANFIS modeller ---------------------------------------------
        if self.anfis_dir.exists():
            for mf in self.anfis_dir.rglob('metrics_*.json'):
                try:
                    with open(mf, encoding='utf-8') as f:
                        m = json.load(f)
                    mtarget = m.get('target') or m.get('Target', '')
                    if mtarget != target:
                        continue
                    val_r2 = float(m.get('val', {}).get('r2')
                                   or m.get('val_r2')
                                   or m.get('Val_R2', -999.0) or -999.0)
                    if val_r2 < -5.0:
                        continue
                    # anfis_models/{dataset}/{config}/metrics_*.json
                    dataset_dir_name = mf.parents[1].name
                    feat_cols = _load_meta(dataset_dir_name)
                    if not feat_cols:
                        continue
                    for mp in mf.parent.glob('*.pkl'):
                        records.append({
                            'val_r2':       val_r2,
                            'model_path':   mp,
                            'feature_names': feat_cols,
                            'source':       'ANFIS',
                            'dataset':      dataset_dir_name,
                        })
                except Exception:
                    continue

        # Sirala, top_n sec
        records.sort(key=lambda x: x['val_r2'], reverse=True)
        if self.top_n_models > 0:
            records = records[:self.top_n_models]

        # Model nesnelerini yukle
        loaded = []
        for rec in records:
            mp = rec['model_path']
            model_id = (f"{rec['source']}_{rec['dataset'][-20:]}"
                        f"_{mp.stem}_R2{rec['val_r2']:.3f}")
            if model_id in self._model_cache:
                model_obj = self._model_cache[model_id]
            else:
                try:
                    if mp.suffix in ('.h5', '.keras'):
                        import tensorflow as tf
                        model_obj = tf.keras.models.load_model(str(mp), compile=False)
                    else:
                        model_obj = joblib.load(str(mp))
                    self._model_cache[model_id] = model_obj
                except Exception as e:
                    logger.debug(f"    Skip {mp.name}: {e}")
                    continue
            loaded.append({
                'model_id':     model_id,
                'model':        model_obj,
                'feature_names': rec['feature_names'],
                'source':       rec['source'],
                'val_r2':       rec['val_r2'],
            })

        n_ai    = sum(1 for r in loaded if r['source'] == 'AI')
        n_anfis = sum(1 for r in loaded if r['source'] == 'ANFIS')
        logger.info(f"  [{target}] {len(loaded)} model yuklendi "
                    f"(AI={n_ai}, ANFIS={n_anfis})")
        return loaded

    # ------------------------------------------------------------------
    # Visualizations
    # ------------------------------------------------------------------

    def _generate_visualizations(
        self,
        all_predictions: List[Dict],
        nucleus_labels: List[str],
        summary: Dict,
    ) -> List[str]:
        """
        Tahmin gorselleri uret:
          - Per-target bar chart (her model tahmini)
          - Multi-target consensus karsilastirma
        """
        try:
            import matplotlib
            matplotlib.use('Agg')
            import matplotlib.pyplot as plt
            import matplotlib.patches as mpatches
        except ImportError:
            logger.warning("[VIS] matplotlib bulunamadi, gorsel uretilmedi")
            return []

        plot_paths: List[str] = []
        plots_dir = self.output_dir / 'plots'
        plots_dir.mkdir(exist_ok=True)

        # ----- 1. Her target icin bar chart -----
        for target in self.targets:
            for i, pred in enumerate(all_predictions):
                tdata = pred.get(target, {})
                if 'error' in tdata:
                    continue

                # Model isimleri ve degerler
                model_items = {k: v for k, v in tdata.items()
                               if k not in ('consensus', 'n_models', 'error')
                               and isinstance(v, (int, float))
                               and not np.isnan(v)}
                if not model_items:
                    continue

                fig, ax = plt.subplots(figsize=(max(8, len(model_items) * 0.6 + 2), 5))

                labels_bar = list(model_items.keys())
                values_bar = list(model_items.values())
                consensus  = tdata.get('consensus', float('nan'))

                colors = ['#2196F3'] * len(labels_bar)
                bars = ax.bar(range(len(labels_bar)), values_bar, color=colors,
                              alpha=0.8, edgecolor='white', linewidth=0.5)

                if not np.isnan(consensus):
                    ax.axhline(consensus, color='#F44336', linewidth=2, linestyle='--',
                               label=f'Konsensus = {consensus:.4f}')
                    ax.legend(fontsize=9)

                ax.set_xticks(range(len(labels_bar)))
                ax.set_xticklabels(
                    [lb.split('_')[1][:15] if '_' in lb else lb[:15]
                     for lb in labels_bar],
                    rotation=45, ha='right', fontsize=7
                )
                ax.set_ylabel(target, fontsize=10)
                ax.set_title(
                    f'{nucleus_labels[i]} — {target} Tahminleri '
                    f'({len(model_items)} model)',
                    fontsize=11
                )
                ax.grid(axis='y', alpha=0.3)
                plt.tight_layout()

                fname = plots_dir / f'{nucleus_labels[i]}_{target}_models.png'
                fname_str = str(fname).replace('/', '_').replace('\\', '_')[-40:]
                fig.savefig(fname, dpi=120, bbox_inches='tight')
                plt.close(fig)
                plot_paths.append(str(fname))

        # ----- 2. Multi-target consensus ozet (tum nukleonlar) -----
        if len(all_predictions) > 0:
            try:
                fig, axes = plt.subplots(
                    1, len(self.targets),
                    figsize=(4 * len(self.targets), 5)
                )
                if len(self.targets) == 1:
                    axes = [axes]

                for ax, target in zip(axes, self.targets):
                    consensus_vals = []
                    for pred in all_predictions:
                        c = pred.get(target, {}).get('consensus', float('nan'))
                        if isinstance(c, float) and not np.isnan(c):
                            consensus_vals.append(c)
                        else:
                            consensus_vals.append(float('nan'))

                    x = range(len(nucleus_labels))
                    valid_x = [i for i, v in enumerate(consensus_vals) if not np.isnan(v)]
                    valid_y = [consensus_vals[i] for i in valid_x]

                    ax.bar([nucleus_labels[i][:10] for i in valid_x], valid_y,
                           color='#1976D2', alpha=0.85, edgecolor='white')

                    smry = summary.get(target, {})
                    if smry.get('mean') is not None:
                        ax.axhline(smry['mean'], color='#E53935', ls='--', lw=1.5,
                                   label=f"Ort={smry['mean']:.3f}")
                        ax.legend(fontsize=8)

                    ax.set_title(target, fontsize=10)
                    ax.set_ylabel('Tahmin', fontsize=9)
                    ax.tick_params(axis='x', rotation=45, labelsize=7)
                    ax.grid(axis='y', alpha=0.3)

                plt.suptitle('Consensus Tahminler — Tum Hedefler', fontsize=12)
                plt.tight_layout()

                summary_path = plots_dir / 'consensus_summary.png'
                fig.savefig(summary_path, dpi=120, bbox_inches='tight')
                plt.close(fig)
                plot_paths.append(str(summary_path))

            except Exception as e:
                logger.warning(f"[VIS] Ozet grafik hatasi: {e}")

        logger.info(f"[VIS] {len(plot_paths)} grafik uretildi: {plots_dir}")
        return plot_paths

    # ------------------------------------------------------------------
    # Save outputs
    # ------------------------------------------------------------------

    def _save_excel(self, predictions: List[Dict], labels: List[str]) -> Optional[Path]:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = 'Predictions'

            HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
            HEADER_FONT = Font(bold=True, color='FFFFFF')
            CONSENSUS_FILL = PatternFill('solid', fgColor='E8F5E9')
            CONSENSUS_FONT = Font(bold=True, color='1B5E20')

            cols = ['Nucleus']
            for target in self.targets:
                model_ids: set = set()
                for pred in predictions:
                    model_ids.update(
                        k for k in pred.get(target, {})
                        if k not in ('consensus', 'n_models', 'error')
                    )
                for mid in sorted(model_ids):
                    cols.append(f'{target}|{mid}')
                cols.append(f'{target}|Consensus')
                cols.append(f'{target}|N_Models')

            for ci, col in enumerate(cols, 1):
                cell = ws.cell(row=1, column=ci, value=col)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal='center')

            for ri, pred in enumerate(predictions, 2):
                ws.cell(row=ri, column=1, value=pred['nucleus'])
                ci = 2
                for target in self.targets:
                    tdata = pred.get(target, {})
                    model_ids = sorted(
                        k for k in tdata
                        if k not in ('consensus', 'n_models', 'error')
                    )
                    for mid in model_ids:
                        v = tdata.get(mid)
                        c = ws.cell(row=ri, column=ci, value=v)
                        ci += 1

                    # Consensus sutunu vurgu
                    c_cell = ws.cell(row=ri, column=ci, value=tdata.get('consensus'))
                    c_cell.font = CONSENSUS_FONT
                    c_cell.fill = CONSENSUS_FILL
                    ws.cell(row=ri, column=ci + 1, value=tdata.get('n_models'))
                    ci += 2

            for col in ws.columns:
                max_len = max((len(str(c.value or '')) for c in col), default=8)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

            out = self.output_dir / 'single_nucleus_predictions.xlsx'
            wb.save(str(out))
            return out
        except Exception as e:
            logger.warning(f"[WARN] Excel kayit hatasi: {e}")
            return None

    def _save_json(self, predictions: List[Dict], summary: Dict) -> Optional[Path]:
        out = self.output_dir / 'single_nucleus_predictions.json'
        try:
            payload = {
                'timestamp':   pd.Timestamp.now().isoformat(),
                'n_nuclei':    len(predictions),
                'targets':     self.targets,
                'top_n_models': self.top_n_models,
                'summary':     summary,
                'predictions': predictions,
            }

            def _serial(obj):
                if isinstance(obj, (np.floating, float)):
                    return None if np.isnan(obj) else float(obj)
                if isinstance(obj, np.integer):
                    return int(obj)
                return str(obj)

            out.write_text(
                json.dumps(payload, indent=2, default=_serial),
                encoding='utf-8'
            )
            return out
        except Exception as e:
            logger.warning(f"[WARN] JSON kayit hatasi: {e}")
            return None


# ---------------------------------------------------------------------------
# Dosya parser: pred_input.txt / aaa2.txt formatli dosyalar
# ---------------------------------------------------------------------------

def _parse_nucleus_txt(path: Path) -> pd.DataFrame:
    """
    Bosluk/tab ayrimli tek cekirdek veya liste dosyasini DataFrame'e cevir.

    Desteklenen formatlar:
    ----------------------
    Format A — sadece sayisal sutunlar (baslik satiri yok):
        Z  N  SPIN  PARITY
        26  30  0.0  1

    Format B — baslik satiri ile:
        # Z  N  SPIN  PARITY
        26  30  0.0  1

    Format C — aaa2.txt genisletilmis (ilk 4 sutun kullanilir):
        Z  N  SPIN  PARITY  MM  QM  ...

    Notlar:
    -------
    - '#' ile baslayan satirlar yorum olarak atlanir
    - Bos satirlar atlanir
    - Sutun siralamasi: Z, N, SPIN, PARITY (ilk 4 sutun)
    - A = Z + N otomatik hesaplanir
    - NUCLEUS adi secimli 5. sutun olabilir (metin ise)
    """
    rows = []
    header_found = False
    col_names_from_header: Optional[List[str]] = None

    with open(path, encoding='utf-8', errors='replace') as f:
        for raw_line in f:
            line = raw_line.strip()
            if not line:
                continue

            # Yorum satiri — ama sutun basligi olabilir
            if line.startswith('#'):
                content = line.lstrip('#').strip()
                parts = content.split()
                # Tum parcalar buyuk harf veya bilinen sutun adlari ise baslik say
                possible_cols = ['Z', 'N', 'A', 'SPIN', 'PARITY', 'MM', 'QM',
                                  'BETA_2', 'NUCLEUS']
                if all(p.upper() in possible_cols for p in parts if p.isalpha()):
                    col_names_from_header = [p.upper() for p in parts]
                continue

            parts = line.split()
            if not parts:
                continue

            # Ilk iki parca Z ve N olarak dene
            try:
                z = int(float(parts[0]))
                n = int(float(parts[1]))
            except (ValueError, IndexError):
                continue  # baslik satiri veya bozuk satir

            row: Dict = {'Z': z, 'N': n, 'A': z + n}

            # SPIN
            if len(parts) > 2:
                try:
                    row['SPIN'] = float(parts[2])
                except ValueError:
                    row['SPIN'] = 0.0
            else:
                row['SPIN'] = 0.0

            # PARITY
            if len(parts) > 3:
                try:
                    pval = parts[3].replace('+', '').replace(' ', '')
                    row['PARITY'] = int(float(pval))
                except ValueError:
                    row['PARITY'] = 1
            else:
                row['PARITY'] = 1

            rows.append(row)

    if not rows:
        raise ValueError(f"Dosyadan hic cekirdek ayristirilmadi: {path}")

    logger.info(f"[PARSE] {path.name}: {len(rows)} cekirdek okundu")
    return pd.DataFrame(rows)
