# -*- coding: utf-8 -*-
"""
PFAZ 4: Generalization Analyzer
================================

Deep analysis of model generalization capabilities

Features:
- Generalization score calculation
- Model ranking
- Statistical analysis
- Excel reports

Author: Nuclear Physics AI Training Pipeline
Version: 1.0.0
Date: 2025-10-15
"""

import numpy as np
import pandas as pd
from pathlib import Path
import logging
from typing import Dict, List
import warnings
warnings.filterwarnings('ignore')

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class GeneralizationAnalyzer:
    """
    Generalization Analyzer
    
    Features:
    - Generalization score calculation
    - Model ranking by generalization
    - Excel reports
    """
    
    def __init__(self, 
                 known_results_file: str = None,
                 unknown_results_file: str = None,
                 output_dir: str = 'generalization_analysis'):
        
        self.known_results_file = Path(known_results_file) if known_results_file else None
        self.unknown_results_file = Path(unknown_results_file) if unknown_results_file else None
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        self.analysis_results = []
        
        logger.info("=" * 80)
        logger.info("GENERALIZATION ANALYZER INITIALIZED")
        logger.info("=" * 80)
    
    def calculate_generalization_scores(self, 
                                        known_df: pd.DataFrame,
                                        unknown_df: pd.DataFrame) -> pd.DataFrame:
        """
        Calculate generalization scores
        
        Generalization Score (GS) = (Unknown_R2 / Known_R2) * 100
        """
        
        logger.info("\nCalculating generalization scores...")
        
        results = []
        
        # Merge on model_id
        for _, known_row in known_df.iterrows():
            model_id = known_row.get('model_id')
            
            unknown_row = unknown_df[unknown_df['model_id'] == model_id]
            
            if unknown_row.empty:
                continue
            
            unknown_row = unknown_row.iloc[0]
            
            # Calculate generalization score
            known_r2 = known_row['r2']
            unknown_r2 = unknown_row['r2']
            
            if known_r2 > 0:
                gs = (unknown_r2 / known_r2) * 100
            else:
                gs = 0
            
            # Classify
            if gs >= 90:
                category = 'Excellent'
            elif gs >= 70:
                category = 'Good'
            elif gs >= 50:
                category = 'Moderate'
            else:
                category = 'Poor'
            
            results.append({
                'model_id': model_id,
                'model_type': known_row.get('type', 'Unknown'),
                'known_r2': known_r2,
                'unknown_r2': unknown_r2,
                'r2_drop': known_r2 - unknown_r2,
                'r2_drop_pct': ((known_r2 - unknown_r2) / known_r2 * 100) if known_r2 > 0 else 0,
                'generalization_score': gs,
                'category': category,
                'known_rmse': known_row['rmse'],
                'unknown_rmse': unknown_row['rmse'],
                'rmse_increase': unknown_row['rmse'] - known_row['rmse']
            })
        
        results_df = pd.DataFrame(results)
        results_df = results_df.sort_values('generalization_score', ascending=False)
        
        logger.info(f"[SUCCESS] Calculated generalization for {len(results_df)} models")
        
        self.analysis_results = results_df
        
        return results_df
    
    def get_best_generalizers(self, top_n: int = 10) -> pd.DataFrame:
        """Get top N models by generalization"""
        
        if self.analysis_results is None or len(self.analysis_results) == 0:
            return pd.DataFrame()
        
        return self.analysis_results.head(top_n)
    
    def generate_excel_report(self, filename: str = 'Generalization_Analysis.xlsx'):
        """Generate Excel report"""
        
        if self.analysis_results is None or len(self.analysis_results) == 0:
            logger.warning("No analysis results")
            return
        
        excel_path = self.output_dir / filename
        logger.info(f"Generating Excel: {excel_path}")
        
        if OPENPYXL_AVAILABLE:
            wb = Workbook()
            wb.remove(wb.active)
            
            # Sheet 1: Full Analysis
            ws1 = wb.create_sheet('Generalization_Analysis')
            self._write_dataframe(ws1, self.analysis_results)
            
            # Sheet 2: Summary by Category
            ws2 = wb.create_sheet('Summary_by_Category')
            summary = self.analysis_results.groupby('category').agg({
                'generalization_score': ['count', 'mean', 'std'],
                'r2_drop_pct': 'mean'
            }).round(2)
            self._write_dataframe(ws2, summary.reset_index())
            
            # Sheet 3: Top 10
            ws3 = wb.create_sheet('Top_10_Generalizers')
            top10 = self.get_best_generalizers(10)
            self._write_dataframe(ws3, top10)
            
            wb.save(excel_path)
        else:
            self.analysis_results.to_csv(self.output_dir / 'generalization_analysis.csv', index=False)
        
        logger.info(f"[SUCCESS] Report saved: {excel_path}")
    
    def _write_dataframe(self, ws, df: pd.DataFrame):
        """Write dataframe to worksheet"""
        if df.empty:
            return
        
        # Headers
        headers = df.columns.tolist()
        ws.append(headers)
        
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for _, row in df.iterrows():
            ws.append(row.tolist())
        
        # Auto-adjust
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width


def main():
    """Test"""
    print("\n" + "=" * 80)
    print("PFAZ 4: GENERALIZATION ANALYZER - TEST")
    print("=" * 80)
    
    # Create mock data
    known_df = pd.DataFrame({
        'model_id': ['model1', 'model2', 'model3'],
        'type': ['AI', 'AI', 'ANFIS'],
        'r2': [0.95, 0.90, 0.85],
        'rmse': [0.10, 0.15, 0.20],
        'mae': [0.08, 0.12, 0.16]
    })
    
    unknown_df = pd.DataFrame({
        'model_id': ['model1', 'model2', 'model3'],
        'type': ['AI', 'AI', 'ANFIS'],
        'r2': [0.88, 0.75, 0.80],
        'rmse': [0.15, 0.22, 0.25],
        'mae': [0.12, 0.18, 0.20]
    })
    
    analyzer = GeneralizationAnalyzer(output_dir='test_generalization')
    
    print("\nCalculating generalization scores...")
    results = analyzer.calculate_generalization_scores(known_df, unknown_df)
    print(results)
    
    print("\nGenerating Excel report...")
    analyzer.generate_excel_report()
    
    print("\n[SUCCESS] TEST COMPLETED!")


if __name__ == "__main__":
    main()
