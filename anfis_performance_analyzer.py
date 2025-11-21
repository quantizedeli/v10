# -*- coding: utf-8 -*-
"""
PFAZ 3: ANFIS Performance Analyzer
===================================

Comprehensive ANFIS performance analysis and config comparison

Features:
- 8 ANFIS configs comparison
- Convergence analysis
- Training time analysis
- Best config identification
- Excel comprehensive reports

Author: Nuclear Physics AI Training Pipeline
Version: 1.0.0
Date: 2025-10-15
"""

import numpy as np
import pandas as pd
from pathlib import Path
import json
import logging
from typing import Dict, List
from collections import defaultdict
import warnings
warnings.filterwarnings('ignore')

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ANFISPerformanceAnalyzer:
    """
    ANFIS Performance Analyzer
    
    Features:
    - Config performance comparison
    - Convergence analysis
    - Training time analysis
    - Excel reports
    """
    
    def __init__(self, trained_models_dir: str, output_dir: str = 'anfis_performance_analysis'):
        self.trained_models_dir = Path(trained_models_dir)
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        self.config_stats = defaultdict(lambda: {
            'count': 0,
            'r2_scores': [],
            'rmse_scores': [],
            'mae_scores': [],
            'training_times': [],
            'convergence_epochs': []
        })
        
        logger.info("=" * 80)
        logger.info("ANFIS PERFORMANCE ANALYZER INITIALIZED")
        logger.info("=" * 80)
    
    def load_training_results(self):
        """Load all ANFIS training results"""
        
        metrics_files = list(self.trained_models_dir.glob('**/metrics_*.json'))
        
        logger.info(f"Found {len(metrics_files)} ANFIS result files")
        
        for metrics_file in metrics_files:
            try:
                with open(metrics_file, 'r') as f:
                    metrics = json.load(f)
                
                config_id = metrics_file.stem.replace('metrics_', '')
                
                # Extract metrics
                r2 = metrics.get('val_r2', 0)
                rmse = metrics.get('val_rmse', 0)
                mae = metrics.get('val_mae', 0)
                training_time = metrics.get('training_time', 0)
                epochs = metrics.get('epochs_trained', 0)
                
                # Update stats
                self.config_stats[config_id]['count'] += 1
                self.config_stats[config_id]['r2_scores'].append(r2)
                self.config_stats[config_id]['rmse_scores'].append(rmse)
                self.config_stats[config_id]['mae_scores'].append(mae)
                self.config_stats[config_id]['training_times'].append(training_time)
                self.config_stats[config_id]['convergence_epochs'].append(epochs)
            
            except Exception as e:
                logger.warning(f"Failed to load {metrics_file}: {e}")
        
        logger.info(f"Loaded results for {len(self.config_stats)} configs")
    
    def get_config_summary(self) -> pd.DataFrame:
        """Get summary statistics per config"""
        
        data = []
        
        for config_id, stats in self.config_stats.items():
            if stats['count'] == 0:
                continue
            
            row = {
                'Config_ID': config_id,
                'Count': stats['count'],
                'Avg_R2': np.mean(stats['r2_scores']),
                'Std_R2': np.std(stats['r2_scores']),
                'Best_R2': np.max(stats['r2_scores']),
                'Worst_R2': np.min(stats['r2_scores']),
                'Avg_RMSE': np.mean(stats['rmse_scores']),
                'Std_RMSE': np.std(stats['rmse_scores']),
                'Avg_MAE': np.mean(stats['mae_scores']),
                'Avg_Training_Time': np.mean(stats['training_times']),
                'Avg_Epochs': np.mean(stats['convergence_epochs'])
            }
            
            data.append(row)
        
        df = pd.DataFrame(data)
        
        if not df.empty:
            df = df.sort_values('Avg_R2', ascending=False)
        
        return df
    
    def generate_excel_report(self, filename: str = 'ANFIS_Performance_Analysis.xlsx'):
        """Generate comprehensive Excel report"""
        
        excel_path = self.output_dir / filename
        
        logger.info(f"Generating Excel report: {excel_path}")
        
        # Get summary
        summary_df = self.get_config_summary()
        
        if OPENPYXL_AVAILABLE:
            wb = Workbook()
            wb.remove(wb.active)
            
            # Sheet 1: Config Summary
            ws = wb.create_sheet('Config_Summary')
            
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            headers = summary_df.columns.tolist()
            ws.append(headers)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            for _, row in summary_df.iterrows():
                ws.append(row.tolist())
            
            # Highlight best config
            if len(summary_df) > 0:
                best_fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(2, col_idx).fill = best_fill
            
            # Auto-adjust columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 25)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(excel_path)
        else:
            # Fallback to simple Excel
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                summary_df.to_excel(writer, sheet_name='Config_Summary', index=False)
        
        logger.info(f"✅ Excel report saved: {excel_path}")


def main():
    """Test execution"""
    
    print("\n" + "=" * 80)
    print("PFAZ 3: ANFIS PERFORMANCE ANALYZER - TEST")
    print("=" * 80)
    
    analyzer = ANFISPerformanceAnalyzer(
        trained_models_dir='test_trained_anfis',
        output_dir='test_anfis_analysis'
    )
    
    print("\nLoading training results...")
    analyzer.load_training_results()
    
    print("\nGenerating summary...")
    summary = analyzer.get_config_summary()
    print(summary)
    
    print("\nGenerating Excel report...")
    analyzer.generate_excel_report()
    
    print("\n✅ TEST COMPLETED!")


if __name__ == "__main__":
    main()
